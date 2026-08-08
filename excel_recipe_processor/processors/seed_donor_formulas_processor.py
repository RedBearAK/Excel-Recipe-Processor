"""
Excel formula transplant processor for seeding formulas from donor files.

This processor samples formulas from designated cells in a source Excel file
and transplants them to corresponding locations in a target Excel file,
creating a starter culture for formula proliferation.

excel_recipe_processor/processors/seed_donor_formulas_processor.py
"""

import logging
import openpyxl
import pandas as pd

from pathlib import Path

from excel_recipe_processor.core.base_processor import FileOpsBaseProcessor, StepProcessorError
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.formula import ArrayFormula

from excel_recipe_processor.processors._helpers.range_patterns   import excel_column_ref_rgx
from excel_recipe_processor.processors._helpers.excel_range_resolver import (
    find_last_data_row, ExcelRangeResolverError
)


logger = logging.getLogger(__name__)


class SeedDonorFormulasProcessor(FileOpsBaseProcessor):
    """
    Transplant formulas from donor file to recipient file.
    
    Samples formulas from specific cells in source Excel file and plants them
    in corresponding cells in target Excel file. Designed for seeding formula
    columns in newly created files with existing data but empty formula columns.
    """
    
    def __init__(self, step_config: dict):
        """Initialize the formula transplant processor."""
        super().__init__(step_config)
        
        # Extract configuration
        self.source_file = self.get_config_value('source_file')
        self.source_sheet = self.get_config_value('source_sheet') 
        self.target_file = self.get_config_value('target_file')
        self.target_sheet = self.get_config_value('target_sheet')
        self.columns = self.get_config_value('columns', [])
        self.start_row = self.get_config_value('start_row', 2)  # 1-indexed
        self.row_count = self.get_config_value('row_count', 3)
        self.force_column_names = self.get_config_value('force_column_names', False)
        self.fill_down = self.get_config_value('fill_down', False)
        self.fill_anchor_columns = self.get_config_value('fill_anchor_columns', None)

        # What to do with a donor cell holding an array (CSE) formula.
        #
        #   "preserve"  write it back as an array formula, faithful to the donor
        #   "convert"   write it as an ordinary formula
        #
        # Converting is safe when the formula is scalar - nothing but single-cell
        # comparisons and ordinary functions - which is the common case for a
        # formula that was entered with Ctrl-Shift-Enter out of habit rather than
        # necessity. Ordinary formulas are also what Excel fills down natively,
        # so they behave predictably for anyone editing afterwards.
        self.array_formula_mode = self.get_config_value('array_formula_mode', 'preserve')

        if self.array_formula_mode not in ('preserve', 'convert'):
            raise StepProcessorError(
                f"Invalid array_formula_mode '{self.array_formula_mode}'. "
                f"Supported: preserve, convert"
            )
        
        # Validate configuration
        self._validate_config()
    
    def _validate_config(self):
        """Validate step configuration."""
        if not self.source_file:
            raise StepProcessorError("source_file is required")
        
        if not self.source_sheet:
            raise StepProcessorError("source_sheet is required")
            
        if not self.target_file:
            raise StepProcessorError("target_file is required")
            
        if not self.target_sheet:
            raise StepProcessorError("target_sheet is required")
            
        if not self.columns or len(self.columns) == 0:
            raise StepProcessorError("columns list cannot be empty")
            
        if self.row_count > 10:
            raise StepProcessorError("row_count cannot exceed 10 (performance limitation)")
            
        if self.start_row < 1:
            raise StepProcessorError("start_row must be 1 or greater (1-indexed)")
    
    def perform_file_operation(self):
        """Execute the formula transplant operation."""
        logger.info(f"🧬 Transplanting formulas from {self.source_file} to {self.target_file}")
        
        # 1. Validate files and sheets exist
        source_wb, source_ws = self._load_workbook_and_sheet(self.source_file, self.source_sheet, "source", read_only=True)
        target_wb, target_ws = self._load_workbook_and_sheet(self.target_file, self.target_sheet, "target", read_only=False)
        
        try:
            # 2. Resolve column specifications to Excel column letters
            resolved_columns = self._resolve_column_specs(source_ws, target_ws)
            logger.info(f"📍 Processing columns: {resolved_columns}")
            
            # 3. Extract and transplant formulas
            transplanted_count = 0
            empty_source_count = 0
            
            for col_letter in resolved_columns:
                for row_offset in range(self.row_count):
                    current_row = self.start_row + row_offset
                    
                    # Extract formula from source
                    source_cell = source_ws[f"{col_letter}{current_row}"]
                    target_cell = target_ws[f"{col_letter}{current_row}"]
                    
                    # Check if source has a formula
                    if source_cell.value is None:
                        logger.debug(f"📍 Source cell {col_letter}{current_row} is empty")
                        empty_source_count += 1
                        continue
                    
                    # Check if target cell is occupied
                    if target_cell.value is not None:
                        raise StepProcessorError(
                            f"Target cell {col_letter}{current_row} already contains data: '{target_cell.value}'. "
                            f"Cannot overwrite existing data."
                        )
                    
                    # Transplant the formula (openpyxl handles making it "live")
                    source_formula = source_cell.value

                    # Honour array_formula_mode on the seeded rows too, so the
                    # whole column is consistent rather than array at the top and
                    # ordinary below it
                    if (isinstance(source_formula, ArrayFormula)
                            and self.array_formula_mode == 'convert'):
                        source_formula = source_formula.text

                    target_cell.value = source_formula
                    
                    # If source had a formula, copy it exactly
                    if hasattr(source_cell, 'formula') and source_cell.formula:
                        target_cell.formula = source_cell.formula
                    
                    logger.debug(f"🧬 Transplanted to {col_letter}{current_row}: {source_cell.value}")
                    transplanted_count += 1
            
            # 4. Optionally continue the seeded formulas to the end of the data
            filled_count = 0
            if self.fill_down:
                filled_count = self._fill_down_columns(target_ws, resolved_columns)

            # 5. Save target workbook (makes formulas "live")
            target_wb.save(self.target_file)

            if transplanted_count == 0:
                raise StepProcessorError(
                    f"No formulas found in donor '{self.source_file}' sheet "
                    f"'{self.source_sheet}' at rows {self.start_row}-"
                    f"{self.start_row + self.row_count - 1} for columns "
                    f"{self.columns}. The donor rows may have been deleted, or "
                    f"start_row may not point at them. Writing a formula-free "
                    f"file silently is worse than stopping here."
                )

            # Log summary
            logger.info(f"✅ Transplanted {transplanted_count} formulas successfully")
            if filled_count:
                logger.info(f"⬇️  Filled {filled_count:,} additional formula cells")
            if empty_source_count > 0:
                logger.warning(f"⚠️ Found {empty_source_count} empty source cells - check column specifications")
            
            # Return description of what was accomplished
            if filled_count:
                return (f"transplanted {transplanted_count} formulas and filled "
                        f"{filled_count:,} more from {self.source_file} to {self.target_file}")

            return f"transplanted {transplanted_count} formulas from {self.source_file} to {self.target_file}"
            
        finally:
            # Clean up workbook resources
            source_wb.close()
            target_wb.close()
    
    def _fill_down_columns(self, target_ws, resolved_columns: list) -> int:
        """
        Continue each seeded formula to the last populated row.

        Uses openpyxl's Translator, which shifts relative cell references the way
        Excel's own fill-down does while leaving named ranges and absolute
        references alone. It manages roughly 100,000 cells a second, so the cost
        is negligible next to loading and saving the workbook.

        Args:
            target_ws:          Worksheet being written
            resolved_columns:   Column letters holding the seeded formulas

        Returns:
            Count of cells filled
        """
        anchors = self.fill_anchor_columns

        if not anchors:
            # Measure the data extent from every column, so a sparse column
            # cannot cut the fill short
            from openpyxl.utils import get_column_letter
            anchors = [get_column_letter(i) for i in range(1, target_ws.max_column + 1)]

        try:
            last_row = find_last_data_row(target_ws, anchors, header_row=1)
        except ExcelRangeResolverError as error:
            raise StepProcessorError(f"Could not determine fill extent: {error}")

        # Where the seeded block was ASKED to end. The donor may hold fewer rows
        # than row_count, in which case some of those cells were never written -
        # so filling from here would step over them and leave a gap. The real
        # starting point is found per column below.
        requested_last_row = self.start_row + self.row_count - 1

        if last_row <= requested_last_row:
            logger.info(f"⬇️  Nothing to fill: data ends at row {last_row}")
            return 0

        filled = 0
        columns_with_formulas = []

        for col_letter in resolved_columns:
            origin = f"{col_letter}{self.start_row}"
            source_value = target_ws[origin].value

            # An array formula arrives as an ArrayFormula object rather than a
            # string, so a plain startswith('=') test skips it. SALE TYPE1 in the
            # VMS workbook is one of these, and skipping it left that column
            # empty below the seeded rows while the log still claimed success.
            is_array = isinstance(source_value, ArrayFormula)

            if is_array:
                formula_text = source_value.text
            elif isinstance(source_value, str) and source_value.startswith('='):
                formula_text = source_value
            else:
                logger.debug(f"⬇️  {origin} holds no formula, not filling this column")
                continue

            columns_with_formulas.append(col_letter)
            translator = Translator(formula_text, origin=origin)
            column_number = target_ws[f"{col_letter}1"].column

            # Fill from the last row that actually received a formula, not from
            # the last row that was requested. A donor with fewer rows than
            # row_count would otherwise leave the difference blank.
            seed_last_row = self.start_row - 1
            for probe_row in range(self.start_row, requested_last_row + 1):
                if target_ws.cell(row=probe_row, column=column_number).value is not None:
                    seed_last_row = probe_row

            if seed_last_row < requested_last_row:
                logger.debug(
                    f"⬇️  {col_letter}: donor supplied {seed_last_row - self.start_row + 1} "
                    f"of {self.row_count} requested rows, filling from "
                    f"{seed_last_row + 1}"
                )

            for row_num in range(seed_last_row + 1, last_row + 1):
                target_reference = f"{col_letter}{row_num}"
                translated = translator.translate_formula(target_reference)

                if is_array and self.array_formula_mode == 'preserve':
                    # Each filled cell needs its own single-cell ref, or Excel
                    # treats them as one spilled block anchored at the origin
                    translated = ArrayFormula(ref=target_reference, text=translated)

                target_ws.cell(row=row_num, column=column_number, value=translated)
                filled += 1

        if filled:
            logger.info(
                f"⬇️  Filled {len(columns_with_formulas)} column(s) from row "
                f"{seed_last_row + 1} to {last_row}"
            )
        else:
            logger.warning(
                "⬇️  Nothing filled: no seeded formulas were found to continue"
            )

        return filled

    def _load_workbook_and_sheet(self, file_path: str, sheet_name: str, context: str, read_only: bool = False):
        """Load workbook and get specified worksheet."""
        # Check file exists
        if not Path(file_path).exists():
            raise StepProcessorError(f"{context.title()} file not found: {file_path}")
        
        try:
            # Load workbook (read_only=True for source files for performance)
            workbook = openpyxl.load_workbook(file_path, read_only=read_only)
            
            # Get worksheet
            if sheet_name not in workbook.sheetnames:
                available_sheets = workbook.sheetnames
                workbook.close()
                raise StepProcessorError(
                    f"{context.title()} sheet '{sheet_name}' not found in {file_path}. "
                    f"Available sheets: {available_sheets}"
                )
            
            worksheet = workbook[sheet_name]
            return workbook, worksheet
            
        except Exception as e:
            if "not found" in str(e) or "Available sheets" in str(e):
                raise  # Re-raise our custom errors
            raise StepProcessorError(f"Error loading {context} file {file_path}: {e}")
    
    def _resolve_column_specs(self, source_ws, target_ws):
        """Resolve column specifications to Excel column letters."""
        resolved_columns = []
        
        for col_spec in self.columns:
            col_spec = str(col_spec).strip()
            
            # Check if it's an Excel column reference (unless forced to treat as name)
            if not self.force_column_names and excel_column_ref_rgx.match(col_spec):
                resolved_columns.append(col_spec)
                logger.debug(f"📋 Column '{col_spec}' treated as Excel reference")
            else:
                # Treat as column name - find in headers
                source_col = self._find_column_by_name(source_ws, col_spec)
                target_col = self._find_column_by_name(target_ws, col_spec)
                
                if source_col and target_col:
                    if source_col == target_col:
                        resolved_columns.append(source_col)
                        logger.debug(f"📋 Column name '{col_spec}' resolved to {source_col}")
                    else:
                        logger.warning(
                            f"⚠️ Column '{col_spec}' found at different positions: "
                            f"source={source_col}, target={target_col}. Using source position."
                        )
                        resolved_columns.append(source_col)
                else:
                    logger.warning(f"⚠️ Column name '{col_spec}' not found, skipping")
        
        if not resolved_columns:
            raise StepProcessorError("No valid columns found after resolution")
        
        return resolved_columns
    
    def _find_column_by_name(self, worksheet, column_name: str):
        """Find Excel column letter by searching for column name in first row."""
        # Search in first row (assuming headers)
        for col_num in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col_num).value
            if cell_value and str(cell_value).strip() == column_name:
                return openpyxl.utils.get_column_letter(col_num)
        
        return None
    
    @classmethod
    def get_minimal_config(cls) -> dict:
        """Get minimal configuration for testing."""
        return {
            'source_file': 'templates/formulas.xlsx',
            'source_sheet': 'Sheet1',
            'target_file': 'output/new_file.xlsx', 
            'target_sheet': 'Sheet1',
            'columns': ['C', 'D'],
            'start_row': 2,
            'row_count': 3
        }
    
    def get_usage_examples(self) -> dict:
        """Get usage examples for this processor."""
        from excel_recipe_processor.utils.processor_examples_loader import load_processor_examples
        return load_processor_examples('seed_donor_formulas')
    
    def get_capabilities(self) -> dict:
        """Get processor capabilities and features."""
        return {
            'description': 'Transplant formulas from donor Excel files to recipient files for seeding calculation columns',
            'operation_type': 'file_operations',
            'column_matching': [
                'excel_column_references',  # A, B, AA, etc.
                'column_name_matching',     # Header text matching
                'mixed_column_types',       # Combination of refs and names
                'auto_type_detection'       # Automatic ref vs name detection
            ],
            'formula_handling': [
                'verbatim_copying',         # Exact formula preservation
                'live_formula_creation',    # Creates working Excel formulas
                'empty_cell_detection',     # Logs when source cells are empty
                'collision_prevention'      # Errors if target cells occupied
            ],
            'file_requirements': [
                'source_file_must_exist',
                'target_file_must_exist',
                'sheet_validation',
                'openpyxl_dependency'
            ],
            'performance_features': [
                'row_count_limiting',       # Max 10 rows for performance
                'efficient_cell_access',
                'resource_cleanup'
            ],
            'integration_features': [
                'fileops_inheritance',      # FileOpsBaseProcessor
                'stage_bypass',             # No stage validation
                'workflow_integration',     # Works with export_file
                'error_handling'            # Comprehensive validation
            ],
            'supported_formats': ['xlsx'],  # Only Excel files supported
            'max_rows_per_operation': 10,   # Performance limitation
            'examples': {
                'template_seeding': 'Copy calculation formulas from budget template to new budget files',
                'model_deployment': 'Deploy formula logic across multiple similar Excel files',
                'starter_cultures': 'Seed empty formula columns with working calculation examples'
            }
        }


# End of file #

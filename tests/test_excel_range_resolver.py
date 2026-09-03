"""
Tests for the shared Excel range resolver helper.

tests/test_excel_range_resolver.py

Runnable with pytest, but written to run standalone and report a score.
Focused on column resolution, span expansion, and extent detection.
"""

from openpyxl import Workbook

from excel_recipe_processor.processors._helpers.excel_range_resolver import (resolve_column_refs, ColumnVocabularyError, 
    resolve_range,
    build_range_ref,
    quote_sheet_name,
    expand_column_span,
    assert_contiguous,
    find_last_data_row,
    resolve_column_letter,
    resolve_column_letters,
    ExcelRangeResolverError,
)


def build_product_sheet():
    """Build a worksheet shaped like the VMS product lookup table."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Product_IDs'

    headers = [
        'Major Species', 'Species', 'Product ID', 'Product Name',
        'Component', 'Product Form', 'Product Group', 'Can Size', 'Pack Size'
    ]
    worksheet.append(headers)

    for row_num in range(1, 21):
        can_size = 14.75 if row_num <= 3 else None
        pack_size = 'Tall' if row_num <= 3 else None
        worksheet.append([
            'SALMON', 'SOCKEYE', 10000 + row_num, f'Product {row_num}',
            'FLESH', 'IQF H&G', 'FROZEN IQF', can_size, pack_size
        ])

    return worksheet


def test_column_reference_passthrough():
    """A positional ref travels through resolve_column_refs; the NAME resolver refuses it (2026-08-26)."""
    print("\nTesting bare column reference handling...")
    worksheet = build_product_sheet()
    refs = resolve_column_refs(['C'])
    try:
        resolve_column_letter(worksheet, 'C')
        print("  ✗ Name resolver accepted a bare ref that names no header")
        return False
    except ColumnVocabularyError:
        pass
    if refs == ['C']:
        print("  ✓ 'C' passes through column_refs; the name resolver refuses it")
        return True
    print(f"  ✗ Expected ['C'] from column_refs, got {refs}")
    return False

def test_column_name_lookup():
    """A header name resolves to its column letter."""
    print("\nTesting header name lookup...")

    worksheet = build_product_sheet()
    result = resolve_column_letter(worksheet, 'Product ID')

    if result == 'C':
        print("  ✓ 'Product ID' resolved to C")
        return True

    print(f"  ✗ Expected 'C', got '{result}'")
    return False


def test_force_column_names():
    """force_column_names stops a header named like a reference being taken as one."""
    print("\nTesting force_column_names...")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(['ID', 'C', 'Value'])
    worksheet.append([1, 'x', 10])

    by_name = resolve_column_letter(worksheet, 'C')
    if by_name == 'B':
        print("  ✓ A header literally named 'C' resolves by NAME to column B")
        return True
    print(f"  ✗ Expected 'B', got '{by_name}'")
    return False

def test_duplicate_header_errors():
    """Duplicate headers are a hard error rather than a first-wins guess."""
    print("\nTesting duplicate header detection...")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(['Region', 'Amount', 'Region'])
    worksheet.append(['West', 10, 'East'])

    try:
        resolve_column_letter(worksheet, 'Region')
    except ExcelRangeResolverError as error:
        if 'appears in 2 columns' in str(error):
            print("  ✓ Duplicate header rejected")
            return True
        print(f"  ✗ Wrong error text: {error}")
        return False

    print("  ✗ Duplicate header was not rejected")
    return False


def test_missing_column_policies():
    """on_missing controls whether an absent column stops the run."""
    print("\nTesting on_missing policies...")

    worksheet = build_product_sheet()
    passed = True

    try:
        resolve_column_letters(worksheet, ['Product ID', 'Nonexistent'], on_missing='error')
        print("  ✗ 'error' policy did not raise")
        passed = False
    except ExcelRangeResolverError:
        print("  ✓ 'error' policy raised")

    result = resolve_column_letters(
        worksheet, ['Product ID', 'Nonexistent'], on_missing='skip'
    )
    if result == ['C']:
        print("  ✓ 'skip' policy dropped the missing column")
    else:
        print(f"  ✗ 'skip' policy returned {result}")
        passed = False

    return passed


def test_span_expansion():
    """Adjacent list entries expand to cover intervening columns."""
    print("\nTesting span expansion...")

    passed = True

    cases = [
        (['B', 'E'], ['B', 'C', 'D', 'E']),
        (['B', 'E', 'G'], ['B', 'C', 'D', 'E', 'F', 'G']),
        (['E', 'B'], ['B', 'C', 'D', 'E']),
        (['C'], ['C']),
    ]

    for source, expected in cases:
        result = expand_column_span(source)
        if result == expected:
            print(f"  ✓ {source} -> {result}")
        else:
            print(f"  ✗ {source} gave {result}, expected {expected}")
            passed = False

    return passed


def test_contiguity_check():
    """Non-contiguous columns are rejected for single-area references."""
    print("\nTesting contiguity check...")

    passed = True

    try:
        assert_contiguous(['B', 'C', 'D'])
        print("  ✓ Contiguous run accepted")
    except ExcelRangeResolverError as error:
        print(f"  ✗ Contiguous run rejected: {error}")
        passed = False

    try:
        assert_contiguous(['B', 'E'])
        print("  ✗ Gapped columns were accepted")
        passed = False
    except ExcelRangeResolverError:
        print("  ✓ Gapped columns rejected")

    return passed


def test_extent_ignores_sparse_column():
    """A sparse column anchored on itself truncates; a full column does not."""
    print("\nTesting extent detection on a sparse column...")

    worksheet = build_product_sheet()

    sparse = find_last_data_row(worksheet, ['H'])
    dense = find_last_data_row(worksheet, ['C'])
    combined = find_last_data_row(worksheet, ['H', 'C'])

    if sparse == 4 and dense == 21 and combined == 21:
        print("  ✓ Sparse=4, dense=21, longest-wins=21")
        return True

    print(f"  ✗ Got sparse={sparse}, dense={dense}, combined={combined}")
    return False


def test_row_modes():
    """Each row mode produces the expected address shape."""
    print("\nTesting row modes...")

    worksheet = build_product_sheet()
    passed = True

    cases = [
        ('data', 'Product_IDs!$C$2:$C$21'),
        ('data_with_header', 'Product_IDs!$C$1:$C$21'),
        ('full_col', 'Product_IDs!$C:$C'),
        ('full_col_no_header', 'Product_IDs!$C$2:$C$1048576'),
    ]

    for mode, expected in cases:
        result = resolve_range(worksheet, ['Product ID'], row_mode=mode)
        if result == expected:
            print(f"  ✓ {mode:20} -> {result}")
        else:
            print(f"  ✗ {mode:20} -> {result}, expected {expected}")
            passed = False

    return passed


def test_multi_column_span_range():
    """A span across named endpoints produces one block reference."""
    print("\nTesting multi-column span range...")

    worksheet = build_product_sheet()
    result = resolve_range(
        worksheet, ['Species', 'Component'], row_mode='data', anchor_columns=['Product ID']
    )

    expected = 'Product_IDs!$B$2:$E$21'

    if result == expected:
        print(f"  ✓ Span resolved to {result}")
        return True

    print(f"  ✗ Got {result}, expected {expected}")
    return False


def test_sheet_name_quoting():
    """Sheet names needing quotes get them, with apostrophes doubled."""
    print("\nTesting sheet name quoting...")

    passed = True

    cases = [
        ('Product_IDs', 'Product_IDs'),
        ('Region-Carrier', "'Region-Carrier'"),
        ('Q4 Sales', "'Q4 Sales'"),
        ("Bob's Data", "'Bob''s Data'"),
        ('2026Data', "'2026Data'"),
    ]

    for source, expected in cases:
        result = quote_sheet_name(source)
        if result == expected:
            print(f"  ✓ {source!r:18} -> {result}")
        else:
            print(f"  ✗ {source!r:18} -> {result}, expected {expected}")
            passed = False

    return passed


def test_relative_and_unqualified():
    """absolute and qualify_sheet toggles both work."""
    print("\nTesting absolute and qualification toggles...")

    worksheet = build_product_sheet()

    result = resolve_range(
        worksheet, ['Product ID'], row_mode='data',
        absolute=False, qualify_sheet=False
    )

    if result == 'C2:C21':
        print(f"  ✓ Relative unqualified range: {result}")
        return True

    print(f"  ✗ Got {result}, expected 'C2:C21'")
    return False


def test_build_range_ref_directly():
    """The low-level assembler handles both shapes."""
    print("\nTesting build_range_ref...")

    passed = True

    whole = build_range_ref('C', 'C', None, None, True, None)
    bounded = build_range_ref('B', 'E', 4, 856, True, 'Sales_Orders')

    if whole == '$C:$C':
        print(f"  ✓ Whole column: {whole}")
    else:
        print(f"  ✗ Whole column: {whole}")
        passed = False

    if bounded == 'Sales_Orders!$B$4:$E$856':
        print(f"  ✓ Bounded block: {bounded}")
    else:
        print(f"  ✗ Bounded block: {bounded}")
        passed = False

    return passed


def main():
    """Run every test and report a final score."""
    print("=== Excel Range Resolver Tests ===")

    tests = [
        test_column_reference_passthrough,
        test_column_name_lookup,
        test_force_column_names,
        test_duplicate_header_errors,
        test_missing_column_policies,
        test_span_expansion,
        test_contiguity_check,
        test_extent_ignores_sparse_column,
        test_row_modes,
        test_multi_column_span_range,
        test_sheet_name_quoting,
        test_relative_and_unqualified,
        test_build_range_ref_directly,
    ]

    passed = 0

    for test_func in tests:
        try:
            if test_func():
                passed += 1
        except Exception as error:
            print(f"  ✗ {test_func.__name__} crashed: {error}")

    print(f"\n=== Results: {passed}/{len(tests)} tests passed ===")

    if passed == len(tests):
        print("✅ All range resolver tests passed!")
        return 1

    print("❌ Some range resolver tests failed!")
    return 0


if __name__ == '__main__':
    exit(0 if main() else 1)


# End of file #

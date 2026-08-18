"""
Tests for save-time inline-string consolidation.

tests/test_inline_string_consolidation.py

openpyxl 3.1+ writes literal strings inline; the session save rewrites
them to Excel's shared-string dialect. These tests pin the doctrine:
verbatim text preservation (escapes, xml:space), dedupe correctness,
existing-table merge with index stability, rich-text refusal,
registration of the part, idempotence, and reload identity.

Runnable standalone or under pytest; the exit code carries the verdict.
"""

import io
import sys
import zipfile

import openpyxl

from excel_recipe_processor.core.inline_string_consolidation import (
    consolidate_inline_strings,
)


def workbook_bytes(rows):
    """An openpyxl-authored package (inline-string dialect) as bytes."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def sheet_xml(package_bytes, part='xl/worksheets/sheet1.xml'):
    with zipfile.ZipFile(io.BytesIO(package_bytes)) as archive:
        return archive.read(part).decode('utf-8')


def test_dedupe_and_dialect():
    """Repeated strings collapse to one entry; cells become t=s refs."""
    print("Testing dedupe and dialect rewrite...")

    package = workbook_bytes([
        ['SALMON', 'Umios Corporation', 1.5],
        ['SALMON', 'Umios Corporation', 2.5],
        ['SALMON', 'ICEBERG CO. LTD', 3.5],
    ])
    out, stats = consolidate_inline_strings(package)
    xml = sheet_xml(out)
    with zipfile.ZipFile(io.BytesIO(out)) as archive:
        sst = archive.read('xl/sharedStrings.xml').decode('utf-8')
    checks = [
        (stats['cells_consolidated'] == 6, '6 string cells consolidated'),
        (stats['unique_strings'] == 3, '3 unique strings'),
        ('t="inlineStr"' not in xml, 'no inline cells remain'),
        (xml.count('t="s"') == 6, 'six shared references'),
        (sst.count('SALMON') == 1, 'SALMON stored once'),
        ('uniqueCount="3"' in sst, 'uniqueCount correct'),
        ('count="6"' in sst, 'reference count correct'),
    ]
    passed = True
    for ok, label in checks:
        print(f"  {'✓' if ok else '✗'} {label}")
        passed &= ok
    return passed


def test_verbatim_text_preservation():
    """Escapes and leading/trailing spaces survive byte-exactly."""
    print("\nTesting verbatim preservation...")

    tricky = ['a < b & c > d', '  leading and trailing  ', 'quote " and \'']
    package = workbook_bytes([[value] for value in tricky])
    out, _ = consolidate_inline_strings(package)
    workbook = openpyxl.load_workbook(io.BytesIO(out))
    worksheet = workbook.active
    passed = True
    for row, expected in enumerate(tricky, start=1):
        actual = worksheet.cell(row=row, column=1).value
        ok = actual == expected
        print(f"  {'✓' if ok else '✗'} {expected!r} -> {actual!r}")
        passed &= ok
    with zipfile.ZipFile(io.BytesIO(out)) as archive:
        sst = archive.read('xl/sharedStrings.xml').decode('utf-8')
    space_ok = 'xml:space="preserve"' in sst
    print(f"  {'✓' if space_ok else '✗'} xml:space=preserve carried into the table")
    return passed and space_ok


def test_existing_table_merged_with_stable_indices():
    """An Excel-lineage table keeps its entries and indices untouched."""
    print("\nTesting existing-table merge...")

    package = workbook_bytes([['new text', 'old text']])
    # graft a pre-existing table + one t="s" cell referencing it
    with zipfile.ZipFile(io.BytesIO(package)) as archive:
        members = {name: archive.read(name) for name in archive.namelist()}
    members['xl/sharedStrings.xml'] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/'
        '2006/main" count="1" uniqueCount="2">'
        '<si><t>old text</t></si><si><t>legacy only</t></si></sst>'
    ).encode()
    rebuilt = io.BytesIO()
    with zipfile.ZipFile(rebuilt, 'w', zipfile.ZIP_DEFLATED) as out:
        for name, data in members.items():
            out.writestr(name, data)
    out_bytes, stats = consolidate_inline_strings(rebuilt.getvalue())
    with zipfile.ZipFile(io.BytesIO(out_bytes)) as archive:
        sst = archive.read('xl/sharedStrings.xml').decode('utf-8')
    xml = sheet_xml(out_bytes)
    first_old = sst.index('old text')
    first_legacy = sst.index('legacy only')
    checks = [
        (first_old < first_legacy, 'existing entry order preserved'),
        ('<v>0</v>' in xml, "'old text' cell reuses existing index 0"),
        (stats['unique_strings'] == 1, "only 'new text' appended"),
    ]
    passed = True
    for ok, label in checks:
        print(f"  {'✓' if ok else '✗'} {label}")
        passed &= ok
    return passed


def test_rich_text_refused():
    """A rich-text inline cell is left byte-untouched and counted."""
    print("\nTesting rich-text refusal...")

    package = workbook_bytes([['plain']])
    with zipfile.ZipFile(io.BytesIO(package)) as archive:
        members = {name: archive.read(name) for name in archive.namelist()}
    xml = members['xl/worksheets/sheet1.xml'].decode()
    rich = ('<c r="B1" t="inlineStr"><is><r><rPr><b/></rPr>'
            '<t>bold bit</t></r><r><t> plain bit</t></r></is></c>')
    xml = xml.replace('</row>', rich + '</row>', 1)
    members['xl/worksheets/sheet1.xml'] = xml.encode()
    rebuilt = io.BytesIO()
    with zipfile.ZipFile(rebuilt, 'w', zipfile.ZIP_DEFLATED) as out:
        for name, data in members.items():
            out.writestr(name, data)
    out_bytes, stats = consolidate_inline_strings(rebuilt.getvalue())
    out_xml = sheet_xml(out_bytes)
    checks = [
        (rich in out_xml, 'rich cell byte-untouched'),
        (stats['cells_skipped'] == 1, 'skip counted'),
        (stats['cells_consolidated'] == 1, 'plain sibling still consolidated'),
    ]
    passed = True
    for ok, label in checks:
        print(f"  {'✓' if ok else '✗'} {label}")
        passed &= ok
    return passed


def test_registration_and_idempotence():
    """Part registered in content types + rels; second pass is a no-op."""
    print("\nTesting registration and idempotence...")

    package = workbook_bytes([['x', 'x', 'y']])
    out_bytes, _ = consolidate_inline_strings(package)
    with zipfile.ZipFile(io.BytesIO(out_bytes)) as archive:
        content_types = archive.read('[Content_Types].xml').decode()
        rels = archive.read('xl/_rels/workbook.xml.rels').decode()
    again, stats = consolidate_inline_strings(out_bytes)
    checks = [
        ('/xl/sharedStrings.xml' in content_types, 'content-type override present'),
        ('sharedStrings.xml' in rels, 'workbook relationship present'),
        (stats['cells_consolidated'] == 0, 'second pass consolidates nothing'),
        (again == out_bytes, 'second pass returns identical bytes'),
    ]
    passed = True
    for ok, label in checks:
        print(f"  {'✓' if ok else '✗'} {label}")
        passed &= ok
    return passed


def main():
    """Run every test and report a final score."""
    print("=== inline-string consolidation tests ===")

    tests = [
        test_dedupe_and_dialect,
        test_verbatim_text_preservation,
        test_existing_table_merged_with_stable_indices,
        test_rich_text_refused,
        test_registration_and_idempotence,
    ]

    passed = 0
    for test_func in tests:
        try:
            if test_func():
                passed += 1
        except Exception as error:
            print(f"✗ {test_func.__name__} crashed: {error}")

    print(f"\n=== Results: {passed}/{len(tests)} tests passed ===")
    return passed == len(tests)


if __name__ == '__main__':
    sys.exit(0 if main() else 1)

# End of file #

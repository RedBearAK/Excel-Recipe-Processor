"""
Tests for format_excel's workbook_theme and pivot_style directives.

tests/test_format_excel_theming.py

Two mechanisms are covered, and the distinction matters: the THEME supplies
the accent colours Excel's style galleries resolve against, while the PIVOT
STYLE is a named entry in styles.xml carrying exact colours and bold total
rows. A regression guard also proves neither disturbs explicit cell
formatting, which is literal RGB and must survive both untouched.

Runnable with pytest, but written to run standalone and report a score.
"""

import re
import zipfile
import tempfile

from pathlib import Path

import openpyxl

from excel_recipe_processor.core.workbook_session import WorkbookSession
from excel_recipe_processor.processors.format_excel_processor import FormatExcelProcessor
from excel_recipe_processor.processors._helpers.format_excel_theme_manager import (
    THEME_PRESETS,
    ERP_DEFAULT_PIVOT_STYLE,
    ThemeManagerError,
    extract_theme_from_file,
    build_theme_with_accents,
)


accent_rgx = re.compile(r'<a:accent(\d)>\s*<a:srgbClr val="([0-9A-Fa-f]{6})"')


def make_workbook(path, header_fill='548235'):
    """A one-sheet workbook with an explicit header colour to guard."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data'
    ws['A1'] = 'Header'
    ws['A2'] = 'value'
    wb.save(path)
    wb.close()


def run_format(path, **extra):
    """Run format_excel on a path with a minimal sheet config plus extras."""
    WorkbookSession.reset()
    config = {
        'processor_type': 'format_excel',
        'target_file': path,
        'formatting': [{
            'sheet': 'Data',
            'header_bold': True,
            'header_background': True,
            'header_background_color': '548235',
            'header_text_color': 'white',
        }],
    }
    config.update(extra)
    FormatExcelProcessor(config).perform_file_operation()
    WorkbookSession.reset()


def read_part(path, part):
    with zipfile.ZipFile(path) as package:
        return package.read(part).decode('utf-8')


def test_default_is_the_purple_builtin_and_no_theme_change():
    """Out of the box: purple built-in pivot style, theme left alone."""
    print("\nTesting the ERP default...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'out.xlsx')
        make_workbook(path)
        theme_before = read_part(path, 'xl/theme/theme1.xml')
        run_format(path)

        styles = read_part(path, 'xl/styles.xml')
        theme_after = read_part(path, 'xl/theme/theme1.xml')

        passed = True

        if f'defaultPivotStyle="{ERP_DEFAULT_PIVOT_STYLE}"' in styles:
            print(f"  ✓ defaultPivotStyle is {ERP_DEFAULT_PIVOT_STYLE} (the purple swatch)")
        else:
            print(f"  ✗ {re.search(chr(39) + 'defaultPivotStyle="([^"]*)"' + chr(39), styles)}")
            passed = False

        if theme_before == theme_after:
            print("  ✓ Theme untouched - no gallery-wide recolouring by default")
        else:
            print("  ✗ Theme changed without being asked")
            passed = False

        if '<tableStyle name=' not in styles:
            print("  ✓ No style definition written - just the built-in name")
        else:
            print("  ✗ A style definition was written")
            passed = False

        return passed


def test_named_builtin_default_pivot_style():
    """A recipe can name any built-in gallery style."""
    print("\nTesting default_pivot_style...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'named.xlsx')
        make_workbook(path)
        run_format(path, default_pivot_style='PivotStyleMedium19')

        if 'defaultPivotStyle="PivotStyleMedium19"' in read_part(path, 'xl/styles.xml'):
            print("  ✓ Named built-in applied")
            return True
        print("  ✗ Not applied")
        return False


def test_theme_injection_is_opt_in():
    """workbook_theme only acts when the recipe asks for it."""
    print("\nTesting opt-in theme injection...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'themed.xlsx')
        make_workbook(path)
        run_format(path, workbook_theme={'preset': 'purple'})

        accents = dict(accent_rgx.findall(read_part(path, 'xl/theme/theme1.xml')))

        if accents.get('1', '').upper() == THEME_PRESETS['purple'][0]:
            print(f"  ✓ Requested palette applied (accent1 {accents['1']})")
            return True
        print(f"  ✗ accent1 is {accents.get('1')}")
        return False


def test_named_preset_and_explicit_accents():
    """preset: and accent_colors: each drive the palette."""
    print("\nTesting preset and explicit accents...")

    passed = True

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'preset.xlsx')
        make_workbook(path)
        run_format(path, workbook_theme={'preset': 'office_modern'})
        accents = dict(accent_rgx.findall(read_part(path, 'xl/theme/theme1.xml')))

        if accents.get('1', '').upper() == THEME_PRESETS['office_modern'][0]:
            print("  ✓ preset applied")
        else:
            print(f"  ✗ preset accent1: {accents.get('1')}")
            passed = False

        path2 = str(Path(temp_dir) / 'explicit.xlsx')
        make_workbook(path2)
        colors = ['1F4E79', '2E75B6', '9DC3E6', '203864', '8FAADC', '444444']
        run_format(path2, workbook_theme={'accent_colors': colors})
        accents2 = dict(accent_rgx.findall(read_part(path2, 'xl/theme/theme1.xml')))

        if [accents2[str(i)].upper() for i in range(1, 7)] == colors:
            print("  ✓ All six explicit accents written in order")
        else:
            print(f"  ✗ got {[accents2.get(str(i)) for i in range(1, 7)]}")
            passed = False

        return passed


def test_theme_from_a_donor_file():
    """from_file copies a whole theme out of another OOXML package."""
    print("\nTesting donor-file extraction...")

    with tempfile.TemporaryDirectory() as temp_dir:
        donor = str(Path(temp_dir) / 'donor.xlsx')
        make_workbook(donor)
        run_format(donor, workbook_theme={'accent_colors':
                                          ['AA1111', 'BB2222', 'CC3333',
                                           'DD4444', 'EE5555', 'FF6666']})

        target = str(Path(temp_dir) / 'target.xlsx')
        make_workbook(target)
        run_format(target, workbook_theme={'from_file': donor})

        accents = dict(accent_rgx.findall(read_part(target, 'xl/theme/theme1.xml')))

        if accents.get('1', '').upper() == 'AA1111' and accents.get('6', '').upper() == 'FF6666':
            print("  ✓ Donor theme adopted whole")
            return True
        print(f"  ✗ accents: {accents}")
        return False


def test_pivot_style_is_registered_and_default():
    """The custom pivot style lands in styles.xml as the workbook default."""
    print("\nTesting the custom pivot style...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'pivot.xlsx')
        make_workbook(path)
        run_format(path, pivot_style={
            'name': 'SBS Pivot',
            'header_background_color': '1F4E79',
            'header_font_color': 'FFFFFF',
            'bold_subtotals': True,
            'bold_grand_totals': True,
        })

        styles = read_part(path, 'xl/styles.xml')
        passed = True

        if 'defaultPivotStyle="SBS Pivot"' in styles:
            print("  ✓ Registered as defaultPivotStyle")
        else:
            print("  ✗ Not the default pivot style")
            passed = False

        for element in ('headerRow', 'firstSubtotalRow', 'totalRow'):
            if f'type="{element}"' not in styles:
                print(f"  ✗ Missing element {element}")
                passed = False
        if passed:
            print("  ✓ Header, subtotal and grand-total elements present")

        if '1F4E79' in styles and styles.count('<dxf>') >= 2:
            print("  ✓ Header colour and a bold differential format written")
        else:
            print(f"  ✗ dxf count {styles.count('<dxf>')}")
            passed = False

        reloaded = openpyxl.load_workbook(path)
        if reloaded._table_styles.defaultPivotStyle == 'SBS Pivot':
            print("  ✓ Round-trips through openpyxl")
        else:
            print("  ✗ Did not round-trip")
            passed = False
        reloaded.close()

        return passed


def test_explicit_cell_formatting_survives_theming():
    """The regression guard: literal cell colours must not shift."""
    print("\nTesting that explicit formatting is untouched...")

    with tempfile.TemporaryDirectory() as temp_dir:
        plain = str(Path(temp_dir) / 'plain.xlsx')
        themed = str(Path(temp_dir) / 'themed.xlsx')
        make_workbook(plain)
        make_workbook(themed)

        run_format(plain, workbook_theme={'apply': False})
        run_format(themed, pivot_style={'name': 'P', 'bold_subtotals': True})

        a = openpyxl.load_workbook(plain)['Data']['A1']
        b = openpyxl.load_workbook(themed)['Data']['A1']

        same = (a.font.bold == b.font.bold
                and str(a.fill.start_color.rgb) == str(b.fill.start_color.rgb)
                and str(a.font.color.rgb) == str(b.font.color.rgb))

        if same:
            print(f"  ✓ Header fill and font identical ({b.fill.start_color.rgb})")
            return True
        print(f"  ✗ plain {a.fill.start_color.rgb}/{a.font.color.rgb} vs "
              f"themed {b.fill.start_color.rgb}/{b.font.color.rgb}")
        return False


def test_configuration_errors_are_loud():
    """Bad colours, wrong accent counts and multiple sources all fail."""
    print("\nTesting configuration validation...")

    passed = True

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'bad.xlsx')

        checks = [
            ({'workbook_theme': {'accent_colors': ['1F4E79']}}, 'too few accents'),
            ({'workbook_theme': {'accent_colors': ['nothex'] * 6}}, 'bad hex'),
            ({'workbook_theme': {'preset': 'chartreuse'}}, 'unknown preset'),
            ({'workbook_theme': {'preset': 'purple', 'from_file': path}}, 'two sources'),
            ({'pivot_style': {'name': 'X'}}, 'pivot style with no formatting'),
        ]

        for extra, label in checks:
            make_workbook(path)
            try:
                run_format(path, **extra)
                print(f"  ✗ Accepted {label}")
                passed = False
            except Exception:
                print(f"  ✓ Rejected {label}")

        return passed


def test_missing_or_invalid_donor_fails_clearly():
    """A donor that does not exist, or carries no theme, is named."""
    print("\nTesting donor error handling...")

    passed = True

    with tempfile.TemporaryDirectory() as temp_dir:
        try:
            extract_theme_from_file(str(Path(temp_dir) / 'nope.xlsx'))
            print("  ✗ Missing donor accepted")
            passed = False
        except ThemeManagerError as error:
            print("  ✓ Missing donor rejected" if 'not found' in str(error) else f"  ✗ {error}")

        not_ooxml = Path(temp_dir) / 'fake.xlsx'
        not_ooxml.write_text('this is not a zip')
        try:
            extract_theme_from_file(str(not_ooxml))
            print("  ✗ Non-OOXML donor accepted")
            passed = False
        except ThemeManagerError:
            print("  ✓ Non-OOXML donor rejected")

        return passed


def test_accent_substitution_preserves_the_rest_of_the_theme():
    """Only the six accents change; fonts and effects survive."""
    print("\nTesting theme substitution is surgical...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'base.xlsx')
        make_workbook(path)
        base = zipfile.ZipFile(path).read('xl/theme/theme1.xml')

        updated = build_theme_with_accents(base, list(THEME_PRESETS['purple']))

        base_text = base.decode('utf-8')
        updated_text = updated.decode('utf-8')

        passed = True

        if len(base_text) == len(updated_text):
            print("  ✓ Same length: only hex digits replaced")
        else:
            print(f"  ✗ Length changed {len(base_text)} -> {len(updated_text)}")
            passed = False

        if '<a:latin typeface=' in updated_text and updated_text.count('<a:effectStyle>') == base_text.count('<a:effectStyle>'):
            print("  ✓ Font scheme and effect styles preserved")
        else:
            print("  ✗ Theme structure disturbed")
            passed = False

        return passed


def test_constructed_files_start_from_the_modern_theme():
    """Every workbook the writer builds carries the modern Office theme."""
    print("\nTesting the modern theme baseline...")

    import pandas as pd
    from excel_recipe_processor.writers.excel_writer import ExcelWriter

    with tempfile.TemporaryDirectory() as temp_dir:
        single = str(Path(temp_dir) / 'single.xlsx')
        multi = str(Path(temp_dir) / 'multi.xlsx')
        frame = pd.DataFrame({'A': [1, 2], 'B': ['x', 'y']})

        writer = ExcelWriter()
        writer.write_file(frame, single)
        writer.write_multiple_sheets({'One': frame, 'Two': frame}, multi)

        passed = True

        for label, path in (('single-sheet', single), ('multi-sheet', multi)):
            theme = read_part(path, 'xl/theme/theme1.xml')
            accents = dict(accent_rgx.findall(theme))

            if accents.get('1', '').upper() == THEME_PRESETS['office_modern'][0]:
                print(f"  ✓ {label} carries the modern palette")
            else:
                print(f"  ✗ {label} accent1 is {accents.get('1')}")
                passed = False

            if 'Aptos' in theme:
                print(f"  ✓ {label} carries the modern font scheme")
            else:
                print(f"  ✗ {label} font scheme is not modern")
                passed = False

            reloaded = pd.read_excel(path, sheet_name=0)
            if len(reloaded) == 2 and list(reloaded['A']) == [1, 2]:
                print(f"  ✓ {label} data intact")
            else:
                print(f"  ✗ {label} data changed")
                passed = False

        return passed


def test_purple_default_tracks_the_modern_palette():
    """The default pivot style names the purple swatch of the BASE theme."""
    print("\nTesting the purple default against modern accents...")

    from excel_recipe_processor.processors._helpers.format_excel_theme_manager import (
        modern_base_theme_bytes,
    )

    accents = [v for _, v in accent_rgx.findall(modern_base_theme_bytes().decode('utf-8'))]

    # PivotStyleLight16 is accent1; each later name steps one accent along
    slot = int(ERP_DEFAULT_PIVOT_STYLE.replace('PivotStyleLight', '')) - 15
    chosen = accents[slot - 1].upper()

    if chosen == 'A02B93':
        print(f"  ✓ {ERP_DEFAULT_PIVOT_STYLE} maps to accent{slot} = {chosen} (the purple swatch)")
        return True
    print(f"  ✗ {ERP_DEFAULT_PIVOT_STYLE} maps to accent{slot} = {chosen}")
    return False


def main():
    """Run every test and report a final score."""
    print("=== format_excel theming tests ===")

    tests = [
        test_constructed_files_start_from_the_modern_theme,
        test_purple_default_tracks_the_modern_palette,
        test_default_is_the_purple_builtin_and_no_theme_change,
        test_named_builtin_default_pivot_style,
        test_theme_injection_is_opt_in,
        test_named_preset_and_explicit_accents,
        test_theme_from_a_donor_file,
        test_pivot_style_is_registered_and_default,
        test_explicit_cell_formatting_survives_theming,
        test_configuration_errors_are_loud,
        test_missing_or_invalid_donor_fails_clearly,
        test_accent_substitution_preserves_the_rest_of_the_theme,
    ]

    passed = 0

    for test_func in tests:
        try:
            if test_func():
                passed += 1
        except Exception as error:
            print(f"  ✗ {test_func.__name__} crashed: {error}")

    print(f"\n=== Results: {passed}/{len(tests)} tests passed ===")

    if passed == len(tests):
        print("✅ All format_excel theming tests passed!")
        return 1

    print("❌ Some format_excel theming tests failed!")
    return 0


if __name__ == '__main__':
    exit(0 if main() else 1)


# End of file #

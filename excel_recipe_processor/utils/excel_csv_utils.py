"""
Simple Excel utilities for the processor (legacy support).

excel_recipe_processor/utils/excel_csv_utils.py

This module is kept for compatibility but the new processor uses direct
Excel reading which is much faster than these CSV conversion utilities.
"""

import logging

logger = logging.getLogger(__name__)


class ExcelCsvUtilsError(Exception):
    """Raised when Excel operations fail."""
    pass


def check_openpyxl_availability():
    """Check if openpyxl is available."""
    try:
        import openpyxl
        return True
    except ImportError:
        raise ExcelCsvUtilsError(
            "openpyxl is required for Excel reading but not available. "
            "Please install it with: pip install openpyxl"
        )


# Legacy compatibility functions - not used by new fast processor
def excel_headers_fast(file_path: str, sheet_name=None, header_row: int = 1) -> list:
    """Legacy function - use processor's direct method instead."""
    logger.warning("Using legacy excel_headers_fast - consider direct processor methods")
    check_openpyxl_availability()
    
    import openpyxl
    
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    worksheet = workbook.active if sheet_name is None else workbook[sheet_name]
    
    headers = []
    max_col = worksheet.max_column or 1
    
    for col in range(1, max_col + 1):
        cell = worksheet.cell(row=header_row, column=col)
        headers.append(str(cell.value) if cell.value is not None else "")
    
    workbook.close()
    return headers


def excel_column_analysis_fast(file_path: str, sheet_name=None, header_row: int = 1, 
                             sample_rows: int = 20) -> dict:
    """Legacy function - use processor's direct method instead.""" 
    logger.warning("Using legacy excel_column_analysis_fast - processor has faster methods")
    
    headers = excel_headers_fast(file_path, sheet_name, header_row)
    
    # Minimal analysis for compatibility
    return {
        'headers': headers,
        'column_has_data': [True] * len(headers),  # Assume all have data
        'empty_column_indices': [],
        'scanned_data_rows': sample_rows,
        'total_columns': len(headers)
    }


# End of file #

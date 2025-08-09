#!/usr/bin/env python3
"""
Test script for format_excel processor sheet targeting functionality.

tests/test_format_excel_sheet_targeting.py

Tests the new explicit sheet targeting format with support for both
sheet names and 1-based numeric indexing.
"""

import sys
import os
import tempfile
import pandas as pd
import openpyxl
from pathlib import Path

# Add the parent directory to the path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from excel_recipe_processor.processors.format_excel_processor import FormatExcelProcessor


def create_multi_sheet_test_file():
    """Create a test Excel file with multiple sheets."""
    test_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    test_file.close()
    
    # Create sample data for different sheets
    summary_data = pd.DataFrame({
        'Metric': ['Revenue', 'Profit', 'Growth'],
        'Q1': [100, 20, 15],
        'Q2': [120, 25, 20],
        'Q3': [140, 30, 25]
    })
    
    detailed_data = pd.DataFrame({
        'Date': ['2024-01-01', '2024-02-01', '2024-03-01'],
        'Sales': [1000, 1200, 1400],
        'Costs': [800, 900, 1000],
        'Region': ['North', 'South', 'East']
    })
    
    charts_data = pd.DataFrame({
        'Category': ['A', 'B', 'C', 'D'],
        'Value': [10, 20, 15, 25]
    })
    
    # Write to Excel with multiple sheets
    with pd.ExcelWriter(test_file.name, engine='openpyxl') as writer:
        summary_data.to_excel(writer, sheet_name='Executive_Summary', index=False)
        detailed_data.to_excel(writer, sheet_name='Detailed_Data', index=False)
        charts_data.to_excel(writer, sheet_name='Charts_Analysis', index=False)
    
    return test_file.name


def test_sheet_name_targeting():
    """Test targeting sheets by name."""
    print("\n1. Testing sheet targeting by name...")
    
    test_file = create_multi_sheet_test_file()
    
    try:
        step_config = {
            'processor_type': 'format_excel',
            'target_file': test_file,
            'active_sheet': 'Executive_Summary',
            'formatting': [
                {
                    'sheet': 'Executive_Summary',
                    'header_background_color': 'navy',
                    'header_text_color': 'white',
                    'header_bold': True,
                    'auto_fit_columns': True
                },
                {
                    'sheet': 'Detailed_Data',
                    'header_background_color': 'darkgreen',
                    'header_text_color': 'white',
                    'auto_fit_columns': True,
                    'auto_filter': True
                }
            ]
        }
        
        processor = FormatExcelProcessor(step_config)
        result = processor.execute()
        
        # Verify file is still readable
        workbook = openpyxl.load_workbook(test_file)
        
        # Check that we have the expected sheets
        expected_sheets = ['Executive_Summary', 'Detailed_Data', 'Charts_Analysis']
        if workbook.sheetnames == expected_sheets:
            print("  ‚úì All sheets present")
        else:
            print(f"  ‚úó Expected {expected_sheets}, got {workbook.sheetnames}")
            return False
        
        # Check active sheet
        if workbook.active.title == 'Executive_Summary':
            print("  ‚úì Active sheet set correctly")
        else:
            print(f"  ‚úó Expected active sheet 'Executive_Summary', got '{workbook.active.title}'")
            return False
        
        workbook.close()
        print("  ‚úì Sheet name targeting successful")
        return True
        
    except Exception as e:
        print(f"  ‚úó Sheet name targeting failed: {e}")
        return False
        
    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)


def test_numeric_sheet_targeting():
    """Test targeting sheets by 1-based index."""
    print("\n2. Testing sheet targeting by number...")
    
    test_file = create_multi_sheet_test_file()
    
    try:
        step_config = {
            'processor_type': 'format_excel',
            'target_file': test_file,
            'active_sheet': 2,  # Set second sheet as active
            'formatting': [
                {
                    'sheet': 1,  # First sheet (Executive_Summary)
                    'header_background_color': 'purple',
                    'header_text_color': 'white',
                    'header_font_size': 16,
                    'row_heights': {1: 30}
                },
                {
                    'sheet': 2,  # Second sheet (Detailed_Data)
                    'header_background_color': 'orange',
                    'header_text_color': 'black',
                    'auto_fit_columns': True
                },
                {
                    'sheet': 3,  # Third sheet (Charts_Analysis)
                    'header_background_color': 'lightblue',
                    'header_text_color': 'darkblue',
                    'freeze_top_row': True
                }
            ]
        }
        
        processor = FormatExcelProcessor(step_config)
        result = processor.execute()
        
        # Verify file and active sheet
        workbook = openpyxl.load_workbook(test_file)
        
        # Check active sheet
        if workbook.active.title == 'Detailed_Data':
            print("  ‚úì Active sheet set correctly by number")
        else:
            print(f"  ‚úó Expected active sheet 'Detailed_Data', got '{workbook.active.title}'")
            print(f"  Debug: Available sheets: {workbook.sheetnames}")
            print(f"  Debug: Active sheet index should be 2 (Detailed_Data)")
            return False
        
        workbook.close()
        print("  ‚úì Numeric sheet targeting successful")
        return True
        
    except Exception as e:
        print(f"  ‚úó Numeric sheet targeting failed: {e}")
        return False
        
    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)


def test_mixed_sheet_targeting():
    """Test mixing sheet names and numbers."""
    print("\n3. Testing mixed sheet targeting...")
    
    test_file = create_multi_sheet_test_file()
    
    try:
        step_config = {
            'processor_type': 'format_excel',
            'target_file': test_file,
            'formatting': [
                {
                    'sheet': 'Executive_Summary',  # By name
                    'header_background_color': 'red',
                    'header_text_color': 'white'
                },
                {
                    'sheet': 2,  # By number
                    'header_background_color': 'blue',
                    'header_text_color': 'white'
                },
                {
                    'sheet': 'Charts_Analysis',  # By name again
                    'header_background_color': 'green',
                    'header_text_color': 'white'
                }
            ]
        }
        
        processor = FormatExcelProcessor(step_config)
        result = processor.execute()
        
        print("  ‚úì Mixed sheet targeting successful")
        return True
        
    except Exception as e:
        print(f"  ‚úó Mixed sheet targeting failed: {e}")
        return False
        
    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)


def test_validation_errors():
    """Test configuration validation errors."""
    print("\n4. Testing validation errors...")
    
    test_file = create_multi_sheet_test_file()
    
    try:
        # Test missing 'sheet' key
        try:
            step_config = {
                'processor_type': 'format_excel',
                'target_file': test_file,
                'formatting': [
                    {
                        'header_bold': True  # Missing 'sheet' key
                    }
                ]
            }
            processor = FormatExcelProcessor(step_config)
            processor.execute()
            print("  ‚úó Should have failed for missing 'sheet' key")
            return False
        except Exception:
            print("  ‚úì Correctly rejected missing 'sheet' key")
        
        # Test invalid sheet index
        try:
            step_config = {
                'processor_type': 'format_excel',
                'target_file': test_file,
                'formatting': [
                    {
                        'sheet': 0,  # Invalid: must be >= 1
                        'header_bold': True
                    }
                ]
            }
            processor = FormatExcelProcessor(step_config)
            processor.execute()
            print("  ‚úó Should have failed for sheet index 0")
            return False
        except Exception:
            print("  ‚úì Correctly rejected sheet index 0")
        
        # Test empty sheet name
        try:
            step_config = {
                'processor_type': 'format_excel',
                'target_file': test_file,
                'formatting': [
                    {
                        'sheet': '',  # Invalid: empty string
                        'header_bold': True
                    }
                ]
            }
            processor = FormatExcelProcessor(step_config)
            processor.execute()
            print("  ‚úó Should have failed for empty sheet name")
            return False
        except Exception:
            print("  ‚úì Correctly rejected empty sheet name")
        
        # Test formatting not being a list
        try:
            step_config = {
                'processor_type': 'format_excel',
                'target_file': test_file,
                'formatting': {  # Should be a list, not dict
                    'sheet': 'Summary',
                    'header_bold': True
                }
            }
            processor = FormatExcelProcessor(step_config)
            processor.execute()
            print("  ‚úó Should have failed for formatting not being a list")
            return False
        except Exception:
            print("  ‚úì Correctly rejected non-list formatting")
        
        print("  ‚úì All validation tests passed")
        return True
        
    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)


def test_nonexistent_sheets():
    """Test handling of nonexistent sheets."""
    print("\n5. Testing nonexistent sheet handling...")
    
    test_file = create_multi_sheet_test_file()
    
    try:
        step_config = {
            'processor_type': 'format_excel',
            'target_file': test_file,
            'formatting': [
                {
                    'sheet': 'Executive_Summary',  # Exists
                    'header_background_color': 'blue'
                },
                {
                    'sheet': 'NonExistent_Sheet',  # Doesn't exist
                    'header_background_color': 'red'
                },
                {
                    'sheet': 10,  # Out of range
                    'header_background_color': 'green'
                }
            ]
        }
        
        processor = FormatExcelProcessor(step_config)
        result = processor.execute()
        
        # Should process successfully, just skip nonexistent sheets
        print("  ‚úì Gracefully handled nonexistent sheets")
        return True
        
    except Exception as e:
        print(f"  ‚úó Failed to handle nonexistent sheets: {e}")
        return False
        
    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)


def test_row_heights():
    """Test row height specifications."""
    print("\n6. Testing row height functionality...")
    
    test_file = create_multi_sheet_test_file()
    
    try:
        step_config = {
            'processor_type': 'format_excel',
            'target_file': test_file,
            'formatting': [
                {
                    'sheet': 1,
                    'header_font_size': 18,  # Large font
                    'row_heights': {
                        1: 35,  # Tall header for large font
                        2: 20,  # Standard data rows
                        3: 20,
                        4: 25   # Slightly taller for specific row
                    },
                    'auto_fit_columns': True
                }
            ]
        }
        
        processor = FormatExcelProcessor(step_config)
        result = processor.execute()
        
        print("  ‚úì Row heights applied successfully")
        return True
        
    except Exception as e:
        print(f"  ‚úó Row heights failed: {e}")
        return False
        
    finally:
        if os.path.exists(test_file):
            os.unlink(test_file)


def main():
    """Run all sheet targeting tests."""
    print("üöÄ Format Excel Processor - Sheet Targeting Tests")
    print("=" * 60)
    print("Testing new explicit sheet targeting with names and numbers")
    
    tests = [
        test_sheet_name_targeting,
        test_numeric_sheet_targeting,
        test_mixed_sheet_targeting,
        test_validation_errors,
        test_nonexistent_sheets,
        test_row_heights
    ]
    
    passed = sum(1 for test in tests if test())
    total = len(tests)
    
    print(f"\nüèÅ SHEET TARGETING TEST RESULTS")
    print("=" * 60)
    
    if passed == total:
        print(f"‚úÖ All {passed}/{total} tests passed!")
        print()
        print("üéØ SHEET TARGETING ACHIEVEMENTS:")
        print("‚úì Sheet targeting by name works correctly")
        print("‚úì Sheet targeting by 1-based index works correctly") 
        print("‚úì Mixed name/number targeting works correctly")
        print("‚úì Configuration validation catches all error cases")
        print("‚úì Graceful handling of nonexistent sheets")
        print("‚úì Row height specifications work correctly")
        print("‚úì Active sheet setting works with both names and numbers")
        print()
        print("üéâ BREAKING CHANGE SUCCESSFUL!")
        print("The new explicit sheet targeting format is ready for production!")
        return 0
    else:
        print(f"‚ùå {passed}/{total} tests passed")
        print("Issues need to be resolved before the new format is ready")
        return 1


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)


# End of file #

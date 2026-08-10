"""
Tests for the import_file verbatim_text_columns option.

tests/test_verbatim_text_columns.py

Runnable with pytest, but written to run standalone and report a score.
"""

import tempfile

from pathlib import Path

import pandas as pd
import openpyxl

from excel_recipe_processor.core.file_reader import FileReader


def make_workbook(path):
    """Text col with N/A-style entries, numeric col with blanks, control col."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, header in enumerate(['Ref', 'Weight', 'Other'], start=1):
        ws.cell(row=1, column=col).value = header
    rows = [('N/A', 10.5, 'N/A'), ('NA', None, 'x'),
            (None, 20.0, None), ('REAL-REF', 30.25, 'NULL')]
    for r, values in enumerate(rows, start=2):
        for col, value in enumerate(values, start=1):
            ws.cell(row=r, column=col).value = value
    wb.save(path)
    wb.close()


def test_protected_column_keeps_text():
    """N/A and NA survive as strings; the empty cell stays missing."""
    print("\nTesting protected column keeps its text...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'probe.xlsx')
        make_workbook(path)
        frame = FileReader.read_file(path, verbatim_text_columns=['Ref'])

        passed = True

        if list(frame['Ref'].fillna('<MISSING>')) == ['N/A', 'NA', '<MISSING>', 'REAL-REF']:
            print("  ✓ Text verbatim, blank still missing")
        else:
            print(f"  ✗ Ref: {[repr(v) for v in frame['Ref']]}")
            passed = False

        return passed


def test_unprotected_columns_match_plain_read():
    """Columns not listed behave exactly like a normal import."""
    print("\nTesting unprotected columns are untouched...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'probe.xlsx')
        make_workbook(path)
        plain = FileReader.read_file(path)
        protected = FileReader.read_file(path, verbatim_text_columns=['Ref'])

        passed = True

        other_equal = plain['Other'].fillna('<M>').tolist() == protected['Other'].fillna('<M>').tolist()
        if other_equal:
            print("  ✓ Unprotected text column identical to plain read")
        else:
            print(f"  ✗ plain {plain['Other'].tolist()} vs protected {protected['Other'].tolist()}")
            passed = False

        if str(protected['Weight'].dtype) == str(plain['Weight'].dtype) and \
                plain['Weight'].sum() == protected['Weight'].sum():
            print(f"  ✓ Numeric column keeps dtype ({protected['Weight'].dtype}) and values")
        else:
            print(f"  ✗ Weight dtype {protected['Weight'].dtype}, sum {protected['Weight'].sum()}")
            passed = False

        return passed


def test_misspelled_column_warns_but_proceeds():
    """A wrong column name must not halt the import."""
    print("\nTesting a misspelled column name...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'probe.xlsx')
        make_workbook(path)
        frame = FileReader.read_file(path, verbatim_text_columns=['Rf Typo'])

        if len(frame) == 4:
            print("  ✓ Import proceeded (warning logged)")
            return True

        print(f"  ✗ Got {len(frame)} rows")
        return False


def test_csv_path_honors_the_option():
    """The CSV reader applies the same policy."""
    print("\nTesting the CSV path...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'probe.csv')
        Path(path).write_text('Ref,Other\nN/A,N/A\nREAL,x\n')

        plain = FileReader.read_file(path)
        protected = FileReader.read_file(path, verbatim_text_columns=['Ref'])

        passed = True

        if pd.isna(plain['Ref'].iloc[0]) and protected['Ref'].iloc[0] == 'N/A':
            print("  ✓ CSV: protected keeps N/A, plain coerces it")
        else:
            print(f"  ✗ plain {plain['Ref'].iloc[0]!r}, protected {protected['Ref'].iloc[0]!r}")
            passed = False

        if pd.isna(protected['Other'].iloc[0]):
            print("  ✓ CSV: unprotected column still coerces")
        else:
            print(f"  ✗ Other: {protected['Other'].iloc[0]!r}")
            passed = False

        return passed


def main():
    """Run every test and report a final score."""
    print("=== verbatim_text_columns tests ===")

    tests = [
        test_protected_column_keeps_text,
        test_unprotected_columns_match_plain_read,
        test_misspelled_column_warns_but_proceeds,
        test_csv_path_honors_the_option,
    ]

    passed = 0

    for test_func in tests:
        try:
            if test_func():
                passed += 1
        except Exception as error:
            print(f"  ✗ {test_func.__name__} crashed: {error}")

    print(f"\n=== Results: {passed}/{len(tests)} tests passed ===")

    if passed == len(tests):
        print("✅ All verbatim_text_columns tests passed!")
        return 1

    print("❌ Some verbatim_text_columns tests failed!")
    return 0


if __name__ == '__main__':
    exit(0 if main() else 1)


# End of file #

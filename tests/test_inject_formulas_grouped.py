"""
Grouped injection shape: sheets_to_receive_formulas entries, exclusivity, naming.

tests/test_inject_formulas_grouped.py

sheets_to_receive_formulas (2026-08-17) is the SOLE live injection
shape: a list of entries, each pairing a sheet_names LIST with its own
formulas. One entry with many sheets is the old broadcast; many
entries give per-sheet formulas; both compose in one step. The
top-level sheet_names + formulas broadcast pair is retired for live
mode with a guided error, the singular per-entry sheet_name key is
guided to the plural list, and per-sheet failures name the entry and
the sheet.

Runnable standalone or under pytest; the exit code carries the verdict.
"""

import os
import sys
import tempfile

import openpyxl
import pandas as pd

from excel_recipe_processor.core.pipeline import registry
from excel_recipe_processor.core.base_processor import StepProcessorError
from excel_recipe_processor.core.workbook_session import WorkbookSession


def make_workbook(path):
    """Two-sheet workbook for injection targets."""
    workbook = openpyxl.Workbook()
    workbook.active.title = 'View_A'
    workbook.create_sheet('View_B')
    for sheet in workbook.worksheets:
        sheet['A1'] = 'Header'
    workbook.save(path)


def make_processor(config):
    """Processor with the step name the errors should carry."""
    full_config = {'processor_type': 'inject_formulas',
                   'step_description': 'grouped drill'}
    full_config.update(config)
    return registry._processors['inject_formulas'](full_config)


def run_step(config):
    """Execute a config against a fresh workbook; return its path."""
    workdir = tempfile.mkdtemp()
    target = os.path.join(workdir, 'target.xlsx')
    make_workbook(target)
    config = dict(config, target_file=target)
    processor = make_processor(config)
    processor.execute(None)
    WorkbookSession.flush_all()
    return target


def test_grouped_injects_per_sheet():
    """Each entry's formulas land on that entry's sheet only."""
    print("Testing per-sheet grouped injection...")

    target = run_step({
        'mode': 'live',
        'sheets_to_receive_formulas': [
            {'sheet_names': ['View_A'],
             'formulas': [{'cell': 'A2', 'excel_formula': '=1+1'}]},
            {'sheet_names': ['View_B'],
             'formulas': [{'cell': 'A2', 'excel_formula': '=2+2'},
                          {'cell': 'B2', 'excel_formula': '=3+3'}]},
        ],
    })
    workbook = openpyxl.load_workbook(target)
    checks = [
        (workbook['View_A']['A2'].value, '=1+1'),
        (workbook['View_A']['B2'].value, None),
        (workbook['View_B']['A2'].value, '=2+2'),
        (workbook['View_B']['B2'].value, '=3+3'),
    ]
    passed = all(actual == expected for actual, expected in checks)
    for actual, expected in checks:
        print(f"  {'✓' if actual == expected else '✗'} {expected!r} -> {actual!r}")
    return passed


def test_broadcast_pair_retired():
    """Top-level sheet_names/formulas in live mode fail with guidance."""
    print("\nTesting the broadcast retirement...")

    try:
        run_step({
            'mode': 'live',
            'sheet_names': ['View_A'],
            'formulas': [{'cell': 'A2', 'excel_formula': '=1'}],
        })
        print("  ✗ retired broadcast pair accepted")
        return False
    except StepProcessorError as error:
        message = str(error)
        ok = ('retired' in message
              and 'sheets_to_receive_formulas' in message
              and 'one entry' in message)
        print(f"  {'✓' if ok else '✗'} guided: {message[:90]}")
        return ok


def test_singular_entry_key_guided():
    """Per-entry sheet_name (singular) is guided to the plural list."""
    print("\nTesting the singular-key guidance...")

    try:
        run_step({
            'mode': 'live',
            'sheets_to_receive_formulas': [
                {'sheet_name': 'View_A',
                 'formulas': [{'cell': 'A2', 'excel_formula': '=1'}]}],
        })
        print("  ✗ singular sheet_name accepted")
        return False
    except StepProcessorError as error:
        message = str(error)
        ok = "unknown key 'sheet_name'" in message and "did you mean 'sheet_names'" in message
        print(f"  {'✓' if ok else '✗'} guided: {message[:90]}")
        return ok


def test_scalar_sheet_names_guided():
    """A bare-string sheet_names is guided to the list form."""
    print("\nTesting the scalar-value guidance...")

    try:
        run_step({
            'mode': 'live',
            'sheets_to_receive_formulas': [
                {'sheet_names': 'View_A',
                 'formulas': [{'cell': 'A2', 'excel_formula': '=1'}]}],
        })
        print("  ✗ bare-string sheet_names accepted")
        return False
    except StepProcessorError as error:
        message = str(error)
        ok = 'sheet_names: expected list of str' in message
        print(f"  {'✓' if ok else '✗'} guided: {message[:90]}")
        return ok


def test_one_entry_many_sheets():
    """One entry's sheet_names list lands the same formulas on each."""
    print("\nTesting the subsumed broadcast (one entry, many sheets)...")

    target = run_step({
        'mode': 'live',
        'sheets_to_receive_formulas': [
            {'sheet_names': ['View_A', 'View_B'],
             'formulas': [{'cell': 'C2', 'excel_formula': '=9'}]}],
    })
    workbook = openpyxl.load_workbook(target)
    ok = (workbook['View_A']['C2'].value == '=9'
          and workbook['View_B']['C2'].value == '=9')
    print(f"  {'✓' if ok else '✗'} same formula on both listed sheets")
    return ok


def test_entry_validation_names_position():
    """A malformed entry fails naming its position in the list."""
    print("\nTesting entry validation...")

    try:
        run_step({
            'mode': 'live',
            'sheets_to_receive_formulas': [
                {'sheet_names': ['View_A'],
                 'formulas': [{'cell': 'A2', 'excel_formula': '=1'}]},
                {'sheet_names': ['View_B']},
            ],
        })
        print("  ✗ entry without formulas accepted")
        return False
    except StepProcessorError as error:
        ok = '[2]' in str(error) and "missing required key 'formulas'" in str(error)
        print(f"  {'✓' if ok else '✗'} named: {str(error)[:90]}")
        return ok


def test_unknown_entry_key_fails_loud():
    """A stray key in an entry is named, not silently ignored."""
    print("\nTesting unknown entry keys...")

    try:
        run_step({
            'mode': 'live',
            'sheets_to_receive_formulas': [
                {'sheet_names': ['View_A'], 'mode': 'awaken',
                 'formulas': [{'cell': 'A2', 'excel_formula': '=1'}]}],
        })
        print("  ✗ stray per-entry key accepted")
        return False
    except StepProcessorError as error:
        ok = "unknown key 'mode'" in str(error)
        print(f"  {'✓' if ok else '✗'} named: {str(error)[:100]}")
        return ok


def test_per_sheet_failure_names_sheet():
    """A bad formula halts naming the entry AND the sheet."""
    print("\nTesting per-sheet failure attribution...")

    try:
        run_step({
            'mode': 'live',
            'sheets_to_receive_formulas': [
                {'sheet_names': ['View_A'],
                 'formulas': [{'cell': 'A2', 'excel_formula': '=1'}]},
                {'sheet_names': ['View_B'],
                 'formulas': [{'cell': 'NOT_A_CELL', 'excel_formula': '=2'}]},
            ],
        })
        print("  ✗ bad cell accepted")
        return False
    except StepProcessorError as error:
        message = str(error)
        ok = 'entry 2' in message and "sheet 'View_B'" in message
        print(f"  {'✓' if ok else '✗'} attributed: {message[:100]}")
        return ok


def main():
    """Run every test and report a final score."""
    print("=== grouped injection shape tests ===")

    tests = [
        test_grouped_injects_per_sheet,
        test_one_entry_many_sheets,
        test_broadcast_pair_retired,
        test_singular_entry_key_guided,
        test_scalar_sheet_names_guided,
        test_entry_validation_names_position,
        test_unknown_entry_key_fails_loud,
        test_per_sheet_failure_names_sheet,
    ]

    passed = 0
    for test_func in tests:
        try:
            if test_func():
                passed += 1
        except Exception as error:
            print(f"✗ {test_func.__name__} crashed: {error}")
        finally:
            try:
                WorkbookSession.discard_all()
            except Exception:
                pass

    print(f"\n=== Results: {passed}/{len(tests)} tests passed ===")
    return passed == len(tests)


if __name__ == '__main__':
    sys.exit(0 if main() else 1)

# End of file #

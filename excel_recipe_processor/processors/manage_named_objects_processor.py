"""
Manage named objects processor for Excel automation recipes.

excel_recipe_processor/processors/manage_named_objects_processor.py

Handles extraction, translation, and restoration of Excel named ranges, 
lambda functions, formulas, and tables. Supports export to/import from
human-readable YAML format with full fidelity preservation.
"""

import re
import yaml
import logging
import openpyxl

from pathlib import Path
from datetime import datetime

from openpyxl.workbook.defined_name import DefinedName

from excel_recipe_processor.core.base_processor import FileOpsBaseProcessor, StepProcessorError
from excel_recipe_processor.processors._helpers.excel_range_resolver import (
    resolve_range, ExcelRangeResolverError
)
from excel_recipe_processor.processors._helpers.defined_name_validator import (
    validate_defined_name, DefinedNameError
)
from excel_recipe_processor.processors._helpers.named_objects_patterns import (
    excel_lambda_detection_rgx, excel_lambda_params_rgx, excel_param_name_rgx,
    excel_lambda_body_rgx, human_lambda_detection_rgx, human_lambda_params_rgx,
    human_lambda_body_rgx, xlfn_function_rgx, xlpm_parameter_rgx,
    excel_prefix_cleanup_rgx, valid_excel_name_rgx, table_reference_rgx,
    sheet_name_from_range_rgx, function_call_rgx, named_reference_in_formula_rgx
)


logger = logging.getLogger(__name__)


class ManageNamedObjectsProcessor(FileOpsBaseProcessor):
    """
    Processor for managing Excel named ranges, formulas, lambda functions, and tables.
    
    Supports exporting to/importing from human-readable YAML format with translation
    between Excel internal format and user-friendly syntax.
    """
    
    SUPPORTED_OPERATIONS = [
        'export_all', 'export_filtered', 'import_all', 'import_filtered',
        'list_objects', 'validate_yaml', 'copy_direct', 'create_from_columns'
    ]

    # What to do when a name already exists in the target workbook. Defaults to
    # 'error' because silently replacing a name is a form of data loss.
    EXISTING_NAME_POLICIES = ['error', 'replace', 'skip']

    # How hard to check names before writing them. Foreign workbooks contain
    # legal names that violate house style, so import defaults to 'excel'
    # while names authored by a recipe default to 'house'.
    NAME_VALIDATION_LEVELS = ['none', 'excel', 'house']

    # Object categories carried in the export structure, in write order.
    OBJECT_SECTIONS = ['named_ranges', 'lambda_functions', 'named_formulas']

    def __init__(self, step_config: dict):
        super().__init__(step_config)
        self.operation = self.get_config_value('operation')
        
        if self.operation not in self.SUPPORTED_OPERATIONS:
            raise StepProcessorError(
                f"Invalid operation '{self.operation}'. "
                f"Supported: {', '.join(self.SUPPORTED_OPERATIONS)}"
            )
    
    def perform_file_operation(self, filename: str) -> dict:
        """
        Required implementation for FileOpsBaseProcessor.
        Delegates to execute() method which handles all operations.
        """
        return self.execute()
    
    def execute(self) -> dict:
        """Execute the named objects management operation."""
        
        if self.operation == 'export_all':
            return self._execute_export_all()
        elif self.operation == 'export_filtered':
            return self._execute_export_filtered()
        elif self.operation == 'import_all':
            return self._execute_import_all()
        elif self.operation == 'import_filtered':
            return self._execute_import_filtered()
        elif self.operation == 'list_objects':
            return self._execute_list_objects()
        elif self.operation == 'validate_yaml':
            return self._execute_validate_yaml()
        elif self.operation == 'create_from_columns':
            return self._execute_create_from_columns()
        elif self.operation == 'copy_direct':
            return self._execute_copy_direct()
        else:
            raise StepProcessorError(f"Operation '{self.operation}' not implemented")
    
    # =============================================================================
    # CORE EXTRACTION METHODS
    # =============================================================================
    
    def extract_all_named_objects(self, workbook) -> dict:
        """Extract all named objects from workbook into structured format."""
        
        # Extract different types of objects
        named_ranges = self.extract_named_ranges(workbook)
        lambda_functions = self.extract_lambda_functions(workbook) 
        named_formulas = self.extract_named_formulas(workbook)
        named_tables = self.extract_named_tables(workbook)
        local_objects = self.extract_local_objects(workbook)
        
        # Build complete structure
        objects_dict = {
            'metadata': {
                'export_date': datetime.now().isoformat(),
                'openpyxl_version': openpyxl.__version__,
                'total_objects': (len(named_ranges) + len(lambda_functions) + 
                                len(named_formulas) + len(named_tables))
            },
            'named_ranges': named_ranges,
            'lambda_functions': lambda_functions,
            'named_formulas': named_formulas,
            'named_tables': named_tables,
            'local_objects': local_objects,
            'export_summary': {
                'named_ranges': len(named_ranges),
                'lambda_functions': len(lambda_functions),
                'named_formulas': len(named_formulas),
                'named_tables': len(named_tables),
                'local_objects': sum(len(sheet_objs.get('named_ranges', [])) 
                                   for sheet_objs in local_objects.values()),
                'total_exported': (len(named_ranges) + len(lambda_functions) + 
                                 len(named_formulas) + len(named_tables))
            }
        }
        
        return objects_dict
    
    def extract_named_ranges(self, workbook) -> list:
        """Extract global named ranges (excluding lambdas and formulas)."""
        
        ranges = []
        
        # Use .values() to get the actual DefinedName objects
        for defined_name in workbook.defined_names.values():
            # Skip local objects (handled separately)
            if defined_name.localSheetId is not None:
                continue
                
            object_type = self.detect_object_type(defined_name)
            
            # Only include simple ranges and constants
            if object_type in ['range', 'constant']:
                range_obj = {
                    'name': defined_name.name,
                    'type': object_type,
                    'scope': 'global',
                    'local_sheet': None,
                    'description': self._generate_description(defined_name),
                    'definition': defined_name.attr_text,
                    'excel_definition': defined_name.attr_text
                }
                ranges.append(range_obj)
        
        return ranges
    
    def extract_lambda_functions(self, workbook) -> list:
        """Extract lambda functions with parameter parsing and translation."""
        
        lambda_funcs = []
        
        # Use .values() to get the actual DefinedName objects
        for defined_name in workbook.defined_names.values():
            # Skip local objects
            if defined_name.localSheetId is not None:
                continue
                
            if self.detect_object_type(defined_name) == 'lambda':
                excel_formula = defined_name.attr_text
                human_formula, parameters = self.translate_lambda_to_human(excel_formula)
                
                lambda_obj = {
                    'name': defined_name.name,
                    'type': 'lambda',
                    'scope': 'global',
                    'local_sheet': None,
                    'description': self._generate_description(defined_name),
                    'definition': human_formula,
                    'parameters': parameters,
                    'excel_definition': excel_formula
                }
                lambda_funcs.append(lambda_obj)
        
        return lambda_funcs
    
    def extract_named_formulas(self, workbook) -> list:
        """Extract named formulas (complex formulas that aren't lambdas)."""
        
        formulas = []
        
        # Use .values() to get the actual DefinedName objects
        for defined_name in workbook.defined_names.values():
            # Skip local objects
            if defined_name.localSheetId is not None:
                continue
                
            if self.detect_object_type(defined_name) == 'formula':
                formula_obj = {
                    'name': defined_name.name,
                    'type': 'formula',
                    'scope': 'global',
                    'local_sheet': None,
                    'description': self._generate_description(defined_name),
                    'definition': self._clean_formula_for_display(defined_name.attr_text),
                    'excel_definition': defined_name.attr_text
                }
                formulas.append(formula_obj)
        
        return formulas
    
    def extract_named_tables(self, workbook) -> list:
        """Extract table definitions with full properties."""
        
        tables = []
        
        for worksheet in workbook.worksheets:
            for table in worksheet.tables.values():
                table_obj = {
                    'name': table.displayName,
                    'type': 'table',
                    'scope': 'local',  # Tables are always worksheet-local
                    'local_sheet': worksheet.title,
                    'description': f"Excel table on sheet '{worksheet.title}'",
                    'table_properties': {
                        'display_name': table.displayName,
                        'range': f"{worksheet.title}!{table.ref}",
                        'has_headers': table.tableStyleInfo.showRowStripes if table.tableStyleInfo else True,
                        'total_row_shown': table.totalsRowShown if hasattr(table, 'totalsRowShown') else False,
                        'style_name': table.tableStyleInfo.name if table.tableStyleInfo else 'TableStyleMedium9',
                        'columns': self._extract_table_columns(worksheet, table),
                        'auto_filter': {
                            'enabled': table.autoFilter is not None,
                            'filter_definitions': []  # Could be expanded
                        },
                        'sort_state': None  # Could be expanded
                    }
                }
                tables.append(table_obj)
        
        return tables
    
    def extract_local_objects(self, workbook) -> dict:
        """Extract worksheet-specific named objects."""
        
        local_objects = {}

        for worksheet in workbook.worksheets:
            sheet_objects = {'named_ranges': []}

            # openpyxl 3.1 moved sheet-scoped names onto the worksheet. They no
            # longer appear in workbook.defined_names at all, so scanning the
            # workbook for a non-null localSheetId finds nothing.
            for defined_name in worksheet.defined_names.values():
                object_type = self.detect_object_type(defined_name)

                local_obj = {
                    'name': defined_name.name,
                    'type': object_type,
                    'scope': 'local',
                    'local_sheet': worksheet.title,
                    'description': self._generate_description(defined_name),
                    'definition': defined_name.attr_text,
                    'excel_definition': defined_name.attr_text
                }
                sheet_objects['named_ranges'].append(local_obj)

            # Only include sheets that have local objects
            if sheet_objects['named_ranges']:
                local_objects[worksheet.title] = sheet_objects

        return local_objects
    
    # =============================================================================
    # OBJECT TYPE DETECTION AND ANALYSIS
    # =============================================================================
    
    def detect_object_type(self, defined_name) -> str:
        """Classify defined name as 'lambda', 'formula', 'range', or 'constant'."""
        
        attr_text = defined_name.attr_text or ""
        
        # Check for lambda function
        if excel_lambda_detection_rgx.search(attr_text):
            return 'lambda'
        
        # Check for formula (contains function calls)
        if function_call_rgx.search(attr_text):
            return 'formula'
            
        # Check for range reference (contains sheet references or cell ranges)
        if '!' in attr_text or ':' in attr_text:
            return 'range'
            
        # Everything else is a constant
        return 'constant'
    
    # =============================================================================
    # LAMBDA TRANSLATION METHODS
    # =============================================================================
    
    def translate_lambda_to_human(self, excel_formula: str) -> tuple:
        """Convert '_xlfn.LAMBDA(_xlpm.param,...)' to 'LAMBDA(param,...)'."""
        
        # Guard clause for input validation
        if not isinstance(excel_formula, str):
            excel_formula = str(excel_formula) if excel_formula else ""
        
        # Extract parameters
        params_match = excel_lambda_params_rgx.search(excel_formula)
        if not params_match:
            return excel_formula, []
        
        # Get parameter names without _xlpm prefix
        param_names = excel_param_name_rgx.findall(params_match.group(1))
        
        # Extract the lambda body
        body_match = excel_lambda_body_rgx.search(excel_formula)
        if not body_match:
            return excel_formula, param_names
        
        lambda_body = body_match.group(2)
        
        # Clean up the body - remove Excel prefixes
        clean_body = self._clean_formula_for_display(lambda_body)
        
        # Build human-readable format
        param_list = ', '.join(param_names)
        human_formula = f"LAMBDA({param_list}, {clean_body})"
        
        return human_formula, param_names
    
    def translate_lambda_to_excel(self, human_formula: str, parameters) -> str:
        """Convert 'LAMBDA(param,...)' to '=_xlfn.LAMBDA(_xlpm.param,...)'."""
        
        # Guard clauses for input validation
        if not isinstance(human_formula, str):
            raise StepProcessorError(f"Human formula must be string, got {type(human_formula)}")
        
        if not isinstance(parameters, list):
            if hasattr(parameters, '__iter__') and not isinstance(parameters, str):
                parameters = list(parameters)
            else:
                raise StepProcessorError(f"Parameters must be list-like, got {type(parameters)}")
        
        # Extract the body from human format
        body_match = human_lambda_body_rgx.search(human_formula)
        if not body_match:
            raise StepProcessorError(f"Invalid lambda format: {human_formula}")
        
        lambda_body = body_match.group(2)
        
        # Add Excel prefixes to the body
        excel_body = self._add_excel_prefixes(lambda_body, parameters)
        
        # Build parameter list with _xlpm prefixes (no spaces)
        excel_params = ','.join(f"_xlpm.{param}" for param in parameters)
        
        # Build Excel format with = prefix and no spaces after commas
        excel_formula = f"=_xlfn.LAMBDA({excel_params},{excel_body})"
        
        return excel_formula
    
    def _clean_formula_for_display(self, formula: str) -> str:
        """Remove Excel internal prefixes for human-readable display."""
        
        # Remove _xlfn and _xlpm prefixes
        clean_formula = excel_prefix_cleanup_rgx.sub('', formula)
        
        return clean_formula.strip()
    
    def _add_excel_prefixes(self, formula: str, parameters) -> str:
        """Add Excel internal prefixes to formula for Excel compatibility."""
        
        # Guard clauses for input validation
        if not isinstance(formula, str):
            formula = str(formula) if formula else ""
        
        if not isinstance(parameters, list):
            if hasattr(parameters, '__iter__') and not isinstance(parameters, str):
                parameters = list(parameters)
            else:
                parameters = []
        
        # Add _xlpm prefix to parameter references
        for param in parameters:
            if isinstance(param, str) and param:
                formula = re.sub(rf'\b{param}\b', f'_xlpm.{param}', formula)
        
        # Add _xlfn prefix to functions and remove spaces in function calls
        common_functions = [
            'SUM', 'SUMIF', 'SUMIFS', 'AVERAGE', 'COUNT', 'COUNTA', 'VLOOKUP',
            'XLOOKUP', 'INDEX', 'MATCH', 'IF', 'IFS', 'AND', 'OR', 'NOT',
            'PV', 'FV', 'PMT', 'RATE', 'NPER', 'NPV', 'IRR', 'XIRR',
            'MAX', 'MIN', 'ABS', 'ROUND', 'ROUNDUP', 'ROUNDDOWN',
            'LEFT', 'RIGHT', 'MID', 'LEN', 'TRIM', 'UPPER', 'LOWER',
            'CONCATENATE', 'TEXTJOIN', 'SUBSTITUTE', 'REPLACE'
        ]
        
        for func in common_functions:
            # First add _xlfn prefix
            formula = re.sub(rf'\b{func}\s*\(', f'_xlfn.{func}(', formula, flags=re.IGNORECASE)
        
        # Remove spaces after commas in function calls (Excel format)
        formula = re.sub(r',\s+', ',', formula)
        
        return formula
    
    # =============================================================================
    # UTILITY METHODS
    # =============================================================================
    
    def _generate_description(self, defined_name) -> str:
        """Generate a description for a defined name based on its properties."""
        
        object_type = self.detect_object_type(defined_name)
        name = defined_name.name
        
        if object_type == 'lambda':
            return f"Lambda function: {name}"
        elif object_type == 'formula':
            return f"Named formula: {name}"
        elif object_type == 'range':
            return f"Named range: {name}"
        else:
            return f"Named constant: {name}"
    
    def _extract_table_columns(self, worksheet, table) -> list:
        """Extract column definitions from Excel table."""
        
        # Get table range
        table_range = worksheet[table.ref]
        
        # Get header row (first row of table)
        if hasattr(table_range, '__iter__') and table_range:
            header_row = next(iter(table_range))
            if not hasattr(header_row, '__iter__'):
                header_row = [header_row]
        else:
            return []
        
        columns = []
        for cell in header_row:
            if cell.value:
                columns.append({
                    'name': str(cell.value),
                    'data_type': 'text'  # Could be enhanced to detect types
                })
        
        return columns
    
    def apply_name_filters(self, objects, include_patterns=None, exclude_patterns=None) -> list:
        """Filter objects by name patterns."""
        
        # Guard clauses for input validation
        if not isinstance(objects, list):
            if hasattr(objects, '__iter__'):
                objects = list(objects)
            else:
                raise StepProcessorError(f"Objects must be list-like, got {type(objects)}")
        
        if include_patterns is not None and not isinstance(include_patterns, list):
            if hasattr(include_patterns, '__iter__') and not isinstance(include_patterns, str):
                include_patterns = list(include_patterns)
            else:
                include_patterns = [include_patterns] if include_patterns else None
        
        if exclude_patterns is not None and not isinstance(exclude_patterns, list):
            if hasattr(exclude_patterns, '__iter__') and not isinstance(exclude_patterns, str):
                exclude_patterns = list(exclude_patterns)
            else:
                exclude_patterns = [exclude_patterns] if exclude_patterns else None
        
        if not include_patterns and not exclude_patterns:
            return objects
        
        filtered = []
        
        for obj in objects:
            if not isinstance(obj, dict) or 'name' not in obj:
                continue
                
            name = obj['name']
            include = True
            
            # Check include patterns
            if include_patterns:
                include = any(self._match_pattern(name, pattern) for pattern in include_patterns)
            
            # Check exclude patterns
            if exclude_patterns and include:
                include = not any(self._match_pattern(name, pattern) for pattern in exclude_patterns)
            
            if include:
                filtered.append(obj)
        
        return filtered
    
    def _match_pattern(self, name: str, pattern: str) -> bool:
        """Match name against pattern (supports * and ? wildcards)."""
        
        # Convert shell-style wildcards to regex
        regex_pattern = pattern.replace('*', '.*').replace('?', '.')
        regex_pattern = f'^{regex_pattern}$'
        
        return bool(re.match(regex_pattern, name, re.IGNORECASE))
    
    # =============================================================================
    # YAML OPERATIONS
    # =============================================================================
    
    def export_to_yaml(self, objects_dict: dict, export_path: str) -> None:
        """Export structured objects to YAML format."""
        
        # Ensure export directory exists
        export_file = Path(export_path)
        export_file.parent.mkdir(parents=True, exist_ok=True)
        
        # Write YAML with proper formatting
        with open(export_file, 'w', encoding='utf-8') as f:
            # Write header comment
            f.write("# Excel Named Objects Export\n")
            f.write(f"# Export date: {objects_dict['metadata']['export_date']}\n")
            f.write(f"# Total objects: {objects_dict['metadata']['total_objects']}\n\n")
            
            # Write YAML content
            yaml.dump(objects_dict, f, default_flow_style=False, allow_unicode=True,
                     sort_keys=False, indent=2)
        
        logger.info(f"Exported {objects_dict['metadata']['total_objects']} objects to {export_path}")
    
    def export_to_vba_format(self, objects_dict: dict, export_path: str) -> None:
        """Export to VBA-compatible simple text format."""
        
        # Ensure export directory exists
        export_file = Path(export_path)
        export_file.parent.mkdir(parents=True, exist_ok=True)
        
        with open(export_file, 'w', encoding='utf-8') as f:
            # Header
            f.write("# Excel Named Objects Export (VBA Compatible)\n")
            f.write(f"# Generated: {objects_dict['metadata']['export_date']}\n")
            f.write(f"# Source: {objects_dict['metadata'].get('source_file', 'unknown')}\n")
            f.write(f"# Total objects: {objects_dict['metadata']['total_objects']}\n\n")
            
            # Named Ranges
            if objects_dict['named_ranges']:
                f.write("[NAMED_RANGES]\n")
                for obj in objects_dict['named_ranges']:
                    line = f"{obj['name']} | {obj['definition']} | {obj['type']} | {obj['scope']} | {obj['description']}\n"
                    f.write(line)
                f.write("\n")
            
            # Lambda Functions (human-readable format)
            if objects_dict['lambda_functions']:
                f.write("[LAMBDA_FUNCTIONS]\n")
                for obj in objects_dict['lambda_functions']:
                    line = f"{obj['name']} | {obj['definition']} | {obj['type']} | {obj['scope']} | {obj['description']}\n"
                    f.write(line)
                f.write("\n")
            
            # Named Formulas
            if objects_dict['named_formulas']:
                f.write("[NAMED_FORMULAS]\n")
                for obj in objects_dict['named_formulas']:
                    line = f"{obj['name']} | {obj['definition']} | {obj['type']} | {obj['scope']} | {obj['description']}\n"
                    f.write(line)
                f.write("\n")
            
            # Named Tables
            if objects_dict['named_tables']:
                f.write("[NAMED_TABLES]\n")
                for obj in objects_dict['named_tables']:
                    props = obj['table_properties']
                    columns = ','.join([col['name'] for col in props['columns']])
                    scope = f"local:{obj['local_sheet']}"
                    style = props.get('style_name', 'TableStyleMedium9')
                    line = f"{obj['name']} | {props['range']} | {obj['type']} | {scope} | {obj['description']} | {style} | {columns}\n"
                    f.write(line)
                f.write("\n")
            
            # Local Ranges
            local_objects = objects_dict.get('local_objects', {})
            if local_objects:
                f.write("[LOCAL_RANGES]\n")
                for sheet_name, sheet_objs in local_objects.items():
                    for obj in sheet_objs.get('named_ranges', []):
                        scope = f"local:{sheet_name}"
                        line = f"{obj['name']} | {obj['definition']} | {obj['type']} | {scope} | {obj['description']}\n"
                        f.write(line)
                f.write("\n")
        
        logger.info(f"Exported {objects_dict['metadata']['total_objects']} objects to VBA format: {export_path}")
    
    def import_from_yaml(self, import_path: str) -> dict:
        """Import and validate objects from YAML format."""
        
        import_file = Path(import_path)
        if not import_file.exists():
            raise StepProcessorError(f"Import file not found: {import_path}")
        
        try:
            with open(import_file, 'r', encoding='utf-8') as f:
                objects_dict = yaml.safe_load(f)
        except yaml.YAMLError as e:
            raise StepProcessorError(f"Invalid YAML format in {import_path}: {e}")
        
        # Validate structure
        self._validate_yaml_structure(objects_dict)
        
        return objects_dict
    
    def _validate_yaml_structure(self, objects_dict: dict) -> None:
        """Validate YAML structure for import compatibility."""
        
        required_sections = ['metadata', 'named_ranges', 'lambda_functions', 
                           'named_formulas', 'named_tables']
        
        for section in required_sections:
            if section not in objects_dict:
                raise StepProcessorError(f"Missing required section: {section}")
        
        # Validate each object has required fields
        all_objects = (objects_dict['named_ranges'] + objects_dict['lambda_functions'] + 
                      objects_dict['named_formulas'] + objects_dict['named_tables'])
        
        for obj in all_objects:
            required_fields = ['name', 'type', 'scope']
            for field in required_fields:
                if field not in obj:
                    raise StepProcessorError(
                        f"Object '{obj.get('name', 'unknown')}' missing required field: {field}"
                    )
    
    # =============================================================================
    # EXECUTION METHODS
    # =============================================================================
    
    def _execute_export_all(self) -> dict:
        """Execute export all operation."""
        
        source_file = self.get_config_value('source_file')
        export_file = self.get_config_value('export_file')
        vba_file = self.get_config_value('vba_file')
        export_formats = self.get_config_value('export_formats')
        
        # Handle multiple file format configuration
        yaml_export_path = None
        vba_export_path = None
        
        if export_formats:
            # New format with explicit format specification
            yaml_export_path = export_formats.get('yaml_file')
            vba_export_path = export_formats.get('vba_file')
        else:
            # Legacy single file format
            if export_file:
                yaml_export_path = export_file
            if vba_file:
                vba_export_path = vba_file
        
        if not source_file:
            raise StepProcessorError("source_file required for export_all operation")
        if not yaml_export_path and not vba_export_path:
            raise StepProcessorError("At least one export file (export_file, vba_file, or export_formats) required")
        
        # Load workbook
        workbook = openpyxl.load_workbook(source_file, data_only=False)
        
        try:
            # Extract all objects
            objects_dict = self.extract_all_named_objects(workbook)
            objects_dict['metadata']['source_file'] = str(source_file)
            
            exports_completed = []
            
            # Export to YAML if requested
            if yaml_export_path:
                self.export_to_yaml(objects_dict, yaml_export_path)
                exports_completed.append(f"YAML: {yaml_export_path}")
            
            # Export to VBA format if requested
            if vba_export_path:
                self.export_to_vba_format(objects_dict, vba_export_path)
                exports_completed.append(f"VBA: {vba_export_path}")
            
            return {
                'operation': 'export_all',
                'source_file': str(source_file),
                'exports_completed': exports_completed,
                'objects_exported': objects_dict['metadata']['total_objects'],
                'summary': objects_dict['export_summary']
            }
            
        finally:
            workbook.close()
    
    def _execute_list_objects(self) -> dict:
        """Execute list objects operation for inventory."""
        
        source_file = self.get_config_value('source_file')
        if not source_file:
            raise StepProcessorError("source_file required for list_objects operation")
        
        workbook = openpyxl.load_workbook(source_file, data_only=False)
        
        try:
            objects_dict = self.extract_all_named_objects(workbook)
            
            # Build inventory summary
            inventory = {
                'source_file': str(source_file),
                'total_objects': objects_dict['metadata']['total_objects'],
                'by_type': objects_dict['export_summary'],
                'object_names': {
                    'named_ranges': [obj['name'] for obj in objects_dict['named_ranges']],
                    'lambda_functions': [obj['name'] for obj in objects_dict['lambda_functions']],
                    'named_formulas': [obj['name'] for obj in objects_dict['named_formulas']],
                    'named_tables': [obj['name'] for obj in objects_dict['named_tables']]
                }
            }
            
            return inventory
            
        finally:
            workbook.close()
    
    # =============================================================================
    # WRITE OPERATIONS - SHARED MACHINERY
    # =============================================================================

    def _get_write_policies(self, default_validation: str = 'excel') -> tuple:
        """
        Read and validate the two policies that govern every write.

        Args:
            default_validation: Validation level when the recipe does not say

        Returns:
            Tuple of (on_existing, name_validation)
        """
        on_existing = self.get_config_value('on_existing', 'error')
        name_validation = self.get_config_value('name_validation', default_validation)

        if on_existing not in self.EXISTING_NAME_POLICIES:
            valid = ', '.join(self.EXISTING_NAME_POLICIES)
            raise StepProcessorError(
                f"Invalid on_existing '{on_existing}'. Supported: {valid}"
            )

        if name_validation not in self.NAME_VALIDATION_LEVELS:
            valid = ', '.join(self.NAME_VALIDATION_LEVELS)
            raise StepProcessorError(
                f"Invalid name_validation '{name_validation}'. Supported: {valid}"
            )

        return on_existing, name_validation

    def _check_name(self, name: str, name_validation: str) -> None:
        """
        Validate a name against the configured level.

        Args:
            name:               Candidate defined name
            name_validation:    'none', 'excel', or 'house'

        Raises:
            StepProcessorError: If the name is unacceptable at this level
        """
        if name_validation == 'none':
            return

        try:
            validate_defined_name(name, enforce_house_style=(name_validation == 'house'))
        except DefinedNameError as error:
            raise StepProcessorError(str(error))

    def _name_exists(self, workbook, name: str, sheet_name=None) -> bool:
        """Report whether a name is already defined at the given scope."""
        if sheet_name is None:
            return name in workbook.defined_names

        if sheet_name not in workbook.sheetnames:
            return False

        return name in workbook[sheet_name].defined_names

    def _remove_name(self, workbook, name: str, sheet_name=None) -> None:
        """Delete an existing name at the given scope."""
        if sheet_name is None:
            del workbook.defined_names[name]
        else:
            del workbook[sheet_name].defined_names[name]

    def write_named_object(self, workbook, name: str, definition: str,
                           sheet_name=None, on_existing: str = 'error',
                           name_validation: str = 'excel') -> str:
        """
        Write a single defined name into a workbook.

        Passing sheet_name makes the name sheet-scoped. Leaving it None makes
        the name workbook-scoped, which is what formulas on other sheets need.

        Args:
            workbook:           openpyxl workbook, opened for writing
            name:               Defined name to create
            definition:         Reference or formula text
            sheet_name:         Sheet for local scope, or None for global
            on_existing:        'error', 'replace', or 'skip'
            name_validation:    'none', 'excel', or 'house'

        Returns:
            'written', 'replaced', or 'skipped'
        """
        if not isinstance(name, str) or not name.strip():
            raise StepProcessorError("Defined name must be a non-empty string")

        if not isinstance(definition, str) or not definition.strip():
            raise StepProcessorError(f"Definition for '{name}' must be a non-empty string")

        clean_name = name.strip()
        clean_definition = definition.strip()

        self._check_name(clean_name, name_validation)

        if sheet_name is not None and sheet_name not in workbook.sheetnames:
            raise StepProcessorError(
                f"Cannot scope '{clean_name}' to sheet '{sheet_name}': "
                f"sheet not found. Available: {workbook.sheetnames}"
            )

        outcome = 'written'

        if self._name_exists(workbook, clean_name, sheet_name):
            scope_label = 'global' if sheet_name is None else f"sheet '{sheet_name}'"

            if on_existing == 'error':
                raise StepProcessorError(
                    f"Defined name '{clean_name}' already exists at {scope_label} scope. "
                    f"Set on_existing to 'replace' or 'skip' to proceed."
                )

            if on_existing == 'skip':
                logger.debug(f"Skipping existing name '{clean_name}' at {scope_label} scope")
                return 'skipped'

            self._remove_name(workbook, clean_name, sheet_name)
            outcome = 'replaced'

        defined_name = DefinedName(clean_name, attr_text=clean_definition)

        if sheet_name is None:
            workbook.defined_names[clean_name] = defined_name
        else:
            workbook[sheet_name].defined_names[clean_name] = defined_name

        logger.debug(f"{outcome.title()} '{clean_name}' -> {clean_definition}")
        return outcome

    def _definition_for_write(self, obj: dict) -> str:
        """
        Choose the text to write for an extracted object.

        Prefers 'excel_definition' because that is the round-trip-faithful form.
        Lambdas edited by hand in the YAML carry a human-readable 'definition'
        instead, and are translated back on the way in.

        Args:
            obj: One object from an export structure

        Returns:
            Definition text ready for Excel
        """
        object_type = obj.get('type', 'range')

        if object_type == 'lambda':
            human = obj.get('definition')
            parameters = obj.get('parameters', [])

            if isinstance(human, str) and human_lambda_detection_rgx.search(human):
                return self.translate_lambda_to_excel(human, parameters)

        excel_definition = obj.get('excel_definition')
        if isinstance(excel_definition, str) and excel_definition.strip():
            return excel_definition

        definition = obj.get('definition')
        if isinstance(definition, str) and definition.strip():
            return definition

        raise StepProcessorError(
            f"Object '{obj.get('name', 'unknown')}' has no usable definition"
        )

    def write_objects_to_workbook(self, workbook, objects_dict: dict,
                                  on_existing: str = 'error',
                                  name_validation: str = 'excel',
                                  include_local: bool = True,
                                  include_patterns=None,
                                  exclude_patterns=None) -> dict:
        """
        Write every named object in an export structure into a workbook.

        Tables are not created. Rebuilding a table requires its data region to
        already exist with matching dimensions, which this processor has no way
        to guarantee, so tables are reported as skipped rather than guessed at.

        Args:
            workbook:           openpyxl workbook, opened for writing
            objects_dict:       Structure produced by extract_all_named_objects
            on_existing:        'error', 'replace', or 'skip'
            name_validation:    'none', 'excel', or 'house'
            include_local:      Also write sheet-scoped names
            include_patterns:   Wildcard patterns to include
            exclude_patterns:   Wildcard patterns to exclude

        Returns:
            Dictionary of counts and per-name outcomes
        """
        results = {
            'written': [],
            'replaced': [],
            'skipped': [],
            'skipped_tables': [],
            'failed': []
        }

        for section in self.OBJECT_SECTIONS:
            objects = objects_dict.get(section, [])

            if not isinstance(objects, list):
                continue

            for obj in self.apply_name_filters(objects, include_patterns, exclude_patterns):
                self._write_object_recording_result(
                    workbook, obj, None, on_existing, name_validation, results
                )

        if include_local:
            local_objects = objects_dict.get('local_objects', {})

            if isinstance(local_objects, dict):
                for sheet_name, sheet_objects in local_objects.items():
                    objects = sheet_objects.get('named_ranges', [])
                    filtered = self.apply_name_filters(
                        objects, include_patterns, exclude_patterns
                    )

                    for obj in filtered:
                        self._write_object_recording_result(
                            workbook, obj, sheet_name, on_existing, name_validation, results
                        )

        tables = objects_dict.get('named_tables', [])
        if isinstance(tables, list):
            for table in self.apply_name_filters(tables, include_patterns, exclude_patterns):
                results['skipped_tables'].append(table.get('name', 'unknown'))

        if results['skipped_tables']:
            logger.warning(
                f"Skipped {len(results['skipped_tables'])} table definitions: table "
                f"creation requires a matching data region and is not attempted"
            )

        return results

    def _write_object_recording_result(self, workbook, obj: dict, sheet_name,
                                       on_existing: str, name_validation: str,
                                       results: dict) -> None:
        """Write one object and record the outcome in the results dictionary."""
        name = obj.get('name', 'unknown')

        try:
            definition = self._definition_for_write(obj)
            outcome = self.write_named_object(
                workbook, name, definition, sheet_name, on_existing, name_validation
            )

            if outcome == 'written':
                results['written'].append(name)
            elif outcome == 'replaced':
                results['replaced'].append(name)
            else:
                results['skipped'].append(name)

        except StepProcessorError as error:
            if on_existing == 'error':
                raise
            logger.warning(f"Could not write '{name}': {error}")
            results['failed'].append({'name': name, 'reason': str(error)})

    def _load_target_workbook(self, target_file: str):
        """Open a workbook for writing, with a clear error when it is absent."""
        target_path = Path(target_file)

        if not target_path.exists():
            raise StepProcessorError(
                f"Target file not found: {target_file}. Named objects are written "
                f"into an existing workbook, so run export_file first."
            )

        return openpyxl.load_workbook(target_path, data_only=False)

    # =============================================================================
    # WRITE OPERATIONS - EXECUTION METHODS
    # =============================================================================

    def _execute_export_filtered(self) -> dict:
        """Export named objects matching the configured name patterns."""

        source_file = self.get_config_value('source_file')
        export_file = self.get_config_value('export_file')
        vba_file = self.get_config_value('vba_file')
        include_patterns = self.get_config_value('include_patterns')
        exclude_patterns = self.get_config_value('exclude_patterns')

        if not source_file:
            raise StepProcessorError("source_file required for export_filtered operation")

        if not export_file and not vba_file:
            raise StepProcessorError(
                "At least one of export_file or vba_file required for export_filtered"
            )

        if not include_patterns and not exclude_patterns:
            raise StepProcessorError(
                "export_filtered requires include_patterns or exclude_patterns. "
                "Use export_all to export everything."
            )

        workbook = openpyxl.load_workbook(source_file, data_only=False)

        try:
            objects_dict = self.extract_all_named_objects(workbook)
            objects_dict['metadata']['source_file'] = str(source_file)

            total_kept = 0

            for section in self.OBJECT_SECTIONS + ['named_tables']:
                kept = self.apply_name_filters(
                    objects_dict.get(section, []), include_patterns, exclude_patterns
                )
                objects_dict[section] = kept
                total_kept += len(kept)

            local_objects = objects_dict.get('local_objects', {})
            filtered_local = {}

            for sheet_name, sheet_objects in local_objects.items():
                kept = self.apply_name_filters(
                    sheet_objects.get('named_ranges', []), include_patterns, exclude_patterns
                )
                if kept:
                    filtered_local[sheet_name] = {'named_ranges': kept}
                    total_kept += len(kept)

            objects_dict['local_objects'] = filtered_local
            objects_dict['metadata']['total_objects'] = total_kept
            objects_dict['export_summary'] = {
                section: len(objects_dict.get(section, []))
                for section in self.OBJECT_SECTIONS + ['named_tables']
            }
            objects_dict['export_summary']['total_exported'] = total_kept

            exports_completed = []

            if export_file:
                self.export_to_yaml(objects_dict, export_file)
                exports_completed.append(f"YAML: {export_file}")

            if vba_file:
                self.export_to_vba_format(objects_dict, vba_file)
                exports_completed.append(f"VBA: {vba_file}")

            return {
                'operation': 'export_filtered',
                'source_file': str(source_file),
                'exports_completed': exports_completed,
                'objects_exported': total_kept,
                'summary': objects_dict['export_summary']
            }

        finally:
            workbook.close()

    def _execute_import_all(self) -> dict:
        """Create every named object described by a YAML file."""
        return self._import_from_yaml_file(filtered=False)

    def _execute_import_filtered(self) -> dict:
        """Create the named objects from a YAML file that match name patterns."""
        return self._import_from_yaml_file(filtered=True)

    def _import_from_yaml_file(self, filtered: bool) -> dict:
        """Shared implementation for import_all and import_filtered."""

        import_file = self.get_config_value('import_file')
        target_file = self.get_config_value('target_file')
        include_local = self.get_config_value('include_local', True)
        include_patterns = self.get_config_value('include_patterns')
        exclude_patterns = self.get_config_value('exclude_patterns')

        operation_name = 'import_filtered' if filtered else 'import_all'

        if not import_file:
            raise StepProcessorError(f"import_file required for {operation_name} operation")

        if not target_file:
            raise StepProcessorError(f"target_file required for {operation_name} operation")

        if filtered and not include_patterns and not exclude_patterns:
            raise StepProcessorError(
                "import_filtered requires include_patterns or exclude_patterns. "
                "Use import_all to import everything."
            )

        if not filtered:
            include_patterns = None
            exclude_patterns = None

        on_existing, name_validation = self._get_write_policies('excel')

        objects_dict = self.import_from_yaml(import_file)
        workbook = self._load_target_workbook(target_file)

        try:
            results = self.write_objects_to_workbook(
                workbook, objects_dict, on_existing, name_validation,
                include_local, include_patterns, exclude_patterns
            )

            workbook.save(target_file)

        finally:
            workbook.close()

        logger.info(
            f"Imported named objects into '{target_file}': "
            f"{len(results['written'])} written, {len(results['replaced'])} replaced, "
            f"{len(results['skipped'])} skipped"
        )

        return {
            'operation': operation_name,
            'import_file': str(import_file),
            'target_file': str(target_file),
            'written': results['written'],
            'replaced': results['replaced'],
            'skipped': results['skipped'],
            'skipped_tables': results['skipped_tables'],
            'failed': results['failed'],
            'objects_written': len(results['written']) + len(results['replaced'])
        }

    def _execute_validate_yaml(self) -> dict:
        """Check a YAML export for structural and naming problems without writing."""

        import_file = self.get_config_value('import_file')

        if not import_file:
            raise StepProcessorError("import_file required for validate_yaml operation")

        name_validation = self.get_config_value('name_validation', 'excel')

        if name_validation not in self.NAME_VALIDATION_LEVELS:
            valid = ', '.join(self.NAME_VALIDATION_LEVELS)
            raise StepProcessorError(
                f"Invalid name_validation '{name_validation}'. Supported: {valid}"
            )

        objects_dict = self.import_from_yaml(import_file)

        problems = []
        checked = 0
        seen_global = set()

        for section in self.OBJECT_SECTIONS + ['named_tables']:
            for obj in objects_dict.get(section, []):
                checked += 1
                name = obj.get('name', 'unknown')

                try:
                    self._check_name(name, name_validation)
                except StepProcessorError as error:
                    problems.append({'name': name, 'section': section, 'problem': str(error)})

                if name in seen_global:
                    problems.append({
                        'name': name,
                        'section': section,
                        'problem': 'Duplicate name at global scope'
                    })
                seen_global.add(name)

                if section != 'named_tables':
                    try:
                        self._definition_for_write(obj)
                    except StepProcessorError as error:
                        problems.append({
                            'name': name, 'section': section, 'problem': str(error)
                        })

        for sheet_name, sheet_objects in objects_dict.get('local_objects', {}).items():
            for obj in sheet_objects.get('named_ranges', []):
                checked += 1
                name = obj.get('name', 'unknown')

                try:
                    self._check_name(name, name_validation)
                except StepProcessorError as error:
                    problems.append({
                        'name': name, 'section': f"local:{sheet_name}", 'problem': str(error)
                    })

        return {
            'operation': 'validate_yaml',
            'import_file': str(import_file),
            'objects_checked': checked,
            'valid': len(problems) == 0,
            'problems': problems
        }

    def _execute_copy_direct(self) -> dict:
        """Copy named objects straight from one workbook into another."""

        source_file = self.get_config_value('source_file')
        target_file = self.get_config_value('target_file')
        include_local = self.get_config_value('include_local', True)
        include_patterns = self.get_config_value('include_patterns')
        exclude_patterns = self.get_config_value('exclude_patterns')

        if not source_file:
            raise StepProcessorError("source_file required for copy_direct operation")

        if not target_file:
            raise StepProcessorError("target_file required for copy_direct operation")

        on_existing, name_validation = self._get_write_policies('excel')

        source_workbook = openpyxl.load_workbook(source_file, data_only=False)

        try:
            objects_dict = self.extract_all_named_objects(source_workbook)
        finally:
            source_workbook.close()

        target_workbook = self._load_target_workbook(target_file)

        try:
            results = self.write_objects_to_workbook(
                target_workbook, objects_dict, on_existing, name_validation,
                include_local, include_patterns, exclude_patterns
            )

            target_workbook.save(target_file)

        finally:
            target_workbook.close()

        logger.info(
            f"Copied named objects from '{source_file}' to '{target_file}': "
            f"{len(results['written'])} written, {len(results['replaced'])} replaced"
        )

        return {
            'operation': 'copy_direct',
            'source_file': str(source_file),
            'target_file': str(target_file),
            'written': results['written'],
            'replaced': results['replaced'],
            'skipped': results['skipped'],
            'skipped_tables': results['skipped_tables'],
            'failed': results['failed'],
            'objects_written': len(results['written']) + len(results['replaced'])
        }

    def _execute_create_from_columns(self) -> dict:
        """
        Build named ranges from column names on generated data.

        This is the operation that makes named ranges survive a regenerated
        file. Rather than copying a fixed address, each range is recomputed
        from the actual data extent on every run, so a range cannot fall behind
        the data the way a hand-maintained one does.
        """

        target_file = self.get_config_value('target_file')
        range_specs = self.get_config_value('ranges')

        if not target_file:
            raise StepProcessorError("target_file required for create_from_columns operation")

        if not isinstance(range_specs, list) or len(range_specs) == 0:
            raise StepProcessorError(
                "create_from_columns requires a non-empty 'ranges' list"
            )

        on_existing, name_validation = self._get_write_policies('house')

        workbook = self._load_target_workbook(target_file)

        created = []
        replaced = []
        skipped = []

        try:
            for index, spec in enumerate(range_specs):
                if not isinstance(spec, dict):
                    raise StepProcessorError(
                        f"Range entry {index + 1} must be a dictionary"
                    )

                name = spec.get('name')
                sheet = spec.get('sheet')
                columns = spec.get('columns')

                if not name:
                    raise StepProcessorError(f"Range entry {index + 1} missing 'name'")

                if not sheet:
                    raise StepProcessorError(f"Range '{name}' missing 'sheet'")

                if not isinstance(columns, list) or len(columns) == 0:
                    raise StepProcessorError(
                        f"Range '{name}' requires a non-empty 'columns' list"
                    )

                if sheet not in workbook.sheetnames:
                    raise StepProcessorError(
                        f"Range '{name}' targets sheet '{sheet}', which was not found. "
                        f"Available: {workbook.sheetnames}"
                    )

                worksheet = workbook[sheet]

                try:
                    reference = resolve_range(
                        worksheet,
                        columns,
                        row_mode=spec.get('row_mode', 'data_with_header'),
                        header_row=spec.get('header_row', 1),
                        anchor_columns=spec.get('anchor_columns'),
                        expand_span=spec.get('expand_span', True),
                        force_column_names=spec.get('force_column_names', False),
                        on_missing=spec.get('on_missing', 'error'),
                        absolute=spec.get('absolute', True),
                        qualify_sheet=True
                    )
                except ExcelRangeResolverError as error:
                    raise StepProcessorError(f"Could not resolve range '{name}': {error}")

                scope_sheet = sheet if spec.get('scope', 'global') == 'local' else None

                outcome = self.write_named_object(
                    workbook, name, reference, scope_sheet, on_existing, name_validation
                )

                record = {'name': name, 'definition': reference, 'sheet': sheet}

                if outcome == 'written':
                    created.append(record)
                elif outcome == 'replaced':
                    replaced.append(record)
                else:
                    skipped.append(record)

                logger.info(f"📐 {outcome.title()} '{name}' -> {reference}")

            workbook.save(target_file)

        finally:
            workbook.close()

        return {
            'operation': 'create_from_columns',
            'target_file': str(target_file),
            'created': created,
            'replaced': replaced,
            'skipped': skipped,
            'ranges_written': len(created) + len(replaced)
        }
    
    def get_operation_type(self) -> str:
        return "named_objects_management"
    
    def get_capabilities(self) -> dict:
        """Get processor capabilities information."""
        return {
            'description': 'Manage Excel named ranges, formulas, lambda functions, and tables',
            'operations': self.SUPPORTED_OPERATIONS,
            'read_operations': ['export_all', 'export_filtered', 'list_objects', 'validate_yaml'],
            'write_operations': ['import_all', 'import_filtered', 'copy_direct', 'create_from_columns'],
            'supported_object_types': ['range', 'constant', 'formula', 'lambda', 'table'],
            'scope_options': ['global', 'local'],
            'export_formats': ['yaml', 'vba_compatible'],
            'dual_export_support': True,
            'write_features': [
                'column_name_addressing', 'automatic_extent_detection',
                'defined_name_validation', 'existing_name_policies',
                'global_and_sheet_scope', 'lambda_round_trip'
            ],
            'existing_name_policies': self.EXISTING_NAME_POLICIES,
            'name_validation_levels': self.NAME_VALIDATION_LEVELS,
            'write_limitations': [
                'tables are not created; they require a matching data region'
            ],
            'lambda_features': [
                'parameter_extraction', 'excel_format_translation', 
                'human_readable_conversion', 'syntax_validation'
            ],
            'table_features': [
                'structure_preservation', 'column_definitions', 
                'style_information', 'filter_state'
            ],
            'filtering_options': [
                'name_patterns', 'object_types', 'scope_filtering', 
                'sheet_filtering', 'include_exclude_patterns'
            ],
            'vba_integration': [
                'simple_text_format', 'pipe_delimited_fields',
                'section_based_organization', 'vba_importable'
            ],
            'file_requirements': ['xlsx', 'xlsm'],
            'dependencies': ['openpyxl', 'pyyaml'],
            'examples': {
                'export_yaml': "Export all named objects to human-readable YAML file",
                'export_vba': "Export to VBA-compatible pipe-delimited text format",
                'dual_export': "Export to both YAML and VBA formats simultaneously",
                'lambda_translation': "Convert Excel LAMBDA functions to human-readable format",
                'table_extraction': "Extract complete table definitions with properties",
                'vba_workflow': "Enable Excel VBA import/export of named objects"
            }
        }


# End of file #

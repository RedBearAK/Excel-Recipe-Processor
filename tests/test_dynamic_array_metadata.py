"""
Tests for the dynamic-array metadata declaration pass.

tests/test_dynamic_array_metadata.py

Runnable with pytest, but written to run standalone and report a score.
Creates real xlsx artifacts and asserts on the raw bytes Excel will read:
cell markers, the metadata part, and the package registrations.
"""

import re
import zipfile
import tempfile

import openpyxl

from pathlib import Path

from excel_recipe_processor.core.dynamic_array_metadata import (
    EXCEL_METADATA_XML,
    DynamicArrayMetadataError,
    declare_dynamic_formulas_in_zip,
    save_workbook_with_declaration,
)


WORLD_REGION_FORMULA = (
    '=_xlfn.IFS(AND(BA2 = "Export", LEN(AV2) > 0), '
    '_xlfn.XLOOKUP(AV2, rng_country, rng_cont), TRUE, "ERROR")'
)
COUNTRY_FORMULA = '=IF(AY2=1,_xlfn.XLOOKUP(AT2,rng_EXdest,rng_country),"")'
SCALAR_SEARCH_FORMULA = '=IFERROR(SEARCH("fresh",AS2),0)'
SCALAR_COUNTIF_FORMULA = '=COUNTIF(A:A,B2)'


def build_workbook_file(work_dir, declare=False):
    """Write a workbook holding the real recipe formulas plus scalar controls."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A2'] = WORLD_REGION_FORMULA
    sheet['B2'] = COUNTRY_FORMULA
    sheet['C2'] = SCALAR_SEARCH_FORMULA
    sheet['D2'] = SCALAR_COUNTIF_FORMULA

    file_path = str(Path(work_dir) / 'built.xlsx')

    if declare:
        report = save_workbook_with_declaration(workbook, file_path)
    else:
        workbook.save(file_path)
        report = None

    workbook.close()
    return file_path, report


def read_cell_xml(file_path, cell_ref):
    """Return the raw <c ...>...</c> element for one cell, or ''."""
    with zipfile.ZipFile(file_path) as archive:
        xml = archive.read('xl/worksheets/sheet1.xml').decode('utf-8')

    match = re.search(rf'<c r="{cell_ref}".*?</c>', xml, re.DOTALL)
    return match.group(0) if match else ''


def test_dynamic_cells_get_the_full_declaration():
    """XLOOKUP-bearing cells gain cm="1" and t="array" ref; the part matches Excel's bytes."""
    print("\nTesting dynamic cells get the full declaration...")

    passed = True

    with tempfile.TemporaryDirectory() as work_dir:
        file_path, report = build_workbook_file(work_dir, declare=True)

        for cell_ref in ('A2', 'B2'):
            cell_xml = read_cell_xml(file_path, cell_ref)
            if f'cm="1"' in cell_xml and f'<f t="array" ref="{cell_ref}">' in cell_xml:
                print(f"  ✓ {cell_ref} carries cm=\"1\" and t=\"array\" ref=\"{cell_ref}\"")
            else:
                print(f"  ✗ {cell_ref} stored as: {cell_xml[:90]}")
                passed = False

        with zipfile.ZipFile(file_path) as archive:
            names = archive.namelist()
            metadata_ok = ('xl/metadata.xml' in names
                           and archive.read('xl/metadata.xml') == EXCEL_METADATA_XML)
            content_types = archive.read('[Content_Types].xml').decode('utf-8')
            rels = archive.read('xl/_rels/workbook.xml.rels').decode('utf-8')

        if metadata_ok:
            print("  ✓ xl/metadata.xml present and byte-identical to Excel's own")
        else:
            print("  ✗ xl/metadata.xml missing or differs from the Excel reference")
            passed = False

        if 'sheetMetadata+xml' in content_types:
            print("  ✓ [Content_Types].xml registers the part")
        else:
            print("  ✗ [Content_Types].xml missing the sheetMetadata override")
            passed = False

        if 'sheetMetadata' in rels:
            print("  ✓ workbook rels point at metadata.xml")
        else:
            print("  ✗ workbook rels missing the sheetMetadata relationship")
            passed = False

        if report['cells_marked'] == 2:
            print("  ✓ Report counts exactly the two dynamic cells")
        else:
            print(f"  ✗ Report counted {report['cells_marked']} marked cells, expected 2")
            passed = False

    return passed


def test_scalar_formulas_stay_untouched():
    """SEARCH and COUNTIF cells keep their plain storage, byte for byte."""
    print("\nTesting scalar formulas stay untouched...")

    passed = True

    with tempfile.TemporaryDirectory() as work_dir:
        plain_path, _ = build_workbook_file(work_dir, declare=False)
        declared_path = str(Path(work_dir) / 'declared.xlsx')
        declare_dynamic_formulas_in_zip(plain_path, declared_path)

        for cell_ref, formula_name in (('C2', 'SEARCH'), ('D2', 'COUNTIF')):
            before = read_cell_xml(plain_path, cell_ref)
            after = read_cell_xml(declared_path, cell_ref)
            if before == after and 'cm=' not in after:
                print(f"  ✓ {cell_ref} ({formula_name}) byte-identical, no marker")
            else:
                print(f"  ✗ {cell_ref} changed: {after[:90]}")
                passed = False

    return passed


def test_pass_is_idempotent_and_reopenable():
    """A second pass marks nothing new, and openpyxl reloads the result."""
    print("\nTesting the pass is idempotent and the file reopens...")

    passed = True

    with tempfile.TemporaryDirectory() as work_dir:
        file_path, _ = build_workbook_file(work_dir, declare=True)
        second_path = str(Path(work_dir) / 'second.xlsx')

        report = declare_dynamic_formulas_in_zip(file_path, second_path)

        if report['cells_marked'] == 0 and report['cells_already_declared'] == 2:
            print("  ✓ Second pass marks nothing; both cells recognized as declared")
        else:
            print(f"  ✗ Second pass: marked {report['cells_marked']}, "
                  f"already {report['cells_already_declared']}")
            passed = False

        if read_cell_xml(file_path, 'A2') == read_cell_xml(second_path, 'A2'):
            print("  ✓ A2 storage unchanged by the second pass")
        else:
            print("  ✗ A2 storage changed on the second pass")
            passed = False

        workbook = openpyxl.load_workbook(second_path)
        a2_value = workbook.active['A2'].value
        workbook.close()

        if a2_value is not None:
            print(f"  ✓ openpyxl reopens the declared file (A2 read as "
                  f"{type(a2_value).__name__})")
        else:
            print("  ✗ openpyxl lost the A2 formula on reload")
            passed = False

    return passed


def test_legacy_cse_cell_gets_completed():
    """A t="array" cell without cm (inherited {braces}) gains cm and keeps its ref."""
    print("\nTesting a legacy CSE cell gets completed into a dynamic one...")

    passed = True

    with tempfile.TemporaryDirectory() as work_dir:
        from openpyxl.worksheet.formula import ArrayFormula

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A2'] = ArrayFormula('A2', '=_xlfn.XLOOKUP(B2,C:C,D:D)')

        plain_path = str(Path(work_dir) / 'cse.xlsx')
        workbook.save(plain_path)
        workbook.close()

        declared_path = str(Path(work_dir) / 'cse_declared.xlsx')
        report = declare_dynamic_formulas_in_zip(plain_path, declared_path)

        cell_xml = read_cell_xml(declared_path, 'A2')

        if report['cells_completed'] == 1 and report['cells_marked'] == 0:
            print("  ✓ Reported as completed, not freshly marked")
        else:
            print(f"  ✗ Report: completed {report['cells_completed']}, "
                  f"marked {report['cells_marked']}")
            passed = False

        if 'cm="1"' in cell_xml and '<f t="array" ref="A2">' in cell_xml:
            print("  ✓ Cell now carries cm=\"1\" alongside its existing array marker")
        else:
            print(f"  ✗ Cell stored as: {cell_xml[:90]}")
            passed = False

    return passed


def test_unrecognized_metadata_part_fails_loud():
    """An alien xl/metadata.xml raises rather than being merged or clobbered."""
    print("\nTesting an unrecognized metadata part fails loud...")

    passed = True

    with tempfile.TemporaryDirectory() as work_dir:
        plain_path, _ = build_workbook_file(work_dir, declare=False)

        alien_path = str(Path(work_dir) / 'alien.xlsx')
        with zipfile.ZipFile(plain_path) as source:
            with zipfile.ZipFile(alien_path, 'w') as target:
                for name in source.namelist():
                    target.writestr(name, source.read(name))
                target.writestr('xl/metadata.xml', b'<metadata><somethingElse/></metadata>')

        try:
            declare_dynamic_formulas_in_zip(alien_path, str(Path(work_dir) / 'out.xlsx'))
            print("  ✗ Alien metadata part was accepted silently")
            passed = False
        except DynamicArrayMetadataError as error:
            if 'XLDAPR' in str(error):
                print("  ✓ Raised DynamicArrayMetadataError naming the XLDAPR expectation")
            else:
                print(f"  ✗ Raised, but with an unhelpful message: {error}")
                passed = False

    return passed


def test_provenance_marks_regardless_of_function():
    """Injected cells are declared even for pre-dynamic-array functions; identical
    unregistered cells are not, and sheet names resolve through workbook.xml."""
    print("\nTesting provenance-based marking...")

    passed = True

    with tempfile.TemporaryDirectory() as work_dir:
        workbook = openpyxl.Workbook()
        first_sheet = workbook.active
        first_sheet.title = 'Front'
        second_sheet = workbook.create_sheet('Data')

        # Identical IFS formulas: B2 is "recipe-injected", D2 is "inherited".
        # IFS predates dynamic arrays, so only provenance can mark it.
        second_sheet['B2'] = '=_xlfn.IFS(A2=1,"yes",TRUE,"no")'
        second_sheet['D2'] = '=_xlfn.IFS(A2=1,"yes",TRUE,"no")'

        plain_path = str(Path(work_dir) / 'prov.xlsx')
        workbook.save(plain_path)
        workbook.close()

        declared_path = str(Path(work_dir) / 'prov_declared.xlsx')
        report = declare_dynamic_formulas_in_zip(
            plain_path, declared_path,
            injected_cells={'Data': [('B', 2, 2)]},
        )

        with zipfile.ZipFile(declared_path) as archive:
            # 'Data' is the SECOND sheet - reading it via the resolver-declared
            # marks below proves name->part resolution rather than tab order.
            data_xml = archive.read('xl/worksheets/sheet2.xml').decode('utf-8')

        b2_match = re.search(r'<c r="B2"[^>]*>', data_xml)
        d2_match = re.search(r'<c r="D2"[^>]*>', data_xml)

        if b2_match and 'cm="1"' in b2_match.group(0):
            print("  ✓ Registered IFS cell B2 declared (provenance)")
        else:
            print(f"  ✗ B2 not declared: {b2_match.group(0) if b2_match else 'missing'}")
            passed = False

        if d2_match and 'cm="1"' not in d2_match.group(0):
            print("  ✓ Identical unregistered IFS cell D2 left alone")
        else:
            print(f"  ✗ D2 wrongly touched: {d2_match.group(0) if d2_match else 'missing'}")
            passed = False

        if report['cells_marked_injected'] == 1 and report['cells_marked'] == 1:
            print("  ✓ Report attributes the mark to injection provenance")
        else:
            print(f"  ✗ Report: marked {report['cells_marked']}, "
                  f"injected {report['cells_marked_injected']}")
            passed = False

        try:
            declare_dynamic_formulas_in_zip(
                plain_path, str(Path(work_dir) / 'bad.xlsx'),
                injected_cells={'NoSuchSheet': [('B', 2, 2)]},
            )
            print("  ✗ Unknown sheet name accepted silently")
            passed = False
        except DynamicArrayMetadataError:
            print("  ✓ Unknown sheet name in injected_cells fails loud")

    return passed


def main():
    tests = [
        test_dynamic_cells_get_the_full_declaration,
        test_scalar_formulas_stay_untouched,
        test_pass_is_idempotent_and_reopenable,
        test_legacy_cse_cell_gets_completed,
        test_unrecognized_metadata_part_fails_loud,
        test_provenance_marks_regardless_of_function,
    ]

    passed = 0

    for test_func in tests:
        try:
            if test_func():
                passed += 1
        except Exception as error:
            print(f"  ✗ {test_func.__name__} crashed: {error}")

    print(f"\n=== Results: {passed}/{len(tests)} tests passed ===")

    if passed == len(tests):
        print("✅ All dynamic-array metadata tests passed!")
        return 1

    print("❌ Some dynamic-array metadata tests failed!")
    return 0


if __name__ == '__main__':
    exit(0 if main() else 1)


# End of file #

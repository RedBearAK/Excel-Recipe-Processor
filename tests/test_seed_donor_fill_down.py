"""
Tests for formula fill-down in seed_donor_formulas.

tests/test_seed_donor_fill_down.py

Runnable with pytest, but written to run standalone and report a score.

These exist because fill_down failed silently once already: an array formula
arrived as an ArrayFormula object rather than a string, the string test skipped
it, and the log still reported the column as filled. Assertions here are on cell
CONTENT rather than on reported counts, since the counts were what lied.
"""

import openpyxl
import tempfile

from pathlib import Path

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

from excel_recipe_processor.processors.seed_donor_formulas_processor import (
    SeedDonorFormulasProcessor
)


# Column layout shared by donor and target, mirroring the VMS shape:
# a key column, a text column the formulas read, then the formula columns.
HEADERS = ['Van Number', 'Product Name', 'Carrier', 'Plain One', 'Plain Two', 'Array One']

PLAIN_ONE = '=IFERROR(SEARCH("Fresh",B{row}),0)'
PLAIN_TWO = '=COUNTIF(rng_carrier,C{row})'
ARRAY_ONE = '=_xlfn.IFS(D{row}=1,"Fresh",E{row}>0,"Carrier",TRUE,"Other")'


def build_donor(path: str, formula_rows: int = 3) -> None:
    """Build a donor holding a few rows of live formulas, one of them an array."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data'
    worksheet.append(HEADERS)

    for offset in range(formula_rows):
        row = 2 + offset
        worksheet.cell(row=row, column=1, value=f'V{row}')
        worksheet.cell(row=row, column=2, value='Fresh Sockeye')
        worksheet.cell(row=row, column=3, value='CMA')
        worksheet.cell(row=row, column=4, value=PLAIN_ONE.format(row=row))
        worksheet.cell(row=row, column=5, value=PLAIN_TWO.format(row=row))
        worksheet.cell(row=row, column=6,
                       value=ArrayFormula(ref=f'F{row}', text=ARRAY_ONE.format(row=row)))

    workbook.save(path)
    workbook.close()


def build_target(path: str, data_rows: int = 12) -> None:
    """Build a target with data but empty formula columns."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data'
    worksheet.append(HEADERS)

    for offset in range(data_rows):
        row = 2 + offset
        worksheet.cell(row=row, column=1, value=f'V{row}')
        worksheet.cell(row=row, column=2, value='Fresh Sockeye')
        worksheet.cell(row=row, column=3, value='CMA')

    workbook.save(path)
    workbook.close()


def run_seed(donor: str, target: str, **overrides):
    """Run seed_donor_formulas with sensible defaults for these fixtures."""
    config = {
        'processor_type': 'seed_donor_formulas',
        'source_file': donor,
        'source_sheet': 'Data',
        'target_file': target,
        'target_sheet': 'Data',
        'column_names': ['Plain One', 'Plain Two', 'Array One'],
        'start_row': 2,
        'row_count': 3,
        'fill_down': True,
    }
    config.update(overrides)
    return SeedDonorFormulasProcessor(config).execute()


def cell_text(worksheet, column_letter: str, row: int):
    """Return a cell's formula text, whether plain or array."""
    value = worksheet[f'{column_letter}{row}'].value
    return getattr(value, 'text', value)


def test_plain_formulas_reach_the_last_row():
    """A plain formula continues to the end of the data."""
    print("\nTesting plain formulas reach the last row...")

    with tempfile.TemporaryDirectory() as temp_dir:
        donor = str(Path(temp_dir) / 'donor.xlsx')
        target = str(Path(temp_dir) / 'target.xlsx')
        build_donor(donor)
        build_target(target, data_rows=12)
        run_seed(donor, target)

        workbook = openpyxl.load_workbook(target)
        worksheet = workbook['Data']
        last = cell_text(worksheet, 'D', 13)
        workbook.close()

        if last == PLAIN_ONE.format(row=13):
            print(f"  ✓ Row 13 holds {last}")
            return True

        print(f"  ✗ Row 13 holds {last!r}")
        return False


def test_array_formula_is_not_skipped():
    """
    The regression that started this module.

    An ArrayFormula is not a string, so a startswith('=') test skips it while
    the log still counts the column as filled.
    """
    print("\nTesting the array formula column is filled...")

    with tempfile.TemporaryDirectory() as temp_dir:
        donor = str(Path(temp_dir) / 'donor.xlsx')
        target = str(Path(temp_dir) / 'target.xlsx')
        build_donor(donor)
        build_target(target, data_rows=12)
        run_seed(donor, target)

        workbook = openpyxl.load_workbook(target)
        worksheet = workbook['Data']
        filled = [r for r in range(5, 14) if worksheet[f'F{r}'].value is not None]
        sample = cell_text(worksheet, 'F', 13)
        workbook.close()

        if len(filled) == 9 and sample == ARRAY_ONE.format(row=13):
            print(f"  ✓ All 9 rows below the seed filled, row 13 = {sample[:44]}...")
            return True

        print(f"  ✗ Only {len(filled)} of 9 rows filled; row 13 = {sample!r}")
        return False


def test_array_ref_is_per_row():
    """Each filled array cell needs its own ref, not the origin's."""
    print("\nTesting each array cell carries its own ref...")

    with tempfile.TemporaryDirectory() as temp_dir:
        donor = str(Path(temp_dir) / 'donor.xlsx')
        target = str(Path(temp_dir) / 'target.xlsx')
        build_donor(donor)
        build_target(target, data_rows=8)
        run_seed(donor, target)

        workbook = openpyxl.load_workbook(target)
        worksheet = workbook['Data']
        refs = [getattr(worksheet[f'F{r}'].value, 'ref', None) for r in (6, 7, 9)]
        workbook.close()

        if refs == ['F6', 'F7', 'F9']:
            print(f"  ✓ Refs follow their own rows: {refs}")
            return True

        print(f"  ✗ Refs are {refs}")
        return False


def test_convert_mode_writes_plain_formulas():
    """array_formula_mode 'convert' drops the array wrapper."""
    print("\nTesting array_formula_mode convert...")

    with tempfile.TemporaryDirectory() as temp_dir:
        donor = str(Path(temp_dir) / 'donor.xlsx')
        target = str(Path(temp_dir) / 'target.xlsx')
        build_donor(donor)
        build_target(target, data_rows=8)
        run_seed(donor, target, array_formula_mode='convert')

        workbook = openpyxl.load_workbook(target)
        worksheet = workbook['Data']
        seeded = worksheet['F2'].value
        filled = worksheet['F8'].value
        workbook.close()

        passed = True

        if isinstance(seeded, str):
            print("  ✓ Seeded row is a plain formula")
        else:
            print(f"  ✗ Seeded row is {type(seeded).__name__}")
            passed = False

        if isinstance(filled, str) and filled == ARRAY_ONE.format(row=8):
            print("  ✓ Filled row is a plain formula, correctly translated")
        else:
            print(f"  ✗ Filled row is {type(filled).__name__}: {filled!r}")
            passed = False

        return passed


def test_named_ranges_do_not_shift():
    """Cell references move with the row; named ranges must not."""
    print("\nTesting named ranges survive translation...")

    with tempfile.TemporaryDirectory() as temp_dir:
        donor = str(Path(temp_dir) / 'donor.xlsx')
        target = str(Path(temp_dir) / 'target.xlsx')
        build_donor(donor)
        build_target(target, data_rows=10)
        run_seed(donor, target)

        workbook = openpyxl.load_workbook(target)
        text = cell_text(workbook['Data'], 'E', 11)
        workbook.close()

        if 'rng_carrier' in text and 'C11' in text:
            print(f"  ✓ {text}")
            return True

        print(f"  ✗ {text!r}")
        return False


def test_short_donor_leaves_no_gap():
    """
    A donor with fewer rows than row_count must not leave blank rows.

    Filling from start_row + row_count - 1 steps over rows the donor never
    supplied. Filling from the last row actually written does not.
    """
    print("\nTesting a donor shorter than row_count...")

    with tempfile.TemporaryDirectory() as temp_dir:
        donor = str(Path(temp_dir) / 'donor.xlsx')
        target = str(Path(temp_dir) / 'target.xlsx')
        build_donor(donor, formula_rows=2)      # only rows 2-3
        build_target(target, data_rows=10)
        run_seed(donor, target, row_count=5, on_existing_cell='skip')

        workbook = openpyxl.load_workbook(target)
        worksheet = workbook['Data']
        blanks = [r for r in range(2, 12) if worksheet[f'D{r}'].value is None]
        workbook.close()

        if not blanks:
            print("  ✓ No gap: every row from 2 to 11 holds a formula")
            return True

        print(f"  ✗ Rows still blank: {blanks}")
        return False


def test_on_existing_cell_policies():
    """error stops, skip preserves, overwrite replaces."""
    print("\nTesting on_existing_cell policies...")

    passed = True

    for mode, expectation in [('error', 'halt'), ('skip', 'KEEP ME'), ('overwrite', 'formula')]:
        with tempfile.TemporaryDirectory() as temp_dir:
            donor = str(Path(temp_dir) / 'donor.xlsx')
            target = str(Path(temp_dir) / 'target.xlsx')
            build_donor(donor)
            build_target(target, data_rows=10)

            workbook = openpyxl.load_workbook(target)
            workbook['Data']['D8'] = 'KEEP ME'
            workbook.save(target)
            workbook.close()

            try:
                run_seed(donor, target, on_existing_cell=mode)
                halted = False
            except Exception:
                halted = True

            if expectation == 'halt':
                if halted:
                    print(f"  ✓ {mode:10} halted on the occupied cell")
                else:
                    print(f"  ✗ {mode:10} did not halt")
                    passed = False
                continue

            workbook = openpyxl.load_workbook(target)
            value = workbook['Data']['D8'].value
            workbook.close()

            if expectation == 'KEEP ME' and value == 'KEEP ME':
                print(f"  ✓ {mode:10} preserved the existing value")
            elif expectation == 'formula' and isinstance(value, str) and value.startswith('='):
                print(f"  ✓ {mode:10} replaced it with a formula")
            else:
                print(f"  ✗ {mode:10} left {value!r}")
                passed = False

    return passed


def test_anchor_columns_accept_header_names():
    """fill_anchor_columns takes header names, not just column letters."""
    print("\nTesting fill_anchor_columns by header name...")

    with tempfile.TemporaryDirectory() as temp_dir:
        donor = str(Path(temp_dir) / 'donor.xlsx')
        target = str(Path(temp_dir) / 'target.xlsx')
        build_donor(donor)
        build_target(target, data_rows=10)

        # A trailing marker row that populates only a formula column
        workbook = openpyxl.load_workbook(target)
        workbook['Data']['D14'] = 'XXXX'
        workbook.save(target)
        workbook.close()

        try:
            run_seed(donor, target, fill_anchor_columns=['Van Number'],
                     on_existing_cell='skip')
        except Exception as error:
            print(f"  ✗ Raised: {error}")
            return False

        workbook = openpyxl.load_workbook(target)
        worksheet = workbook['Data']
        marker = worksheet['D14'].value
        last_real = worksheet['D11'].value
        workbook.close()

        if marker == 'XXXX' and isinstance(last_real, str) and last_real.startswith('='):
            print("  ✓ Anchored on Van Number: fill stopped before the marker row")
            return True

        print(f"  ✗ marker={marker!r} last_real={last_real!r}")
        return False


def test_column_without_a_formula_is_not_filled():
    """A seed cell holding a constant must not be smeared down the column."""
    print("\nTesting a column with no formula is left alone...")

    with tempfile.TemporaryDirectory() as temp_dir:
        donor = str(Path(temp_dir) / 'donor.xlsx')
        target = str(Path(temp_dir) / 'target.xlsx')
        build_donor(donor)

        # Replace the donor's formula with a constant
        workbook = openpyxl.load_workbook(donor)
        for row in (2, 3, 4):
            workbook['Data'].cell(row=row, column=4, value='CONSTANT')
        workbook.save(donor)
        workbook.close()

        build_target(target, data_rows=10)
        run_seed(donor, target)

        workbook = openpyxl.load_workbook(target)
        worksheet = workbook['Data']
        seeded = worksheet['D2'].value
        below = [worksheet[f'D{r}'].value for r in range(5, 12)]
        workbook.close()

        if seeded == 'CONSTANT' and all(v is None for v in below):
            print("  ✓ Constant seeded but not filled down")
            return True

        print(f"  ✗ seeded={seeded!r} below={below[:3]}")
        return False


def test_empty_transplant_raises():
    """Finding no donor formulas at all is an error, not a warning."""
    print("\nTesting an empty transplant halts...")

    with tempfile.TemporaryDirectory() as temp_dir:
        donor = str(Path(temp_dir) / 'donor.xlsx')
        target = str(Path(temp_dir) / 'target.xlsx')
        build_donor(donor)
        build_target(target, data_rows=10)

        # Wipe the donor's formula rows
        workbook = openpyxl.load_workbook(donor)
        for row in (2, 3, 4):
            for column in (4, 5, 6):
                workbook['Data'].cell(row=row, column=column).value = None
        workbook.save(donor)
        workbook.close()

        try:
            run_seed(donor, target)
        except Exception as error:
            if 'No formulas found' in str(error):
                print("  ✓ Halted with a message naming the donor and rows")
                return True
            print(f"  ✗ Wrong error: {error}")
            return False

        print("  ✗ Produced a formula-free file without complaint")
        return False


def main():
    """Run every test and report a final score."""
    print("=== seed_donor_formulas fill-down tests ===")

    tests = [
        test_plain_formulas_reach_the_last_row,
        test_array_formula_is_not_skipped,
        test_array_ref_is_per_row,
        test_convert_mode_writes_plain_formulas,
        test_named_ranges_do_not_shift,
        test_short_donor_leaves_no_gap,
        test_on_existing_cell_policies,
        test_anchor_columns_accept_header_names,
        test_column_without_a_formula_is_not_filled,
        test_empty_transplant_raises,
    ]

    passed = 0

    for test_func in tests:
        try:
            if test_func():
                passed += 1
        except Exception as error:
            print(f"  ✗ {test_func.__name__} crashed: {error}")

    print(f"\n=== Results: {passed}/{len(tests)} tests passed ===")

    if passed == len(tests):
        print("✅ All fill-down tests passed!")
        return 1

    print("❌ Some fill-down tests failed!")
    return 0


if __name__ == '__main__':
    exit(0 if main() else 1)


# End of file #

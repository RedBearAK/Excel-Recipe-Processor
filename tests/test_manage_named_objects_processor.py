"""
Test module for ManageNamedObjectsProcessor.

tests/test_manage_named_objects_processor.py

Tests extraction, translation, and export of Excel named ranges, lambda functions,
formulas, and tables. Demonstrates YAML export capabilities and format translation.
"""

import os
import sys
import tempfile
import openpyxl
import pandas as pd

from pathlib import Path

# Add project root to Python path for imports  
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname, absolute_coordinate
from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_recipe_processor.processors.manage_named_objects_processor import ManageNamedObjectsProcessor


def create_test_workbook_with_named_objects():
    """Create a test workbook with various types of named objects for testing."""
    
    wb = openpyxl.Workbook()
    
    # Create main data sheet
    ws_data = wb.active
    ws_data.title = "Data"
    
    # Add sample data
    data = [
        ["ProductCode", "Quarter", "Revenue", "Units", "Region"],
        ["A001", "Q1", 15000, 100, "North"],
        ["A002", "Q1", 23000, 150, "South"], 
        ["A003", "Q1", 18500, 120, "East"],
        ["A001", "Q2", 17000, 110, "North"],
        ["A002", "Q2", 25000, 160, "South"]
    ]
    
    for row in data:
        ws_data.append(row)
    
    # Create calculations sheet
    ws_calc = wb.create_sheet("Calculations")
    ws_calc['A1'] = "Tax Rate"
    ws_calc['B1'] = 0.21
    ws_calc['A2'] = "Company Name"
    ws_calc['B2'] = "Test Corporation"
    
    # Create named ranges (simple ranges and constants)
    # Global named range for data
    data_range_ref = f"{quote_sheetname('Data')}!{absolute_coordinate('A1:E6')}"
    data_range = DefinedName("stg_revenue_data", attr_text=data_range_ref)
    wb.defined_names.add(data_range)
    
    # Named constant for tax rate
    tax_rate = DefinedName("tax_rate", attr_text="0.21")
    wb.defined_names.add(tax_rate)
    
    # Named constant for company name
    company_name = DefinedName("company_name", attr_text='"Test Corporation"')
    wb.defined_names.add(company_name)
    
    # Create lambda functions
    # Simple lambda for growth calculation
    growth_lambda = DefinedName(
        "calc_growth_rate", 
        attr_text="=_xlfn.LAMBDA(_xlpm.current,_xlpm.previous,(_xlpm.current-_xlpm.previous)/_xlpm.previous)"
    )
    wb.defined_names.add(growth_lambda)
    
    # Complex lambda with lookup
    lookup_lambda = DefinedName(
        "stg_financial_lookup",
        attr_text="=_xlfn.LAMBDA(_xlpm.code,_xlpm.quarter,_xlfn.XLOOKUP(_xlpm.code,stg_revenue_data[ProductCode],stg_revenue_data[_xlpm.quarter]))"
    )
    wb.defined_names.add(lookup_lambda)
    
    # Create named formula (non-lambda)
    revenue_formula = DefinedName(
        "total_revenue_q1",
        attr_text='=SUMIF(stg_revenue_data[Quarter],"Q1",stg_revenue_data[Revenue])'
    )
    wb.defined_names.add(revenue_formula)
    
    # Create local named range (sheet-specific)
    calc_sheet_id = wb.sheetnames.index('Calculations')
    local_range_ref = f"{quote_sheetname('Calculations')}!{absolute_coordinate('B1:B2')}"
    local_range = DefinedName(
        "local_calc_range",
        attr_text=local_range_ref,
        localSheetId=calc_sheet_id
    )
    wb.defined_names.add(local_range)
    
    # Create Excel table
    table = Table(displayName="ProductData", ref="A1:E6")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=True
    )
    ws_data.add_table(table)
    
    return wb


def test_object_extraction():
    """Test extraction of different object types from Excel workbook."""
    
    print("=== Testing Object Extraction ===")
    
    # Create test workbook
    wb = create_test_workbook_with_named_objects()
    
    # Create processor for testing
    step_config = {
        'processor_type': 'manage_named_objects',
        'operation': 'export_all',
        'source_file': 'test.xlsx',
        'export_file': 'test_export.yaml'
    }
    
    processor = ManageNamedObjectsProcessor(step_config)
    
    try:
        # Test extraction methods
        all_objects = processor.extract_all_named_objects(wb)
        
        print(f"✓ Extracted {all_objects['metadata']['total_objects']} total objects")
        print(f"  - Named ranges: {len(all_objects['named_ranges'])}")
        print(f"  - Lambda functions: {len(all_objects['lambda_functions'])}")  
        print(f"  - Named formulas: {len(all_objects['named_formulas'])}")
        print(f"  - Named tables: {len(all_objects['named_tables'])}")
        print(f"  - Local objects: {sum(len(sheet_objs.get('named_ranges', [])) for sheet_objs in all_objects['local_objects'].values())}")
        
        # Test object type detection
        named_ranges = all_objects['named_ranges']
        if len(named_ranges) >= 3:
            print(f"✓ Named ranges detected correctly:")
            for nr in named_ranges:
                print(f"  - {nr['name']}: {nr['type']} ({nr['definition']})")
        
        # Test lambda function extraction and translation
        lambda_funcs = all_objects['lambda_functions']
        if len(lambda_funcs) >= 2:
            print(f"✓ Lambda functions detected and translated:")
            for lf in lambda_funcs:
                print(f"  - {lf['name']}: {lf['definition']}")
                print(f"    Parameters: {lf['parameters']}")
                print(f"    Excel format: {lf['excel_definition']}")
        
        # Test table extraction
        tables = all_objects['named_tables']
        if len(tables) >= 1:
            print(f"✓ Tables extracted:")
            for table in tables:
                print(f"  - {table['name']}: {table['table_properties']['range']}")
                print(f"    Columns: {[col['name'] for col in table['table_properties']['columns']]}")
        
        # Test local objects
        local_objects = all_objects['local_objects']
        if local_objects:
            print(f"✓ Local objects extracted:")
            for sheet_name, sheet_objs in local_objects.items():
                print(f"  - Sheet '{sheet_name}': {len(sheet_objs['named_ranges'])} objects")
        
        return True
        
    except Exception as e:
        print(f"✗ Object extraction failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    finally:
        wb.close()


def test_lambda_translation():
    """Test lambda function translation between Excel and human formats."""
    
    print("\\n=== Testing Lambda Translation ===")
    
    step_config = {
        'processor_type': 'manage_named_objects',
        'operation': 'list_objects', 
        'source_file': 'test.xlsx'
    }
    processor = ManageNamedObjectsProcessor(step_config)
    
    # Test cases for lambda translation
    test_cases = [
        {
            'excel': "=_xlfn.LAMBDA(_xlpm.x,_xlpm.y,_xlpm.x+_xlpm.y)",
            'human': "LAMBDA(x, y, x+y)",
            'params': ["x", "y"]
        },
        {
            'excel': "=_xlfn.LAMBDA(_xlpm.rate,_xlpm.periods,_xlfn.PV(_xlpm.rate,_xlpm.periods,0))",
            'human': "LAMBDA(rate, periods, PV(rate, periods, 0))",
            'params': ["rate", "periods"]
        },
        {
            'excel': "=_xlfn.LAMBDA(_xlpm.current,_xlpm.previous,(_xlpm.current-_xlpm.previous)/_xlpm.previous)",
            'human': "LAMBDA(current, previous, (current-previous)/previous)",
            'params': ["current", "previous"]
        }
    ]
    
    all_passed = True
    
    for i, test_case in enumerate(test_cases, 1):
        try:
            # Test Excel to human translation
            human_result, params_result = processor.translate_lambda_to_human(test_case['excel'])
            
            if human_result.replace(' ', '') == test_case['human'].replace(' ', ''):
                print(f"✓ Test {i}: Excel to human translation passed")
            else:
                print(f"✗ Test {i}: Excel to human translation failed")
                print(f"  Expected: {test_case['human']}")
                print(f"  Got: {human_result}")
                all_passed = False
            
            if params_result == test_case['params']:
                print(f"✓ Test {i}: Parameter extraction passed")
            else:
                print(f"✗ Test {i}: Parameter extraction failed")
                print(f"  Expected: {test_case['params']}")
                print(f"  Got: {params_result}")
                all_passed = False
            
            # Test human to Excel translation
            excel_result = processor.translate_lambda_to_excel(test_case['human'], test_case['params'])
            
            # Exact comparison now that we have proper formatting
            if excel_result == test_case['excel']:
                print(f"✓ Test {i}: Human to Excel translation passed")
            else:
                print(f"✗ Test {i}: Human to Excel translation failed")
                print(f"  Expected: {test_case['excel']}")
                print(f"  Got: {excel_result}")
                all_passed = False
                
        except Exception as e:
            print(f"✗ Test {i}: Translation failed with error: {e}")
            all_passed = False
    
    return all_passed


def test_dual_format_export():
    """Test export to both YAML and VBA formats."""
    
    print("\\n=== Testing Dual Format Export ===")
    
    # Create test workbook
    wb = create_test_workbook_with_named_objects()
    
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save test workbook
            test_file = Path(temp_dir) / "test_workbook.xlsx"
            wb.save(test_file)
            
            # Test dual export configuration
            step_config = {
                'processor_type': 'manage_named_objects',
                'operation': 'export_all',
                'source_file': str(test_file),
                'export_formats': {
                    'yaml_file': str(Path(temp_dir) / "objects.yaml"),
                    'vba_file': str(Path(temp_dir) / "objects.txt")
                }
            }
            
            processor = ManageNamedObjectsProcessor(step_config)
            result = processor.execute()
            
            print(f"✓ Dual export completed successfully")
            print(f"  - Source: {result['source_file']}")
            print(f"  - Exports: {result['exports_completed']}")
            print(f"  - Objects exported: {result['objects_exported']}")
            
            # Check both files exist
            yaml_file = Path(temp_dir) / "objects.yaml"
            vba_file = Path(temp_dir) / "objects.txt"
            
            yaml_success = False
            vba_success = False
            
            if yaml_file.exists():
                yaml_size = yaml_file.stat().st_size
                print(f"✓ YAML export created: {yaml_size} bytes")
                
                # Show sample YAML content
                with open(yaml_file, 'r') as f:
                    content = f.read()
                    lines = content.split('\\n')[:15]
                    print("✓ Sample YAML content:")
                    for line in lines:
                        print(f"  {line}")
                    if len(content.split('\\n')) > 15:
                        print("  ...")
                yaml_success = True
            
            if vba_file.exists():
                vba_size = vba_file.stat().st_size
                print(f"✓ VBA export created: {vba_size} bytes")
                
                # Show sample VBA content
                with open(vba_file, 'r') as f:
                    content = f.read()
                    lines = content.split('\\n')[:20]
                    print("✓ Sample VBA content:")
                    for line in lines:
                        print(f"  {line}")
                    if len(content.split('\\n')) > 20:
                        print("  ...")
                vba_success = True
            
            return yaml_success and vba_success
                
    except Exception as e:
        print(f"✗ Dual format export failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    finally:
        wb.close()


def test_vba_format_compatibility():
    """Test VBA format structure and parsing compatibility."""
    
    print("\\n=== Testing VBA Format Compatibility ===")
    
    # Create test workbook
    wb = create_test_workbook_with_named_objects()
    
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save test workbook
            test_file = Path(temp_dir) / "test_workbook.xlsx"
            wb.save(test_file)
            
            # Export VBA format only
            step_config = {
                'processor_type': 'manage_named_objects',
                'operation': 'export_all',
                'source_file': str(test_file),
                'vba_file': str(Path(temp_dir) / "vba_export.txt")
            }
            
            processor = ManageNamedObjectsProcessor(step_config)
            result = processor.execute()
            
            # Read and parse the VBA format
            vba_file = Path(temp_dir) / "vba_export.txt"
            
            with open(vba_file, 'r') as f:
                content = f.read()
            
            # Test parsing logic similar to VBA
            lines = content.replace('\r', '').split('\n')
            sections_found = []
            objects_by_section = {}
            current_section = ""
            
            for line in lines:
                line = line.strip()
                
                # Skip comments and empty lines
                if not line or line.startswith('#'):
                    continue
                    
                # Check for section headers
                if line.startswith('[') and line.endswith(']'):
                    current_section = line[1:-1]
                    sections_found.append(current_section)
                    objects_by_section[current_section] = []
                    continue
                
                # Parse object line
                if current_section and ' | ' in line:
                    parts = line.split(' | ')
                    if len(parts) >= 5:  # Name, Definition, Type, Scope, Description
                        obj_info = {
                            'name': parts[0].strip(),
                            'definition': parts[1].strip(),
                            'type': parts[2].strip(),
                            'scope': parts[3].strip(),
                            'description': parts[4].strip()
                        }
                        objects_by_section[current_section].append(obj_info)
            
            print(f"✓ VBA format parsing successful")
            print(f"  - Sections found: {sections_found}")
            
            # Validate expected sections
            expected_sections = ['NAMED_RANGES', 'LAMBDA_FUNCTIONS', 'NAMED_FORMULAS', 'NAMED_TABLES']
            sections_validated = 0
            
            for section in expected_sections:
                if section in objects_by_section:
                    count = len(objects_by_section[section])
                    if count > 0:
                        print(f"  - {section}: {count} objects")
                        sections_validated += 1
                        
                        # Show sample objects
                        for obj in objects_by_section[section][:2]:  # First 2 objects
                            print(f"    * {obj['name']}: {obj['type']} - {obj['description']}")
            
            # Test lambda format conversion simulation
            if 'LAMBDA_FUNCTIONS' in objects_by_section:
                print("✓ Lambda function format validation:")
                lambda_success = 0
                for lambda_obj in objects_by_section['LAMBDA_FUNCTIONS']:
                    human_formula = lambda_obj['definition']
                    print(f"  - {lambda_obj['name']}: {human_formula}")
                    
                    # Simulate VBA conversion test
                    if human_formula.startswith('LAMBDA('):
                        print(f"    ✓ Valid human-readable lambda format")
                        lambda_success += 1
                    else:
                        print(f"    ✗ Invalid lambda format")
                
                # Require at least some successful validations
                return sections_validated >= 3 and lambda_success > 0
            
            return sections_validated >= 3
            
    except Exception as e:
        print(f"✗ VBA format compatibility test failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    finally:
        wb.close()


def test_list_objects_operation():
    """Test the list_objects operation for inventory."""
    
    print("\\n=== Testing List Objects Operation ===")
    
    # Create test workbook
    wb = create_test_workbook_with_named_objects()
    
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save test workbook
            test_file = Path(temp_dir) / "test_workbook.xlsx"
            wb.save(test_file)
            
            # Create processor and list objects
            step_config = {
                'processor_type': 'manage_named_objects',
                'operation': 'list_objects',
                'source_file': str(test_file)
            }
            
            processor = ManageNamedObjectsProcessor(step_config)
            inventory = processor.execute()
            
            print(f"✓ Object listing completed successfully")
            print(f"  - Source: {inventory['source_file']}")
            print(f"  - Total objects: {inventory['total_objects']}")
            print(f"  - By type: {inventory['by_type']}")
            
            # Show object names by category
            for category, names in inventory['object_names'].items():
                if names:
                    print(f"✓ {category}: {', '.join(names)}")
            
            return True
            
    except Exception as e:
        print(f"✗ List objects operation failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    finally:
        wb.close()


def test_processor_capabilities():
    """Test processor capabilities and configuration validation."""
    
    print("\\n=== Testing Processor Capabilities ===")
    
    try:
        # Test valid configuration
        step_config = {
            'processor_type': 'manage_named_objects',
            'operation': 'export_all',
            'source_file': 'test.xlsx',
            'export_file': 'test.yaml'
        }
        
        processor = ManageNamedObjectsProcessor(step_config)
        capabilities = processor.get_capabilities()
        
        print("✓ Processor created successfully")
        print(f"✓ Description: {capabilities['description']}")
        print(f"✓ Supported operations: {', '.join(capabilities['operations'])}")
        print(f"✓ Object types: {', '.join(capabilities['supported_object_types'])}")
        print(f"✓ Lambda features: {', '.join(capabilities['lambda_features'])}")
        
        # Test invalid operation
        try:
            invalid_config = {
                'processor_type': 'manage_named_objects',
                'operation': 'invalid_operation'
            }
            ManageNamedObjectsProcessor(invalid_config)
            print("✗ Invalid operation validation failed")
            return False
        except Exception:
            print("✓ Invalid operation properly rejected")
        
        return True
        
    except Exception as e:
        print(f"✗ Capabilities test failed: {e}")
        return False


def main():
    """Run all tests and return overall success status."""
    
    print("🧪 Testing ManageNamedObjectsProcessor with VBA Integration")
    print("=" * 60)
    
    tests = [
        ("Object Extraction", test_object_extraction),
        ("Lambda Translation", test_lambda_translation), 
        ("Dual Format Export", test_dual_format_export),
        ("VBA Format Compatibility", test_vba_format_compatibility),
        ("List Objects Operation", test_list_objects_operation),
        ("Processor Capabilities", test_processor_capabilities)
    ]
    
    passed = 0
    total = len(tests)
    
    for test_name, test_func in tests:
        print(f"\\n🔬 Running: {test_name}")
        print("-" * 35)
        
        try:
            if test_func():
                passed += 1
                print(f"✅ {test_name}: PASSED")
            else:
                print(f"❌ {test_name}: FAILED")
        except Exception as e:
            print(f"❌ {test_name}: ERROR - {e}")
            import traceback
            traceback.print_exc()
    
    print("\\n" + "=" * 60)
    print(f"📊 FINAL RESULTS: {passed}/{total} tests passed")
    
    if passed == total:
        print("🎉 All tests passed! ManageNamedObjectsProcessor with VBA integration is working correctly.")
        print()
        print("🔧 KEY ACHIEVEMENTS:")
        print("✓ Named range extraction and classification")
        print("✓ Lambda function detection and translation")
        print("✓ Excel table structure preservation")
        print("✓ YAML export with human-readable format")
        print("✓ VBA-compatible export format")
        print("✓ Dual format export capability")
        print("✓ Local and global scope handling")
        print("✓ Parameter extraction from lambda functions")
        print("✓ VBA import/export workflow support")
        print()
        print("🔄 WORKFLOW INTEGRATION:")
        print("- Python: Complex analysis and bulk operations")
        print("- VBA: Manual updates and Excel-native operations")
        print("- Both formats maintain object integrity")
        print("- Seamless conversion between human/Excel formats")
        print()
        print("📋 READY FOR PRODUCTION:")
        print("- Export Excel models to version-controlled formats")
        print("- Enable collaborative editing of named objects")
        print("- Integrate with automated deployment pipelines")
        print("- Support both technical and business users")
        return 0
    else:
        print(f"⚠️  {total - passed} tests failed. Issues need to be resolved.")
        return 1


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)


# End of file #

"""
Tests for profile_named_objects, including the name-drift alarm.

tests/test_profile_named_objects.py

Facts are asserted against a workbook built with one of everything:
a global range name, a harvest-grammar stored lambda, a sheet-scoped
name, a hidden name, a constant, and a worksheet table. Then the
anchor consumer replays the 2026-08-14 incident: a copy of the
workbook with the lambda DELETED (what Excel's repair did to
fn_blank_safe) is profiled and diffed against the original - the
vanished name must surface as a diff row. Runnable directly or with
pytest; direct runs are the authoritative score.
"""

import os
import sys
import tempfile

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

from excel_recipe_processor.core.stage_manager import StageManager
from excel_recipe_processor.processors.profile_named_objects_processor import (
    ProfileNamedObjectsProcessor,
)

STORED_LAMBDA = '_xlfn.LAMBDA(_xlpm.v,IF(_xlpm.v="","",_xlpm.v))'


def build_workbook(path):
    """One of every named-object kind, with known values."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Data'
    sheet.append(['Key', 'Value'])
    sheet.append(['a', 1])
    sheet.append(['b', 2])

    names = workbook.defined_names
    names['rng_keys'] = DefinedName('rng_keys', attr_text='Data!$A$2:$A$3')
    names['fn_blank_safe'] = DefinedName('fn_blank_safe',
                                         attr_text=STORED_LAMBDA)
    names['answer'] = DefinedName('answer', attr_text='42')
    hidden_name = DefinedName('fml_secret', attr_text='SUM(Data!$B$2:$B$3)')
    hidden_name.hidden = True
    names['fml_secret'] = hidden_name

    scoped = DefinedName('rng_local', attr_text='Data!$B$2:$B$3',
                         localSheetId=0)
    sheet.defined_names['rng_local'] = scoped

    table = Table(displayName='tbl_data', ref='A1:B3')
    sheet.add_table(table)
    workbook.save(path)


def run_processor(workbooks):
    config = {'processor_type': 'profile_named_objects',
              'workbooks': workbooks,
              'save_to_stage': 'stg_t_names'}
    return ProfileNamedObjectsProcessor(config).load_data()


def test_catalog_facts():
    """Every object kind lands with correct contract values."""
    print("\nTesting named-object catalog facts...")

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as handle:
        path = handle.name
    try:
        build_workbook(path)
        profile = run_processor([path]).set_index('Name')

        checks = [
            ('all six objects cataloged', len(profile) == 6),
            ('range classified',
             profile.loc['rng_keys', 'Object_Type'] == 'range'),
            ('lambda classified',
             profile.loc['fn_blank_safe', 'Object_Type'] == 'lambda'),
            ('constant classified',
             profile.loc['answer', 'Object_Type'] == 'constant'),
            ('formula classified',
             profile.loc['fml_secret', 'Object_Type'] == 'formula'),
            ('table classified',
             profile.loc['tbl_data', 'Object_Type'] == 'table'),
            ('stored definition verbatim',
             profile.loc['fn_blank_safe', 'Definition'] == STORED_LAMBDA),
            ('lambda human-translated',
             profile.loc['fn_blank_safe', 'Human_Definition']
             == 'LAMBDA(v, IF(v="","",v))'),
            ('lambda parameters listed',
             profile.loc['fn_blank_safe', 'Parameters'] == 'v'),
            ('formula prefixes stripped for display',
             profile.loc['fml_secret', 'Human_Definition']
             == 'SUM(Data!$B$2:$B$3)'),
            ('hidden flag captured',
             profile.loc['fml_secret', 'Hidden'] == True
             and profile.loc['rng_keys', 'Hidden'] == False),
            ('global scope', profile.loc['rng_keys', 'Scope'] == 'global'),
            ('sheet scope', profile.loc['rng_local', 'Scope'] == 'Data'),
            ('table scoped to its sheet with ref',
             profile.loc['tbl_data', 'Scope'] == 'Data'
             and profile.loc['tbl_data', 'Definition'] == 'Data!A1:B3'),
        ]
        for name, ok in checks:
            if not ok:
                print(f"✗ {name}")
                print(profile.to_string())
                return False
            print(f"✓ {name}")
        return True
    finally:
        os.unlink(path)


def test_name_drift_alarm():
    """The fn_blank_safe scenario: a repair-deleted name surfaces in a diff."""
    print("\nTesting the name-drift alarm (the fn_blank_safe scenario)...")

    from excel_recipe_processor.processors.diff_data_processor import (
        DiffDataProcessor,
    )

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as h1, \
         tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as h2:
        previous_path, repaired_path = h1.name, h2.name
    try:
        build_workbook(previous_path)
        # What Excel's repair did on 2026-08-14: the lambda deleted whole
        build_workbook(repaired_path)
        damaged = load_workbook(repaired_path)
        del damaged.defined_names['fn_blank_safe']
        damaged.save(repaired_path)

        StageManager.initialize_stages(max_stages=30)
        baseline = run_processor([previous_path]).drop(columns='Workbook')
        current = run_processor([repaired_path]).drop(columns='Workbook')
        StageManager.save_stage('stg_t_names_baseline', baseline, 'test')

        diff = DiffDataProcessor({
            'processor_type': 'diff_data',
            'reference_stage': 'stg_t_names_baseline',
            'source_stage': 'stg_t_names_current',
            'key_columns': 'Name',
            'save_to_stage': 'stg_t_names_drift',
        }).execute(current)

        text = diff.to_string()
        if 'fn_blank_safe' not in text:
            print("✗ the vanished lambda did not surface in the diff")
            print(text[:600])
            return False
        print("✓ fn_blank_safe's disappearance surfaced as a diff row -")
        print("  the 2026-08-14 incident would have been caught pre-eyeball")
        return True
    finally:
        os.unlink(previous_path)
        os.unlink(repaired_path)


def main():
    """Run all tests and report results."""
    print("profile_named_objects tests")
    print("=" * 50)

    tests = [
        test_catalog_facts,
        test_name_drift_alarm,
    ]

    passed = 0
    for test in tests:
        try:
            if test():
                passed += 1
            else:
                print(f"FAILED: {test.__name__}")
        except Exception as error:
            print(f"FAILED with exception: {test.__name__}: {error}")

    print("=" * 50)
    print(f"Final score: {passed}/{len(tests)} tests passed")
    return passed == len(tests)


if __name__ == '__main__':
    sys.exit(0 if main() else 1)

# End of file #

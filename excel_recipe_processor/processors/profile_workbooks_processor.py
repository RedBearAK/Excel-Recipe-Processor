"""
Profile workbooks: per-sheet workbook metadata as an ordinary stage.

excel_recipe_processor/processors/profile_workbooks_processor.py

Third member of the profile_* family (profile_files, profile_sheets,
profile_workbooks; profile_named_objects planned). Same principles:
plural inputs, one identity-keyed output stage, NO apply siblings -
consumers are existing processors reading the contract by column name.
The anchor consumer is the DRIFT ALARM shape: profile the previous
run's output and the current one, diff_data the two frames on
Workbook-relative keys, and a vanished tab, an unhidden sheet, a
row-count jump, or a zoom drift becomes an itemized diff row instead
of a surprise.

Named objects are deliberately COUNTED here, never cataloged: the
catalog is Name-keyed, which under the family's own key-split
principle makes it a different contract (profile_named_objects,
planned - it will share manage_named_objects' extraction helpers).
Has_VBA is the same kind of tripwire: presence only.

OUTPUT CONTRACT (columns by name; future facts APPEND, never rename).
Sheet-keyed rows, one per sheet per workbook:
    Workbook            The path as configured
    Sheet               Sheet name
    Position            1-based tab position
    State               visible / hidden / veryHidden (openpyxl spelling)
    Tab_Color           RGB string, or '' when uncolored
    Max_Row, Max_Col    openpyxl-reported extents
    Frozen_Panes        e.g. 'A2', or '' when unfrozen
    Zoom_Percent        Sheet view zoom (100 when unset)
    DV_Count            dataValidation rule count on the sheet
    Named_Object_Count  WORKBOOK-level count, repeated per row (tripwire)
    Has_VBA             WORKBOOK-level flag, repeated per row (tripwire)

v1 reads workbooks FROM DISK (openpyxl load): the drift-alarm shape
reads the PREVIOUS run's file, which only exists on disk. Profiling
the session's cached in-flight workbook is a documented follow-up for
when a real consumer needs it.
"""

import logging
import zipfile

import pandas as pd
from openpyxl import load_workbook

from excel_recipe_processor.core.base_processor import ImportBaseProcessor, StepProcessorError
from excel_recipe_processor.core.config_schema import Key, Schema, name_list


logger = logging.getLogger(__name__)


class ProfileWorkbooksProcessor(ImportBaseProcessor):
    """Profile one or more workbooks into a per-sheet metadata stage."""

    @classmethod
    def config_schema(cls) -> Schema:
        """Declared keys (2026-09-03); see core/config_schema.py."""
        return Schema([Key('workbooks', 'list', item_kind='str', required=True)])

    @classmethod
    def get_minimal_config(cls):
        return {
            'workbooks': ['some_workbook.xlsx'],
            'save_to_stage': 'stg_workbook_profiles',
        }

    def __init__(self, step_config: dict):
        super().__init__(step_config)

        self.workbooks = self.get_config_value('workbooks', None)
        if not isinstance(self.workbooks, list) or len(self.workbooks) == 0:
            raise StepProcessorError(
                f"Step '{self.step_name}': 'workbooks' must be a non-empty "
                f"list of workbook paths"
            )

    def get_capabilities(self) -> dict:
        """Processor capabilities information."""
        return {
            'description': 'Per-sheet workbook metadata discovery (state, color, extents, counts)',
            'profile_columns': ['Workbook', 'Sheet', 'Position', 'State',
                                'Tab_Color', 'Max_Row', 'Max_Col',
                                'Frozen_Panes', 'Zoom_Percent', 'DV_Count',
                                'Named_Object_Count', 'Has_VBA'],
            'named_objects': 'COUNT only - the catalog is Name-keyed and belongs to the planned profile_named_objects',
            'anchor_consumer': 'drift alarm: diff_data on profiles of consecutive run outputs',
        }

    def load_data(self) -> pd.DataFrame:
        """Profile every configured workbook, one row per sheet."""
        rows = []
        for path in self.workbooks:
            try:
                workbook = load_workbook(str(path))
            except Exception as error:
                raise StepProcessorError(
                    f"Step '{self.step_name}': could not open workbook "
                    f"'{path}': {error}"
                )

            named_object_count = len(workbook.defined_names)
            has_vba = self._has_vba(str(path))

            for position, sheet_name in enumerate(workbook.sheetnames, 1):
                sheet = workbook[sheet_name]
                tab_color = ''
                color = sheet.sheet_properties.tabColor
                if color is not None and color.rgb:
                    tab_color = str(color.rgb)
                zoom = sheet.sheet_view.zoomScale
                rows.append({
                    'Workbook': str(path),
                    'Sheet': sheet_name,
                    'Position': position,
                    'State': sheet.sheet_state,
                    'Tab_Color': tab_color,
                    'Max_Row': sheet.max_row,
                    'Max_Col': sheet.max_column,
                    'Frozen_Panes': sheet.freeze_panes or '',
                    'Zoom_Percent': int(zoom) if zoom else 100,
                    'DV_Count': len(sheet.data_validations.dataValidation),
                    'Named_Object_Count': named_object_count,
                    'Has_VBA': has_vba,
                })
            workbook.close()

        profile = pd.DataFrame(rows, columns=[
            'Workbook', 'Sheet', 'Position', 'State', 'Tab_Color',
            'Max_Row', 'Max_Col', 'Frozen_Panes', 'Zoom_Percent',
            'DV_Count', 'Named_Object_Count', 'Has_VBA'])
        logger.info(
            f"📚 Profiled {len(self.workbooks)} workbook(s): "
            f"{len(profile)} sheet rows")
        return profile

    def _has_vba(self, path: str) -> bool:
        """Macro presence: the vbaProject part exists in the package."""
        try:
            with zipfile.ZipFile(path) as archive:
                return any(name.endswith('vbaProject.bin')
                           for name in archive.namelist())
        except Exception:
            return False

# End of file #

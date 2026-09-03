"""
Profile named objects: a Name-keyed workbook catalog as an ordinary stage.

excel_recipe_processor/processors/profile_named_objects_processor.py

Fourth member of the profile_* family. The split from
manage_named_objects is by OUTPUT CURRENCY, not by capability overlap:
manage owns workbook<->YAML library round trips and the storage
translation machinery; this processor emits a Name-keyed STAGE for
in-pipeline consumers - filter it, export it to a reference tab, or
diff it across runs. The anchor consumer is the name-drift alarm: the
2026-08-14 incident where Excel's repair silently DELETED fn_blank_safe
(orphaning its callers into #NAME?) would have surfaced as a
"name vanished since last run" diff row instead of a user report.

Classification and human translation come from the SAME shared helpers
manage delegates to (_helpers/named_objects_extraction.py) - one truth
for what a defined name is and how its stored grammar reads.

OUTPUT CONTRACT (columns by name; future facts APPEND, never rename).
One row per defined name plus one per worksheet table:
    Workbook          Path as configured
    Name              Object name
    Object_Type       lambda / formula / range / constant / table
    Scope             'global', or the sheet name for sheet-scoped names
                      (tables always carry their sheet)
    Hidden            definedName hidden flag (tables: False)
    Definition        STORED text verbatim (attr_text; table ref) - the
                      byte truth, right for drift comparison
    Human_Definition  Display syntax: lambdas translated, storage
                      prefixes stripped from formulas; ranges/refs as-is
    Parameters        Comma-joined lambda parameter list, '' otherwise
"""

import logging

import pandas as pd
from openpyxl import load_workbook

from excel_recipe_processor.core.base_processor import ImportBaseProcessor, StepProcessorError
from excel_recipe_processor.core.config_schema import Key, Schema, name_list
from excel_recipe_processor.processors._helpers.named_objects_extraction import (
    detect_object_type,
    translate_lambda_to_human,
    clean_formula_for_display,
)


logger = logging.getLogger(__name__)


class ProfileNamedObjectsProcessor(ImportBaseProcessor):
    """Catalog defined names and tables into a Name-keyed metadata stage."""

    @classmethod
    def config_schema(cls) -> Schema:
        """Declared keys (2026-09-03); see core/config_schema.py."""
        return Schema([Key('workbooks', 'list', item_kind='str', required=True)])

    @classmethod
    def get_minimal_config(cls):
        return {
            'workbooks': ['some_workbook.xlsx'],
            'save_to_stage': 'stg_named_object_profiles',
        }

    def __init__(self, step_config: dict):
        super().__init__(step_config)

        self.workbooks = self.get_config_value('workbooks', None)
        if not isinstance(self.workbooks, list) or len(self.workbooks) == 0:
            raise StepProcessorError(
                f"Step '{self.step_name}': 'workbooks' must be a non-empty "
                f"list of workbook paths"
            )

    def get_capabilities(self) -> dict:
        """Processor capabilities information."""
        return {
            'description': 'Per-name workbook object discovery (ranges, lambdas, formulas, tables)',
            'profile_columns': ['Workbook', 'Name', 'Object_Type', 'Scope',
                                'Hidden', 'Definition', 'Human_Definition',
                                'Parameters'],
            'shares_helpers_with': 'manage_named_objects (named_objects_extraction)',
            'anchor_consumer': 'name-drift alarm: diff_data on profiles of consecutive run outputs',
        }

    def load_data(self) -> pd.DataFrame:
        """Catalog every configured workbook's named objects."""
        rows = []
        for path in self.workbooks:
            try:
                workbook = load_workbook(str(path))
            except Exception as error:
                raise StepProcessorError(
                    f"Step '{self.step_name}': could not open workbook "
                    f"'{path}': {error}"
                )

            for defined_name in workbook.defined_names.values():
                rows.append(self._defined_name_row(str(path), workbook,
                                                   defined_name))
            for sheet_name in workbook.sheetnames:
                # openpyxl 3.x keeps SHEET-SCOPED names on the worksheet's
                # own collection, not the workbook's - both must be walked
                for defined_name in workbook[sheet_name].defined_names.values():
                    rows.append(self._defined_name_row(
                        str(path), workbook, defined_name,
                        scope_override=sheet_name))
            for sheet_name in workbook.sheetnames:
                # openpyxl's TableList.items() yields (name, REF STRING)
                for table_name, table_ref in workbook[sheet_name].tables.items():
                    rows.append({
                        'Workbook': str(path),
                        'Name': table_name,
                        'Object_Type': 'table',
                        'Scope': sheet_name,
                        'Hidden': False,
                        'Definition': f'{sheet_name}!{table_ref}',
                        'Human_Definition': f'{sheet_name}!{table_ref}',
                        'Parameters': '',
                    })
            workbook.close()

        profile = pd.DataFrame(rows, columns=[
            'Workbook', 'Name', 'Object_Type', 'Scope', 'Hidden',
            'Definition', 'Human_Definition', 'Parameters'])
        logger.info(
            f"🏷️  Profiled {len(self.workbooks)} workbook(s): "
            f"{len(profile)} named objects")
        return profile

    def _defined_name_row(self, path: str, workbook, defined_name,
                          scope_override=None) -> dict:
        """One contract row for a defined name."""
        stored = defined_name.attr_text or ''
        object_type = detect_object_type(defined_name)

        parameters = ''
        if object_type == 'lambda':
            human, param_names = translate_lambda_to_human(stored)
            parameters = ', '.join(param_names)
        elif object_type == 'formula':
            human = clean_formula_for_display(stored)
        else:
            human = stored

        scope = scope_override or 'global'
        if scope_override is None and defined_name.localSheetId is not None:
            scope = workbook.sheetnames[defined_name.localSheetId]

        return {
            'Workbook': path,
            'Name': defined_name.name,
            'Object_Type': object_type,
            'Scope': scope,
            'Hidden': bool(defined_name.hidden),
            'Definition': stored,
            'Human_Definition': human,
            'Parameters': parameters,
        }

# End of file #

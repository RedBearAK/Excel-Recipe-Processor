"""
Functional tests for the manage_named_objects write operations.

tests/test_manage_named_objects_write.py

Runnable with pytest, but written to run standalone and report a score.
Exercises real workbooks on disk rather than mocks, because the whole point
of these operations is what Excel ends up seeing.
"""

import openpyxl
import tempfile

from pathlib import Path

from excel_recipe_processor.processors.manage_named_objects_processor import (
    ManageNamedObjectsProcessor
)


def build_lookup_workbook(path: str) -> None:
    """Build a workbook shaped like the VMS product lookup table."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Product_IDs'

    worksheet.append([
        'Major Species', 'Species', 'Product ID', 'Product Name',
        'Component', 'Product Form', 'Product Group', 'Can Size', 'Pack Size'
    ])

    for row_num in range(1, 31):
        can_size = 14.75 if row_num <= 5 else None
        worksheet.append([
            'SALMON', 'SOCKEYE', 10000 + row_num, f'Product {row_num}',
            'FLESH', 'IQF H&G', 'FROZEN IQF', can_size, None
        ])

    orders = workbook.create_sheet('Region-Carrier')
    orders.append(['Plant Origin', 'Region'])
    for name, region in [('Sitka', 'Southeast'), ('Craig', 'Southeast'),
                         ('Valdez', 'Prince William Sound')]:
        orders.append([name, region])

    workbook.save(path)
    workbook.close()


def test_create_from_columns():
    """Named ranges are computed from column names and real data extent."""
    print("\nTesting create_from_columns...")

    with tempfile.TemporaryDirectory() as temp_dir:
        target = str(Path(temp_dir) / 'lookup.xlsx')
        build_lookup_workbook(target)

        processor = ManageNamedObjectsProcessor({
            'processor_type': 'manage_named_objects',
            'operation': 'create_from_columns',
            'target_file': target,
            'ranges': [
                {'name': 'rng_PID', 'sheet_name': 'Product_IDs',
                 'columns': ['Product ID'], 'row_mode': 'data'},
                {'name': 'rng_Prod_Form', 'sheet_name': 'Product_IDs',
                 'columns': ['Product Form'], 'row_mode': 'data'},
                {'name': 'rng_PlantOrig', 'sheet_name': 'Region-Carrier',
                 'columns': ['Plant Origin'], 'row_mode': 'data'},
            ]
        })

        result = processor.execute()

        if result['ranges_written'] != 3:
            print(f"  ✗ Expected 3 ranges, wrote {result['ranges_written']}")
            return False

        workbook = openpyxl.load_workbook(target)
        actual = {name: dn.attr_text for name, dn in workbook.defined_names.items()}
        workbook.close()

        expected = {
            'rng_PID': 'Product_IDs!$C$2:$C$31',
            'rng_Prod_Form': 'Product_IDs!$F$2:$F$31',
            'rng_PlantOrig': "'Region-Carrier'!$A$2:$A$4",
        }

        passed = True

        for name, reference in expected.items():
            if actual.get(name) == reference:
                print(f"  ✓ {name:16} -> {reference}")
            else:
                print(f"  ✗ {name:16} -> {actual.get(name)}, expected {reference}")
                passed = False

        return passed


def test_sparse_column_needs_anchor():
    """A sparse column truncates unless anchored on a dense one."""
    print("\nTesting sparse column anchoring...")

    with tempfile.TemporaryDirectory() as temp_dir:
        target = str(Path(temp_dir) / 'lookup.xlsx')
        build_lookup_workbook(target)

        processor = ManageNamedObjectsProcessor({
            'processor_type': 'manage_named_objects',
            'operation': 'create_from_columns',
            'target_file': target,
            'ranges': [
                {'name': 'rng_CanSize_bad', 'sheet_name': 'Product_IDs',
                 'columns': ['Can Size'], 'row_mode': 'data'},
                {'name': 'rng_CanSize_good', 'sheet_name': 'Product_IDs',
                 'columns': ['Can Size'], 'row_mode': 'data',
                 'anchor_columns': ['Product ID']},
            ]
        })

        processor.execute()

        workbook = openpyxl.load_workbook(target)
        actual = {name: dn.attr_text for name, dn in workbook.defined_names.items()}
        workbook.close()

        bad = actual.get('rng_CanSize_bad')
        good = actual.get('rng_CanSize_good')

        if bad == 'Product_IDs!$H$2:$H$6' and good == 'Product_IDs!$H$2:$H$31':
            print(f"  ✓ Unanchored truncates to {bad}")
            print(f"  ✓ Anchored extends to    {good}")
            return True

        print(f"  ✗ Got unanchored={bad}, anchored={good}")
        return False


def test_column_span():
    """A span across named endpoints yields one block reference."""
    print("\nTesting column span...")

    with tempfile.TemporaryDirectory() as temp_dir:
        target = str(Path(temp_dir) / 'lookup.xlsx')
        build_lookup_workbook(target)

        processor = ManageNamedObjectsProcessor({
            'processor_type': 'manage_named_objects',
            'operation': 'create_from_columns',
            'target_file': target,
            'ranges': [
                {'name': 'rng_prodinfo', 'sheet_name': 'Product_IDs',
                 'columns': ['Species', 'Component'], 'row_mode': 'data'},
            ]
        })

        processor.execute()

        workbook = openpyxl.load_workbook(target)
        actual = workbook.defined_names['rng_prodinfo'].attr_text
        workbook.close()

        if actual == 'Product_IDs!$B$2:$E$31':
            print(f"  ✓ Span resolved to {actual}")
            return True

        print(f"  ✗ Got {actual}, expected Product_IDs!$B$2:$E$31")
        return False


def test_house_style_rejected():
    """create_from_columns enforces house style on names by default."""
    print("\nTesting house style enforcement...")

    with tempfile.TemporaryDirectory() as temp_dir:
        target = str(Path(temp_dir) / 'lookup.xlsx')
        build_lookup_workbook(target)

        processor = ManageNamedObjectsProcessor({
            'processor_type': 'manage_named_objects',
            'operation': 'create_from_columns',
            'target_file': target,
            'ranges': [
                {'name': 'rng_PID2026', 'sheet_name': 'Product_IDs',
                 'columns': ['Product ID'], 'row_mode': 'data'},
            ]
        })

        try:
            processor.execute()
        except Exception as error:
            if 'separator before digits' in str(error):
                print("  ✓ 'rng_PID2026' rejected under house style")
                return True
            print(f"  ✗ Wrong rejection reason: {error}")
            return False

        print("  ✗ 'rng_PID2026' was accepted")
        return False


def test_on_existing_policies():
    """error, replace, and skip all behave as documented."""
    print("\nTesting on_existing policies...")

    with tempfile.TemporaryDirectory() as temp_dir:
        target = str(Path(temp_dir) / 'lookup.xlsx')
        build_lookup_workbook(target)

        base_spec = [{'name': 'rng_PID', 'sheet_name': 'Product_IDs',
                      'columns': ['Product ID'], 'row_mode': 'data'}]

        def run(policy):
            return ManageNamedObjectsProcessor({
                'processor_type': 'manage_named_objects',
                'operation': 'create_from_columns',
                'target_file': target,
                'on_existing': policy,
                'ranges': base_spec
            }).execute()

        run('error')
        passed = True

        try:
            run('error')
            print("  ✗ 'error' policy did not raise on a second write")
            passed = False
        except Exception:
            print("  ✓ 'error' policy raised on a second write")

        result = run('skip')
        if len(result['skipped']) == 1:
            print("  ✓ 'skip' policy left the existing name alone")
        else:
            print(f"  ✗ 'skip' policy gave {result}")
            passed = False

        result = run('replace')
        if len(result['replaced']) == 1:
            print("  ✓ 'replace' policy replaced the name")
        else:
            print(f"  ✗ 'replace' policy gave {result}")
            passed = False

        return passed


def test_sheet_scoped_names():
    """Local scope writes to the worksheet, not the workbook."""
    print("\nTesting sheet-scoped names...")

    with tempfile.TemporaryDirectory() as temp_dir:
        target = str(Path(temp_dir) / 'lookup.xlsx')
        build_lookup_workbook(target)

        ManageNamedObjectsProcessor({
            'processor_type': 'manage_named_objects',
            'operation': 'create_from_columns',
            'target_file': target,
            'ranges': [
                {'name': 'rng_local_pid', 'sheet_name': 'Product_IDs',
                 'columns': ['Product ID'], 'row_mode': 'data', 'scope': 'local'},
                {'name': 'rng_global_pid', 'sheet_name': 'Product_IDs',
                 'columns': ['Product ID'], 'row_mode': 'data'},
            ]
        }).execute()

        workbook = openpyxl.load_workbook(target)
        global_names = list(workbook.defined_names.keys())
        local_names = list(workbook['Product_IDs'].defined_names.keys())
        workbook.close()

        if 'rng_global_pid' in global_names and 'rng_local_pid' in local_names:
            if 'rng_local_pid' not in global_names:
                print("  ✓ Local name on the sheet, global name on the workbook")
                return True

        print(f"  ✗ global={global_names}, local={local_names}")
        return False


def test_export_then_import_round_trip():
    """Names survive an export to YAML and an import into a fresh workbook."""
    print("\nTesting export/import round trip...")

    with tempfile.TemporaryDirectory() as temp_dir:
        source = str(Path(temp_dir) / 'source.xlsx')
        target = str(Path(temp_dir) / 'target.xlsx')
        yaml_path = str(Path(temp_dir) / 'names.yaml')

        build_lookup_workbook(source)
        build_lookup_workbook(target)

        ManageNamedObjectsProcessor({
            'processor_type': 'manage_named_objects',
            'operation': 'create_from_columns',
            'target_file': source,
            'ranges': [
                {'name': 'rng_PID', 'sheet_name': 'Product_IDs',
                 'columns': ['Product ID'], 'row_mode': 'data'},
                {'name': 'rng_region', 'sheet_name': 'Region-Carrier',
                 'columns': ['Region'], 'row_mode': 'data'},
            ]
        }).execute()

        export_result = ManageNamedObjectsProcessor({
            'processor_type': 'manage_named_objects',
            'operation': 'export_all',
            'source_file': source,
            'yaml_file': yaml_path
        }).execute()

        if export_result['objects_exported'] < 2:
            print(f"  ✗ Exported only {export_result['objects_exported']} objects")
            return False

        print(f"  ✓ Exported {export_result['objects_exported']} objects to YAML")

        import_result = ManageNamedObjectsProcessor({
            'processor_type': 'manage_named_objects',
            'operation': 'import_all',
            'yaml_file': yaml_path,
            'target_file': target
        }).execute()

        workbook = openpyxl.load_workbook(target)
        actual = {name: dn.attr_text for name, dn in workbook.defined_names.items()}
        workbook.close()

        if actual.get('rng_PID') == 'Product_IDs!$C$2:$C$31':
            if actual.get('rng_region') == "'Region-Carrier'!$B$2:$B$4":
                print(f"  ✓ Imported {import_result['objects_written']} names intact")
                return True

        print(f"  ✗ Round trip produced {actual}")
        return False


def test_copy_direct():
    """Names copy straight from one workbook into another."""
    print("\nTesting copy_direct...")

    with tempfile.TemporaryDirectory() as temp_dir:
        source = str(Path(temp_dir) / 'source.xlsx')
        target = str(Path(temp_dir) / 'target.xlsx')

        build_lookup_workbook(source)
        build_lookup_workbook(target)

        ManageNamedObjectsProcessor({
            'processor_type': 'manage_named_objects',
            'operation': 'create_from_columns',
            'target_file': source,
            'ranges': [
                {'name': 'rng_PID', 'sheet_name': 'Product_IDs',
                 'columns': ['Product ID'], 'row_mode': 'data'},
                {'name': 'rng_ProdGrp', 'sheet_name': 'Product_IDs',
                 'columns': ['Product Group'], 'row_mode': 'data'},
            ]
        }).execute()

        result = ManageNamedObjectsProcessor({
            'processor_type': 'manage_named_objects',
            'operation': 'copy_direct',
            'source_file': source,
            'target_file': target
        }).execute()

        workbook = openpyxl.load_workbook(target)
        names = set(workbook.defined_names.keys())
        workbook.close()

        if {'rng_PID', 'rng_ProdGrp'}.issubset(names):
            print(f"  ✓ Copied {result['objects_written']} names")
            return True

        print(f"  ✗ Target has {names}")
        return False


def test_validate_yaml_finds_problems():
    """validate_yaml reports bad names without touching any workbook."""
    print("\nTesting validate_yaml...")

    with tempfile.TemporaryDirectory() as temp_dir:
        yaml_path = Path(temp_dir) / 'bad.yaml'
        yaml_path.write_text(
            "metadata:\n"
            "  export_date: '2026-08-06'\n"
            "named_ranges:\n"
            "  - name: Q1\n"
            "    type: range\n"
            "    scope: global\n"
            "    excel_definition: Sheet1!$A$1:$A$5\n"
            "lambda_functions: []\n"
            "named_formulas: []\n"
            "named_tables: []\n"
            "local_objects: {}\n"
        )

        result = ManageNamedObjectsProcessor({
            'processor_type': 'manage_named_objects',
            'operation': 'validate_yaml',
            'yaml_file': str(yaml_path)
        }).execute()

        if not result['valid'] and len(result['problems']) >= 1:
            reason = result['problems'][0]['problem']
            if 'cell reference' in reason:
                print(f"  ✓ Flagged 'Q1': {reason[:60]}...")
                return True

        print(f"  ✗ Got {result}")
        return False


def main():
    """Run every test and report a final score."""
    print("=== Manage Named Objects Write Operation Tests ===")

    tests = [
        test_create_from_columns,
        test_sparse_column_needs_anchor,
        test_column_span,
        test_house_style_rejected,
        test_on_existing_policies,
        test_sheet_scoped_names,
        test_export_then_import_round_trip,
        test_copy_direct,
        test_validate_yaml_finds_problems,
    ]

    passed = 0

    for test_func in tests:
        try:
            if test_func():
                passed += 1
        except Exception as error:
            print(f"  ✗ {test_func.__name__} crashed: {error}")

    print(f"\n=== Results: {passed}/{len(tests)} tests passed ===")

    if passed == len(tests):
        print("✅ All named object write tests passed!")
        return 1

    print("❌ Some named object write tests failed!")
    return 0


if __name__ == '__main__':
    exit(0 if main() else 1)


# End of file #

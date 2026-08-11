"""
Tests for the WorkbookSession shared-workbook cache.

tests/test_workbook_session.py

Runnable with pytest, but written to run standalone and report a score.
"""

import tempfile

from pathlib import Path

import openpyxl

from excel_recipe_processor.core.workbook_session import WorkbookSession, WorkbookSessionError


def make_workbook(path, value='original'):
    wb = openpyxl.Workbook()
    wb.active['A1'] = value
    wb.save(path)
    wb.close()


def test_second_get_returns_the_same_object():
    """Identity is the whole point: mutations accumulate across steps."""
    print("\nTesting load-once identity...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'w.xlsx')
        make_workbook(path)
        WorkbookSession.reset()
        WorkbookSession.set_deferred(True)

        first = WorkbookSession.get_workbook(path)
        first.active['B1'] = 'mutation from step one'
        second = WorkbookSession.get_workbook(path)

        passed = second is first and second.active['B1'].value == 'mutation from step one'
        print("  ✓ Same object, earlier mutation visible" if passed else "  ✗ Different objects")

        WorkbookSession.reset()
        return passed


def test_flush_writes_and_empties_the_session():
    """Dirty workbooks reach disk exactly at flush; the session then forgets."""
    print("\nTesting flush...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'w.xlsx')
        make_workbook(path)
        WorkbookSession.reset()
        WorkbookSession.set_deferred(True)

        wb = WorkbookSession.get_workbook(path)
        wb.active['A1'] = 'changed'
        WorkbookSession.mark_dirty(path)

        on_disk_before = openpyxl.load_workbook(path).active['A1'].value
        written = WorkbookSession.flush_all()
        on_disk_after = openpyxl.load_workbook(path).active['A1'].value

        passed = True

        if on_disk_before == 'original' and on_disk_after == 'changed':
            print("  ✓ Disk unchanged until flush, changed after")
        else:
            print(f"  ✗ before {on_disk_before!r}, after {on_disk_after!r}")
            passed = False

        if written == 1 and not WorkbookSession.is_open(path):
            print("  ✓ One file written; session empty")
        else:
            print(f"  ✗ written {written}, still open {WorkbookSession.is_open(path)}")
            passed = False

        return passed


def test_discard_leaves_disk_untouched():
    """The failure path: mutations vanish, the file stays as last written."""
    print("\nTesting discard...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'w.xlsx')
        make_workbook(path)
        WorkbookSession.reset()
        WorkbookSession.set_deferred(True)

        wb = WorkbookSession.get_workbook(path)
        wb.active['A1'] = 'doomed mutation'
        WorkbookSession.mark_dirty(path)
        discarded = WorkbookSession.discard_all()

        on_disk = openpyxl.load_workbook(path).active['A1'].value

        if discarded == 1 and on_disk == 'original':
            print("  ✓ One workbook discarded; disk untouched")
            return True
        print(f"  ✗ discarded {discarded}, disk {on_disk!r}")
        return False


def test_multiple_files_tracked_independently():
    """A recipe may touch several files; only dirty ones write."""
    print("\nTesting multiple files...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path_a = str(Path(temp_dir) / 'a.xlsx')
        path_b = str(Path(temp_dir) / 'b.xlsx')
        make_workbook(path_a, 'a-original')
        make_workbook(path_b, 'b-original')
        WorkbookSession.reset()
        WorkbookSession.set_deferred(True)

        WorkbookSession.get_workbook(path_a).active['A1'] = 'a-changed'
        WorkbookSession.mark_dirty(path_a)
        WorkbookSession.get_workbook(path_b)  # opened, never dirtied

        written = WorkbookSession.flush_all()

        a_disk = openpyxl.load_workbook(path_a).active['A1'].value
        b_disk = openpyxl.load_workbook(path_b).active['A1'].value

        if written == 1 and a_disk == 'a-changed' and b_disk == 'b-original':
            print("  ✓ Dirty file written, clean file untouched")
            return True
        print(f"  ✗ written {written}, a {a_disk!r}, b {b_disk!r}")
        return False


def test_adopt_defers_the_export_write():
    """An adopted workbook exists nowhere on disk until flush."""
    print("\nTesting adopt_workbook (the export bridge)...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'bridged.xlsx')
        WorkbookSession.reset()
        WorkbookSession.set_deferred(True)

        wb = openpyxl.Workbook()
        wb.active['A1'] = 'exported value'
        WorkbookSession.adopt_workbook(path, wb)

        passed = True

        if not Path(path).exists():
            print("  ✓ Nothing on disk before flush")
        else:
            print("  ✗ File exists before flush")
            passed = False

        WorkbookSession.flush_all()

        on_disk = openpyxl.load_workbook(path).active['A1'].value
        if on_disk == 'exported value':
            print("  ✓ Flush produced the file with the exported data")
        else:
            print(f"  ✗ disk {on_disk!r}")
            passed = False

        return passed


def test_adopt_refuses_a_held_path():
    """Two exports to one path without a flush is a loud error."""
    print("\nTesting adopt collision...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'twice.xlsx')
        WorkbookSession.reset()
        WorkbookSession.set_deferred(True)

        WorkbookSession.adopt_workbook(path, openpyxl.Workbook())

        try:
            WorkbookSession.adopt_workbook(path, openpyxl.Workbook())
            print("  ✗ Second adoption accepted")
            WorkbookSession.reset()
            return False
        except WorkbookSessionError:
            print("  ✓ Refused with WorkbookSessionError")
            WorkbookSession.reset()
            return True


def test_adopt_saves_immediately_when_standalone():
    """Outside a pipeline, adoption degenerates to a plain save."""
    print("\nTesting standalone adoption...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'standalone.xlsx')
        WorkbookSession.reset()   # deferred defaults to False

        wb = openpyxl.Workbook()
        wb.active['A1'] = 'direct'
        WorkbookSession.adopt_workbook(path, wb)

        exists = Path(path).exists()
        held = WorkbookSession.is_open(path)
        WorkbookSession.reset()

        if exists and not held:
            print("  ✓ Saved immediately, session holds nothing")
            return True
        print(f"  ✗ exists {exists}, held {held}")
        return False


def test_immediate_mode_is_the_standalone_default():
    """Outside a pipeline, mark_dirty saves right away - legacy semantics."""
    print("\nTesting immediate mode (the default)...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'w.xlsx')
        make_workbook(path)
        WorkbookSession.reset()   # deferred defaults to False

        wb = WorkbookSession.get_workbook(path)
        wb.active['A1'] = 'standalone change'
        WorkbookSession.mark_dirty(path)

        on_disk = openpyxl.load_workbook(path).active['A1'].value

        WorkbookSession.reset()

        if on_disk == 'standalone change':
            print("  \u2713 Saved immediately, no flush needed")
            return True
        print(f"  \u2717 disk {on_disk!r}")
        return False


def test_mark_dirty_requires_a_loaded_path():
    """Dirtying a never-loaded path is a programming error and fails loudly."""
    print("\nTesting mark_dirty guard...")

    WorkbookSession.reset()

    try:
        WorkbookSession.mark_dirty('/nonexistent/never_loaded.xlsx')
        print("  ✗ Accepted a never-loaded path")
        return False
    except WorkbookSessionError:
        print("  ✓ Refused with WorkbookSessionError")
        return True


def main():
    """Run every test and report a final score."""
    print("=== WorkbookSession tests ===")

    tests = [
        test_second_get_returns_the_same_object,
        test_flush_writes_and_empties_the_session,
        test_discard_leaves_disk_untouched,
        test_multiple_files_tracked_independently,
        test_adopt_defers_the_export_write,
        test_adopt_refuses_a_held_path,
        test_adopt_saves_immediately_when_standalone,
        test_immediate_mode_is_the_standalone_default,
        test_mark_dirty_requires_a_loaded_path,
    ]

    passed = 0

    for test_func in tests:
        try:
            if test_func():
                passed += 1
        except Exception as error:
            print(f"  ✗ {test_func.__name__} crashed: {error}")

    print(f"\n=== Results: {passed}/{len(tests)} tests passed ===")

    if passed == len(tests):
        print("✅ All WorkbookSession tests passed!")
        return 1

    print("❌ Some WorkbookSession tests failed!")
    return 0


if __name__ == '__main__':
    exit(0 if main() else 1)


# End of file #

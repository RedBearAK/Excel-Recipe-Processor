"""
Sheet-level visual features for format_excel: banded rows, outline borders.

excel_recipe_processor/processors/_helpers/format_excel_sheet_features.py

Banded rows paint every second data row for row tracking, and banding WINS
over column tints by design: the tint marks a column's provenance and still
reads through on the off-band rows, while the band's whole purpose is
keeping the eye on a row, so it must be continuous. The processor enforces
that ordering by calling apply_banded_rows AFTER the column-formats pass
and BEFORE cell_formats, so explicit spot rules still beat everything.

Outline borders draw a box around a range - the whole used range by
default, or explicit A1-style ranges. Only the outward-facing sides of the
perimeter cells are touched, and existing sides are preserved, so a box
can land on top of per-cell gridline-gray borders without erasing them.
"""

import logging

from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries

from excel_recipe_processor.processors._helpers.range_patterns import range_ref_rgx
from excel_recipe_processor.processors._helpers.excel_range_resolver import (
    find_last_data_row, ExcelRangeResolverError
)


logger = logging.getLogger(__name__)


def sheet_data_extent(worksheet, header_row: int) -> int:
    """
    Last data row of the sheet, measured from the widest column.

    The same one-extent-for-the-whole-sheet measurement the column
    formats pass uses, so banding and outline borders agree with the
    column rules about where the data ends.
    """
    if worksheet.max_column < 1:
        return header_row
    anchor_letters = [
        get_column_letter(i) for i in range(1, worksheet.max_column + 1)
    ]
    try:
        return find_last_data_row(worksheet, anchor_letters, header_row)
    except ExcelRangeResolverError as error:
        raise SheetFeatureError(f"Could not determine data extent: {error}")

# The gridline remedy color, same doctrine as column background_color:
# a fill suppresses Excel's gridlines, thin D9D9D9 restores the ruling
GRIDLINE_GRAY = 'D9D9D9'


class SheetFeatureError(ValueError):
    """A sheet-level feature spec is invalid."""
    pass


def apply_banded_rows(worksheet, band_color: str, header_row: int,
                      last_row: int, border_style=None,
                      border_color=None) -> int:
    """
    Fill every second data row with band_color, across all used columns.

    The first data row stays unbanded, the second is banded, and so on -
    the same rhythm Excel's own table styles use. Optionally rules the
    banded cells with a border (fills suppress gridlines; a thin border
    in gridline gray restores the ruled look on the banded rows).

    Args:
        worksheet:      openpyxl worksheet object
        band_color:     6-digit hex fill for the banded rows
        header_row:     Row holding the headers
        last_row:       Last data row (inclusive)
        border_style:   Optional openpyxl border style name
        border_color:   6-digit hex for the border; defaults to
                        gridline gray when border_style is given

    Returns:
        Number of rows banded
    """
    if last_row <= header_row:
        logger.warning(
            f"[{worksheet.title}] banded_row_color skipped: no data rows"
        )
        return 0

    band_fill = PatternFill(
        start_color=band_color, end_color=band_color, fill_type='solid'
    )

    band_border = None
    if border_style is not None:
        side_color = border_color if border_color else GRIDLINE_GRAY
        edge = Side(style=border_style, color=side_color)
        band_border = Border(left=edge, right=edge, top=edge, bottom=edge)

    last_column = worksheet.max_column
    banded_count = 0

    # Second data row, fourth, sixth... (header_row + 2, + 4, ...)
    for row_num in range(header_row + 2, last_row + 1, 2):
        for col_num in range(1, last_column + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.fill = band_fill
            if band_border is not None:
                cell.border = band_border
        banded_count += 1

    return banded_count


def _merged_side(new_side, existing_side):
    """Prefer the new side; keep the existing one where no new side."""
    if new_side is not None:
        return new_side
    return existing_side


def apply_outline_border(worksheet, ranges, style: str, color: str,
                         header_row: int, last_row: int) -> list:
    """
    Draw a box around each given range, preserving interior cell borders.

    Only the outward-facing sides of perimeter cells change: a top-row
    cell gets the box's top side, a corner cell gets two sides, and any
    sides that cell already had (gridline-gray ruling, say) survive.

    Args:
        worksheet:      openpyxl worksheet object
        ranges:         None (whole used range), an A1-style range
                        string, or a list of range strings
        style:          openpyxl border style name for the box
        color:          6-digit hex for the box
        header_row:     Row holding the headers (used range default)
        last_row:       Last data row (used range default)

    Returns:
        List of the range strings actually boxed

    Raises:
        SheetFeatureError: a range string is malformed
    """
    if ranges is None:
        if last_row < header_row or worksheet.max_column < 1:
            logger.warning(
                f"[{worksheet.title}] outline border skipped: empty sheet"
            )
            return []
        last_letter = get_column_letter(worksheet.max_column)
        range_list = [f"A{header_row}:{last_letter}{last_row}"]
    elif isinstance(ranges, str):
        range_list = [ranges]
    elif isinstance(ranges, list):
        range_list = [str(entry) for entry in ranges]
    else:
        raise SheetFeatureError(
            f"outline_border_range must be a range string like 'B2:F40' "
            f"or a list of them, got: {type(ranges).__name__}"
        )

    edge = Side(style=style, color=color)
    boxed = []

    for range_text in range_list:
        candidate = range_text.strip()
        if not range_ref_rgx.match(candidate):
            raise SheetFeatureError(
                f"outline_border_range entry {candidate!r} is not an "
                f"A1-style range (like 'B2:F40')"
            )

        min_col, min_row, max_col, max_row = range_boundaries(candidate)

        for col_num in range(min_col, max_col + 1):
            for row_num in (min_row, max_row):
                cell = worksheet.cell(row=row_num, column=col_num)
                existing = cell.border
                cell.border = Border(
                    top=_merged_side(
                        edge if row_num == min_row else None, existing.top),
                    bottom=_merged_side(
                        edge if row_num == max_row else None, existing.bottom),
                    left=_merged_side(
                        edge if col_num == min_col else None, existing.left),
                    right=_merged_side(
                        edge if col_num == max_col else None, existing.right),
                )

        for row_num in range(min_row + 1, max_row):
            for col_num in (min_col, max_col):
                cell = worksheet.cell(row=row_num, column=col_num)
                existing = cell.border
                cell.border = Border(
                    top=existing.top,
                    bottom=existing.bottom,
                    left=_merged_side(
                        edge if col_num == min_col else None, existing.left),
                    right=_merged_side(
                        edge if col_num == max_col else None, existing.right),
                )

        boxed.append(candidate)

    return boxed


# End of file #

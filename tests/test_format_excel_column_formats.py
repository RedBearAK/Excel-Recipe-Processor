"""
Tests for column-addressed formatting in format_excel.

tests/test_format_excel_column_formats.py

Runnable with pytest, but written to run standalone and report a score.
Covers number formats, column hiding, per-column alignment, and the shared
column-name resolution they all depend on.
"""

import openpyxl
import pandas as pd
import tempfile

from pathlib import Path

from excel_recipe_processor.processors.format_excel_processor import FormatExcelProcessor
from excel_recipe_processor.processors._helpers.format_excel_column_formats import (
    resolve_number_format, NUMBER_FORMAT_ALIASES, ColumnFormatError
)


def build_workbook(path: str) -> None:
    """Build a sheet shaped like the VMS output."""
    frame = pd.DataFrame({
        'Product ID': [10001, 10002, 10003],
        'Cases (24)': [1234, 56, 7890],
        'Gross Wt': [4500, 220, 33000],
        'Price': [123.45, 0.0, -99.99],
        'Total Price': [1000.0, 2000.5, -30.0],
        'Notes': ['a', 'b', 'c'],
        'Original Van Numbers': ['V1', 'V2', 'V3'],
    })
    frame.to_excel(path, index=False, sheet_name='VMS Data')


def run_format(path: str, sheet_config: dict):
    """Apply one sheet formatting configuration."""
    config = {'sheet_names': ['VMS Data']}
    config.update(sheet_config)
    return FormatExcelProcessor({
        'processor_type': 'format_excel',
        'target_file': path,
        'formatting': [config]
    }).execute()


def test_number_format_aliases():
    """Aliases resolve to Excel codes, literals pass through untouched."""
    print("\nTesting number format aliases...")

    passed = True

    if resolve_number_format('thousands') == '#,##0':
        print("  ✓ 'thousands' -> '#,##0'")
    else:
        print(f"  ✗ 'thousands' -> {resolve_number_format('thousands')}")
        passed = False

    accounting = resolve_number_format('accounting')
    if accounting.startswith('_($*') and '(' in accounting:
        print("  ✓ 'accounting' produced a parenthesised-negative code")
    else:
        print(f"  ✗ 'accounting' -> {accounting}")
        passed = False

    if resolve_number_format('0.000') == '0.000':
        print("  ✓ literal code '0.000' passed through")
    else:
        print("  ✗ literal code was altered")
        passed = False

    return passed


def test_number_formats_applied():
    """Number formats land on data cells."""
    print("\nTesting number formats on columns...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'wb.xlsx')
        build_workbook(path)

        run_format(path, {'column_formats': [
            {'columns': ['Cases (24)', 'Gross Wt'], 'number_format': 'thousands'},
            {'columns': ['Price', 'Total Price'], 'number_format': 'accounting'},
        ]})

        workbook = openpyxl.load_workbook(path)
        worksheet = workbook['VMS Data']
        results = {
            'Cases (24)': worksheet['B2'].number_format,
            'Gross Wt': worksheet['C2'].number_format,
            'Price': worksheet['D2'].number_format,
        }
        workbook.close()

        passed = True

        if results['Cases (24)'] == '#,##0' and results['Gross Wt'] == '#,##0':
            print("  ✓ Both quantity columns got '#,##0'")
        else:
            print(f"  ✗ Quantity columns: {results}")
            passed = False

        if results['Price'] == NUMBER_FORMAT_ALIASES['accounting']:
            print("  ✓ Price got the accounting code")
        else:
            print(f"  ✗ Price: {results['Price']}")
            passed = False

        return passed


def test_header_row_untouched():
    """Formatting applies below the header, not to it."""
    print("\nTesting the header row is left alone...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'wb.xlsx')
        build_workbook(path)

        run_format(path, {'column_formats': [
            {'columns': ['Cases (24)'], 'number_format': 'thousands'},
        ]})

        workbook = openpyxl.load_workbook(path)
        header_format = workbook['VMS Data']['B1'].number_format
        data_format = workbook['VMS Data']['B2'].number_format
        workbook.close()

        if header_format == 'General' and data_format == '#,##0':
            print("  ✓ Header still General, data formatted")
            return True

        print(f"  ✗ header={header_format} data={data_format}")
        return False


def test_column_hiding():
    """Named columns are hidden, others are not."""
    print("\nTesting column hiding...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'wb.xlsx')
        build_workbook(path)

        run_format(path, {'hidden_columns': ['Notes', 'Original Van Numbers']})

        workbook = openpyxl.load_workbook(path)
        worksheet = workbook['VMS Data']
        hidden = {c: bool(worksheet.column_dimensions[c].hidden) for c in 'ABCDEFG'}
        workbook.close()

        if hidden['F'] and hidden['G'] and not any(hidden[c] for c in 'ABCDE'):
            print("  ✓ Notes (F) and Original Van Numbers (G) hidden, rest visible")
            return True

        print(f"  ✗ {hidden}")
        return False


def test_per_column_alignment():
    """Alignment applies to one column without disturbing others."""
    print("\nTesting per-column alignment...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'wb.xlsx')
        build_workbook(path)

        run_format(path, {'column_formats': [
            {'columns': ['Product ID'], 'alignment_horizontal': 'center'},
        ]})

        workbook = openpyxl.load_workbook(path)
        worksheet = workbook['VMS Data']
        target = worksheet['A2'].alignment.horizontal
        neighbour = worksheet['B2'].alignment.horizontal
        workbook.close()

        if target == 'center' and neighbour != 'center':
            print("  ✓ Product ID centred, neighbour untouched")
            return True

        print(f"  ✗ Product ID={target} neighbour={neighbour}")
        return False


def test_missing_column_policy():
    """An unknown column warns by default and raises when asked to."""
    print("\nTesting missing column handling...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'wb.xlsx')
        build_workbook(path)
        passed = True

        try:
            run_format(path, {'column_formats': [
                {'columns': ['Product ID', 'Nonexistent'], 'number_format': 'thousands'},
            ]})
            print("  ✓ 'warn' default did not halt the run")
        except Exception as error:
            print(f"  ✗ Default policy raised: {error}")
            passed = False

        try:
            run_format(path, {
                'on_missing_column': 'error',
                'column_formats': [
                    {'columns': ['Nonexistent'], 'number_format': 'thousands'},
                ]})
            print("  ✗ 'error' policy did not raise")
            passed = False
        except Exception:
            print("  ✓ 'error' policy raised")

        return passed


def test_rule_with_no_effect_is_rejected():
    """A rule specifying no action is a configuration mistake."""
    print("\nTesting empty rule rejection...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'wb.xlsx')
        build_workbook(path)

        try:
            run_format(path, {'column_formats': [{'columns': ['Price']}]})
        except Exception as error:
            if 'does nothing' in str(error):
                print("  ✓ Rule with no action rejected with a clear message")
                return True
            print(f"  ✗ Wrong error: {error}")
            return False

        print("  ✗ Empty rule was accepted")
        return False


def test_per_column_header_styling():
    """A rule can restyle only its own columns' headers."""
    print("\nTesting per-column header styling...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'wb.xlsx')
        build_workbook(path)

        run_format(path, {
            'header_background': True,
            'header_background_color': '1F4E79',
            'header_text_color': 'white',
            'column_formats': [
                {'columns': ['Cases (24)', 'Price'],
                 'font_color': 'red',
                 'header_font_color': 'white',
                 'header_background_color': 'red',
                 'header_bold': True},
            ]})

        workbook = openpyxl.load_workbook(path)
        worksheet = workbook['VMS Data']

        def fill(ref):
            colour = worksheet[ref].fill.start_color.rgb
            return colour if isinstance(colour, str) else None

        def font(ref):
            colour = worksheet[ref].font.color
            return colour.rgb if colour and isinstance(colour.rgb, str) else None

        targeted_header = fill('B1')
        other_header = fill('A1')
        targeted_data = font('B2')
        other_data = font('A2')
        workbook.close()

        passed = True

        if targeted_header and targeted_header.endswith('FF0000'):
            print("  ✓ Targeted header went red")
        else:
            print(f"  ✗ Targeted header fill is {targeted_header}")
            passed = False

        if other_header and other_header.endswith('1F4E79'):
            print("  ✓ Untargeted header kept the sheet-wide colour")
        else:
            print(f"  ✗ Untargeted header fill is {other_header}")
            passed = False

        if targeted_data and targeted_data.endswith('FF0000'):
            print("  ✓ Targeted data font went red")
        else:
            print(f"  ✗ Targeted data font is {targeted_data}")
            passed = False

        if not (other_data and other_data.endswith('FF0000')):
            print("  ✓ Untargeted data font left alone")
        else:
            print(f"  ✗ Untargeted data font is {other_data}")
            passed = False

        return passed


def test_css_colour_names_resolve():
    """CSS names work here the same way they do in header_* options."""
    print("\nTesting CSS colour names...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'wb.xlsx')
        build_workbook(path)

        run_format(path, {'column_formats': [
            {'columns': ['Price'], 'font_color': 'forestgreen'},
        ]})

        workbook = openpyxl.load_workbook(path)
        colour = workbook['VMS Data']['D2'].font.color
        value = colour.rgb if colour and isinstance(colour.rgb, str) else None
        workbook.close()

        if value and value.endswith('228B22'):
            print(f"  ✓ 'forestgreen' resolved to {value}")
            return True

        print(f"  ✗ Got {value}")
        return False


def test_number_format_and_font_combine():
    """One rule can set a number format and a font colour together."""
    print("\nTesting combined number format and font colour...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'wb.xlsx')
        build_workbook(path)

        run_format(path, {'column_formats': [
            {'columns': ['Cases (24)'], 'number_format': 'thousands', 'font_color': 'red'},
        ]})

        workbook = openpyxl.load_workbook(path)
        cell = workbook['VMS Data']['B2']
        fmt = cell.number_format
        colour = cell.font.color
        value = colour.rgb if colour and isinstance(colour.rgb, str) else None
        workbook.close()

        if fmt == '#,##0' and value and value.endswith('FF0000'):
            print("  ✓ Both applied to the same column")
            return True

        print(f"  ✗ format={fmt} colour={value}")
        return False


def main():
    """Run every test and report a final score."""
    print("=== format_excel column formatting tests ===")

    tests = [
        test_number_format_aliases,
        test_number_formats_applied,
        test_header_row_untouched,
        test_column_hiding,
        test_per_column_alignment,
        test_missing_column_policy,
        test_rule_with_no_effect_is_rejected,
        test_per_column_header_styling,
        test_css_colour_names_resolve,
        test_number_format_and_font_combine,
    ]

    passed = 0

    for test_func in tests:
        try:
            if test_func():
                passed += 1
        except Exception as error:
            print(f"  ✗ {test_func.__name__} crashed: {error}")

    print(f"\n=== Results: {passed}/{len(tests)} tests passed ===")

    if passed == len(tests):
        print("✅ All column formatting tests passed!")
        return 1

    print("❌ Some column formatting tests failed!")
    return 0


if __name__ == '__main__':
    exit(0 if main() else 1)


# End of file #

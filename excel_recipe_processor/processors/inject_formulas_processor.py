"""
Inject formulas step processor for Excel automation recipes.

excel_recipe_processor/processors/inject_formulas_processor.py

Handles injecting formulas into Excel files with support for both "live" (dynamic) 
and "dead" (text) formulas. Can work at the cell, range, or auto-scan level.
Supports both stage-to-stage operations (dead formulas) and file operations (live/dead/awaken).
"""

import re
import pandas as pd
import logging
import openpyxl

from pathlib import Path

# try:
#     OPENPYXL_AVAILABLE = True
# except ImportError:
#     OPENPYXL_AVAILABLE = False

from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from excel_recipe_processor.core.base_processor import FileOpsBaseProcessor, BaseStepProcessor, StepProcessorError
from excel_recipe_processor.core.workbook_session import WorkbookSession
from excel_recipe_processor.processors._helpers.inject_formulas_rgx import (
    column_placeholder_rgx,
    function_call_rgx,
)
from excel_recipe_processor.processors._helpers.inject_formulas_functions import (
    FUTURE_FUNCTION_PREFIXES,
)


logger = logging.getLogger(__name__)


class InjectFormulasProcessor(FileOpsBaseProcessor):
    """
    Processor for injecting formulas into existing Excel files.
    
    Supports both "live" formulas (dynamic calculations) and "dead" formulas 
    (text documentation). Can target specific cells, ranges, or auto-scan 
    entire sheets for formula-like text to awaken.
    """
    
class InjectFormulasProcessor(FileOpsBaseProcessor):
    """
    Processor for injecting formulas into Excel files or DataFrame stages.
    
    Supports both "live" formulas (dynamic calculations) and "dead" formulas 
    (text documentation). Can target specific cells, ranges, or auto-scan 
    entire sheets for formula-like text to awaken.
    
    Operating Modes:
    - dead: Stage-to-stage operation injecting formula text into DataFrames
    - live/awaken: File operations manipulating existing Excel files
    """
    
    def __init__(self, step_config: dict):
        # Initialize stage directives to None to detect what gets set
        self.source_stage = None
        self.save_to_stage = None
        
        # Let BaseStepProcessor read config and set stage directives if present
        BaseStepProcessor.__init__(self, step_config)
        
        # Validate configuration based on mode
        self._validate_configuration()
    
    def _validate_configuration(self):
        """Validate processor configuration based on operating mode."""
        mode = self.get_config_value('mode', 'live')
        has_source = bool(self.source_stage)
        has_save = bool(self.save_to_stage)
        has_target = bool(self.get_config_value('target_file'))
        
        # Validate mode
        if mode not in ['live', 'dead', 'awaken']:
            raise StepProcessorError(f"Invalid mode '{mode}'. Must be 'live', 'dead', or 'awaken'")
        
        # Mode-specific validation
        if mode == 'dead':
            # Dead formulas: must be stage-to-stage
            if not (has_source and has_save):
                raise StepProcessorError("Dead mode requires 'source_stage' and 'save_to_stage'")
            if has_target:
                raise StepProcessorError("Dead mode cannot use 'target_file' - it operates on stages")
        
        elif mode in ['live', 'awaken']:
            # Live/awaken: must have target_file, cannot use save_to_stage
            if not has_target:
                raise StepProcessorError(f"'{mode}' mode requires 'target_file'")
            if has_save:
                raise StepProcessorError(f"'{mode}' mode cannot use 'save_to_stage' - it operates on files")
    
    @classmethod
    def get_minimal_config(cls):
        return {
            'mode': 'live',
            'target_file': 'output.xlsx'
        }
    
    def execute(self, data=None):
        """Execute the appropriate operation based on mode."""
        mode = self.get_config_value('mode', 'live')
        
        if mode == 'dead':
            return self._execute_stage_operations()
        else:
            return self._execute_file_operations()
    
    def _execute_stage_operations(self):
        """Execute stage-to-stage operations for dead formulas."""
        self.log_step_start()
        
        # Load input data
        data = self._load_input_data()
        
        # Inject dead formulas into DataFrame
        modified_data = self._inject_dead_formulas_to_dataframe(data)
        
        # Save to output stage
        self._save_output_data(modified_data)
        
        formulas = self.get_config_value('formulas', [])
        self.log_step_complete(f"injected {len(formulas)} dead formulas into stage")
        return modified_data
    
    def _execute_file_operations(self):
        """Execute file operations for live/awaken formulas."""
        result = self.perform_file_operation()
        return pd.DataFrame()  # File operations return empty DataFrame
    
    def _load_input_data(self):
        """Load data from source_stage."""
        from excel_recipe_processor.core.stage_manager import StageManager
        return StageManager.load_stage(self.source_stage)
    
    def _save_output_data(self, data):
        """Save data to save_to_stage."""
        from excel_recipe_processor.core.stage_manager import StageManager
        StageManager.save_stage(
            stage_name=self.save_to_stage,
            data=data,
            description=f"Data with injected formulas from step: '{self.step_name}'",
            step_name=self.step_name,
            confirm_replacement=self.confirm_stage_replacement
        )
    
    def _inject_dead_formulas_to_dataframe(self, df):
        """Inject dead formulas as text into DataFrame cells."""
        import pandas as pd
        
        # Create a copy and convert to object dtype to allow mixed types
        result_df = df.copy().astype('object')
        formulas = self.get_config_value('formulas', [])
        
        for formula_def in formulas:
            if 'cell' in formula_def:
                # Handle cell reference like 'A1' -> row 0, col 0
                cell_ref = formula_def['cell']
                formula = formula_def['formula']
                
                # Convert Excel cell reference to pandas coordinates
                row_idx, col_idx = self._excel_ref_to_pandas(cell_ref, result_df)
                
                # Inject as text with single quote prefix
                formula_text = f"'{formula}" if not formula.startswith("'") else formula
                result_df.iloc[row_idx, col_idx] = formula_text
                
                logger.debug(f"Injected dead formula in {cell_ref}: {formula}")
        
        return result_df
    
    def _excel_ref_to_pandas(self, cell_ref, df):
        """Convert Excel cell reference like 'A1' to pandas row/col indices."""
        try:
            col_letter, row_num = coordinate_from_string(cell_ref)
            
            # Convert to 0-based indices (Excel is 1-based)
            row_idx = row_num - 1  # -1 for 0-based
            col_idx = column_index_from_string(col_letter) - 1  # -1 for 0-based
            
            # Validate indices are within DataFrame bounds
            if row_idx < 0 or row_idx >= len(df):
                raise StepProcessorError(f"Row {row_num} in cell '{cell_ref}' is outside DataFrame bounds")
            if col_idx < 0 or col_idx >= len(df.columns):
                raise StepProcessorError(f"Column {col_letter} in cell '{cell_ref}' is outside DataFrame bounds")
            
            return row_idx, col_idx
            
        except Exception as e:
            raise StepProcessorError(f"Invalid cell reference '{cell_ref}': {e}")
    
    def perform_file_operation(self) -> str:
        """Inject formulas into the target Excel file."""
        # # Check openpyxl availability
        # if not OPENPYXL_AVAILABLE:
        #     raise StepProcessorError("openpyxl is required for formula injection but not installed")
        
        target_file = self.get_config_value('target_file')
        mode = self.get_config_value('mode', 'live')
        formulas = self.get_config_value('formulas', [])
        auto_scan = self.get_config_value('auto_scan', False)
        sheets = self.get_config_value('sheets', None)  # None = active sheet, 'all' = all sheets
        
        # Apply variable substitution to target filename
        if hasattr(self, 'variable_substitution') and self.variable_substitution:
            resolved_file = self.variable_substitution.substitute(target_file)
        else:
            resolved_file = target_file
        
        # Check file exists
        # "Exists" means on disk OR live in the workbook session - under the
        # export bridge the file has not touched disk yet.
        if not Path(resolved_file).exists() and not WorkbookSession.is_open(resolved_file):
            raise StepProcessorError(f"Target file not found: '{resolved_file}'")
        
        # Process the file based on mode
        if mode == 'awaken':
            result = self._awaken_formulas(resolved_file, sheets, auto_scan)
        else:
            result = self._inject_formulas(resolved_file, mode, formulas, sheets)
        
        return result
    
    def _inject_formulas(self, filename: str, mode: str, formulas: list, sheets) -> str:
        """
        Inject specific formulas into the Excel file.
        
        Args:
            filename: Excel file to modify
            mode: 'live' or 'dead'
            formulas: List of formula definitions
            sheets: Sheet selection (None, sheet name, or 'all')
            
        Returns:
            Description of operation performed
        """
        workbook = WorkbookSession.get_workbook(filename)
        formulas_injected = 0
        sheets_processed = 0

        # Per-cell live writes collected by _store_formula, keyed by sheet
        # title; compressed to ranges and registered with the session below.
        self._written_live_cells = {}

        # Determine which sheets to process
        target_sheets = self._get_target_sheets(workbook, sheets)
        
        for sheet_name in target_sheets:
            worksheet = workbook[sheet_name]
            sheet_formulas = 0
            
            for formula_def in formulas:
                sheet_formulas += self._apply_formula_to_sheet(worksheet, formula_def, mode)
            
            formulas_injected += sheet_formulas
            sheets_processed += 1
            
            logger.debug(f"Injected {sheet_formulas} formulas in sheet '{sheet_name}'")

        # Register the recipe-authored cells BEFORE mark_dirty: in
        # standalone (non-deferred) mode, mark_dirty saves immediately and
        # the declaration pass reads the registry during that save.
        for sheet_name, written_cells in self._written_live_cells.items():
            ranges = self._compress_cells_to_ranges(written_cells)
            WorkbookSession.register_injected_formulas(filename, sheet_name, ranges)

        # Save the modified workbook
        WorkbookSession.mark_dirty(filename)
        
        mode_desc = "live" if mode == "live" else "dead"
        return f"injected {formulas_injected} {mode_desc} formulas across {sheets_processed} sheets in {filename}"
    
    def _awaken_formulas(self, filename: str, sheets, auto_scan: bool) -> str:
        """
        Awaken dead formulas in the Excel file.
        
        Args:
            filename: Excel file to modify
            sheets: Sheet selection
            auto_scan: Whether to scan entire sheets for dead formulas
            
        Returns:
            Description of operation performed
        """
        workbook = WorkbookSession.get_workbook(filename)
        formulas_awakened = 0
        sheets_processed = 0
        
        # Determine which sheets to process
        target_sheets = self._get_target_sheets(workbook, sheets)
        
        for sheet_name in target_sheets:
            worksheet = workbook[sheet_name]
            sheet_awakened = self._awaken_sheet_formulas(worksheet)
            
            formulas_awakened += sheet_awakened
            sheets_processed += 1
            
            logger.debug(f"Awakened {sheet_awakened} formulas in sheet '{sheet_name}'")
        
        # Save the modified workbook
        WorkbookSession.mark_dirty(filename)
        
        return f"awakened {formulas_awakened} dead formulas across {sheets_processed} sheets in {filename}"
    
    @staticmethod
    def _compress_cells_to_ranges(written_cells) -> list:
        """
        Collapse (column_letters, row) cells into (column, first, last) runs.

        A fill-down of eight thousand rows becomes one entry; scattered
        single cells stay single-row ranges. Keeps the session registry and
        the declaration pass proportional to the number of injected
        COLUMNS rather than cells.

        Args:
            written_cells: List of (column_letters, row_number)

        Returns:
            List of (column_letters, first_row, last_row)
        """
        by_column = {}
        for column_letters, row_number in written_cells:
            by_column.setdefault(column_letters, set()).add(row_number)

        ranges = []
        for column_letters in sorted(by_column):
            rows = sorted(by_column[column_letters])
            run_start = rows[0]
            previous = rows[0]
            for row_number in rows[1:]:
                if row_number == previous + 1:
                    previous = row_number
                    continue
                ranges.append((column_letters, run_start, previous))
                run_start = row_number
                previous = row_number
            ranges.append((column_letters, run_start, previous))

        return ranges

    def _get_target_sheets(self, workbook, sheets) -> list:
        """
        Determine which sheets to process.
        
        Args:
            workbook: openpyxl workbook
            sheets: Sheet specification (None, sheet name, or 'all')
            
        Returns:
            List of sheet names to process
        """
        if sheets is None:
            # Use active sheet
            return [workbook.active.title]
        elif sheets == 'all':
            # All sheets
            return workbook.sheetnames
        elif isinstance(sheets, str):
            # Single sheet by name
            if sheets not in workbook.sheetnames:
                raise StepProcessorError(f"Sheet '{sheets}' not found in workbook")
            return [sheets]
        elif isinstance(sheets, list):
            # Multiple specific sheets
            for sheet in sheets:
                if sheet not in workbook.sheetnames:
                    raise StepProcessorError(f"Sheet '{sheet}' not found in workbook")
            return sheets
        else:
            raise StepProcessorError(f"Invalid 'sheets' specification: '{sheets}'")
    
    def _apply_formula_to_sheet(self, worksheet, formula_def: dict, mode: str) -> int:
        """
        Apply a single formula definition to a worksheet.
        
        Args:
            worksheet: openpyxl worksheet
            formula_def: Dictionary with 'cell'/'range' and 'formula' keys.
                         The formula may use {col:Header Name} placeholders,
                         resolved against the sheet's header row, and a cell
                         target may set fill_down: true
            mode: 'live' or 'dead'
            
        Returns:
            Number of cells modified
        """
        if not isinstance(formula_def, dict):
            raise StepProcessorError("Formula definition must be a dictionary")
        
        if 'formula' not in formula_def:
            raise StepProcessorError("Formula definition must include 'formula' key")
        
        formula = formula_def['formula']
        
        # Ensure formula starts with = for live mode
        if mode == 'live' and not formula.startswith('='):
            formula = '=' + formula
        
        # Functions added after the 2007 format must be STORED with an
        # _xlfn. prefix or Excel shows #NAME?; see the prefix map for why.
        formula = self._prefix_future_functions(formula)

        # Column names resolve to letters against THIS sheet's header row -
        # in the formula AND in the target, since naming the column a formula
        # is written INTO is the same fragility as naming ones it reads.
        formula = self._resolve_column_placeholders(worksheet, formula)

        # Handle cell vs range specification
        if 'cell' in formula_def:
            cell_ref = self._resolve_column_placeholders(worksheet, str(formula_def['cell']))
            as_array = formula_def.get('array_formula', False)
            if formula_def.get('fill_down', False):
                return self._apply_formula_with_fill_down(
                    worksheet, cell_ref, formula, mode, as_array
                )
            if as_array:
                self._store_formula(worksheet, cell_ref, formula, mode, True)
                return 1
            return self._apply_formula_to_cell(worksheet, cell_ref, formula, mode)
        elif 'range' in formula_def:
            range_ref = self._resolve_column_placeholders(worksheet, str(formula_def['range']))
            return self._apply_formula_to_range(worksheet, range_ref, formula, mode)
        else:
            raise StepProcessorError("Formula definition must include either 'cell' or 'range' key")

    def _prefix_future_functions(self, formula: str) -> str:
        """
        Give post-2007 Excel functions the storage prefix they require.

        Excel stores XLOOKUP as _xlfn.XLOOKUP and displays it as XLOOKUP. A
        formula written with the plain name is read as an unknown defined
        name, so the cell shows #NAME? and the formula bar renders it with an
        implicit-intersection marker (=@IFS(...)). Recipes therefore write
        ordinary Excel syntax and this adds the prefixes.

        Only names present in the map are touched, and a name that already
        carries a prefix is left alone, so re-running is safe.

        Args:
            formula: Formula text as written in the recipe

        Returns:
            Formula with future-function names prefixed for storage
        """
        def substitute(match):
            name = match.group(1)
            prefix = FUTURE_FUNCTION_PREFIXES.get(name.upper())

            if prefix is None:
                return match.group(0)

            return match.group(0).replace(name, f"{prefix}{name}", 1)

        prefixed = function_call_rgx.sub(substitute, formula)

        if prefixed != formula:
            changed = sorted({
                name for name in FUTURE_FUNCTION_PREFIXES
                if f"{FUTURE_FUNCTION_PREFIXES[name]}{name}" in prefixed
            })
            logger.debug(f"Prefixed future function(s) for storage: {changed}")

        return prefixed

    def _resolve_column_placeholders(self, worksheet, formula: str) -> str:
        """
        Replace {col:Header Name} with that column's letter on THIS sheet.

        Excel formulas address columns by letter, which makes a recipe
        fragile: inserting one column shifts every letter after it and the
        formula silently reads the wrong data. Naming the column instead
        means the letter is resolved from the header row at injection time,
        so a layout change costs nothing.

        Args:
            worksheet: Sheet whose header row supplies the positions
            formula:   Formula text, possibly containing placeholders

        Returns:
            Formula with every placeholder replaced by a column letter
        """
        if '{col:' not in formula:
            return formula

        headers = {}
        for column_index, cell in enumerate(worksheet[1], start=1):
            if cell.value is not None:
                headers[str(cell.value).strip()] = get_column_letter(column_index)

        def substitute(match):
            name = match.group(1).strip()
            if name not in headers:
                raise StepProcessorError(
                    f"Formula references column '{name}', which is not in the header row "
                    f"of sheet '{worksheet.title}'. Available: {sorted(headers)[:8]}..."
                )
            return headers[name]

        return column_placeholder_rgx.sub(substitute, formula)

    def _store_formula(self, worksheet, target_ref: str, formula_text: str,
                       mode: str, as_array: bool) -> None:
        """
        Write one formula cell, optionally marked as an array formula.

        Excel applies IMPLICIT INTERSECTION to a plain formula whose result
        could be an array - a formula containing XLOOKUP, say - and displays
        it with a leading @.

        CORRECTED CLAIM (2026-08-13): the array marker alone is NOT enough
        to retire the @ - it trades it for legacy {CSE} braces, because
        t="array" without the cm="1"/xl/metadata.xml declaration means "old
        Ctrl+Shift+Enter formula". Use array_formula only for a formula
        that genuinely IS an array formula. The honest fix for the @ is the
        dynamic-array declaration pass (settings: declare_dynamic_formulas,
        or the declare_dynamic_formulas processor); see
        core/dynamic_array_metadata.py.

        Args:
            worksheet:   Target sheet
            target_ref:  Cell to write, e.g. 'AV2'
            formula_text: Formula, already translated for this cell
            mode:        'live' or 'dead'
            as_array:    Store with the array marker
        """
        if mode != 'live':
            worksheet[target_ref].value = f"'{formula_text}"
            return

        if as_array:
            worksheet[target_ref].value = ArrayFormula(target_ref, formula_text)
        else:
            worksheet[target_ref].value = formula_text

        # Record the write for the dynamic-array declaration's provenance
        # registry: cells the recipe authored are declared aware at save
        # regardless of function vocabulary. Collected per cell here (the
        # one live-write funnel), compressed to ranges at registration.
        if not hasattr(self, '_written_live_cells'):
            self._written_live_cells = {}
        column_letters, row_number = coordinate_from_string(target_ref)
        self._written_live_cells.setdefault(worksheet.title, []).append(
            (column_letters, row_number)
        )

    def _apply_formula_with_fill_down(self, worksheet, cell_ref: str, formula: str, mode: str,
                                      as_array: bool = False) -> int:
        """
        Write the formula at cell_ref and continue it to the last data row.

        The row extent comes from the sheet itself, so the fill follows
        however many rows this run produced.

        Args:
            worksheet: openpyxl worksheet
            cell_ref:  Where the formula starts, e.g. 'AV2'
            formula:   Formula text, already placeholder-resolved
            mode:      'live' or 'dead'

        Returns:
            Number of cells written
        """
        origin_cell = worksheet[cell_ref]
        last_row = worksheet.max_row

        if last_row < origin_cell.row:
            logger.warning(f"⚠️  Nothing to fill: sheet ends at row {last_row}")
            return 0

        written = 0
        column_letter = origin_cell.column_letter

        for row_number in range(origin_cell.row, last_row + 1):
            target = f"{column_letter}{row_number}"
            adjusted = self._adjust_formula_for_cell(formula, target, cell_ref)
            self._store_formula(worksheet, target, adjusted, mode, as_array)
            written += 1

        logger.info(f"⬇️  Filled {column_letter}{origin_cell.row}:{column_letter}{last_row} "
                    f"({written:,} cells)")
        return written
    
    def _apply_formula_to_cell(self, worksheet, cell_ref: str, formula: str, mode: str) -> int:
        """
        Apply formula to a single cell.
        
        Args:
            worksheet: openpyxl worksheet
            cell_ref: Cell reference like 'A1', 'B5', etc.
            formula: Formula to inject
            mode: 'live' or 'dead'
            
        Returns:
            Number of cells modified (always 1 for single cell)
        """
        # Validate cell reference
        if not self._is_valid_cell_reference(cell_ref):
            raise StepProcessorError(f"Invalid cell reference: {cell_ref}")

        # All writes go through the one live-write funnel so the provenance
        # registry sees every recipe-authored cell. (Until 2026-08-13 this
        # wrote cell.value directly, invisibly to the declaration pass.)
        self._store_formula(worksheet, cell_ref, formula, mode, False)

        logger.debug(f"Set {mode} formula in {cell_ref}: {formula}")
        return 1
    
    def _apply_formula_to_range(self, worksheet, range_ref: str, formula: str, mode: str) -> int:
        """
        Apply formula to a range of cells.

        The range's FIRST cell is the origin the formula was written for;
        every other cell gets the formula translated relative to it, the
        way Excel's own fill does.

        Args:
            worksheet: openpyxl worksheet
            range_ref: Range reference like 'A1:A10', 'B2:D5', etc.
            formula: Base formula to inject (will be adjusted for each cell)
            mode: 'live' or 'dead'

        Returns:
            Number of cells modified
        """
        # Validate range reference
        if not self._is_valid_range_reference(range_ref):
            raise StepProcessorError(f"Invalid range reference: {range_ref}")

        range_origin = range_ref.split(':')[0].replace('$', '')
        cells_modified = 0

        cell_range = worksheet[range_ref]

        # worksheet[range] yields a tuple of row tuples for a real range,
        # or a bare Cell for a single-cell "range" like "B2:B2" - normalize
        # to a flat cell walk so every write goes through the one funnel
        # (which also feeds the provenance registry).
        if hasattr(cell_range, 'coordinate'):
            flat_cells = [cell_range]
        else:
            flat_cells = []
            for row in cell_range:
                if hasattr(row, 'coordinate'):
                    flat_cells.append(row)
                else:
                    flat_cells.extend(row)

        for cell in flat_cells:
            adjusted_formula = self._adjust_formula_for_cell(
                formula, cell.coordinate, range_origin
            )
            self._store_formula(worksheet, cell.coordinate, adjusted_formula, mode, False)
            cells_modified += 1

        logger.debug(f"Applied {mode} formula to range {range_ref}: {cells_modified} cells")
        return cells_modified
    
    def _awaken_sheet_formulas(self, worksheet) -> int:
        """
        Scan a worksheet for dead formulas and awaken them.
        
        Args:
            worksheet: openpyxl worksheet
            
        Returns:
            Number of formulas awakened
        """
        formulas_awakened = 0
        
        # Scan all cells with data
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Check if it looks like a dead formula
                    cell_value = cell.value.strip()
                    if self._looks_like_formula(cell_value):
                        # Remove single quote prefix if present
                        if cell_value.startswith("'="):
                            formula = cell_value[1:]  # Remove leading quote
                        elif cell_value.startswith("="):
                            formula = cell_value
                        else:
                            continue
                        
                        # Awaken the formula
                        cell.value = formula
                        formulas_awakened += 1
                        logger.debug(f"Awakened formula in {cell.coordinate}: {formula}")
        
        return formulas_awakened
    
    def _adjust_formula_for_cell(self, base_formula: str, cell_coord: str, origin: str = None) -> str:
        """
        Adjust a base formula for a specific cell location.
        
        For now, this is a simple implementation. Could be enhanced
        to automatically adjust relative references.
        
        Args:
            base_formula: Base formula template
            cell_coord: Target cell coordinate like 'A1', 'B5'
            
        Returns:
            Adjusted formula for the specific cell
        """
        if not base_formula.startswith('='):
            base_formula = '=' + base_formula

        if origin is None or origin == cell_coord:
            return base_formula

        # openpyxl's Translator shifts relative references the way Excel does
        # when a formula is copied: A2 written at row 2 becomes A3 at row 3,
        # while $A$2 and named ranges stay put. Without this, a formula
        # applied to a range repeated the ORIGIN's references in every cell,
        # so every row silently read row 2's data.
        return Translator(base_formula, origin=origin).translate_formula(cell_coord)
    
    def _is_valid_cell_reference(self, cell_ref: str) -> bool:
        """Check if a string is a valid Excel cell reference."""
        try:
            coordinate_from_string(cell_ref)
            return True
        except:
            return False
    
    def _is_valid_range_reference(self, range_ref: str) -> bool:
        """Check if a string is a valid Excel range reference."""
        if ':' not in range_ref:
            # Single cell reference
            return self._is_valid_cell_reference(range_ref)
        
        try:
            start_cell, end_cell = range_ref.split(':')
            return (self._is_valid_cell_reference(start_cell) and 
                   self._is_valid_cell_reference(end_cell))
        except:
            return False
    
    def _looks_like_formula(self, text: str) -> bool:
        """
        Check if text looks like it could be a formula.
        
        Args:
            text: Text to check
            
        Returns:
            True if text looks like a formula
        """
        if not isinstance(text, str):
            return False
        
        text = text.strip()
        
        # Check for formula patterns
        formula_patterns = [
            r"^'?=",  # Starts with = or '=
            r"=\s*[A-Z]+\d+",  # Contains cell references
            r"=\s*[A-Z]+\(",  # Contains function calls
            r"=\s*(SUM|AVERAGE|COUNT|MIN|MAX|IF|VLOOKUP|INDEX|MATCH)",  # Common functions
        ]
        
        for pattern in formula_patterns:
            if re.match(pattern, text, re.IGNORECASE):
                return True
        
        return False
    
    def get_operation_type(self) -> str:
        """Get the type of file operation this processor performs."""
        return "formula_injection"
    
    def get_capabilities(self) -> dict:
        """Get processor capabilities information."""
        return {
            'description': 'Inject formulas into existing Excel files with live/dead '
                           'modes, name-addressed cells, and Excel-style fill-down',
            'operation_type': 'formula_injection',
            'supported_modes': ['live', 'dead', 'awaken'],
            'targeting_options': ['single_cell', 'cell_range', 'auto_scan'],
            'column_placeholders': '{col:Header Name} in cell refs and formulas '
                                   'resolves to the column letter from the header '
                                   'row at injection time, so upstream column '
                                   'insertions cannot silently repoint a formula',
            'fill_down': 'fill_down: true fills a formula from its origin cell to '
                         'the last data row, translating relative references per '
                         'row the way Excel\'s own fill handle does',
            'function_name_translation': 'modern function names (XLOOKUP, IFS, '
                                         'FILTER, ...) are stored with the _xlfn. '
                                         'prefixes Excel requires internally, and '
                                         'display without them',
            'dynamic_array_declaration': 'live-mode cells register with the '
                                         'workbook session as recipe-authored; with '
                                         'settings declare_dynamic_formulas: true '
                                         'they are declared dynamic-array-aware at '
                                         'save and open without the implicit-'
                                         'intersection @ (any function, by '
                                         'provenance)',
            'array_formula_caveat': 'array_formula: true stores a legacy CSE array '
                                    'formula ({braces}); use it only for formulas '
                                    'that genuinely are array formulas - it is NOT '
                                    'the fix for the @ display',
            'sheet_support': ['single_sheet', 'multiple_sheets', 'all_sheets'],
            'file_requirements': ['xlsx', 'xlsm'],
            'dependencies': ['openpyxl'],
            'stage_requirements': 'none',
        }

    def get_usage_examples(self) -> dict:
        """Get usage examples from the external YAML file."""
        from excel_recipe_processor.utils.processor_examples_loader import load_processor_examples
        return load_processor_examples('inject_formulas')

# End of file #

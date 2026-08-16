"""
Tests for profile_sheets and the format_excel width-inheritance consumer.

tests/test_profile_sheets.py

Covers the profile facts (widths via the SHARED clamp, dtypes, blank
counting incl. empty strings, distincts), multi-input Source
identities, the guided config refusals, the auto-fit parity promise
(scan of a frame == format_excel auto-fit of the same data written
plain), and the consumer directive end to end including its selector
rules. Runnable directly or with pytest; direct runs are the
authoritative score.
"""

import os
import sys
import tempfile

import pandas as pd
from openpyxl import Workbook, load_workbook

from excel_recipe_processor.core.stage_manager import StageManager
from excel_recipe_processor.processors.profile_sheets_processor import (
    ProfileSheetsProcessor,
)
from excel_recipe_processor.processors._helpers.column_width_scan import (
    fitted_width,
    scan_frame_column_widths,
)


def make_frame():
    """A small frame with known measurable properties."""
    return pd.DataFrame({
        'Short': ['a', 'b', 'c', None],
        'A_Much_Longer_Header': [1, 2, 2, 2],
        'Values': ['medium text here', '', 'x', 'yy'],
    })


def run_processor(config):
    config = dict(config)
    config['processor_type'] = 'profile_sheets'
    processor = ProfileSheetsProcessor(config)
    return processor.load_data()


def test_profile_facts():
    """Widths, dtypes, blanks (NA + empty string), distincts, positions."""
    print("\nTesting profile facts...")
    StageManager.initialize_stages(max_stages=30)
    StageManager.save_stage('stg_t_facts', make_frame(), 'test')

    profile = run_processor({
        'sheets': [{'source_stage': 'stg_t_facts'}],
        'save_to_stage': 'stg_t_profile',
        'min_width': 8, 'max_width': 40, 'padding': 4,
    })

    by_col = profile.set_index('Column')
    checks = [
        ('Short width = header+padding (floor does not bind: 5+4=9)',
         by_col.loc['Short', 'Width'] == fitted_width(5, 8, 40, 4)),
        ('Header dominates width',
         by_col.loc['A_Much_Longer_Header', 'Width']
         == fitted_width(len('A_Much_Longer_Header'), 8, 40, 4)),
        ('Values width from longest cell',
         by_col.loc['Values', 'Width'] == fitted_width(16, 8, 40, 4)),
        ('NA counted blank', by_col.loc['Short', 'Blank_Count'] == 1),
        ('Empty string counted blank', by_col.loc['Values', 'Blank_Count'] == 1),
        ('Distinct over non-blank', by_col.loc['A_Much_Longer_Header',
                                               'Distinct_Count'] == 2),
        ('Empty string excluded from distinct',
         by_col.loc['Values', 'Distinct_Count'] == 3),
        ('Positions 1-based ordered',
         list(profile['Position']) == [1, 2, 3]),
        ('Row_Count', (profile['Row_Count'] == 4).all()),
        ('Source labeled', (profile['Source'] == 'stg_t_facts').all()),
    ]
    for name, ok in checks:
        if not ok:
            print(f"✗ {name}")
            print(profile)
            return False
        print(f"✓ {name}")
    return True


def test_multi_input_and_refusals():
    """Multiple inputs get distinct Sources; config mistakes refuse."""
    print("\nTesting multi-input identities and guided refusals...")
    StageManager.initialize_stages(max_stages=30)
    StageManager.save_stage('stg_t_one', make_frame(), 'test')

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as handle:
        path = handle.name
    try:
        make_frame().to_excel(path, index=False, sheet_name='Tab_A')
        profile = run_processor({
            'sheets': [{'source_stage': 'stg_t_one'},
                       {'input_file': path, 'sheet_name': 'Tab_A'}],
            'save_to_stage': 'stg_t_profile2',
        })
        sources = set(profile['Source'].unique())
        file_sources = {s for s in sources if s.endswith('!Tab_A')}
        if len(sources) != 2 or 'stg_t_one' not in sources \
                or len(file_sources) != 1:
            print(f"✗ Source identities wrong: {sources}")
            return False
        print(f"✓ Two inputs, two Source identities")

        # Stage-vs-file parity: identical data, identical facts
        stage_half = profile[profile['Source'] == 'stg_t_one'] \
            .drop(columns='Source').reset_index(drop=True)
        file_half = profile[profile['Source'] != 'stg_t_one'] \
            .drop(columns='Source').reset_index(drop=True)
        if not stage_half.equals(file_half):
            print("✗ Stage-input and file-input facts differ for same data")
            return False
        print("✓ Stage and file inputs yield identical facts")
    finally:
        os.unlink(path)

    for label, sheets in [
        ("both kinds in one entry",
         [{'source_stage': 's', 'input_file': 'f.xlsx'}]),
        ("neither kind", [{'sheet_name': 'X'}]),
        ("empty list", []),
    ]:
        try:
            run_processor({'sheets': sheets, 'save_to_stage': 'stg_x'})
            print(f"✗ {label}: should have refused")
            return False
        except Exception as error:
            if 'sheets' not in str(error) and 'exactly ONE' not in str(error):
                print(f"✗ {label}: unguided error: {error}")
                return False
            print(f"✓ {label}: refused with guidance")
    return True


def test_autofit_parity():
    """Scan of a frame == format_excel auto-fit of the same data, plain."""
    print("\nTesting the shared-math parity promise...")

    from excel_recipe_processor.processors.format_excel_processor import (
        FormatExcelProcessor,
    )

    frame = make_frame()
    scanned = {name: width for name, width, _ in
               scan_frame_column_widths(frame, min_width=10, max_width=40)}

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(frame.columns))
    for row in frame.itertuples(index=False):
        sheet.append(['' if value is None else value for value in row])

    config = FormatExcelProcessor.get_minimal_config()
    config['processor_type'] = 'format_excel'
    processor = FormatExcelProcessor(config)
    processor._auto_fit_columns(
        sheet, {'min_column_width': 10, 'max_column_width': 40})

    for index, name in enumerate(frame.columns):
        letter = sheet.cell(row=1, column=index + 1).column_letter
        fitted = sheet.column_dimensions[letter].width
        if fitted != scanned[name]:
            print(f"✗ {name}: scan {scanned[name]} vs auto-fit {fitted}")
            return False
    print("✓ Scan widths equal auto-fit widths on plain data")
    return True


def test_consumer_directive():
    """format_excel inherits widths by header name, selector rules hold."""
    print("\nTesting column_widths_from_stage...")

    from excel_recipe_processor.processors.format_excel_processor import (
        FormatExcelProcessor,
    )
    from excel_recipe_processor.core.base_processor import StepProcessorError

    StageManager.initialize_stages(max_stages=30)
    profile = pd.DataFrame({
        'Source': ['stg_seed'] * 2 + ['stg_other'],
        'Column': ['Alpha', 'Beta', 'Alpha'],
        'Width': [33, 21, 11],
    })
    StageManager.save_stage('stg_t_widths', profile, 'test')

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Alpha', 'Beta', 'Not_Profiled'])
    sheet.append(['x', 'y', 'z'])

    config = FormatExcelProcessor.get_minimal_config()
    config['processor_type'] = 'format_excel'
    processor = FormatExcelProcessor(config)

    # Ambiguous without a selector
    try:
        processor._apply_widths_from_profile_stage(
            sheet, {'column_widths_from_stage': 'stg_t_widths'})
        print("✗ multi-Source stage without selector should refuse")
        return False
    except StepProcessorError as error:
        if 'column_widths_source' not in str(error):
            print(f"✗ refusal lacks the selector hint: {error}")
            return False
        print("✓ multi-Source without selector refused, hint given")

    inherited = processor._apply_widths_from_profile_stage(
        sheet, {'column_widths_from_stage': 'stg_t_widths',
                'column_widths_source': 'stg_seed'})
    widths = {sheet.cell(row=1, column=i + 1).value:
              sheet.column_dimensions[
                  sheet.cell(row=1, column=i + 1).column_letter].width
              for i in range(3)}
    if inherited != 2 or widths['Alpha'] != 33 or widths['Beta'] != 21:
        print(f"✗ inheritance wrong: {inherited} applied, {widths}")
        return False
    if widths['Not_Profiled'] not in (None, 13):  # openpyxl default, untouched
        print(f"✗ unprofiled column was touched: {widths['Not_Profiled']}")
        return False
    print("✓ widths inherited by header name; unprofiled column untouched")

    try:
        processor._apply_widths_from_profile_stage(
            sheet, {'column_widths_from_stage': 'stg_t_widths',
                    'column_widths_source': 'stg_nonexistent'})
        print("✗ unknown selector should refuse")
        return False
    except StepProcessorError as error:
        if 'stg_seed' not in str(error):
            print(f"✗ refusal does not list available sources: {error}")
            return False
        print("✓ unknown selector refused, available sources listed")
    return True


def main():
    """Run all tests and report results."""
    print("profile_sheets and width-inheritance tests")
    print("=" * 50)

    tests = [
        test_profile_facts,
        test_multi_input_and_refusals,
        test_autofit_parity,
        test_consumer_directive,
    ]

    passed = 0
    for test in tests:
        try:
            if test():
                passed += 1
            else:
                print(f"FAILED: {test.__name__}")
        except Exception as error:
            print(f"FAILED with exception: {test.__name__}: {error}")

    print("=" * 50)
    print(f"Final score: {passed}/{len(tests)} tests passed")
    return passed == len(tests)


if __name__ == '__main__':
    sys.exit(0 if main() else 1)

# End of file #

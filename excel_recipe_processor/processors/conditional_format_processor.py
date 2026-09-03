"""
Conditional formatting for Excel outputs, in canonical ERP vocabulary.

excel_recipe_processor/processors/conditional_format_processor.py

Writes native Excel conditional-formatting rules - the live kind that keep
re-evaluating as the human edits the file - using the same condition names
the rest of the framework speaks (filter_data's vocabulary: equals,
greater_than, is_empty, ...) plus a few Excel-only concepts extended in the
same style (between, duplicates, unique). Excel's own operator spellings
(greaterThan, containsText, ...) are accepted as aliases but draw a
warning naming the canonical form, so recipes converge on one language.

Formula rules use the SAME authoring convention as inject_formulas: write
the formula as if for data row 2, with {col:Header Name} placeholders. Two
deliberate differences from injection, both because conditional formatting
translates one formula across a whole range: placeholders here resolve to
$-LOCKED columns ($AX2, not AX2) so a row-wise test cannot drift sideways,
and modern function names get the _xlfn. storage prefix because openpyxl
stores rule formulas verbatim and an unprefixed modern function would make
the rule silently never fire.

Rules are written in list order, which is Excel's priority order. Text,
blank, duplicate and unique conditions are emitted one rule per column,
because their formulas anchor to a range's first cell (and a duplicates
domain should be per-column unless a literal range: says otherwise).
"""

import logging

from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule, CellIsRule, FormulaRule, DataBarRule, ColorScaleRule

from excel_recipe_processor.core.workbook_session import WorkbookSession
from excel_recipe_processor.core.base_processor import FileOpsBaseProcessor, StepProcessorError
from excel_recipe_processor.processors._helpers.excel_color_support import normalize_color
from excel_recipe_processor.processors._helpers.sheet_addressing import resolve_sheet_ref
from excel_recipe_processor.processors._helpers.inject_formulas_rgx import column_placeholder_rgx
from excel_recipe_processor.processors._helpers.inject_formulas_functions import prefix_future_functions


logger = logging.getLogger(__name__)


# Canonical condition -> how it becomes an Excel rule. 'cell_is' conditions
# share one CellIsRule shape; the others are distinct Excel rule types with
# per-column anchored formulas built in _build_anchored_formula.
CELL_IS_OPERATORS = {
    'equals':        'equal',
    'not_equals':    'notEqual',
    'greater_than':  'greaterThan',
    'greater_equal': 'greaterThanOrEqual',
    'less_than':     'lessThan',
    'less_equal':    'lessThanOrEqual',
    'between':       'between',
    'not_between':   'notBetween',
}

TEXT_RULE_TYPES = {
    'contains':     'containsText',
    'not_contains': 'notContainsText',
    'starts_with':  'beginsWith',
    'ends_with':    'endsWith',
}

BLANK_RULE_TYPES = {
    'is_empty':  'containsBlanks',
    'not_empty': 'notContainsBlanks',
}

VALUE_SET_RULE_TYPES = {
    'duplicates': 'duplicateValues',
    'unique':     'uniqueValues',
}

# Excel's own spellings, accepted with a warning naming the canonical form.
CONDITION_ALIASES = {
    'equal':              'equals',
    'notEqual':           'not_equals',
    'greaterThan':        'greater_than',
    'greaterThanOrEqual': 'greater_equal',
    'lessThan':           'less_than',
    'lessThanOrEqual':    'less_equal',
    'notBetween':         'not_between',
    'containsText':       'contains',
    'notContainsText':    'not_contains',
    'notContains':        'not_contains',
    'beginsWith':         'starts_with',
    'endsWith':           'ends_with',
    'containsBlanks':     'is_empty',
    'notContainsBlanks':  'not_empty',
    'duplicateValues':    'duplicates',
    'uniqueValues':       'unique',
}

ALL_CANONICAL_CONDITIONS = sorted(
    set(CELL_IS_OPERATORS) | set(TEXT_RULE_TYPES)
    | set(BLANK_RULE_TYPES) | set(VALUE_SET_RULE_TYPES)
)


class ConditionalFormatProcessor(FileOpsBaseProcessor):
    """Write native conditional-formatting rules using canonical ERP vocabulary."""

    @classmethod
    def get_minimal_config(cls) -> dict:
        """Smallest configuration that constructs and validates."""
        return {
            'target_file': 'output.xlsx',
            'sheet_name': 'Data',
            'rules': [
                {'when_cell': {'column_names': ['Status'], 'condition': 'equals', 'value': 'Bad'},
                 'style': {'fill': 'FFC7CE'}},
            ],
        }

    def _validate_file_operation_config(self):
        """Validate structure, conditions (normalizing aliases), and colors up front."""
        if not self.get_config_value('target_file'):
            raise StepProcessorError(
                f"Conditional format step '{self.step_name}' requires 'target_file'"
            )
        if self.get_config_value('sheet', None):
            raise StepProcessorError(
                f"Conditional format step '{self.step_name}': 'sheet' was "
                f"replaced by 'sheet_name' (2026-08-14 sheet-addressing "
                f"doctrine)"
            )
        if not self.get_config_value('sheet_name'):
            raise StepProcessorError(
                f"Conditional format step '{self.step_name}' requires "
                f"'sheet_name' (a tab name, or '?sheet_001?' by position)"
            )

        rules = self.get_config_value('rules', [])
        if not isinstance(rules, list) or not rules:
            raise StepProcessorError(
                f"Conditional format step '{self.step_name}' requires a non-empty 'rules' list"
            )

        for rule_index, rule in enumerate(rules, start=1):
            self._validate_rule(rule, rule_index)

    def _validate_rule(self, rule, rule_index: int) -> None:
        """One rule: exactly one kind, valid condition, valid colors."""
        context = f"rule {rule_index} in step '{self.step_name}'"

        if not isinstance(rule, dict):
            raise StepProcessorError(f"{context}: each rule must be a mapping")

        kinds = [k for k in ('when_formula', 'when_cell', 'color_scale', 'data_bar') if k in rule]
        if len(kinds) != 1:
            raise StepProcessorError(
                f"{context}: exactly one of when_formula / when_cell / "
                f"color_scale / data_bar, got {kinds or 'none'}"
            )

        if kinds[0] == 'when_cell':
            spec = rule['when_cell']
            condition = self._canonical_condition(spec.get('condition'), context)
            value = spec.get('value')
            if condition in CELL_IS_OPERATORS and condition in ('between', 'not_between'):
                if not isinstance(value, list) or len(value) != 2:
                    raise StepProcessorError(
                        f"{context}: '{condition}' needs value: [low, high]"
                    )
            elif condition in CELL_IS_OPERATORS or condition in TEXT_RULE_TYPES:
                if value is None:
                    raise StepProcessorError(f"{context}: '{condition}' needs a 'value'")
            if spec.get('columns'):
                raise StepProcessorError(
                    f"{context}: 'columns' was renamed 'column_names' "
                    f"(2026-08-26) - header NAME strings only. Rename "
                    f"the key."
                )
            if not spec.get('column_names'):
                raise StepProcessorError(f"{context}: when_cell needs 'column_names' (list of header names)")

        if kinds[0] == 'when_formula':
            spec = rule['when_formula'] if isinstance(rule['when_formula'], dict) else rule
            formula = spec.get('formula') if isinstance(spec, dict) else None
            if kinds[0] == 'when_formula' and isinstance(rule['when_formula'], str):
                formula = rule['when_formula']
            if not formula:
                raise StepProcessorError(f"{context}: when_formula needs the formula text")
            if rule.get('columns'):
                raise StepProcessorError(
                    f"{context}: 'columns' was renamed 'column_names' "
                    f"(2026-08-26) - header NAME strings only. Rename "
                    f"the key."
                )
            targets = [k for k in ('apply_to', 'column_names', 'range') if rule.get(k)]
            if len(targets) != 1:
                raise StepProcessorError(
                    f"{context}: when_formula needs exactly one target - "
                    f"apply_to: \"entire_row\", columns: [...], or range: \"A2:B99\""
                )

        style = rule.get('style', {})
        for color_key in ('fill', 'font_color'):
            if color_key in style:
                try:
                    normalize_color(style[color_key])
                except ValueError as error:
                    raise StepProcessorError(f"{context}: invalid {color_key}: {error}")

        for scale_key in ('min_color', 'mid_color', 'max_color'):
            for kind_key in ('color_scale', 'data_bar'):
                spec = rule.get(kind_key, {})
                if isinstance(spec, dict) and scale_key in spec:
                    try:
                        normalize_color(spec[scale_key])
                    except ValueError as error:
                        raise StepProcessorError(f"{context}: invalid {scale_key}: {error}")

    def _canonical_condition(self, condition, context: str, warn: bool = True) -> str:
        """Normalize a condition name, warning on Excel-native aliases (once, at validation)."""
        if condition is None:
            raise StepProcessorError(f"{context}: when_cell needs a 'condition'")

        name = str(condition)

        if name in ALL_CANONICAL_CONDITIONS:
            return name

        if name in CONDITION_ALIASES:
            canonical = CONDITION_ALIASES[name]
            if not warn:
                return canonical
            logger.warning(
                f"⚠️ {context}: '{name}' accepted, but the canonical ERP "
                f"condition name is '{canonical}' - consider updating the recipe"
            )
            return canonical

        raise StepProcessorError(
            f"{context}: unknown condition '{name}'. Canonical conditions: "
            f"{ALL_CANONICAL_CONDITIONS}"
        )

    # ------------------------------------------------------------------
    # Execution
    # ------------------------------------------------------------------

    def perform_file_operation(self):
        """Apply the rules to the target sheet via the workbook session."""
        target_file = self.get_config_value('target_file')
        sheet_name = self.get_config_value('sheet_name')
        rules = self.get_config_value('rules')

        resolved_file = self._resolve_path(target_file)

        workbook = WorkbookSession.get_workbook(resolved_file)
        try:
            sheet_name = resolve_sheet_ref(
                sheet_name, workbook.sheetnames,
                f"Conditional format step '{self.step_name}'"
            )
        except ValueError as error:
            raise StepProcessorError(str(error))
        worksheet = workbook[sheet_name]

        headers = self._build_header_map(worksheet)
        last_row = worksheet.max_row
        last_col_letter = get_column_letter(worksheet.max_column)

        rules_written = 0
        for rule_index, rule in enumerate(rules, start=1):
            rules_written += self._apply_rule(
                worksheet, rule, rule_index, headers, last_row, last_col_letter
            )

        WorkbookSession.mark_dirty(resolved_file)

        return (
            f"wrote {rules_written} conditional-formatting rule(s) "
            f"on '{sheet_name}' (rows 2-{last_row})"
        )

    def _apply_rule(self, worksheet, rule, rule_index, headers, last_row, last_col_letter) -> int:
        """Write one recipe rule; may emit several Excel rules (per-column anchors)."""
        context = f"rule {rule_index} in step '{self.step_name}'"
        style = rule.get('style', {})
        stop_if_true = bool(rule.get('stop_if_true', False))

        if 'when_formula' in rule:
            formula_text = rule['when_formula']
            if isinstance(formula_text, dict):
                formula_text = formula_text.get('formula')

            formula_text = self._prepare_formula(str(formula_text), headers, context)
            target_range = self._build_formula_target(
                rule, headers, last_row, last_col_letter, context
            )
            worksheet.conditional_formatting.add(
                target_range,
                FormulaRule(formula=[formula_text], stopIfTrue=stop_if_true,
                            **self._style_as_rule_kwargs(style))
            )
            return 1

        if 'when_cell' in rule:
            return self._apply_when_cell(
                worksheet, rule, headers, last_row, style, stop_if_true, context
            )

        if 'color_scale' in rule:
            spec = rule['color_scale']
            scale_kwargs = {
                'start_type': 'min', 'start_color': normalize_color(spec.get('min_color', 'FFFFFF')),
                'end_type': 'max', 'end_color': normalize_color(spec.get('max_color', '63BE7B')),
            }
            if 'mid_color' in spec:
                scale_kwargs.update({
                    'mid_type': 'percentile', 'mid_value': 50,
                    'mid_color': normalize_color(spec['mid_color']),
                })
            written = 0
            for column_range, _ in self._column_ranges(spec, headers, last_row, context):
                worksheet.conditional_formatting.add(column_range, ColorScaleRule(**scale_kwargs))
                written += 1
            return written

        if 'data_bar' in rule:
            spec = rule['data_bar']
            bar_color = normalize_color(spec.get('color', '638EC6'))
            written = 0
            for column_range, _ in self._column_ranges(spec, headers, last_row, context):
                worksheet.conditional_formatting.add(
                    column_range,
                    DataBarRule(start_type='min', end_type='max', color=bar_color)
                )
                written += 1
            return written

        raise StepProcessorError(f"{context}: no recognized rule kind")

    def _apply_when_cell(self, worksheet, rule, headers, last_row, style,
                         stop_if_true, context) -> int:
        """Canonical-condition rule; emits one Excel rule per column where anchored."""
        spec = rule['when_cell']
        condition = self._canonical_condition(spec.get('condition'), context, warn=False)
        value = spec.get('value')

        written = 0

        if condition in CELL_IS_OPERATORS:
            operands = value if condition in ('between', 'not_between') else [value]
            formula_operands = [self._value_as_formula_operand(operand) for operand in operands]
            # No anchored references in the operands, so all columns can
            # share one rule across a combined multi-range.
            ranges = [column_range for column_range, _
                      in self._column_ranges(spec, headers, last_row, context)]
            worksheet.conditional_formatting.add(
                ' '.join(ranges),
                CellIsRule(operator=CELL_IS_OPERATORS[condition],
                           formula=formula_operands, stopIfTrue=stop_if_true,
                           **self._style_as_rule_kwargs(style))
            )
            return 1

        # Anchored kinds: one rule per column, each with its own anchor cell.
        for column_range, column_letter in self._column_ranges(spec, headers, last_row, context):
            anchor = f"{column_letter}2"

            if condition in TEXT_RULE_TYPES:
                text = str(value)
                anchored = self._build_text_formula(condition, text, anchor)
                excel_rule = Rule(
                    type=TEXT_RULE_TYPES[condition],
                    operator=TEXT_RULE_TYPES[condition] if condition in ('contains', 'not_contains') else
                             ('beginsWith' if condition == 'starts_with' else 'endsWith'),
                    text=text, formula=[anchored],
                    dxf=self._style_as_dxf(style), stopIfTrue=stop_if_true,
                )
            elif condition in BLANK_RULE_TYPES:
                comparison = '=0' if condition == 'is_empty' else '>0'
                excel_rule = Rule(
                    type=BLANK_RULE_TYPES[condition],
                    formula=[f"LEN(TRIM({anchor})){comparison}"],
                    dxf=self._style_as_dxf(style), stopIfTrue=stop_if_true,
                )
            else:  # duplicates / unique - per column so each column is its own domain
                excel_rule = Rule(
                    type=VALUE_SET_RULE_TYPES[condition],
                    dxf=self._style_as_dxf(style), stopIfTrue=stop_if_true,
                )

            worksheet.conditional_formatting.add(column_range, excel_rule)
            written += 1

        return written

    # ------------------------------------------------------------------
    # Formula and range plumbing
    # ------------------------------------------------------------------

    def _prepare_formula(self, formula_text: str, headers, context) -> str:
        """{col:} placeholders to $-locked letters, then _xlfn prefixes; no leading =."""
        def substitute(match):
            header_name = match.group(1).strip()
            if header_name not in headers:
                raise StepProcessorError(
                    f"{context}: column '{header_name}' not found in the header "
                    f"row. Available: {sorted(headers)}"
                )
            return f"${headers[header_name]}"

        resolved = column_placeholder_rgx.sub(substitute, formula_text)
        prefixed = prefix_future_functions(resolved)
        return prefixed[1:] if prefixed.startswith('=') else prefixed

    def _build_formula_target(self, rule, headers, last_row, last_col_letter, context) -> str:
        """Resolve a when_formula rule's one target into an A1 range."""
        if rule.get('apply_to'):
            if rule['apply_to'] != 'entire_row':
                raise StepProcessorError(
                    f"{context}: apply_to only supports \"entire_row\"; "
                    f"use columns: or range: for anything else"
                )
            return f"A2:{last_col_letter}{last_row}"

        if rule.get('column_names'):
            ranges = [column_range for column_range, _
                      in self._column_ranges(rule, headers, last_row, context)]
            return ' '.join(ranges)

        return str(rule['range'])

    def _column_ranges(self, spec, headers, last_row, context):
        """Yield (range_text, column_letter) for each named column, data rows only."""
        column_names = spec.get('column_names', [])
        if not column_names:
            raise StepProcessorError(f"{context}: needs 'column_names' (list of header names)")

        for column_name in column_names:
            if column_name not in headers:
                raise StepProcessorError(
                    f"{context}: column '{column_name}' not found in the header "
                    f"row. Available: {sorted(headers)}"
                )
            letter = headers[column_name]
            yield f"{letter}2:{letter}{last_row}", letter

    @staticmethod
    def _build_header_map(worksheet) -> dict:
        """Header text -> column letter, from row 1."""
        headers = {}
        for column_index, cell in enumerate(worksheet[1], start=1):
            if cell.value is not None:
                headers[str(cell.value).strip()] = get_column_letter(column_index)
        return headers

    @staticmethod
    def _value_as_formula_operand(value) -> str:
        """Numbers pass through; strings get the formula quoting Excel needs."""
        if isinstance(value, bool):
            return 'TRUE' if value else 'FALSE'
        if isinstance(value, (int, float)):
            return str(value)
        text = str(value).replace('"', '""')
        return f'"{text}"'

    @staticmethod
    def _build_text_formula(condition: str, text: str, anchor: str) -> str:
        """The anchored formula Excel expects each text rule type to carry."""
        escaped = text.replace('"', '""')
        if condition == 'contains':
            return f'NOT(ISERROR(SEARCH("{escaped}",{anchor})))'
        if condition == 'not_contains':
            return f'ISERROR(SEARCH("{escaped}",{anchor}))'
        if condition == 'starts_with':
            return f'LEFT({anchor},LEN("{escaped}"))="{escaped}"'
        return f'RIGHT({anchor},LEN("{escaped}"))="{escaped}"'

    def _style_as_rule_kwargs(self, style: dict) -> dict:
        """Style dict -> the font/fill kwargs the convenience rule classes take."""
        kwargs = {}
        font_options = {}
        if 'font_color' in style:
            font_options['color'] = normalize_color(style['font_color'])
        if style.get('bold'):
            font_options['bold'] = True
        if style.get('italic'):
            font_options['italic'] = True
        if font_options:
            kwargs['font'] = Font(**font_options)
        if 'fill' in style:
            fill_hex = normalize_color(style['fill'])
            kwargs['fill'] = PatternFill(start_color=fill_hex, end_color=fill_hex,
                                         fill_type='solid')
        return kwargs

    def _style_as_dxf(self, style: dict) -> DifferentialStyle:
        """Style dict -> the differential style the raw Rule class takes."""
        kwargs = self._style_as_rule_kwargs(style)
        return DifferentialStyle(font=kwargs.get('font'), fill=kwargs.get('fill'))

    def _resolve_path(self, filename: str) -> str:
        """Apply recipe variable substitution to a configured path."""
        if hasattr(self, 'variable_substitution') and self.variable_substitution:
            return self.variable_substitution.substitute(filename)
        return filename

    def get_usage_examples(self) -> dict:
        """Get usage examples from the external YAML file."""
        from excel_recipe_processor.utils.processor_examples_loader import load_processor_examples
        return load_processor_examples('conditional_format')

    def get_capabilities(self) -> dict:
        """
        Get processor capabilities information.

        Returns:
            Dictionary with processor capabilities
        """
        return {
            'description': 'Write native Excel conditional-formatting rules that '
                           'stay live in the file',
            'vocabulary': 'canonical ERP condition names (filter_data\'s equals, '
                          'greater_than, is_empty, ... plus between, duplicates, '
                          'unique in the same style); Excel-native spellings '
                          '(greaterThan, containsText, ...) accepted as aliases '
                          'with a warning naming the canonical form',
            'rule_kinds': ['when_formula', 'when_cell', 'color_scale', 'data_bar'],
            'formula_convention': 'same as inject_formulas - write for data row 2 '
                                  'with {col:Header Name} placeholders - except '
                                  'placeholders resolve $-LOCKED here so row-wise '
                                  'tests cannot drift sideways as Excel translates '
                                  'the rule across its range; modern function names '
                                  'get the _xlfn storage prefix automatically',
            'targeting': 'apply_to: entire_row, columns: [names], or a literal range',
            'ordering': 'list order is Excel priority order; stop_if_true supported',
            'per_column_anchoring': 'text, blank, duplicate and unique conditions '
                                    'emit one Excel rule per named column, each '
                                    'anchored to its own range (and duplicates are '
                                    'per-column domains by design)',
        }

# End of file #

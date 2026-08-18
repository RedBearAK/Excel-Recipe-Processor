"""
Tests for the calamine Excel read path.

tests/test_calamine_reader.py

The engine switch is only acceptable if a calamine read is value-for-value
identical to an openpyxl read of the same file. These tests prove that
surface directly, engine against engine, and skip gracefully on machines
without the python-calamine wheel (where FileReader serves the openpyxl
path unchanged).

Runnable with pytest, but written to run standalone and report a score.
"""

import tempfile

from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl

from excel_recipe_processor.core.file_reader import FileReader, FileReaderError, CALAMINE_AVAILABLE


def make_mixed_workbook(path):
    """Numeric, text, mixed, datetime, blank and NA-text columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['Num', 'Text', 'Mixed', 'Date', 'Blank', 'NAtext']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    rows = [
        (1.5, 'alpha', 'x', datetime(2026, 3, 1), None, 'N/A'),
        (2, 'beta', 7, datetime(2026, 3, 2), None, 'NA'),
        (3.25, 'gamma', None, None, None, 'real'),
    ]
    for r, values in enumerate(rows, 2):
        for col, value in enumerate(values, 1):
            ws.cell(row=r, column=col).value = value
    formulas = wb.create_sheet('Formulas')
    formulas['A1'] = 'F'
    formulas['A2'] = '=1+1'
    formulas['B1'] = 'Plain'
    formulas['B2'] = 5
    wb.save(path)
    wb.close()


def test_engines_agree_on_a_mixed_frame():
    """dtypes, values, datetimes and blanks identical across engines."""
    print("\nTesting engine agreement on mixed data...")

    if not CALAMINE_AVAILABLE:
        print("  ~ python-calamine not installed; openpyxl path in use (skip)")
        return True

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'mixed.xlsx')
        make_mixed_workbook(path)

        via_openpyxl = pd.read_excel(path)
        via_calamine = pd.read_excel(path, engine='calamine')

        passed = True

        if via_openpyxl.equals(via_calamine):
            print("  ✓ Frames identical, value for value")
        else:
            print("  ✗ Frames differ")
            passed = False

        if dict(via_openpyxl.dtypes.astype(str)) == dict(via_calamine.dtypes.astype(str)):
            print(f"  ✓ dtypes identical: {dict(via_calamine.dtypes.astype(str))}")
        else:
            print(f"  ✗ dtypes differ")
            passed = False

        if isinstance(via_calamine['Date'].iloc[0], pd.Timestamp) and pd.isna(via_calamine['Date'].iloc[2]):
            print("  ✓ Datetimes are Timestamps, missing date is NaT")
        else:
            print("  ✗ Datetime handling differs")
            passed = False

        return passed


def test_engines_agree_on_the_raw_na_path():
    """keep_default_na=False (the verbatim_text_columns substrate) matches."""
    print("\nTesting the raw-NA path...")

    if not CALAMINE_AVAILABLE:
        print("  ~ python-calamine not installed (skip)")
        return True

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'mixed.xlsx')
        make_mixed_workbook(path)

        via_openpyxl = pd.read_excel(path, keep_default_na=False)
        via_calamine = pd.read_excel(path, engine='calamine', keep_default_na=False)

        if via_openpyxl.equals(via_calamine) and list(via_calamine['NAtext']) == ['N/A', 'NA', 'real']:
            print("  ✓ Raw frames identical; literal N/A text intact")
            return True
        print(f"  ✗ NAtext: {list(via_calamine['NAtext'])}")
        return False


def test_engines_agree_on_formula_cells():
    """A formula cell without a cached result reads the same both ways."""
    print("\nTesting formula-cell agreement...")

    if not CALAMINE_AVAILABLE:
        print("  ~ python-calamine not installed (skip)")
        return True

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'mixed.xlsx')
        make_mixed_workbook(path)

        via_openpyxl = pd.read_excel(path, sheet_name='Formulas')
        via_calamine = pd.read_excel(path, sheet_name='Formulas', engine='calamine')

        if via_openpyxl.equals(via_calamine):
            print("  ✓ Formula sheet identical across engines")
            return True
        print(f"  ✗ openpyxl {via_openpyxl.to_dict('list')} vs calamine {via_calamine.to_dict('list')}")
        return False


def test_verbatim_columns_ride_the_new_engine():
    """FileReader's protected-text policy behaves identically on calamine."""
    print("\nTesting verbatim_text_columns through FileReader...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'mixed.xlsx')
        make_mixed_workbook(path)

        frame = FileReader.read_file(path, verbatim_text_columns=['NAtext'])

        passed = True

        if list(frame['NAtext']) == ['N/A', 'NA', 'real']:
            print("  ✓ Protected column keeps its literal text")
        else:
            print(f"  ✗ NAtext: {list(frame['NAtext'])}")
            passed = False

        if frame['Blank'].isna().all():
            print("  ✓ Blank column still imports as missing")
        else:
            print("  ✗ Blank column corrupted")
            passed = False

        if str(frame['Num'].dtype) == 'float64':
            print("  ✓ Numeric dtype preserved")
        else:
            print(f"  ✗ Num dtype: {frame['Num'].dtype}")
            passed = False

        return passed


def test_positional_sheets_through_filereader():
    """
    Sheet addressing per the 2026-08-14 doctrine: ints are positional for
    internal callers; STRINGS are always NAMES. The numeric-string
    coercion that once shadowed tabs literally named "1" is gone, so
    sheet='1' on a workbook without such a tab must fail loud.
    """
    print("\nTesting positional sheet addressing...")

    with tempfile.TemporaryDirectory() as temp_dir:
        path = str(Path(temp_dir) / 'mixed.xlsx')
        make_mixed_workbook(path)

        by_int = FileReader.read_file(path, sheet=1)
        second = FileReader.read_file(path, sheet=2)

        try:
            FileReader.read_file(path, sheet='1')
            print("  ✗ sheet='1' should be a NAME and fail loud on this workbook")
            return False
        except FileReaderError as error:
            if 'not found' not in str(error):
                print(f"  ✗ wrong error for name '1': {error}")
                return False

        if len(by_int) > 0 and list(second.columns) == ['F', 'Plain']:
            print("  ✓ int addresses by position; string '1' is a name and fails loud")
            return True
        print(f"  ✗ second sheet columns: {list(second.columns)}")
        return False


def main():
    """Run every test and report a final score."""
    print("=== calamine reader tests ===")

    if not CALAMINE_AVAILABLE:
        print("(python-calamine absent: engine-comparison tests will skip, "
              "FileReader tests exercise the openpyxl fallback)")

    tests = [
        test_engines_agree_on_a_mixed_frame,
        test_engines_agree_on_the_raw_na_path,
        test_engines_agree_on_formula_cells,
        test_verbatim_columns_ride_the_new_engine,
        test_positional_sheets_through_filereader,
    ]

    passed = 0

    for test_func in tests:
        try:
            if test_func():
                passed += 1
        except Exception as error:
            print(f"  ✗ {test_func.__name__} crashed: {error}")

    print(f"\n=== Results: {passed}/{len(tests)} tests passed ===")

    if passed == len(tests):
        print("✅ All calamine reader tests passed!")
        return 1

    print("❌ Some calamine reader tests failed!")
    return 0


if __name__ == '__main__':
    exit(0 if main() else 1)


# End of file #

"""
Format Excel step processor for Excel automation recipes - Enhanced Version with Template Support.

excel_recipe_processor/processors/format_excel_processor.py

Handles formatting existing Excel files with auto-fit columns, header styling, and other presentation features.
Enhanced with comprehensive header formatting including text colors and font sizes.
NEW: Template support for reusable formatting configurations.
"""

import logging
import openpyxl
import webcolors

from pathlib import Path

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from excel_recipe_processor.core.variable_substitution import VariableSubstitution
from excel_recipe_processor.core.base_processor import FileOpsBaseProcessor, StepProcessorError


logger = logging.getLogger(__name__)


class FormatExcelProcessor(FileOpsBaseProcessor):
    """
    Processor for formatting existing Excel files.
    
    Applies formatting like auto-fit columns, header styling, freeze panes,
    and other presentation features to make Excel files look professional.
    
    NEW: Supports reusable formatting templates to reduce configuration redundancy.
    """
    
    @classmethod
    def get_minimal_config(cls) -> dict:
        return {
            'target_file': 'output.xlsx'
        }

    def perform_file_operation(self) -> str:
        """Format the target Excel file with template support."""
        target_file = self.get_config_value('target_file')
        sheet_configs = self.get_config_value('formatting', [])
        active_sheet = self.get_config_value('active_sheet')
        templates = self.get_config_value('templates', [])
        
        # Validate configuration
        self._validate_format_config(target_file, sheet_configs, active_sheet, templates)
        
        # Apply variable substitution BEFORE file operations
        if hasattr(self, 'variable_substitution') and self.variable_substitution:
            resolved_file = self.variable_substitution.substitute(target_file)
        else:
            resolved_file = target_file
        
        # Check file exists
        if not Path(resolved_file).exists():
            raise StepProcessorError(f"Target file not found: {resolved_file}")
        
        # Load and format the workbook
        formatted_sheets = self._format_excel_file(resolved_file, sheet_configs, active_sheet, templates)
        
        return f"formatted {resolved_file} ({formatted_sheets} sheets processed)"

    def _validate_file_operation_config(self):
        """Validate format_excel specific configuration with template support."""
        if not self.get_config_value('target_file'):
            raise StepProcessorError(f"Format Excel step '{self.step_name}' requires 'target_file'")
        
        # Get and validate the formatting configuration
        sheet_configs = self.get_config_value('formatting', [])
        active_sheet = self.get_config_value('active_sheet')
        target_file = self.get_config_value('target_file')
        templates = self.get_config_value('templates', [])
        
        # Validate using the enhanced validation method
        self._validate_format_config(target_file, sheet_configs, active_sheet, templates)

    def _validate_format_config(self, target_file: str, sheet_configs: list, active_sheet=None, templates: list = None) -> None:
        """
        Validate the explicit sheet targeting configuration format with template support.
        
        Args:
            target_file: Target Excel file path
            sheet_configs: List of sheet configuration dicts
            active_sheet: Active sheet specification (optional)
            templates: List of template definitions (optional)
        """
        # Validate templates first
        if templates is not None:
            self._validate_templates(templates)
        
        # Require sheet_configs to be a list
        if not isinstance(sheet_configs, list):
            raise StepProcessorError("'formatting' must be a list of sheet configurations, each with a 'sheet' key")
        
        if not sheet_configs:
            logger.warning("Empty formatting list - no sheets will be formatted")
            return
        
        # Build template lookup for validation
        template_lookup = {}
        if templates:
            for template in templates:
                if 'template_name' in template:
                    template_lookup[template['template_name']] = template
        
        # Validate each sheet configuration
        for i, sheet_config in enumerate(sheet_configs):
            if not isinstance(sheet_config, dict):
                raise StepProcessorError(f"Formatting entry {i+1} must be a dictionary")
            
            if 'sheet' not in sheet_config:
                raise StepProcessorError(f"Formatting entry {i+1} must have a 'sheet' key")
            
            sheet_spec = sheet_config['sheet']
            if not isinstance(sheet_spec, (str, int)):
                raise StepProcessorError(f"Sheet specification must be string (name) or integer (1-based index), got: {type(sheet_spec).__name__}")
            
            if isinstance(sheet_spec, int) and sheet_spec < 1:
                raise StepProcessorError(f"Sheet index must be >= 1, got: {sheet_spec}")
            
            if isinstance(sheet_spec, str) and not sheet_spec.strip():
                raise StepProcessorError("Sheet name cannot be empty")
            
            # Validate apply_templates if present
            if 'apply_templates' in sheet_config:
                apply_templates = sheet_config['apply_templates']
                if not isinstance(apply_templates, list):
                    raise StepProcessorError(f"'apply_templates' must be a list of template names for sheet {sheet_spec}")
                
                for template_name in apply_templates:
                    if not isinstance(template_name, str):
                        raise StepProcessorError(f"Template names must be strings, got {type(template_name).__name__} for sheet {sheet_spec}")
                    
                    if template_name not in template_lookup:
                        logger.warning(f"⚠️ Template '{template_name}' referenced by sheet {sheet_spec} not found - will be skipped")
            
            # Validate the formatting options for this sheet
            self._validate_sheet_formatting_options(sheet_config, f"sheet {sheet_spec}")
        
        # Validate active_sheet specification
        if active_sheet is not None:
            if not isinstance(active_sheet, (str, int)):
                raise StepProcessorError(f"active_sheet must be string (name) or integer (1-based index), got: {type(active_sheet).__name__}")
            
            if isinstance(active_sheet, int) and active_sheet < 1:
                raise StepProcessorError(f"active_sheet index must be >= 1, got: {active_sheet}")
            
            if isinstance(active_sheet, str) and not active_sheet.strip():
                raise StepProcessorError("active_sheet name cannot be empty")

    def _validate_templates(self, templates: list) -> None:
        """
        Validate template definitions.
        
        Args:
            templates: List of template dictionaries
        """
        if not isinstance(templates, list):
            raise StepProcessorError("'templates' must be a list of template definitions")
        
        template_names = set()
        
        for i, template in enumerate(templates):
            if not isinstance(template, dict):
                raise StepProcessorError(f"Template {i+1} must be a dictionary")
            
            if 'template_name' not in template:
                raise StepProcessorError(f"Template {i+1} must have a 'template_name' key")
            
            template_name = template['template_name']
            if not isinstance(template_name, str):
                raise StepProcessorError(f"Template name must be a string, got {type(template_name).__name__}")
            
            if not template_name.strip():
                raise StepProcessorError(f"Template {i+1} name cannot be empty")
            
            template_name = template_name.strip()
            
            if template_name in template_names:
                raise StepProcessorError(f"Duplicate template name '{template_name}' found")
            
            template_names.add(template_name)
            
            # Validate the formatting options in this template
            try:
                self._validate_sheet_formatting_options(template, f"template '{template_name}'")
            except StepProcessorError as e:
                raise StepProcessorError(f"Template '{template_name}' validation failed: {e}")

    def _build_template_lookup(self, templates: list) -> dict:
        """
        Build a lookup dictionary for templates.
        
        Args:
            templates: List of template definitions
            
        Returns:
            Dictionary mapping template names to template configurations
        """
        template_lookup = {}
        
        if not templates:
            return template_lookup
            
        for template in templates:
            if 'template_name' in template:
                template_name = template['template_name']
                # Create a copy without the template_name key for formatting
                template_config = {k: v for k, v in template.items() if k != 'template_name'}
                template_lookup[template_name] = template_config
        
        return template_lookup

    def _apply_templates_to_sheet_config(self, sheet_config: dict, template_lookup: dict) -> dict:
        """
        Apply templates to a sheet configuration.
        
        Args:
            sheet_config: Original sheet configuration
            template_lookup: Dictionary of available templates
            
        Returns:
            Enhanced sheet configuration with template rules applied
        """
        if 'apply_templates' not in sheet_config:
            return sheet_config
        
        # Start with an empty config, then apply templates in order
        enhanced_config = {}
        
        # Apply templates in order (later templates override earlier ones)
        apply_templates = sheet_config.get('apply_templates', [])
        for template_name in apply_templates:
            if template_name in template_lookup:
                template_config = template_lookup[template_name]
                logger.debug(f"Applying template '{template_name}' with {len(template_config)} rules")
                enhanced_config.update(template_config)
            else:
                logger.warning(f"⚠️ Template '{template_name}' not found - skipping")
        
        # Apply direct sheet rules (these override template rules)
        for key, value in sheet_config.items():
            if key not in ['apply_templates']:  # Don't copy the apply_templates directive itself
                enhanced_config[key] = value
        
        return enhanced_config

    def _validate_sheet_formatting_options(self, sheet_config: dict, context: str) -> None:
        """
        Complete validation for sheet formatting options with header alignment support.
        
        Args:
            sheet_config: Sheet configuration dictionary
            context: Context string for error messages (e.g., "sheet 1", "template 'basic'")
        """
        # List of known formatting options (this helps catch typos)
        known_options = {
            # Sheet targeting
            'sheet', 'apply_templates', 'template_name',
            
            # Phase 1: Basic formatting
            'auto_fit_columns', 'header_bold', 'header_background', 'header_background_color',
            'freeze_top_row', 'auto_filter', 'max_column_width', 'min_column_width',
            'row_heights',
            
            # Phase 1 Enhanced: Header text formatting
            'header_text_color', 'header_font_size',
            
            # Phase 1 Enhanced: Header alignment
            'header_alignment_horizontal', 'header_alignment_vertical',
            
            # Phase 2: General cell formatting
            'general_text_color', 'general_font_size', 'general_font_name',
            'general_alignment_horizontal', 'general_alignment_vertical',
            
            # Phase 3: Cell range targeting and advanced colors
            'cell_ranges'
        }
        
        # Check for unknown options (helps catch typos)
        unknown_options = set(sheet_config.keys()) - known_options
        if unknown_options:
            unknown_list = ', '.join(sorted(unknown_options))
            logger.warning(f"⚠️ Unknown formatting options in {context}: {unknown_list}")
        
        # Validate specific options
        if 'header_background_color' in sheet_config:
            try:
                self._normalize_color(sheet_config['header_background_color'])
            except ValueError as e:
                raise StepProcessorError(f"Invalid header_background_color in {context}: {e}")
        
        if 'header_text_color' in sheet_config:
            try:
                self._normalize_color(sheet_config['header_text_color'])
            except ValueError as e:
                raise StepProcessorError(f"Invalid header_text_color in {context}: {e}")
        
        if 'general_text_color' in sheet_config:
            try:
                self._normalize_color(sheet_config['general_text_color'])
            except ValueError as e:
                raise StepProcessorError(f"Invalid general_text_color in {context}: {e}")
        
        # Validate alignment values
        valid_h_alignments = ['left', 'center', 'right', 'justify', 'distributed']
        valid_v_alignments = ['top', 'center', 'bottom', 'justify', 'distributed']
        
        for alignment_key, valid_values, alignment_type in [
            ('header_alignment_horizontal', valid_h_alignments, 'horizontal header'),
            ('header_alignment_vertical', valid_v_alignments, 'vertical header'),
            ('general_alignment_horizontal', valid_h_alignments, 'horizontal general'),
            ('general_alignment_vertical', valid_v_alignments, 'vertical general')
        ]:
            if alignment_key in sheet_config:
                alignment_value = sheet_config[alignment_key]
                if alignment_value not in valid_values:
                    raise StepProcessorError(
                        f"Invalid {alignment_type} alignment '{alignment_value}' in {context}. "
                        f"Valid values: {', '.join(valid_values)}"
                    )
        
        # Validate numeric values
        if 'header_font_size' in sheet_config:
            font_size = sheet_config['header_font_size']
            if not isinstance(font_size, (int, float)) or font_size <= 0:
                raise StepProcessorError(f"header_font_size must be a positive number in {context}, got: {font_size}")
        
        if 'general_font_size' in sheet_config:
            font_size = sheet_config['general_font_size']
            if not isinstance(font_size, (int, float)) or font_size <= 0:
                raise StepProcessorError(f"general_font_size must be a positive number in {context}, got: {font_size}")
        
        if 'max_column_width' in sheet_config:
            width = sheet_config['max_column_width']
            if not isinstance(width, (int, float)) or width <= 0:
                raise StepProcessorError(f"max_column_width must be a positive number in {context}, got: {width}")
        
        if 'min_column_width' in sheet_config:
            width = sheet_config['min_column_width']
            if not isinstance(width, (int, float)) or width <= 0:
                raise StepProcessorError(f"min_column_width must be a positive number in {context}, got: {width}")
        
        # Validate row_heights
        if 'row_heights' in sheet_config:
            row_heights = sheet_config['row_heights']
            if not isinstance(row_heights, dict):
                raise StepProcessorError(f"row_heights must be a dictionary in {context}, got: {type(row_heights).__name__}")
            
            for row_num, height in row_heights.items():
                if not isinstance(row_num, int) or row_num < 1:
                    raise StepProcessorError(f"row_heights keys must be positive integers in {context}, got: {row_num}")
                
                if not isinstance(height, (int, float)) or height <= 0:
                    raise StepProcessorError(f"row_heights values must be positive numbers in {context}, got: {height}")
        
        # Validate cell_ranges
        if 'cell_ranges' in sheet_config:
            cell_ranges = sheet_config['cell_ranges']
            if not isinstance(cell_ranges, dict):
                raise StepProcessorError(f"cell_ranges must be a dictionary in {context}, got: {type(cell_ranges).__name__}")
            
            for range_spec, range_formatting in cell_ranges.items():
                if not isinstance(range_spec, str):
                    raise StepProcessorError(f"cell_ranges keys must be strings in {context}, got: {type(range_spec).__name__}")
                
                if not isinstance(range_formatting, dict):
                    raise StepProcessorError(f"cell_ranges values must be dictionaries in {context}, got: {type(range_formatting).__name__}")
                
                # Validate the formatting options for this range
                # (Recursively validate, but exclude range-specific validation to avoid infinite recursion)
                try:
                    self._validate_range_formatting_options(range_formatting, f"{context} range '{range_spec}'")
                except StepProcessorError as e:
                    raise StepProcessorError(f"Invalid formatting for range '{range_spec}' in {context}: {e}")

    def _validate_range_formatting_options(self, range_formatting: dict, context: str) -> None:
        """
        Validate formatting options for a cell range.
        
        Args:
            range_formatting: Range formatting dictionary
            context: Context string for error messages
        """
        # Known range formatting options
        known_range_options = {
            'text_color', 'background_color', 'font_size', 'font_name', 'bold', 'italic',
            'alignment_horizontal', 'alignment_vertical', 'border'
        }
        
        # Check for unknown options
        unknown_options = set(range_formatting.keys()) - known_range_options
        if unknown_options:
            unknown_list = ', '.join(sorted(unknown_options))
            logger.warning(f"⚠️ Unknown range formatting options in {context}: {unknown_list}")
        
        # Validate colors
        for color_key in ['text_color', 'background_color']:
            if color_key in range_formatting:
                try:
                    self._normalize_color(range_formatting[color_key])
                except ValueError as e:
                    raise StepProcessorError(f"Invalid {color_key} in {context}: {e}")
        
        # Validate alignment values
        valid_h_alignments = ['left', 'center', 'right', 'justify', 'distributed']
        valid_v_alignments = ['top', 'center', 'bottom', 'justify', 'distributed']
        
        if 'alignment_horizontal' in range_formatting:
            alignment_value = range_formatting['alignment_horizontal']
            if alignment_value not in valid_h_alignments:
                raise StepProcessorError(f"Invalid horizontal alignment '{alignment_value}' in {context}")
        
        if 'alignment_vertical' in range_formatting:
            alignment_value = range_formatting['alignment_vertical']
            if alignment_value not in valid_v_alignments:
                raise StepProcessorError(f"Invalid vertical alignment '{alignment_value}' in {context}")
        
        # Validate numeric values
        if 'font_size' in range_formatting:
            font_size = range_formatting['font_size']
            if not isinstance(font_size, (int, float)) or font_size <= 0:
                raise StepProcessorError(f"font_size must be a positive number in {context}")

    def _normalize_color(self, color) -> str:
        """
        Normalize color to 6-digit uppercase hex format with webcolors support.
        
        Supports multiple color formats:
        - Hex with hash: #FF0000, #F00
        - Hex without hash: FF0000, F00  
        - CSS color names: red, blue, forestgreen (if webcolors available)
        - RGB format: rgb(255, 0, 0) (if webcolors available)
        
        Args:
            color: Color in various formats
            
        Returns:
            6-digit uppercase hex color (without #)
            
        Raises:
            ValueError: If color format is invalid
        """
        if color is None:
            raise ValueError("Color cannot be None")
        
        # Convert to string and clean whitespace
        color_str = str(color).strip()
        
        if not color_str:
            raise ValueError("Color cannot be empty")
        
        # Try webcolors first if available
        try:
            import webcolors
            return self._normalize_color_with_webcolors(color_str)
        except ImportError:
            # Fall back to basic hex color processing
            return self._normalize_color_basic(color_str)

    def _normalize_color_with_webcolors(self, color_str: str) -> str:
        """
        Normalize color using webcolors library for advanced color format support.
        
        Args:
            color_str: Color string in various formats
            
        Returns:
            6-digit uppercase hex color
        """
        # Handle RGB format: rgb(255, 0, 0)
        if color_str.lower().startswith('rgb(') and color_str.endswith(')'):
            try:
                # Extract RGB values
                rgb_content = color_str[4:-1]  # Remove 'rgb(' and ')'
                rgb_parts = [int(x.strip()) for x in rgb_content.split(',')]
                
                if len(rgb_parts) != 3:
                    raise ValueError("RGB format must have exactly 3 values")
                
                for val in rgb_parts:
                    if not 0 <= val <= 255:
                        raise ValueError("RGB values must be between 0 and 255")
                
                # Convert to hex
                hex_color = '%02X%02X%02X' % tuple(rgb_parts)
                return hex_color
                
            except (ValueError, TypeError) as e:
                raise ValueError(f"Invalid RGB format '{color_str}': {e}")
        
        # Try CSS color name
        try:
            # webcolors.name_to_hex returns hex with #, so we remove it
            hex_with_hash = webcolors.name_to_hex(color_str.lower())
            return hex_with_hash[1:].upper()  # Remove # and convert to uppercase
        except ValueError:
            pass  # Not a CSS color name, try hex format
        
        # Handle hex format (with or without #)
        if self._contains_non_hex_chars(color_str):
            raise ValueError(f"Unrecognized color format: '{color_str}'. Supported formats: hex (#FF0000), CSS names (red, blue), RGB (rgb(255,0,0))")
        
        return self._normalize_color_basic(color_str)

    def _contains_non_hex_chars(self, text: str) -> bool:
        """
        Check if text contains characters that aren't valid hex.
        
        Args:
            text: Text to check
            
        Returns:
            True if contains non-hex characters, False if only hex characters
        """
        hex_chars = set('0123456789ABCDEFabcdef')
        return any(c not in hex_chars for c in text)

    def _normalize_color_basic(self, color_clean: str) -> str:
        """
        Basic color normalization (fallback when webcolors unavailable).
        
        Args:
            color_clean: Cleaned color string
            
        Returns:
            6-digit uppercase hex color
        """
        # Remove # if present
        if color_clean.startswith('#'):
            color_clean = color_clean[1:]
        
        # Check for empty string after removing hash
        if not color_clean:
            raise ValueError("Color cannot be just a hash symbol")
        
        if not all(c in '0123456789ABCDEFabcdef' for c in color_clean):
            raise ValueError("Must be a valid hex color (e.g., 'FF0000', '#FF0000')")
        
        if len(color_clean) == 3:
            # Expand 3-digit to 6-digit hex
            color_clean = ''.join([c*2 for c in color_clean])
        elif len(color_clean) != 6:
            raise ValueError("Hex color must be 3 or 6 digits")
        
        return color_clean.upper()

    def _format_excel_file(self, filename: str, sheet_configs: list, active_sheet=None, templates: list = None) -> int:
        """
        Apply formatting to an Excel file with explicit sheet targeting and template support.
        
        Args:
            filename: Path to Excel file
            sheet_configs: List of sheet configuration dicts
            active_sheet: Sheet to set as active (name or number), optional
            templates: List of template definitions, optional
            
        Returns:
            Number of sheets processed
        """
        logger.info(f"📋 Loading Excel file: {Path(filename).name}")
        
        # Build template lookup
        template_lookup = self._build_template_lookup(templates or [])
        if template_lookup:
            template_names = ', '.join(f"'{name}'" for name in template_lookup.keys())
            logger.info(f"📝 Available templates: {template_names}")
        
        # Load workbook
        workbook = openpyxl.load_workbook(filename)
        sheets_processed = 0
        total_sheets = len(workbook.worksheets)
        
        logger.info(f"📊 Found {total_sheets} worksheet(s): {', '.join(workbook.sheetnames)}")
        
        # Check if we have sheet configurations
        if not sheet_configs:
            logger.warning("⚠️ No sheet formatting configurations found - no formatting applied")
            workbook.close()
            return 0
            
        logger.info(f"🎯 Processing {len(sheet_configs)} explicit sheet configuration(s)")
        
        # Process each sheet configuration
        for i, sheet_config in enumerate(sheet_configs):
            if 'sheet' not in sheet_config:
                workbook.close()
                raise StepProcessorError(f"Formatting entry {i+1} must have a 'sheet' key")
            
            sheet_spec = sheet_config['sheet']
            sheet_name = self._resolve_sheet_name(workbook, sheet_spec)
            
            if sheet_name:
                worksheet = workbook[sheet_name]
                logger.info(f"🔧 Processing sheet: '{sheet_name}' (specified as: {sheet_spec})")
                
                # Apply templates to sheet configuration
                enhanced_config = self._apply_templates_to_sheet_config(sheet_config, template_lookup)
                
                # Log template application if templates were used
                if 'apply_templates' in sheet_config and sheet_config['apply_templates']:
                    applied_template_names = ', '.join(f"'{name}'" for name in sheet_config['apply_templates'])
                    logger.info(f"📝 Applied templates: {applied_template_names}")
                
                self._apply_sheet_formatting(worksheet, enhanced_config)
                sheets_processed += 1
            else:
                logger.warning(f"⚠️ Sheet '{sheet_spec}' not found, skipping")
        
        # Set active sheet if specified (supports both names and numbers)
        if active_sheet is not None:
            active_sheet_name = self._resolve_sheet_name(workbook, active_sheet)
            if active_sheet_name:
                workbook.active = workbook[active_sheet_name]
                logger.info(f"📌 Set active sheet to '{active_sheet_name}' (specified as: {active_sheet})")
            else:
                logger.warning(f"⚠️ Active sheet '{active_sheet}' not found")
        
        # Save workbook
        logger.info(f"💾 Saving formatted workbook...")
        workbook.save(filename)
        workbook.close()
        
        logger.info(f"✅ Excel formatting completed: {sheets_processed}/{total_sheets} sheets processed")
        return sheets_processed

    def _resolve_sheet_name(self, workbook, sheet_spec) -> str:
        """
        Resolve a sheet specification (name or number) to actual sheet name.
        
        Args:
            workbook: openpyxl workbook object
            sheet_spec: Sheet name (string) or 1-based index (integer)
            
        Returns:
            Actual sheet name if found, None if not found
        """
        if isinstance(sheet_spec, str):
            # Sheet specified by name
            if sheet_spec in workbook.sheetnames:
                return sheet_spec
            else:
                logger.debug(f"Sheet name '{sheet_spec}' not found in {workbook.sheetnames}")
                return None
                
        elif isinstance(sheet_spec, int):
            # Sheet specified by 1-based index
            if 1 <= sheet_spec <= len(workbook.worksheets):
                # Convert to 0-based index and get sheet name
                resolved_name = workbook.worksheets[sheet_spec - 1].title
                logger.debug(f"Sheet index {sheet_spec} resolved to '{resolved_name}'")
                return resolved_name
            else:
                logger.debug(f"Sheet index {sheet_spec} out of range (1-{len(workbook.worksheets)})")
                return None
                
        else:
            # Invalid sheet specification type
            logger.debug(f"Invalid sheet specification type: {type(sheet_spec)}")
            return None

    def _apply_sheet_formatting(self, worksheet, formatting: dict) -> None:
        """
        Apply formatting to a specific worksheet with correct operation order.
        
        Args:
            worksheet: openpyxl worksheet object
            formatting: Formatting configuration for this sheet (may include resolved templates)
        """
        sheet_name = worksheet.title
        applied_operations = []
        
        # STEP 1: Apply text formatting FIRST so auto-fit can measure correctly
        
        # Apply general formatting to all cells
        if self._has_general_formatting(formatting):
            logger.info(f"🎨 [{sheet_name}] Applying general cell formatting")
            self._apply_general_formatting(worksheet, formatting)
            applied_operations.append("general formatting")
        
        # Enhanced header formatting (BEFORE auto-fit so it can measure bold/larger text)
        if self._has_header_formatting(formatting):
            header_details = self._get_header_formatting_details(formatting)
            logger.info(f"👑 [{sheet_name}] Header formatting: {header_details}")
            self._apply_header_formatting(worksheet, formatting)
            applied_operations.append("header formatting")
        
        # Apply cell range specific formatting (BEFORE auto-fit)
        cell_ranges = formatting.get('cell_ranges', {})
        if cell_ranges:
            logger.info(f"🎯 [{sheet_name}] Applying cell range formatting to {len(cell_ranges)} ranges")
            self._apply_cell_range_formatting(worksheet, cell_ranges)
            applied_operations.append(f"range formatting ({len(cell_ranges)} ranges)")
        
        # STEP 2: Set row heights (might affect auto-fit calculations)
        row_heights = formatting.get('row_heights', {})
        if row_heights:
            logger.info(f"📏 [{sheet_name}] Setting custom row heights for {len(row_heights)} rows")
            self._apply_row_heights(worksheet, row_heights)
            applied_operations.append(f"row heights ({len(row_heights)} rows)")
        
        # STEP 3: Column sizing (auto-fit or explicit sizing)
        if formatting.get('auto_fit_columns'):
            logger.info(f"📐 [{sheet_name}] Auto-fitting column widths")
            self._auto_fit_columns(worksheet, formatting)
            applied_operations.append("auto-fit columns")
        
        # STEP 4: Worksheet-level features
        if formatting.get('freeze_top_row'):
            logger.info(f"🧊 [{sheet_name}] Freezing top row")
            worksheet.freeze_panes = 'A2'
            applied_operations.append("freeze top row")
        
        if formatting.get('auto_filter'):
            logger.info(f"🔍 [{sheet_name}] Adding auto-filter")
            self._add_auto_filter(worksheet)
            applied_operations.append("auto-filter")
        
        # Log summary of applied operations
        if applied_operations:
            operations_summary = ', '.join(applied_operations)
            logger.info(f"✅ [{sheet_name}] Applied: {operations_summary}")
        else:
            logger.info(f"ℹ️ [{sheet_name}] No formatting operations applied")

    def _has_general_formatting(self, formatting: dict) -> bool:
        """Check if general cell formatting is needed."""
        return any([
            formatting.get('general_text_color') is not None,
            formatting.get('general_font_size') is not None,
            formatting.get('general_font_name') is not None,
            formatting.get('general_alignment_horizontal') is not None,
            formatting.get('general_alignment_vertical') is not None
        ])

    def _has_header_formatting(self, formatting: dict) -> bool:
        """Check if header formatting is needed."""
        return any([
            formatting.get('header_bold'),
            formatting.get('header_background'),
            formatting.get('header_text_color') is not None,
            formatting.get('header_font_size') is not None,
            formatting.get('header_alignment_horizontal') is not None,
            formatting.get('header_alignment_vertical') is not None
        ])

    def _get_header_formatting_details(self, formatting: dict) -> str:
        """Get a descriptive string of header formatting operations."""
        details = []
        if formatting.get('header_bold'):
            details.append("bold")
        if formatting.get('header_background'):
            color = formatting.get('header_background_color', 'default')
            details.append(f"background ({color})")
        if formatting.get('header_text_color'):
            details.append(f"text color ({formatting['header_text_color']})")
        if formatting.get('header_font_size'):
            details.append(f"font size ({formatting['header_font_size']})")
        if formatting.get('header_alignment_horizontal'):
            details.append(f"h-align ({formatting['header_alignment_horizontal']})")
        if formatting.get('header_alignment_vertical'):
            details.append(f"v-align ({formatting['header_alignment_vertical']})")
        
        return ', '.join(details) if details else "basic"

    def _apply_general_formatting(self, worksheet, formatting: dict) -> None:
        """
        Apply general formatting to all data cells (excluding header row).
        
        Args:
            worksheet: openpyxl worksheet object
            formatting: Formatting configuration
        """
        # Check if any general formatting is needed
        needs_general_formatting = any([
            formatting.get('general_text_color') is not None,
            formatting.get('general_font_size') is not None,
            formatting.get('general_font_name') is not None,
            formatting.get('general_alignment_horizontal') is not None,
            formatting.get('general_alignment_vertical') is not None
        ])
        
        if not needs_general_formatting:
            return
        
        # Build font formatting if needed
        font_kwargs = {}
        general_text_color = formatting.get('general_text_color')
        if general_text_color:
            normalized_color = self._normalize_color(general_text_color)
            font_kwargs['color'] = normalized_color
        
        general_font_size = formatting.get('general_font_size')
        if general_font_size:
            font_kwargs['size'] = general_font_size
        
        general_font_name = formatting.get('general_font_name')
        if general_font_name:
            font_kwargs['name'] = general_font_name
        
        # Build alignment formatting if needed
        alignment_kwargs = {}
        general_h_align = formatting.get('general_alignment_horizontal')
        if general_h_align:
            alignment_kwargs['horizontal'] = general_h_align
        
        general_v_align = formatting.get('general_alignment_vertical')
        if general_v_align:
            alignment_kwargs['vertical'] = general_v_align
        
        # Apply formatting to all data cells (excluding headers which are handled separately)
        start_row = 2 if worksheet.max_row > 1 else 1  # Skip header row if present
        
        for row in range(start_row, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                
                # Apply font formatting
                if font_kwargs:
                    # Merge with existing font properties to avoid overwriting header formatting
                    existing_font = cell.font
                    merged_font_kwargs = {
                        'name': existing_font.name,
                        'size': existing_font.size,
                        'bold': existing_font.bold,
                        'italic': existing_font.italic,
                        'color': existing_font.color
                    }
                    # Override with general formatting
                    merged_font_kwargs.update(font_kwargs)
                    cell.font = Font(**merged_font_kwargs)
                
                # Apply alignment formatting
                if alignment_kwargs:
                    # Merge with existing alignment properties
                    existing_alignment = cell.alignment
                    merged_alignment_kwargs = {
                        'horizontal': existing_alignment.horizontal,
                        'vertical': existing_alignment.vertical,
                        'text_rotation': existing_alignment.text_rotation,
                        'wrap_text': existing_alignment.wrap_text,
                        'shrink_to_fit': existing_alignment.shrink_to_fit,
                        'indent': existing_alignment.indent
                    }
                    # Override with general alignment
                    merged_alignment_kwargs.update(alignment_kwargs)
                    cell.alignment = Alignment(**merged_alignment_kwargs)

    def _apply_header_formatting(self, worksheet, formatting: dict) -> None:
        """
        Apply enhanced header formatting to the first row.
        
        Args:
            worksheet: openpyxl worksheet object
            formatting: Formatting configuration
        """
        if worksheet.max_row < 1:
            return  # No data to format
        
        # Build font formatting
        font_kwargs = {}
        
        # Header text color
        header_text_color = formatting.get('header_text_color')
        if header_text_color:
            normalized_color = self._normalize_color(header_text_color)
            font_kwargs['color'] = normalized_color
        
        # Header font size
        header_font_size = formatting.get('header_font_size')
        if header_font_size:
            font_kwargs['size'] = header_font_size
        
        # Header bold (legacy support)
        if formatting.get('header_bold'):
            font_kwargs['bold'] = True
        
        # Build background fill
        fill = None
        if formatting.get('header_background'):
            background_color = formatting.get('header_background_color', 'D3D3D3')
            normalized_bg_color = self._normalize_color(background_color)
            fill = PatternFill(start_color=normalized_bg_color, end_color=normalized_bg_color, fill_type="solid")
        
        # Build alignment
        alignment_kwargs = {}
        header_h_align = formatting.get('header_alignment_horizontal')
        if header_h_align:
            alignment_kwargs['horizontal'] = header_h_align
        
        header_v_align = formatting.get('header_alignment_vertical')
        if header_v_align:
            alignment_kwargs['vertical'] = header_v_align
        
        # Apply to first row
        for cell in worksheet[1]:  # First row
            # Apply font formatting
            if font_kwargs:
                # Preserve existing font properties and merge with new ones
                existing_font = cell.font
                merged_font_kwargs = {
                    'name': existing_font.name,
                    'size': existing_font.size,
                    'bold': existing_font.bold,
                    'italic': existing_font.italic,
                    'color': existing_font.color
                }
                merged_font_kwargs.update(font_kwargs)
                cell.font = Font(**merged_font_kwargs)
            
            # Apply background fill
            if fill:
                cell.fill = fill
            
            # Apply alignment
            if alignment_kwargs:
                existing_alignment = cell.alignment
                merged_alignment_kwargs = {
                    'horizontal': existing_alignment.horizontal,
                    'vertical': existing_alignment.vertical,
                    'text_rotation': existing_alignment.text_rotation,
                    'wrap_text': existing_alignment.wrap_text,
                    'shrink_to_fit': existing_alignment.shrink_to_fit,
                    'indent': existing_alignment.indent
                }
                merged_alignment_kwargs.update(alignment_kwargs)
                cell.alignment = Alignment(**merged_alignment_kwargs)

    def _apply_cell_range_formatting(self, worksheet, cell_ranges: dict) -> None:
        """
        Apply formatting to specific cell ranges.
        
        Args:
            worksheet: openpyxl worksheet object
            cell_ranges: Dictionary of range specifications and their formatting
        """
        for range_spec, range_formatting in cell_ranges.items():
            try:
                # Get the range of cells
                cell_range = worksheet[range_spec]
                
                # Handle both single cells and ranges
                if hasattr(cell_range, '__iter__') and not hasattr(cell_range, 'value'):
                    # It's a range of cells
                    cells_to_format = []
                    for row in cell_range:
                        if hasattr(row, '__iter__'):
                            # Multiple rows
                            cells_to_format.extend(row)
                        else:
                            # Single row
                            cells_to_format.append(row)
                else:
                    # It's a single cell
                    cells_to_format = [cell_range]
                
                # Apply formatting to each cell in the range
                for cell in cells_to_format:
                    self._apply_cell_formatting(cell, range_formatting)
                    
            except Exception as e:
                logger.warning(f"⚠️ Could not apply formatting to range '{range_spec}': {e}")

    def _apply_cell_formatting(self, cell, formatting: dict) -> None:
        """
        Apply formatting to a single cell.
        
        Args:
            cell: openpyxl cell object
            formatting: Dictionary of formatting options
        """
        # Font formatting
        font_kwargs = {}
        
        if 'text_color' in formatting:
            normalized_color = self._normalize_color(formatting['text_color'])
            font_kwargs['color'] = normalized_color
        
        if 'font_size' in formatting:
            font_kwargs['size'] = formatting['font_size']
        
        if 'font_name' in formatting:
            font_kwargs['name'] = formatting['font_name']
        
        if 'bold' in formatting:
            font_kwargs['bold'] = formatting['bold']
        
        if 'italic' in formatting:
            font_kwargs['italic'] = formatting['italic']
        
        if font_kwargs:
            # Merge with existing font
            existing_font = cell.font
            merged_font_kwargs = {
                'name': existing_font.name,
                'size': existing_font.size,
                'bold': existing_font.bold,
                'italic': existing_font.italic,
                'color': existing_font.color
            }
            merged_font_kwargs.update(font_kwargs)
            cell.font = Font(**merged_font_kwargs)
        
        # Background color
        if 'background_color' in formatting:
            normalized_bg_color = self._normalize_color(formatting['background_color'])
            fill = PatternFill(start_color=normalized_bg_color, end_color=normalized_bg_color, fill_type="solid")
            cell.fill = fill
        
        # Alignment
        alignment_kwargs = {}
        
        if 'alignment_horizontal' in formatting:
            alignment_kwargs['horizontal'] = formatting['alignment_horizontal']
        
        if 'alignment_vertical' in formatting:
            alignment_kwargs['vertical'] = formatting['alignment_vertical']
        
        if alignment_kwargs:
            existing_alignment = cell.alignment
            merged_alignment_kwargs = {
                'horizontal': existing_alignment.horizontal,
                'vertical': existing_alignment.vertical,
                'text_rotation': existing_alignment.text_rotation,
                'wrap_text': existing_alignment.wrap_text,
                'shrink_to_fit': existing_alignment.shrink_to_fit,
                'indent': existing_alignment.indent
            }
            merged_alignment_kwargs.update(alignment_kwargs)
            cell.alignment = Alignment(**merged_alignment_kwargs)
        
        # Border formatting (simplified version)
        if 'border' in formatting:
            border_spec = formatting['border']
            if isinstance(border_spec, str):
                # Simple border style for all sides
                side = Side(style=border_spec)
                cell.border = Border(top=side, bottom=side, left=side, right=side)
            elif isinstance(border_spec, dict):
                # Complex border formatting
                self._apply_complex_border(cell, border_spec)

    def _apply_complex_border(self, cell, border_spec: dict) -> None:
        """
        Apply complex border formatting to a cell.
        
        Args:
            cell: openpyxl cell object
            border_spec: Dictionary specifying border formatting
        """
        border_kwargs = {}
        
        # Handle different border specification formats
        for side_name in ['top', 'bottom', 'left', 'right']:
            if side_name in border_spec:
                side_spec = border_spec[side_name]
                if isinstance(side_spec, str):
                    # Simple style
                    border_kwargs[side_name] = Side(style=side_spec)
                elif isinstance(side_spec, dict):
                    # Complex style with color
                    style = side_spec.get('style', 'thin')
                    color = side_spec.get('color')
                    side_kwargs = {'style': style}
                    if color:
                        side_kwargs['color'] = self._normalize_color(color)
                    border_kwargs[side_name] = Side(**side_kwargs)
        
        # Handle 'all' specification
        if 'all' in border_spec:
            all_spec = border_spec['all']
            if isinstance(all_spec, str):
                side = Side(style=all_spec)
                border_kwargs.update({
                    'top': side, 'bottom': side, 'left': side, 'right': side
                })
            elif isinstance(all_spec, dict):
                style = all_spec.get('style', 'thin')
                color = all_spec.get('color')
                side_kwargs = {'style': style}
                if color:
                    side_kwargs['color'] = self._normalize_color(color)
                side = Side(**side_kwargs)
                border_kwargs.update({
                    'top': side, 'bottom': side, 'left': side, 'right': side
                })
        
        if border_kwargs:
            cell.border = Border(**border_kwargs)

    def _apply_row_heights(self, worksheet, row_heights: dict) -> None:
        """
        Apply custom row heights.
        
        Args:
            worksheet: openpyxl worksheet object
            row_heights: Dictionary mapping row numbers to heights
        """
        for row_num, height in row_heights.items():
            worksheet.row_dimensions[row_num].height = height

    def _auto_fit_columns(self, worksheet, formatting: dict) -> None:
        """
        Auto-fit column widths based on content with optional constraints.
        
        Args:
            worksheet: openpyxl worksheet object
            formatting: Formatting configuration (may contain width constraints)
        """
        # Tunable constants for auto-fit behavior
        BASE_PADDING = 4           # Base extra width beyond content (tune this for general spacing)
        AUTO_FILTER_EXTRA = 3      # Additional width when auto-filter dropdowns are present
        
        max_width = formatting.get('max_column_width', 100)
        min_width = formatting.get('min_column_width', 8)
        
        # Check if auto-filter is enabled (adds dropdown arrows that need space)
        has_auto_filter = (
            formatting.get('auto_filter', False) or 
            worksheet.auto_filter.ref is not None
        )
        
        # Calculate total padding
        auto_filter_padding = AUTO_FILTER_EXTRA if has_auto_filter else 0
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        # Consider font size for width calculation
                        content_length = len(str(cell.value))
                        
                        # Adjust for font size (rough approximation)
                        if hasattr(cell.font, 'size') and cell.font.size:
                            size_factor = cell.font.size / 11  # 11 is default font size
                            content_length = int(content_length * size_factor)
                        
                        # Adjust for bold text (roughly 10% wider)
                        if hasattr(cell.font, 'bold') and cell.font.bold:
                            content_length = int(content_length * 1.2)
                        
                        if content_length > max_length:
                            max_length = content_length
                            
                except Exception:
                    pass  # Skip cells that can't be measured
            
            # Calculate final width with all padding
            total_padding = BASE_PADDING + auto_filter_padding
            adjusted_width = max(min_width, min(max_length + total_padding, max_width))
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _add_auto_filter(self, worksheet) -> None:
        """
        Add auto-filter to the data range.
        
        Args:
            worksheet: openpyxl worksheet object
        """
        if worksheet.max_row > 1:
            data_range = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            worksheet.auto_filter.ref = data_range

    def get_operation_type(self) -> str:
        return "excel_formatting"
    
    def get_capabilities(self) -> dict:
        """Get processor capabilities information."""
        return {
            'description': 'Format existing Excel files with professional presentation features',
            'formatting_features': self.get_supported_features(),
            'formatting_categories': [
                'column_sizing', 'enhanced_header_styling', 'general_cell_formatting',
                'cell_range_targeting', 'advanced_color_support', 'border_formatting',
                'pane_freezing', 'row_sizing', 'data_filtering', 'sheet_activation'
            ],
            'template_support': True,  # NEW: Template feature
            'file_requirements': ['xlsx', 'xls'],
            'dependencies': ['openpyxl'],
            'optional_dependencies': ['webcolors'],
            'phase_1_enhancements': ['header_text_color', 'header_font_size'],
            'phase_2_enhancements': ['general_text_color', 'general_font_size', 'general_font_name', 'general_alignment_horizontal', 'general_alignment_vertical'],
            'phase_3_enhancements': ['cell_ranges', 'webcolors_integration', 'border_formatting', 'css_color_names', 'rgb_color_support'],
            'template_enhancements': ['reusable_templates', 'template_composition', 'template_override'],  # NEW
            'color_formats_supported': [
                'hex_with_hash (#FF0000)', 'hex_without_hash (FF0000)', 'short_hex (#F00)', 
                'css_color_names (red, blue, forestgreen)', 'rgb_format (rgb(255, 0, 0))'
            ],
            'examples': {
                'auto_fit': "Automatically size columns to fit content",
                'enhanced_headers': "Bold headers with custom text color and font size",
                'general_formatting': "Apply font and alignment to all data cells",
                'cell_range_targeting': "Format specific ranges with custom styles",
                'css_colors': "Use CSS color names like 'red', 'forestgreen', 'navy'",
                'rgb_colors': "Use RGB format like 'rgb(255, 0, 0)' for red",
                'borders': "Add professional borders to specific cell ranges",
                'dark_theme_fix': "White text on dark backgrounds for readability",
                'freeze_panes': "Freeze top row for easier navigation",
                'professional': "Apply comprehensive formatting for business reports",
                'templates': "Create reusable formatting templates to reduce configuration redundancy"  # NEW
            }
        }
    
    def get_supported_features(self) -> list:
        """Get list of supported formatting features."""
        return [
            'auto_fit_columns', 'header_formatting', 'general_formatting',
            'cell_range_targeting', 'freeze_panes', 'auto_filter', 'row_sizing',
            'advanced_colors', 'border_formatting', 'template_support'  # NEW: template_support
        ]
    
    def get_usage_examples(self) -> dict:
        """Get usage examples for this processor."""
        from excel_recipe_processor.utils.processor_examples_loader import load_processor_examples
        return load_processor_examples('format_excel')


# End of file #

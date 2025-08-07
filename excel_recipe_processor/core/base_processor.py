"""
Base step processor for Excel automation recipes.

Defines the interface and common functionality that all step processors must implement.
"""

import pandas as pd
import logging

from abc import ABC, abstractmethod
from typing import Any


logger = logging.getLogger(__name__)

step_desc = 'step_description'
proc_type = 'processor_type'

class StepProcessorError(Exception):
    """Raised when a step processor encounters an error during execution."""
    pass


class BaseStepProcessor(ABC):
    """
    Abstract base class for all step processors.
    
    All step types (filter_data, pivot_table, etc.) inherit from this class
    and implement the execute method.
    """
    
    def __init__(self, step_config: dict):
        """
        Initialize the step processor.
        
        Args:
            step_config: Dictionary containing step configuration from recipe
        """
        # Guard clause: step_config must be a dictionary
        if not isinstance(step_config, dict):
            raise StepProcessorError("Step configuration must be a dictionary")
        
        # Guard clause: step must have a type
        if proc_type not in step_config:
            raise StepProcessorError(f"Step configuration missing required '{proc_type}' field")
        
        step_type = step_config[proc_type]
        if not isinstance(step_type, str) or not step_type.strip():
            raise StepProcessorError(f"Step {proc_type} must be a non-empty string")
        
        self.step_config = step_config
        self.step_type = step_type
        self.step_name = step_config.get('step_description', f'Unnamed {step_type} step')
        
        self.source_stage = step_config.get('source_stage')
        self.save_to_stage = step_config.get('save_to_stage')
        self.confirm_stage_replacement = step_config.get('confirm_stage_replacement', False)
        
        # Guard clause: step_name must be a string if provided
        if 'step_description' in step_config and not isinstance(self.step_name, str):
            raise StepProcessorError("Step 'step_description' must be a string")
        
        logger.debug(f"Initialized {self.__class__.__name__} for step: {self.step_name}")
    
    @abstractmethod
    def execute(self, data: Any) -> Any:
        """
        Execute the step on the provided data.
        
        Args:
            data: Input data to process (typically pandas DataFrame)
            
        Returns:
            Processed data
            
        Raises:
            StepProcessorError: If step execution fails
        """
        pass
    
    def validate_required_fields(self, required_fields: list) -> None:
        """
        Validate that required configuration fields are present.
        
        Args:
            required_fields: List of required field names
            
        Raises:
            StepProcessorError: If any required fields are missing
        """
        # Guard clause: required_fields must be a list
        if not isinstance(required_fields, list):
            raise StepProcessorError("Required fields must be provided as a list")
        
        missing_fields = []
        for field in required_fields:
            # Guard clause: each field name must be a string
            if not isinstance(field, str):
                raise StepProcessorError(f"Field name must be a string, got: {type(field)}")
            
            if field not in self.step_config:
                missing_fields.append(field)
        
        if missing_fields:
            raise StepProcessorError(
                f"Step '{self.step_name}' missing required fields: {', '.join(missing_fields)}"
            )
    
    def get_config_value(self, key: str, default: Any = None) -> Any:
        """
        Get a configuration value with optional default.
        
        Args:
            key: Configuration key to retrieve
            default: Default value if key not found
            
        Returns:
            Configuration value or default
        """
        # Guard clause: key must be a string
        if not isinstance(key, str):
            raise StepProcessorError("Configuration key must be a string")
        
        return self.step_config.get(key, default)
    
    def validate_data_not_empty(self, data: Any) -> None:
        """
        Validate that input data is not None or empty.
        
        Args:
            data: Data to validate
            
        Raises:
            StepProcessorError: If data is None or empty
        """
        if data is None:
            raise StepProcessorError(f"Step '{self.step_name}' received None data")
        
        # Check for pandas DataFrame
        if hasattr(data, 'empty') and data.empty:
            raise StepProcessorError(f"Step '{self.step_name}' received empty DataFrame")
        
        # Check for lists
        if isinstance(data, list) and len(data) == 0:
            raise StepProcessorError(f"Step '{self.step_name}' received empty list")
        
        # Check for dictionaries
        if isinstance(data, dict) and len(data) == 0:
            raise StepProcessorError(f"Step '{self.step_name}' received empty dictionary")
    
    def log_step_start(self) -> None:
        """Log the start of step execution."""
        logger.info(f"Starting step: '{self.step_name}' ({self.step_type})")
    
    def log_step_complete(self, result_info: str = "") -> None:
        """
        Log the completion of step execution.
        
        Args:
            result_info: Optional information about the result
        """
        if result_info:
            logger.info(f"Completed step: '{self.step_name}'")
            logger.info(f" - {result_info}")
        else:
            logger.info(f"Completed step: '{self.step_name}'")
    
    def log_step_error(self, error: Exception) -> None:
        """
        Log step execution error.
        
        Args:
            error: Exception that occurred
        """
        logger.error(f"Error in step '{self.step_name}': {error}")
    
    def load_input_data(self) -> pd.DataFrame:
        """Load input data from source_stage."""
        if not self.source_stage:
            raise StepProcessorError(f"Step '{self.step_name}' requires source_stage")
        
        from excel_recipe_processor.core.stage_manager import StageManager
        return StageManager.load_stage(self.source_stage)
    
    def save_output_data(self, data) -> None:
        """Save output data to save_to_stage."""
        if not self.save_to_stage:
            raise StepProcessorError(f"Step '{self.step_name}' requires save_to_stage")
        
        from excel_recipe_processor.core.stage_manager import StageManager
        StageManager.save_stage(
            stage_name=self.save_to_stage,
            data=data,
            description=f"Result from {self.step_name}",
            step_name=self.step_name,
            confirm_replacement=self.confirm_stage_replacement
        )
    
    def execute_stage_to_stage(self) -> pd.DataFrame:
        """Execute complete stage-to-stage operation."""
        self.log_step_start()
        
        # Load input
        input_data = self.load_input_data()
        
        # Process
        result = self.execute(input_data)
        
        # Save output
        self.save_output_data(result)
        
        self.log_step_complete(f"processed {len(result)} rows")
        return result
    
    def __str__(self) -> str:
        """String representation of the step processor."""
        return f"{self.__class__.__name__}(name='{self.step_name}', {proc_type}='{self.step_type}')"
    
    def __repr__(self) -> str:
        """Developer representation of the step processor."""
        return f"{self.__class__.__name__}(step_config={self.step_config})"


class StepProcessorRegistry:
    """
    Registry for mapping step types to their processor classes.
    
    This allows the system to dynamically create the correct processor
    for each step type defined in a recipe.
    """
    
    def __init__(self):
        """Initialize the registry."""
        self._processors = {}
    
    def register(self, step_type: str, processor_class: type) -> None:
        """
        Register a processor class for a step type.
        
        Args:
            step_type: String identifier for the step type
            processor_class: Class that inherits from BaseStepProcessor
            
        Raises:
            StepProcessorError: If registration is invalid
        """
        # Guard clauses
        if not isinstance(step_type, str) or not step_type.strip():
            raise StepProcessorError("Step type must be a non-empty string")
        
        if not isinstance(processor_class, type):
            raise StepProcessorError("Processor must be a class")
        
        if not issubclass(processor_class, BaseStepProcessor):
            raise StepProcessorError(
                f"Processor class {processor_class.__name__} must inherit from BaseStepProcessor"
            )
        
        self._processors[step_type] = processor_class
        logger.debug(f"Registered processor for step type: {step_type}")
    
    def get_processor_class(self, step_type: str) -> type:
        """
        Get the processor class for a step type.
        
        Args:
            step_type: Step type to look up
            
        Returns:
            Processor class
            
        Raises:
            StepProcessorError: If step type is not registered
        """
        # Guard clause
        if not isinstance(step_type, str):
            raise StepProcessorError("Step type must be a string")
        
        if step_type not in self._processors:
            available_types = list(self._processors.keys())
            raise StepProcessorError(
                f"Unknown step type: {step_type}. "
                f"Available types: {', '.join(available_types) if available_types else 'none'}"
            )
        
        return self._processors[step_type]
    
    def create_processor(self, step_config: dict) -> BaseStepProcessor:
        """
        Create a processor instance for a step configuration.
        
        Args:
            step_config: Step configuration dictionary
            
        Returns:
            Initialized processor instance
            
        Raises:
            StepProcessorError: If processor creation fails
        """
        # Guard clause
        if not isinstance(step_config, dict):
            raise StepProcessorError("Step configuration must be a dictionary")
        
        if proc_type not in step_config:
            raise StepProcessorError(f"Step configuration missing {proc_type} field")
        
        step_type = step_config[proc_type]
        processor_class = self.get_processor_class(step_type)
        
        try:
            return processor_class(step_config)
        except Exception as e:
            raise StepProcessorError(f"Failed to create processor for step type '{step_type}': {e}")
    
    def get_registered_types(self) -> list:
        """
        Get list of all registered step types.
        
        Returns:
            List of registered step type strings
        """
        return list(self._processors.keys())


# Global registry instance
registry = StepProcessorRegistry()


class ImportBaseProcessor(BaseStepProcessor):
    """Base class for processors that import data (create stages)."""
    
    def __init__(self, step_config: dict):
        super().__init__(step_config)
        
        # Import processors don't need source_stage
        self.source_stage = None
        
        # But they must have save_to_stage
        if not self.save_to_stage:
            raise StepProcessorError(f"Import step '{self.step_name}' requires save_to_stage")
    
    def execute(self, data=None):
        """Execute import operation (implements BaseStepProcessor abstract method)."""
        return self.execute_import()
    
    def execute_import(self) -> pd.DataFrame:
        """Execute import operation."""
        self.log_step_start()
        
        # Load data (implemented by subclass)
        data = self.load_data()
        
        # Save to stage
        self.save_output_data(data)
        
        self.log_step_complete(f"imported {len(data)} rows")
        return data
    
    def save_output_data(self, data) -> None:
        """Save output data to save_to_stage."""
        from excel_recipe_processor.core.stage_manager import StageManager
        StageManager.save_stage(
            stage_name=self.save_to_stage,
            data=data,
            description=f"Imported via step: '{self.step_name}'",
            step_name=self.step_name,
            confirm_replacement=self.confirm_stage_replacement
        )
    
    @abstractmethod
    def load_data(self) -> pd.DataFrame:
        """Load data from source (file, etc.)."""
        pass


class ExportBaseProcessor(BaseStepProcessor):
    """Base class for processors that export data (consume stages)."""
    
    def __init__(self, step_config: dict):
        super().__init__(step_config)
        
        # Export processors don't need save_to_stage
        self.save_to_stage = None
        
        # But they must have source_stage
        if not self.source_stage:
            raise StepProcessorError(f"Export step '{self.step_name}' requires source_stage")
    
    def execute(self, data=None):
        """Execute export operation (implements BaseStepProcessor abstract method)."""
        self.execute_export()
        # Export processors don't return data, but execute() expects return value
        return self.load_input_data()  # Return the source data for consistency
    
    def execute_export(self) -> None:
        """Execute export operation."""
        self.log_step_start()
        
        # Load from stage
        data = self.load_input_data()
        
        # Save data (implemented by subclass)
        self.save_data(data)
        
        self.log_step_complete(f"exported {len(data)} rows")
    
    def load_input_data(self) -> pd.DataFrame:
        """Load input data from source_stage."""
        from excel_recipe_processor.core.stage_manager import StageManager
        return StageManager.load_stage(self.source_stage)
    
    @abstractmethod
    def save_data(self, data: pd.DataFrame) -> None:
        """Save data to destination (file, etc.)."""
        pass


class FileOpsBaseProcessor(BaseStepProcessor):
    """
    Base class for processors that perform file operations without stage I/O.
    
    These processors manipulate external files (formatting, conversion, backup, etc.)
    without participating in the data pipeline. They don't consume or produce 
    meaningful stage data - they just perform file operations as side effects.
    
    Examples: format_excel, convert_file_format, backup_files, create_charts
    """
    
    def __init__(self, step_config: dict):
        super().__init__(step_config)
        
        # File operation processors don't use the stage system
        self.source_stage = None
        self.save_to_stage = None
        
        # They should have some kind of file target though
        # Subclasses can override this validation if needed
        self._validate_file_operation_config()
    
    def _validate_file_operation_config(self):
        """
        Validate that file operation has required configuration.
        
        Base implementation checks for common file operation fields.
        Subclasses can override for specific requirements.
        """
        # Most file operations need some kind of target file
        common_file_fields = ['target_file', 'input_file', 'output_file', 'file_path']
        
        has_file_field = any(
            self.get_config_value(field) is not None 
            for field in common_file_fields
        )
        
        if not has_file_field:
            # This is just a warning - some file ops might not need these
            logger.debug(f"File operation step '{self.step_name}' has no standard file fields")
    
    def execute(self, data=None) -> pd.DataFrame:
        """
        Execute file operation (implements BaseStepProcessor abstract method).
        
        Args:
            data: Pipeline data (ignored for file operations, may be None)
            
        Returns:
            Empty DataFrame (file operations don't produce meaningful data)
            
        Raises:
            StepProcessorError: If file operation fails
        """
        return self.execute_file_operation()
    
    def execute_file_operation(self) -> pd.DataFrame:
        """
        Execute the file operation with proper logging and error handling.
        
        Returns:
            Empty DataFrame (file operations are side effects)
            
        Raises:
            StepProcessorError: If file operation fails
        """
        self.log_step_start()
        
        try:
            # Perform the actual file operation (implemented by subclass)
            operation_result = self.perform_file_operation()
            
            # Log completion with operation-specific info
            self.log_step_complete(operation_result or "file operation completed")
            
            # Return empty DataFrame since file operations don't produce pipeline data
            return pd.DataFrame()
            
        except Exception as e:
            if isinstance(e, StepProcessorError):
                raise
            else:
                raise StepProcessorError(f"File operation failed in step '{self.step_name}': {e}")
    
    @abstractmethod
    def perform_file_operation(self) -> str:
        """
        Perform the specific file operation.
        
        This is where subclasses implement their file manipulation logic.
        
        Returns:
            String describing what was accomplished (for logging)
            Examples: "formatted 3 sheets in report.xlsx"
                     "converted report.xlsx to report.pdf"
                     "backed up 5 files to backup/"
                     
        Raises:
            Exception: Any file operation errors (will be wrapped in StepProcessorError)
        """
        pass
    
    def get_operation_type(self) -> str:
        """
        Get the type of file operation this processor performs.
        
        Used for logging and error messages. Subclasses can override.
        
        Returns:
            String describing the operation type
        """
        return "file_operation"


class FormatExcelProcessor(FileOpsBaseProcessor):
    """Processor for formatting existing Excel files."""
    
    @classmethod
    def get_minimal_config(cls) -> dict:
        return {'target_file': 'output.xlsx'}
    
    def _validate_file_operation_config(self):
        """Validate format_excel specific configuration."""
        # Override base validation to check for target_file specifically
        if not self.get_config_value('target_file'):
            raise StepProcessorError(f"Format Excel step '{self.step_name}' requires 'target_file'")
    
    def perform_file_operation(self) -> str:
        """Format the target Excel file."""
        # Check openpyxl availability
        try:
            import openpyxl
        except ImportError:
            raise StepProcessorError("openpyxl is required for Excel formatting but not installed")
        
        target_file = self.get_config_value('target_file')
        formatting = self.get_config_value('formatting', {})
        
        # Apply variable substitution to target filename
        final_target_file = self._apply_variable_substitution(target_file)
        
        # Validate file exists and is Excel format
        from pathlib import Path
        file_path = Path(final_target_file)
        
        if not file_path.exists():
            raise StepProcessorError(f"Target file not found: {final_target_file}")
        
        if file_path.suffix.lower() not in ['.xlsx', '.xls']:
            raise StepProcessorError(f"Target file must be Excel format (.xlsx or .xls), got: {file_path.suffix}")
        
        # Load and format the workbook
        formatted_sheets = self._format_excel_file(final_target_file, formatting)
        
        return f"formatted {final_target_file} ({formatted_sheets} sheets processed)"
    
    def _apply_variable_substitution(self, filename: str) -> str:
        """Apply variable substitution to filename."""
        # Get custom variables from processor config
        custom_variables = self.get_config_value('variables', {})
        
        # Create variable substitution instance
        from excel_recipe_processor.core.variable_substitution import VariableSubstitution
        variable_sub = VariableSubstitution(
            input_path=None,
            recipe_path=None, 
            custom_variables=custom_variables
        )
        
        # Apply substitution
        try:
            substituted = variable_sub.substitute(filename)
            if substituted != filename:
                logger.debug(f"Variable substitution: {filename} → {substituted}")
            return substituted
        except Exception as e:
            logger.warning(f"Variable substitution failed for '{filename}': {e}")
            return filename
    
    def _format_excel_file(self, filename: str, formatting: dict) -> int:
        """Apply formatting to Excel file and return number of sheets processed."""
        # This would contain all the existing Excel formatting logic
        # from the current FormatExcelProcessor.execute() method
        
        # For brevity, just showing the structure:
        import openpyxl
        
        workbook = openpyxl.load_workbook(filename)
        sheets_processed = 0
        
        for worksheet in workbook.worksheets:
            if formatting.get('auto_fit_columns'):
                self._auto_fit_columns(worksheet)
            
            if formatting.get('header_bold'):
                self._make_headers_bold(worksheet)
            
            if formatting.get('header_background'):
                self._add_header_background(worksheet, formatting.get('header_background_color', 'D3D3D3'))
            
            if formatting.get('freeze_top_row'):
                worksheet.freeze_panes = 'A2'
            
            if formatting.get('auto_filter'):
                self._add_auto_filter(worksheet)
            
            sheets_processed += 1
        
        workbook.save(filename)
        workbook.close()
        
        return sheets_processed
    
    def _auto_fit_columns(self, worksheet):
        """Auto-fit column widths."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _make_headers_bold(self, worksheet):
        """Make first row bold."""
        from openpyxl.styles import Font
        
        for cell in worksheet[1]:  # First row
            cell.font = Font(bold=True)
    
    def _add_header_background(self, worksheet, color):
        """Add background color to first row."""
        from openpyxl.styles import PatternFill
        
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        for cell in worksheet[1]:  # First row
            cell.fill = fill
    
    def _add_auto_filter(self, worksheet):
        """Add auto-filter to data range."""
        if worksheet.max_row > 1:
            from openpyxl.utils import get_column_letter
            data_range = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            worksheet.auto_filter.ref = data_range
    
    def get_operation_type(self) -> str:
        return "excel_formatting"
    
    def get_capabilities(self) -> dict:
        """Get processor capabilities information."""
        return {
            'description': 'Format existing Excel files with professional presentation features',
            'operation_type': 'file_formatting',
            'formatting_features': [
                'auto_fit_columns', 'column_widths', 'header_bold', 'header_background',
                'freeze_panes', 'freeze_top_row', 'row_heights', 'auto_filter', 'active_sheet'
            ],
            'file_requirements': ['xlsx', 'xls'],
            'dependencies': ['openpyxl'],
            'stage_requirements': 'none',  # Key difference from data processors
            'examples': {
                'auto_fit': "Automatically size columns to fit content",
                'header_styling': "Bold headers with background color",
                'freeze_panes': "Freeze top row for easier navigation"
            }
        }

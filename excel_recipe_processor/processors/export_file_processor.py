
"""
Export file step processor for Excel automation recipes.

excel_recipe_processor/processors/export_file_processor.py

Pure stage-based file export - consumes stages, saves to files.
"""

import shutil
import logging

import openpyxl

from pathlib import Path

from excel_recipe_processor.writers.excel_writer import ExcelWriter, DEFAULT_DELETE_BACKUPS_BEYOND

from excel_recipe_processor.core.file_writer import FileWriter, FileWriterError
from excel_recipe_processor.processors._helpers.sheet_addressing import reject_token_for_creation
from excel_recipe_processor.core.base_processor import ExportBaseProcessor, StepProcessorError

logger = logging.getLogger(__name__)


class ExportFileProcessor(ExportBaseProcessor):
    """
    Processor for exporting data from stages to files.
    
    Supports Excel, CSV, and TSV output with variable substitution
    and multi-sheet Excel export capabilities.
    """
    
    @classmethod
    def get_minimal_config(cls):
        return {
            'source_stage': 'final_data',
            'output_file': 'output.xlsx'
        }
    
    # def save_data(self, data):
    #     """Save data to file (implements ExportBaseProcessor abstract method)."""
    #     output_file = self.get_config_value('output_file')
    #     sheet_name = self.get_config_value('sheet_name', 'Data')
    #     explicit_format = self.get_config_value('format', None)
    #     sheets = self.get_config_value('sheets', None)
        
    #     # Apply variable substitution if available
    #     if hasattr(self, 'variable_substitution') and self.variable_substitution:
    #         substituted_path = self.variable_substitution.substitute(output_file)
    #     else:
    #         substituted_path = output_file
        
    #     try:
    #         if sheets:
    #             # Multi-sheet export
    #             sheets_data = self._build_sheets_data(sheets)
    #             FileWriter.write_multi_sheet_excel(sheets_data, substituted_path)
    #         else:
    #             # Single file export
    #             FileWriter.write_file(
    #                 data,
    #                 substituted_path,
    #                 sheet_name=sheet_name,
    #                 explicit_format=explicit_format
    #             )
            
    #         logger.info(f"Exported {len(data)} rows to {substituted_path}")
            
    #     except FileWriterError as e:
    #         raise StepProcessorError(f"Failed to export to '{output_file}': {e}")


    def _export_into_template(self, data, template_file: str, output_file: str,
                              sheet_name: str, create_backup: bool,
                              delete_backups_beyond: int) -> str:
        """
        Copy a template workbook and replace one sheet's contents with the data.

        Args:
            data:           DataFrame to write
            template_file:  Workbook to copy
            output_file:    Where the copy goes
            sheet_name:     Sheet within the copy to replace
            create_backup:  Back up output_file if it already exists

        Returns:
            Description of what was written
        """
        template_path = Path(template_file)
        output_path = Path(output_file)

        if not template_path.exists():
            raise StepProcessorError(f"Template file not found: {template_file}")

        if output_path.exists() and create_backup:
            ExcelWriter().create_backup(output_path, delete_backups_beyond)

        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Byte copy first, so nothing is touched that openpyxl would have to
        # reinterpret. Only the subsequent load/save round-trip can alter the
        # file, and that happens exactly once.
        if template_path.resolve() != output_path.resolve():
            shutil.copy2(template_path, output_path)

        workbook = openpyxl.load_workbook(output_path)

        if sheet_name not in workbook.sheetnames:
            raise StepProcessorError(
                f"Template has no sheet '{sheet_name}'. Available: {workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        sheet_position = workbook.sheetnames.index(sheet_name)
        previous_rows = worksheet.max_row

        # Clear the sheet in place rather than deleting and recreating it, which
        # would move it to the end of the tab order.
        if worksheet.max_row > 0:
            worksheet.delete_rows(1, worksheet.max_row)

        worksheet.append([str(column) for column in data.columns])

        # openpyxl cannot write pandas NA sentinels - pd.NA, NaT and numpy nan
        # all raise "Cannot convert <NA> to Excel". The normal export path goes
        # through pandas, which converts them to blanks on the way out; writing
        # cells directly has to do it here instead.
        cleaned = data.astype(object).where(data.notna(), None)

        for row in cleaned.itertuples(index=False, name=None):
            worksheet.append(list(row))

        workbook.save(output_path)
        workbook.close()

        logger.info(
            f"📋 Copied template '{template_path.name}' and replaced sheet "
            f"'{sheet_name}' (position {sheet_position + 1}): "
            f"{previous_rows} rows -> {len(data) + 1} rows"
        )

        return (f"wrote {len(data)} rows into sheet '{sheet_name}' of a copy of "
                f"'{template_path.name}'")

    def get_capabilities(self) -> dict:
        """
        Get processor capabilities information.

        Returns:
            Dictionary with processor capabilities
        """
        return {
            'description': 'Export stages to Excel or CSV, including multi-sheet workbooks, backing up any file being replaced',
            'file_formats': ['xlsx', 'csv', 'tsv'],
            'excel_options': ['multi-sheet export from named stages', 'sheet naming', 'active sheet selection', 'template-based export'],
            'safety': [
                'timestamped backup of an existing output file, extension preserved',
                'create_backup: false to disable; delete_backups_beyond keeps the newest N and deletes older',
            ],
        }

    def save_data(self, data):
        """Save data to file (implements ExportBaseProcessor abstract method)."""
        output_file = self.get_config_value('output_file')
        sheet_name = self.get_config_value('sheet_name', 'Data')
        try:
            reject_token_for_creation(sheet_name, f"Export step '{self.step_name}'")
        except ValueError as error:
            raise StepProcessorError(str(error))
        explicit_format = self.get_config_value('format', None)
        sheets = self.get_config_value('sheets', None)
        # See if user wants to disable the creation of a backup file to avoid clobbering same name
        create_backup = self.get_config_value('create_backup', True)

        # OPT delete_backups_beyond: how many of the NEWEST timestamped
        # backups to keep, counting the one about to be made; every older
        # one is deleted. Named for what it does to the surplus rather than
        # for a ceiling, because "max allowed" could be misread as refusing
        # to make new backups once the count is reached - the opposite
        # behaviour, and a dangerous one to assume.
        delete_backups_beyond = self.get_config_value(
            'delete_backups_beyond', DEFAULT_DELETE_BACKUPS_BEYOND)
        
        # Apply variable substitution BEFORE calling FileWriter
        if hasattr(self, 'variable_substitution') and self.variable_substitution:
            resolved_file = self.variable_substitution.substitute(output_file)
        else:
            resolved_file = output_file
        
        # Template mode: copy an existing workbook and replace one sheet inside
        # it, instead of building a new workbook from nothing.
        #
        # This is what makes a "copy the last good file and swap the data"
        # workflow possible. Everything the template carries - lookup sheets,
        # named ranges, formatting on other sheets, charts, images - rides along
        # untouched, because only the one named sheet is rewritten.
        template_file = self.get_config_value('template_file', None)

        if template_file:
            if hasattr(self, 'variable_substitution') and self.variable_substitution:
                resolved_template = self.variable_substitution.substitute(template_file)
            else:
                resolved_template = template_file

            return self._export_into_template(
                data, resolved_template, resolved_file, sheet_name, create_backup,
                delete_backups_beyond
            )

        try:

            if sheets:
                # Multi-sheet export
                sheets_data = self._build_sheets_data(sheets)

                logger.info(f"📄 Writing {len(sheets_data)} sheets:")
                for sheet_label, sheet_frame in sheets_data.items():
                    logger.info(
                        f"   • {sheet_label}: {len(sheet_frame):,} rows × "
                        f"{len(sheet_frame.columns)} columns"
                    )
                FileWriter.write_multi_sheet_excel(
                    sheets_data,
                    resolved_file,
                    create_backup=create_backup
                )  # No variables parameter
            else:
                # Single file export
                FileWriter.write_file(
                    data,
                    resolved_file,  # No variables parameter needed
                    sheet_name=sheet_name,
                    explicit_format=explicit_format,
                    create_backup=create_backup
                )
            
            logger.info(f"Exported {len(data)} rows to '{resolved_file}'")
            
        except FileWriterError as e:
            raise StepProcessorError(f"Failed to export to '{output_file}': {e}")



    def _build_sheets_data(self, sheets):
        """Build dictionary of sheet data for multi-sheet export."""
        from excel_recipe_processor.core.stage_manager import StageManager
        
        sheets_data = {}
        
        for sheet_config in sheets:
            sheet_name = sheet_config['sheet_name']
            # Export CREATES tabs; a ?sheet_NNN? pseudo-name addresses tabs
            # that already exist, so it is meaningless here - and its
            # characters are illegal in a real title anyway. Fail loud.
            try:
                reject_token_for_creation(
                    sheet_name, f"Export step '{self.step_name}'"
                )
            except ValueError as error:
                raise StepProcessorError(str(error))
            data_source = sheet_config.get('data_source')
            
            if not data_source:
                raise StepProcessorError(f"Sheet '{sheet_name}' missing data_source")
            
            try:
                sheet_data = StageManager.load_stage(data_source)
                sheets_data[sheet_name] = sheet_data
            except Exception as e:
                available_stages = list(StageManager.list_stages().keys())
                raise StepProcessorError(
                    f"Cannot load data_source '{data_source}' for sheet '{sheet_name}': {e}. "
                    f"Available stages: {available_stages}"
                )
        
        return sheets_data

# End of file #

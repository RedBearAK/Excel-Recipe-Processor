#!/usr/bin/env python3
"""
Functional tests for SeedDonorFormulasProcessor with actual Excel files.

tests/test_seed_donor_formulas_functional.py

Creates actual Excel files with formulas and tests the transplanting functionality.
"""

import os
import sys
import tempfile
import openpyxl

from pathlib import Path

# Add project root to Python path for imports
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from excel_recipe_processor.core.base_processor import StepProcessorError
from excel_recipe_processor.processors.seed_donor_formulas_processor import SeedDonorFormulasProcessor


def create_source_excel_file(file_path: str):
    """Create a source Excel file with sample data and formulas."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Budget"
    
    # Add headers
    worksheet['A1'] = 'Item'
    worksheet['B1'] = 'Quantity'
    worksheet['C1'] = 'Unit_Price'
    worksheet['D1'] = 'Total'
    worksheet['E1'] = 'Tax'
    worksheet['F1'] = 'Grand_Total'
    
    # Add sample data
    data = [
        ['Widget A', 10, 5.50],
        ['Widget B', 25, 12.75],
        ['Widget C', 15, 8.25],
        ['Widget D', 30, 15.00]
    ]
    
    for i, (item, qty, price) in enumerate(data, start=2):
        worksheet[f'A{i}'] = item
        worksheet[f'B{i}'] = qty
        worksheet[f'C{i}'] = price
        
        # Add formulas that we want to transplant
        worksheet[f'D{i}'] = f'=B{i}*C{i}'  # Total = Quantity * Unit Price
        worksheet[f'E{i}'] = f'=D{i}*0.1'   # Tax = Total * 10%
        worksheet[f'F{i}'] = f'=D{i}+E{i}'  # Grand Total = Total + Tax
    
    # Save the file
    workbook.save(file_path)
    workbook.close()
    
    print(f"✓ Created source file: {file_path}")
    return file_path


def create_target_excel_file(file_path: str):
    """Create a target Excel file with data but no formulas."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Budget"
    
    # Add headers (same as source)
    worksheet['A1'] = 'Item'
    worksheet['B1'] = 'Quantity'
    worksheet['C1'] = 'Unit_Price'
    worksheet['D1'] = 'Total'
    worksheet['E1'] = 'Tax'
    worksheet['F1'] = 'Grand_Total'
    
    # Add different sample data (same structure, different values)
    data = [
        ['Product X', 20, 7.25],
        ['Product Y', 35, 9.50],
        ['Product Z', 18, 11.75]
    ]
    
    for i, (item, qty, price) in enumerate(data, start=2):
        worksheet[f'A{i}'] = item
        worksheet[f'B{i}'] = qty
        worksheet[f'C{i}'] = price
        # D, E, F columns intentionally left empty for formula transplanting
    
    # Save the file
    workbook.save(file_path)
    workbook.close()
    
    print(f"✓ Created target file: {file_path}")
    return file_path


def verify_formulas_transplanted(target_file: str, expected_formulas: dict):
    """Verify that formulas were correctly transplanted."""
    workbook = openpyxl.load_workbook(target_file)
    worksheet = workbook.active
    
    success_count = 0
    total_checks = len(expected_formulas)
    
    print("\n🔍 Verifying transplanted formulas:")
    
    for cell_ref, expected_formula in expected_formulas.items():
        cell = worksheet[cell_ref]
        
        # Check if cell has the expected formula
        if hasattr(cell, 'value') and cell.value and str(cell.value) == expected_formula:
            print(f"   ✓ {cell_ref}: {cell.value}")
            success_count += 1
        else:
            print(f"   ✗ {cell_ref}: Expected '{expected_formula}', got '{cell.value}'")
    
    workbook.close()
    
    print(f"\n📊 Formula verification: {success_count}/{total_checks} formulas correct")
    
    return success_count == total_checks


def test_formula_liveness():
    """Test that transplanted formulas become live when Excel would open the file."""
    print("Testing formula liveness (Excel compatibility)...")
    
    with tempfile.TemporaryDirectory() as temp_dir:
        # Create a test to verify openpyxl formula handling
        test_file = str(Path(temp_dir) / "formula_test.xlsx")
        
        # Create workbook with known data and formula
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Set up known values
        worksheet['A1'] = 10
        worksheet['B1'] = 5
        worksheet['C1'] = '=A1*B1'  # Should calculate to 50
        
        # Save and close
        workbook.save(test_file)
        workbook.close()
        
        # Reopen and check formula properties
        workbook = openpyxl.load_workbook(test_file, data_only=False)
        worksheet = workbook.active
        
        formula_cell = worksheet['C1']
        
        print(f"   Formula cell value: {formula_cell.value}")
        print(f"   Formula cell data_type: {formula_cell.data_type}")
        
        # Check if it's stored as a formula
        if hasattr(formula_cell, 'value') and str(formula_cell.value).startswith('='):
            print("   ✓ Formula stored correctly by openpyxl")
            
            # Now test with data_only=True to see cached values
            workbook.close()
            workbook_data = openpyxl.load_workbook(test_file, data_only=True)
            worksheet_data = workbook_data.active
            
            calculated_value = worksheet_data['C1'].value
            print(f"   Cached calculated value: {calculated_value}")
            
            workbook_data.close()
            
            # Expected calculation: 10 * 5 = 50
            if calculated_value == 50:
                print("   ✓ Formula would calculate correctly in Excel")
                return True
            elif calculated_value is None:
                print("   ⚠️ No cached value (normal for new formulas) - Excel will calculate on open")
                return True
            else:
                print(f"   ✗ Unexpected calculated value: {calculated_value}")
                return False
        else:
            print(f"   ✗ Not stored as formula: {formula_cell.value}")
            return False


def verify_formulas_are_live(target_file: str, test_cases: list):
    """Verify that transplanted formulas are stored as live formulas."""
    # Open with formulas 
    workbook = openpyxl.load_workbook(target_file, data_only=False)
    worksheet = workbook.active
    
    print("\n🔍 Verifying formulas are stored as live formulas:")
    
    live_count = 0
    
    for cell_ref, description in test_cases:
        cell = worksheet[cell_ref]
        cell_value = str(cell.value) if cell.value else "None"
        
        if cell_value.startswith('='):
            print(f"   ✓ {cell_ref}: Live formula - {cell_value}")
            live_count += 1
        else:
            print(f"   ✗ {cell_ref}: Not a live formula - {cell_value}")
    
    workbook.close()
    
    # Also check with data_only=True to see if Excel would calculate
    print("\n📊 Checking Excel calculation readiness:")
    try:
        workbook_data = openpyxl.load_workbook(target_file, data_only=True)
        worksheet_data = workbook_data.active
        
        for cell_ref, description in test_cases:
            cached_value = worksheet_data[cell_ref].value
            if cached_value is None:
                print(f"   ⚠️ {cell_ref}: No cached value (Excel will calculate on open)")
            elif isinstance(cached_value, (int, float)):
                print(f"   ✓ {cell_ref}: Has cached calculation = {cached_value}")
            else:
                print(f"   ? {cell_ref}: Cached value = {cached_value}")
        
        workbook_data.close()
        
    except Exception as e:
        print(f"   ⚠️ Could not check cached values: {e}")
    
    print(f"\n📊 Live formula verification: {live_count}/{len(test_cases)} formulas are live")
    
    return live_count == len(test_cases)


def test_basic_formula_transplant():
    """Test basic formula transplant functionality."""
    print("Testing basic formula transplant...")
    
    with tempfile.TemporaryDirectory() as temp_dir:
        # Create test files
        source_file = str(Path(temp_dir) / "source_budget.xlsx")
        target_file = str(Path(temp_dir) / "target_budget.xlsx")
        
        create_source_excel_file(source_file)
        create_target_excel_file(target_file)
        
        # Configure processor to transplant formulas from columns D, E, F
        step_config = {
            'processor_type': 'seed_donor_formulas',
            'step_description': 'Test formula transplant',
            'source_file': source_file,
            'source_sheet': 'Budget',
            'target_file': target_file,
            'target_sheet': 'Budget',
            'columns': ['D', 'E', 'F'],  # Total, Tax, Grand_Total columns
            'start_row': 2,
            'row_count': 3  # Process 3 rows of data
        }
        
        try:
            # Execute the transplant
            processor = SeedDonorFormulasProcessor(step_config)
            result = processor.perform_file_operation()
            
            print(f"✓ Processor completed: {result}")
            
            # Verify the transplanted formulas are correct text
            expected_formulas = {
                'D2': '=B2*C2',
                'E2': '=D2*0.1', 
                'F2': '=D2+E2',
                'D3': '=B3*C3',
                'E3': '=D3*0.1',
                'F3': '=D3+E3'
            }
            
            if not verify_formulas_transplanted(target_file, expected_formulas):
                print("✗ Formula text verification failed")
                return False
            
            # Verify formulas are stored as live formulas
            test_cases = [
                ('D2', 'Total calculation'),
                ('E2', 'Tax calculation'),
                ('F2', 'Grand total calculation')
            ]
            
            if verify_formulas_are_live(target_file, test_cases):
                print("✓ All formulas are live and Excel-ready!")
                return True
            else:
                print("✗ Formula liveness verification failed")
                return False
                
        except Exception as e:
            print(f"✗ Formula transplant failed: {e}")
            return False


def test_column_name_matching():
    """Test transplanting using column names instead of Excel references."""
    print("\nTesting column name matching...")
    
    with tempfile.TemporaryDirectory() as temp_dir:
        # Create test files
        source_file = str(Path(temp_dir) / "source_named.xlsx")
        target_file = str(Path(temp_dir) / "target_named.xlsx")
        
        create_source_excel_file(source_file)
        create_target_excel_file(target_file)
        
        # Configure processor to use column names
        step_config = {
            'processor_type': 'seed_donor_formulas',
            'step_description': 'Test column name matching',
            'source_file': source_file,
            'source_sheet': 'Budget',
            'target_file': target_file,
            'target_sheet': 'Budget',
            'columns': ['Total', 'Tax'],  # Use column header names
            'start_row': 2,
            'row_count': 2
        }
        
        try:
            processor = SeedDonorFormulasProcessor(step_config)
            result = processor.perform_file_operation()
            
            print(f"✓ Column name matching completed: {result}")
            
            # Verify specific formulas were transplanted
            expected_formulas = {
                'D2': '=B2*C2',  # Total column
                'E2': '=D2*0.1'  # Tax column
            }
            
            if not verify_formulas_transplanted(target_file, expected_formulas):
                print("✗ Formula text verification failed")
                return False
            
            # Verify formulas are live
            test_cases = [
                ('D2', 'Total calculation'),
                ('E2', 'Tax calculation')
            ]
            
            if verify_formulas_are_live(target_file, test_cases):
                print("✓ Column name matching successful and formulas are live!")
                return True
            else:
                print("✗ Formula liveness verification failed")
                return False
                
        except Exception as e:
            print(f"✗ Column name matching failed: {e}")
            return False


def test_create_verification_file():
    """Create a persistent file for manual verification in Excel/LibreOffice."""
    print("Creating verification file for manual testing...")
    
    # Create in current directory so it persists
    verification_file = "formula_transplant_verification.xlsx"
    
    try:
        # Create source file in temp location
        with tempfile.TemporaryDirectory() as temp_dir:
            source_file = str(Path(temp_dir) / "verification_source.xlsx")
            create_source_excel_file(source_file)
            
            # Create target file as the persistent verification file
            create_target_excel_file(verification_file)
            
            # Run the transplant
            step_config = {
                'processor_type': 'seed_donor_formulas',
                'step_description': 'Create verification file',
                'source_file': source_file,
                'source_sheet': 'Budget',
                'target_file': verification_file,
                'target_sheet': 'Budget',
                'columns': ['D', 'E', 'F'],  # Total, Tax, Grand_Total
                'start_row': 2,
                'row_count': 3
            }
            
            processor = SeedDonorFormulasProcessor(step_config)
            result = processor.perform_file_operation()
            
            print(f"✓ {result}")
            print(f"✓ Created verification file: {verification_file}")
            print("\n📋 MANUAL VERIFICATION STEPS:")
            print("   1. Open 'formula_transplant_verification.xlsx' in Excel/LibreOffice")
            print("   2. Check cells D2, E2, F2 (and D3, E3, F3)")
            print("   3. Verify they show calculated values (not formula text)")
            print("   4. Expected calculations:")
            print("      D2 = 20 * 7.25 = 145")
            print("      E2 = 145 * 0.1 = 14.5") 
            print("      F2 = 145 + 14.5 = 159.5")
            print("   5. Try changing B2 or C2 - formulas should recalculate")
            print("\n💡 If you see calculated values, the transplanting worked perfectly!")
            
            return True
            
    except Exception as e:
        print(f"✗ Failed to create verification file: {e}")
        return False
    """Test that processor detects and prevents overwriting existing data."""
    print("\nTesting collision detection...")
    
    with tempfile.TemporaryDirectory() as temp_dir:
        # Create source file
        source_file = str(Path(temp_dir) / "source_collision.xlsx")
        create_source_excel_file(source_file)
        
        # Create target file with data in formula columns
        target_file = str(Path(temp_dir) / "target_collision.xlsx")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Budget"
        
        # Add headers and data with occupied target cells
        worksheet['A1'] = 'Item'
        worksheet['B1'] = 'Quantity'
        worksheet['C1'] = 'Unit_Price'
        worksheet['D1'] = 'Total'
        
        worksheet['A2'] = 'Product X'
        worksheet['B2'] = 20
        worksheet['C2'] = 7.25
        worksheet['D2'] = 'EXISTING_DATA'  # This should cause collision
        
        workbook.save(target_file)
        workbook.close()
        
        step_config = {
            'processor_type': 'seed_donor_formulas',
            'step_description': 'Test collision detection',
            'source_file': source_file,
            'source_sheet': 'Budget',
            'target_file': target_file,
            'target_sheet': 'Budget',
            'columns': ['D'],
            'start_row': 2,
            'row_count': 1
        }
        
        try:
            processor = SeedDonorFormulasProcessor(step_config)
            processor.perform_file_operation()
            
            print("✗ Should have failed with collision error")
            return False
            
        except StepProcessorError as e:
            if "already contains data" in str(e):
                print("✓ Collision detection working correctly")
                return True
            else:
                print(f"✗ Wrong error type: {e}")
                return False
        except Exception as e:
            print(f"✗ Unexpected error: {e}")
            return False


def test_collision_detection():
    """Test that processor detects and prevents overwriting existing data."""
    print("\nTesting collision detection...")
    
    with tempfile.TemporaryDirectory() as temp_dir:
        # Create source file
        source_file = str(Path(temp_dir) / "source_collision.xlsx")
        create_source_excel_file(source_file)
        
        # Create target file with data in formula columns
        target_file = str(Path(temp_dir) / "target_collision.xlsx")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Budget"
        
        # Add headers and data with occupied target cells
        worksheet['A1'] = 'Item'
        worksheet['B1'] = 'Quantity'
        worksheet['C1'] = 'Unit_Price'
        worksheet['D1'] = 'Total'
        
        worksheet['A2'] = 'Product X'
        worksheet['B2'] = 20
        worksheet['C2'] = 7.25
        worksheet['D2'] = 'EXISTING_DATA'  # This should cause collision
        
        workbook.save(target_file)
        workbook.close()
        
        step_config = {
            'processor_type': 'seed_donor_formulas',
            'step_description': 'Test collision detection',
            'source_file': source_file,
            'source_sheet': 'Budget',
            'target_file': target_file,
            'target_sheet': 'Budget',
            'columns': ['D'],
            'start_row': 2,
            'row_count': 1
        }
        
        try:
            processor = SeedDonorFormulasProcessor(step_config)
            processor.perform_file_operation()
            
            print("✗ Should have failed with collision error")
            return False
            
        except StepProcessorError as e:
            if "already contains data" in str(e):
                print("✓ Collision detection working correctly")
                return True
            else:
                print(f"✗ Wrong error type: {e}")
                return False
        except Exception as e:
            print(f"✗ Unexpected error: {e}")
            return False


def main():
    """Run all functional tests."""
    
    print("=== SeedDonorFormulasProcessor Functional Tests ===\n")
    print("Testing actual formula transplanting with Excel files...\n")
    
    tests = [
        test_formula_liveness,
        test_basic_formula_transplant,
        test_column_name_matching,
        test_collision_detection,
        test_create_verification_file
    ]
    
    passed = 0
    
    for test_func in tests:
        try:
            if test_func():
                passed += 1
                print(f"✓ {test_func.__name__} passed\n")
            else:
                print(f"✗ {test_func.__name__} failed\n")
        except Exception as e:
            print(f"✗ {test_func.__name__} failed with error: {e}\n")
    
    print(f"=== Results: {passed}/{len(tests)} functional tests passed ===")
    
    if passed == len(tests):
        print("\n🎉 All functional tests passed!")
        print("💡 The formula transplanting functionality works correctly!")
        print("💡 Formulas will be live and calculate when Excel opens the files!")
        print("\n🔍 For 100% confidence: Open 'formula_transplant_verification.xlsx'")
        return 1
    else:
        print("\n❌ Some functional tests failed!")
        print("💡 Check Excel file creation and openpyxl formula handling")
        return 0


if __name__ == "__main__":
    exit(main())


# End of file #

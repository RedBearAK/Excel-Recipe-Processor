"""
Column-addressed formatting for format_excel.

excel_recipe_processor/processors/_helpers/format_excel_column_formats.py

Number formats, font colours, column hiding and per-column alignment all need
the same thing:
turn a column NAME into a column letter, then act on that whole column below the
header. Resolution is delegated to excel_range_resolver, which already does this
for named-range creation, so a column name means the same thing everywhere.

Excel number format codes are cryptic and easy to get subtly wrong, so the
common ones are available under readable aliases. Any literal Excel format code
is still accepted.
"""

import logging

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string

from excel_recipe_processor.processors._helpers.range_patterns import cell_ref_rgx, range_ref_rgx
from excel_recipe_processor.processors._helpers.excel_range_resolver import (
    resolve_column_letters, find_last_data_row, ExcelRangeResolverError
)
from excel_recipe_processor.processors._helpers.format_excel_hyperlink_utils import (
    build_hyperlink_target, HyperlinkTargetError,
    HYPERLINK_KINDS, DEFAULT_HYPERLINK_COLOR
)


logger = logging.getLogger(__name__)


# Readable names for the format codes this project actually uses. Anything not
# listed here is passed to Excel verbatim, so custom codes still work.
NUMBER_FORMAT_ALIASES = {
    # Whole numbers with thousands separators - Cases, Packages, Units, weights
    'thousands': '#,##0',
    'thousands_2dp': '#,##0.00',

    # Accounting: currency symbol pinned left, negatives in parentheses, zero as
    # a dash. This is what Excel's own "Accounting" button produces.
    'accounting': '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)',
    'accounting_0dp': '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)',

    'currency': '$#,##0.00',
    'currency_0dp': '$#,##0',

    'percent': '0%',
    'percent_2dp': '0.00%',

    'date': 'mm/dd/yyyy',
    'date_iso': 'yyyy-mm-dd',
    'datetime': 'mm/dd/yyyy hh:mm',

    'integer': '0',
    'text': '@',
}

VALID_HORIZONTAL = (
    'left', 'center', 'right', 'justify', 'distributed', 'centerContinuous',
    'fill', 'general'
)

VALID_VERTICAL = ('top', 'center', 'bottom', 'justify', 'distributed')

VALID_UNDERLINE = ('single', 'double')


def _normalize_underline(value, context: str) -> str:
    """
    Turn a font_underline spec into an openpyxl underline value.

    true means 'single' (the underline everyone means); 'single' and
    'double' pass through. Anything else raises with the legal values
    named, per the guided-error doctrine.
    """
    if value is True:
        return 'single'
    if value in VALID_UNDERLINE:
        return value
    raise ColumnFormatError(
        f"{context}: 'font_underline' must be true, 'single', or "
        f"'double', got {value!r}"
    )


def _default_color_normalizer(color) -> str:
    """Fall back to plain hex handling when the processor supplies nothing."""
    text = str(color).strip().lstrip('#').upper()

    if len(text) == 3:
        text = ''.join(char * 2 for char in text)

    if len(text) != 6:
        raise ColumnFormatError(
            f"Cannot interpret colour '{color}'. Use 6-digit hex, or run this "
            f"through format_excel so CSS names are available."
        )

    return text


class ColumnFormatError(Exception):
    """Raised when a column formatting rule cannot be applied."""
    pass


def resolve_number_format(spec: str) -> str:
    """
    Turn a format alias into an Excel format code.

    Args:
        spec: An alias from NUMBER_FORMAT_ALIASES, or a literal Excel code

    Returns:
        Excel number format code
    """
    if not isinstance(spec, str) or not spec.strip():
        raise ColumnFormatError("number_format must be a non-empty string")

    key = spec.strip()

    if key in NUMBER_FORMAT_ALIASES:
        return NUMBER_FORMAT_ALIASES[key]

    # Not an alias, so treat it as a literal Excel format code
    return key


def apply_column_formats(worksheet, rules: list, header_row: int = 1,
                         on_missing: str = 'warn', color_normalizer=None) -> list:
    """
    Apply number formats, fonts and alignment to whole columns, by name.

    A rule may style the data rows, that column's header cell, or both. Header
    styling here is per column, unlike the sheet-wide header_* options, which
    is what makes it possible to mark a subset of columns - the ten this recipe
    inserts, for instance - without touching the rest.

    A rule may instead set whole_column: true, which applies number_format,
    font and alignment at the COLUMN-DIMENSION level (a col-level style in
    the file) rather than per cell. That is the correct mechanism for
    columns whose values arrive at Excel calculation time - dynamic-array
    spills - because those cells do not exist in the file when formats are
    written, so per-cell formatting up to the current data extent cannot
    reach them, while a column style is inherited by every cell Excel
    creates. Header-cell styling stays per-cell either way (an explicit
    cell style overrides the column style, keeping the header clean).
    NOTE: openpyxl serializes a col-level style with a width attribute; a
    whole_column rule without an explicit 'width' therefore leaves the
    column at openpyxl's default 13 - pair whole_column with width.

    Args:
        worksheet:          openpyxl worksheet object
        rules:              List of rule dictionaries
        header_row:         Row holding the headers
        on_missing:         'error', 'warn', or 'skip' for unresolvable columns
        color_normalizer:   Callable turning a colour spec into 6-digit hex.
                            Supplied by the processor so CSS names such as
                            "red" work the same way they do elsewhere.

    Returns:
        List of human-readable descriptions of what was applied
    """
    if color_normalizer is None:
        color_normalizer = _default_color_normalizer
    if not isinstance(rules, list):
        raise ColumnFormatError("column_formats must be a list of rules")

    applied = []

    # One extent for the whole sheet, measured from the widest column. Measuring
    # per column would truncate a rule applied to a sparsely populated column.
    try:
        anchor_letters = [get_column_letter(i) for i in range(1, worksheet.max_column + 1)]
        last_row = find_last_data_row(worksheet, anchor_letters, header_row)
    except ExcelRangeResolverError as error:
        raise ColumnFormatError(f"Could not determine data extent: {error}")

    sheet_has_data = last_row > header_row
    if not sheet_has_data and not any(
            isinstance(rule, dict) and rule.get('whole_column')
            for rule in rules):
        logger.warning(f"[{worksheet.title}] No data rows below header, skipping column formats")
        return applied

    for index, rule in enumerate(rules):
        if not isinstance(rule, dict):
            raise ColumnFormatError(f"column_formats rule {index + 1} must be a dictionary")

        columns = rule.get('columns')

        if not isinstance(columns, list) or len(columns) == 0:
            raise ColumnFormatError(
                f"column_formats rule {index + 1} requires a non-empty 'columns' list"
            )

        number_format = rule.get('number_format')
        horizontal = rule.get('alignment_horizontal')
        vertical = rule.get('alignment_vertical')
        wrap_text = rule.get('wrap_text')
        font_color = rule.get('font_color')
        font_bold = rule.get('font_bold')
        font_italic = rule.get('font_italic')
        font_size = rule.get('font_size')
        font_underline = rule.get('font_underline')
        font_strikethrough = rule.get('font_strikethrough')
        background_color = rule.get('background_color')
        border_style = rule.get('border_style')
        border_color = rule.get('border_color')
        make_hyperlinks = rule.get('make_hyperlinks')
        hyperlink_color = rule.get('hyperlink_color')
        header_font_color = rule.get('header_font_color')
        header_background_color = rule.get('header_background_color')
        header_bold = rule.get('header_bold')
        width = rule.get('width')
        whole_column = rule.get('whole_column', False)

        if not isinstance(whole_column, bool):
            raise ColumnFormatError(
                f"column_formats rule {index + 1}: 'whole_column' must be "
                f"true or false, got {whole_column!r}"
            )

        if not whole_column and not sheet_has_data:
            logger.warning(
                f"[{worksheet.title}] column_formats rule {index + 1} "
                f"skipped: no data rows below the header (whole_column "
                f"rules still apply on empty sheets)"
            )
            continue

        actionable = (number_format, horizontal, vertical, wrap_text, font_color,
                      background_color, border_style, font_bold, font_italic,
                      font_size, font_underline, font_strikethrough,
                      make_hyperlinks, header_font_color,
                      header_background_color, header_bold, width)

        if all(value is None for value in actionable):
            raise ColumnFormatError(
                f"column_formats rule {index + 1} does nothing: supply number_format, "
                f"alignment_horizontal, alignment_vertical, wrap_text, font_color, "
                f"background_color, font_bold, font_italic, font_underline, "
                f"font_strikethrough, make_hyperlinks, header_font_color, "
                f"header_background_color, "
                f"or header_bold"
            )

        if horizontal is not None and horizontal not in VALID_HORIZONTAL:
            valid = ', '.join(VALID_HORIZONTAL)
            raise ColumnFormatError(
                f"column_formats rule {index + 1} invalid alignment_horizontal "
                f"'{horizontal}'. Valid: {valid}"
            )

        if vertical is not None and vertical not in VALID_VERTICAL:
            valid = ', '.join(VALID_VERTICAL)
            raise ColumnFormatError(
                f"column_formats rule {index + 1} invalid alignment_vertical "
                f"'{vertical}'. Valid: {valid}"
            )

        try:
            letters = resolve_column_letters(
                worksheet, columns, header_row,
                force_column_names=rule.get('force_column_names', False),
                on_missing=on_missing
            )
        except ExcelRangeResolverError as error:
            if on_missing == 'error':
                raise ColumnFormatError(f"column_formats rule {index + 1}: {error}")
            logger.warning(f"column_formats rule {index + 1}: {error}")
            continue

        format_code = resolve_number_format(number_format) if number_format else None
        data_color = color_normalizer(font_color) if font_color is not None else None
        head_color = color_normalizer(header_font_color) if header_font_color is not None else None
        head_fill = (color_normalizer(header_background_color)
                     if header_background_color is not None else None)
        # Data-cell fill (2026-08-23): tints the DATA rows of the named
        # columns, leaving the header to header_background_color. Built for
        # marking hand-maintained columns as more tenuous than lookups.
        data_fill = (color_normalizer(background_color)
                     if background_color is not None else None)

        # Data-cell borders (2026-08-23). A cell FILL suppresses Excel's
        # gridlines wherever it lands - not a bug of ours, just how Excel
        # paints - so a tinted column looks like a solid slab. A thin
        # border in gridline gray (D9D9D9) restores the ruled look. Any
        # openpyxl border style name works; border_color defaults to the
        # gridline gray when only border_style is given.
        data_border = None
        if border_style is not None:
            side_color = color_normalizer(border_color) if border_color else 'D9D9D9'
            edge = Side(style=border_style, color=side_color)
            data_border = Border(left=edge, right=edge, top=edge, bottom=edge)
        elif border_color is not None:
            raise ColumnFormatError(
                f"column_formats rule {index + 1}: 'border_color' requires "
                f"'border_style' (e.g. thin, hair, medium)."
            )

        if font_size is not None and (
                not isinstance(font_size, (int, float)) or font_size <= 0):
            raise ColumnFormatError(
                f"column_formats rule {index + 1}: 'font_size' must be a "
                f"positive number, got {font_size!r}"
            )

        underline_value = None
        if font_underline is not None:
            underline_value = _normalize_underline(
                font_underline, f"column_formats rule {index + 1}")

        if font_strikethrough is not None and not isinstance(font_strikethrough, bool):
            raise ColumnFormatError(
                f"column_formats rule {index + 1}: 'font_strikethrough' "
                f"must be true or false, got {font_strikethrough!r}"
            )

        # Hyperlinks (2026-08-23): the rule DECLARES what the bare cell
        # text is - a formatter sniffing content to guess between a path
        # and a URL would be implicit behavior. Real cell.hyperlink
        # relationships, not HYPERLINK() formulas: no recalc, no caches,
        # no dynamic-array machinery for what is static metadata.
        link_color = None
        if make_hyperlinks is not None:
            if make_hyperlinks not in HYPERLINK_KINDS:
                legal = ', '.join(HYPERLINK_KINDS)
                raise ColumnFormatError(
                    f"column_formats rule {index + 1}: 'make_hyperlinks' "
                    f"must be one of: {legal}. Got: {make_hyperlinks!r}"
                )
            if whole_column:
                raise ColumnFormatError(
                    f"column_formats rule {index + 1}: 'make_hyperlinks' "
                    f"writes per-cell link relationships and does not "
                    f"combine with whole_column. Remove one side."
                )
            link_color = (color_normalizer(hyperlink_color)
                          if hyperlink_color is not None
                          else DEFAULT_HYPERLINK_COLOR)
        elif hyperlink_color is not None:
            raise ColumnFormatError(
                f"column_formats rule {index + 1}: 'hyperlink_color' "
                f"requires 'make_hyperlinks' (file_paths, web_urls, or "
                f"email_addresses)."
            )

        touches_data_font = (data_color is not None or font_bold is not None
                             or font_italic is not None or font_size is not None
                             or underline_value is not None
                             or font_strikethrough is not None)
        touches_alignment = horizontal is not None or vertical is not None or wrap_text is not None
        touches_header = head_color is not None or head_fill is not None or header_bold is not None

        link_count = 0

        for letter in letters:
            col_index = column_index_from_string(letter)

            if whole_column:
                if data_fill is not None or data_border is not None:
                    raise ColumnFormatError(
                        f"column_formats rule {index + 1}: 'background_color' "
                        f"and 'border_style' are per-cell and do not combine "
                        f"with whole_column (a column-dimension fill or border "
                        f"would paint a million empty rows). Remove one side."
                    )
                dimension = worksheet.column_dimensions[letter]
                if format_code:
                    dimension.number_format = format_code
                if touches_data_font:
                    dimension.font = Font(
                        bold=font_bold, italic=font_italic,
                        size=font_size, color=data_color,
                        underline=underline_value, strike=font_strikethrough
                    )
                if touches_alignment:
                    dimension.alignment = Alignment(
                        horizontal=horizontal, vertical=vertical,
                        wrap_text=wrap_text
                    )
                if width is None:
                    logger.info(
                        f"[{worksheet.title}] whole_column style on {letter} "
                        f"carries openpyxl's default width 13; add 'width:' "
                        f"to the rule to control it"
                    )

            if touches_header:
                header_cell = worksheet.cell(row=header_row, column=col_index)
                existing_font = header_cell.font
                header_cell.font = Font(
                    name=existing_font.name,
                    size=existing_font.size,
                    bold=header_bold if header_bold is not None else existing_font.bold,
                    italic=existing_font.italic,
                    # Color OBJECT passthrough: .rgb on a theme-based
                    # default is an RGB descriptor Font() rejects
                    color=head_color if head_color is not None else existing_font.color
                )
                if head_fill is not None:
                    header_cell.fill = PatternFill(
                        start_color=head_fill, end_color=head_fill, fill_type='solid'
                    )

            if whole_column:
                continue

            for row_num in range(header_row + 1, last_row + 1):
                cell = worksheet.cell(row=row_num, column=col_index)

                if format_code:
                    cell.number_format = format_code

                if touches_data_font:
                    existing = cell.font
                    cell.font = Font(
                        name=existing.name,
                        size=font_size if font_size is not None else existing.size,
                        bold=font_bold if font_bold is not None else existing.bold,
                        italic=font_italic if font_italic is not None else existing.italic,
                        underline=underline_value if underline_value is not None else existing.underline,
                        strike=font_strikethrough if font_strikethrough is not None else existing.strike,
                        color=data_color if data_color is not None else existing.color
                    )

                if touches_alignment:
                    existing = cell.alignment
                    cell.alignment = Alignment(
                        horizontal=horizontal if horizontal is not None else existing.horizontal,
                        vertical=vertical if vertical is not None else existing.vertical,
                        wrap_text=wrap_text if wrap_text is not None else existing.wrap_text
                    )

                if data_fill is not None:
                    cell.fill = PatternFill(
                        start_color=data_fill, end_color=data_fill, fill_type='solid'
                    )

                if data_border is not None:
                    cell.border = data_border

                if make_hyperlinks is not None:
                    raw_value = cell.value
                    if raw_value is None or (isinstance(raw_value, str)
                                             and not raw_value.strip()):
                        # Sparse link columns are a designed shape - same
                        # reasoning as the low-match warning suppression
                        continue
                    if not isinstance(raw_value, str):
                        raise ColumnFormatError(
                            f"column_formats rule {index + 1}, cell "
                            f"{cell.coordinate}: make_hyperlinks needs text "
                            f"cells, got {type(raw_value).__name__} "
                            f"{raw_value!r}"
                        )
                    try:
                        target = build_hyperlink_target(make_hyperlinks, raw_value)
                    except HyperlinkTargetError as error:
                        raise ColumnFormatError(
                            f"column_formats rule {index + 1}, cell "
                            f"{cell.coordinate}: {error}"
                        )
                    cell.hyperlink = target
                    # Excel's own link presentation, on top of whatever
                    # font the cell already carries
                    existing = cell.font
                    cell.font = Font(
                        name=existing.name,
                        size=existing.size,
                        bold=existing.bold,
                        italic=existing.italic,
                        underline='single',
                        strike=existing.strike,
                        color=link_color
                    )
                    link_count += 1

        parts = []
        if format_code:
            label = number_format if number_format in NUMBER_FORMAT_ALIASES else 'custom'
            parts.append(f"number format {label}")
        if horizontal is not None:
            parts.append(f"h-align {horizontal}")
        if vertical is not None:
            parts.append(f"v-align {vertical}")
        if wrap_text is not None:
            parts.append(f"wrap {wrap_text}")
        if font_color is not None:
            parts.append(f"font {font_color}")
        if font_bold is not None:
            parts.append(f"bold {font_bold}")
        if font_size is not None:
            parts.append(f"size {font_size}")
        if font_underline is not None:
            parts.append(f"underline {underline_value}")
        if font_strikethrough is not None:
            parts.append(f"strikethrough {font_strikethrough}")
        if make_hyperlinks is not None:
            parts.append(f"hyperlinks {make_hyperlinks} ({link_count} cells)")
        if header_font_color is not None or header_background_color is not None:
            parts.append("header styling")
        if width is not None:
            parts.append(f"width {width}")

        mechanism = ' (whole column)' if whole_column else ''
        description = f"{', '.join(parts)}{mechanism} on {len(letters)} column(s): {', '.join(columns[:4])}"
        if len(columns) > 4:
            description += ' ...'

        applied.append(description)
        logger.info(f"🔢 [{worksheet.title}] {description}")

    return applied


def apply_column_widths(worksheet, rules: list, header_row: int = 1,
                        on_missing: str = 'warn') -> list:
    """
    Apply explicit column widths from any rule carrying a 'width'.

    Kept separate from apply_column_formats because widths must be set AFTER
    auto-fit, or auto-fit overwrites them. Everything else in a rule has to be
    applied BEFORE auto-fit so that widths are measured against formatted text.

    Args:
        worksheet:      openpyxl worksheet object
        rules:          The same rule list passed to apply_column_formats
        header_row:     Row holding the headers
        on_missing:     'error', 'warn', or 'skip' for unresolvable columns

    Returns:
        List of descriptions of what was set
    """
    applied = []

    for index, rule in enumerate(rules):
        if not isinstance(rule, dict):
            continue

        width = rule.get('width')

        if width is None:
            continue

        columns = rule.get('columns', [])

        try:
            letters = resolve_column_letters(
                worksheet, columns, header_row,
                force_column_names=rule.get('force_column_names', False),
                on_missing=on_missing
            )
        except ExcelRangeResolverError as error:
            if on_missing == 'error':
                raise ColumnFormatError(f"column_formats rule {index + 1} width: {error}")
            logger.warning(f"column_formats rule {index + 1} width: {error}")
            continue

        for letter in letters:
            worksheet.column_dimensions[letter].width = float(width)

        description = f"width {width} on {len(letters)} column(s)"
        applied.append(description)
        logger.info(f"📏 [{worksheet.title}] {description}: {', '.join(columns[:4])}")

    return applied


def apply_hidden_columns(worksheet, columns: list, header_row: int = 1,
                         on_missing: str = 'warn') -> list:
    """
    Hide whole columns, addressed by name.

    Hiding sets the column dimension's hidden flag, which is what Excel's own
    Hide command does. The data stays in the file and stays available to
    formulas; it simply is not displayed.

    Args:
        worksheet:      openpyxl worksheet object
        columns:        Column names or letters to hide
        header_row:     Row holding the headers
        on_missing:     'error', 'warn', or 'skip' for unresolvable columns

    Returns:
        List of the column letters hidden
    """
    if not isinstance(columns, list) or len(columns) == 0:
        raise ColumnFormatError("hidden_columns must be a non-empty list")

    try:
        letters = resolve_column_letters(worksheet, columns, header_row, False, on_missing)
    except ExcelRangeResolverError as error:
        if on_missing == 'error':
            raise ColumnFormatError(f"hidden_columns: {error}")
        logger.warning(f"hidden_columns: {error}")
        return []

    for letter in letters:
        worksheet.column_dimensions[letter].hidden = True

    logger.info(
        f"🙈 [{worksheet.title}] Hid {len(letters)} column(s): "
        f"{', '.join(f'{c} ({l})' for c, l in zip(columns, letters))}"
    )

    return letters


def apply_cell_formats(worksheet, rules: list, color_normalizer=None) -> list:
    """
    Apply fonts, alignment and number formats to SPECIFIC cells or ranges.

    The spot-styling counterpart of apply_column_formats: a rule names
    explicit A1-style cells or ranges ("B2", "A4:D4") instead of columns,
    for the styling that column vocabulary cannot express - a single
    prompt cell, a label row on a sheet whose headers live elsewhere.
    Explicit cell styles also override column-dimension (whole_column)
    styles, so a spot rule wins where they overlap.

    Args:
        worksheet:          openpyxl worksheet object
        rules:              List of rule dictionaries with a 'cells' list
                            plus number_format / font_color / font_bold /
                            font_italic / alignment_horizontal /
                            alignment_vertical / wrap_text
        color_normalizer:   Callable turning a colour spec into 6-digit hex

    Returns:
        List of human-readable descriptions of what was applied
    """
    if color_normalizer is None:
        color_normalizer = _default_color_normalizer
    if not isinstance(rules, list):
        raise ColumnFormatError("cell_formats must be a list of rules")

    applied = []

    for index, rule in enumerate(rules):
        if not isinstance(rule, dict):
            raise ColumnFormatError(f"cell_formats rule {index + 1} must be a dictionary")

        cells = rule.get('cells')
        if not isinstance(cells, list) or len(cells) == 0:
            raise ColumnFormatError(
                f"cell_formats rule {index + 1} requires a non-empty 'cells' "
                f"list of A1-style cells or ranges, e.g. ['B2'] or ['A4:D4']"
            )
        for cell_text in cells:
            candidate = str(cell_text).strip()
            if not (cell_ref_rgx.match(candidate) or range_ref_rgx.match(candidate)):
                raise ColumnFormatError(
                    f"cell_formats rule {index + 1}: {cell_text!r} is not an "
                    f"A1-style cell or range (like 'B2' or 'A4:D4')"
                )

        number_format = rule.get('number_format')
        horizontal = rule.get('alignment_horizontal')
        vertical = rule.get('alignment_vertical')
        wrap_text = rule.get('wrap_text')
        font_color = rule.get('font_color')
        font_bold = rule.get('font_bold')
        font_italic = rule.get('font_italic')
        font_size = rule.get('font_size')
        font_underline = rule.get('font_underline')
        font_strikethrough = rule.get('font_strikethrough')

        if font_size is not None and (
                not isinstance(font_size, (int, float)) or font_size <= 0):
            raise ColumnFormatError(
                f"cell_formats rule {index + 1}: 'font_size' must be a "
                f"positive number, got {font_size!r}"
            )

        underline_value = None
        if font_underline is not None:
            underline_value = _normalize_underline(
                font_underline, f"cell_formats rule {index + 1}")

        if font_strikethrough is not None and not isinstance(font_strikethrough, bool):
            raise ColumnFormatError(
                f"cell_formats rule {index + 1}: 'font_strikethrough' "
                f"must be true or false, got {font_strikethrough!r}"
            )

        actionable = (number_format, horizontal, vertical, wrap_text,
                      font_color, font_bold, font_italic, font_size,
                      font_underline, font_strikethrough)
        if all(value is None for value in actionable):
            raise ColumnFormatError(
                f"cell_formats rule {index + 1} does nothing: supply "
                f"number_format, alignment_horizontal, alignment_vertical, "
                f"wrap_text, font_color, font_bold, font_italic, font_size, "
                f"font_underline, or font_strikethrough"
            )

        if horizontal is not None and horizontal not in VALID_HORIZONTAL:
            raise ColumnFormatError(
                f"cell_formats rule {index + 1} invalid alignment_horizontal "
                f"'{horizontal}'. Valid: {', '.join(VALID_HORIZONTAL)}"
            )
        if vertical is not None and vertical not in VALID_VERTICAL:
            raise ColumnFormatError(
                f"cell_formats rule {index + 1} invalid alignment_vertical "
                f"'{vertical}'. Valid: {', '.join(VALID_VERTICAL)}"
            )

        format_code = resolve_number_format(number_format) if number_format else None
        data_color = color_normalizer(font_color) if font_color is not None else None
        touches_font = (data_color is not None or font_bold is not None
                        or font_italic is not None or font_size is not None
                        or underline_value is not None
                        or font_strikethrough is not None)
        touches_alignment = horizontal is not None or vertical is not None or wrap_text is not None

        cell_count = 0
        for cell_text in cells:
            block = worksheet[str(cell_text).strip().replace('$', '')]
            # A single cell comes back bare; ranges come back as row tuples
            if not isinstance(block, tuple):
                block = ((block,),)
            for row_cells in block:
                for cell in row_cells:
                    if format_code:
                        cell.number_format = format_code
                    if touches_font:
                        existing = cell.font
                        cell.font = Font(
                            name=existing.name,
                            size=font_size if font_size is not None else existing.size,
                            bold=font_bold if font_bold is not None else existing.bold,
                            italic=font_italic if font_italic is not None else existing.italic,
                            underline=underline_value if underline_value is not None else existing.underline,
                            strike=font_strikethrough if font_strikethrough is not None else existing.strike,
                            # Pass the Color OBJECT through when preserving:
                            # .rgb on a theme-based default is an RGB
                            # descriptor, which Font(color=...) rejects.
                            color=data_color if data_color is not None else existing.color
                        )
                    if touches_alignment:
                        existing = cell.alignment
                        cell.alignment = Alignment(
                            horizontal=horizontal if horizontal is not None else existing.horizontal,
                            vertical=vertical if vertical is not None else existing.vertical,
                            wrap_text=wrap_text if wrap_text is not None else existing.wrap_text
                        )
                    cell_count += 1

        parts = []
        if format_code:
            label = number_format if number_format in NUMBER_FORMAT_ALIASES else 'custom'
            parts.append(f"number format {label}")
        if horizontal is not None:
            parts.append(f"h-align {horizontal}")
        if vertical is not None:
            parts.append(f"v-align {vertical}")
        if wrap_text is not None:
            parts.append(f"wrap {wrap_text}")
        if font_color is not None:
            parts.append(f"font {font_color}")
        if font_bold is not None:
            parts.append(f"bold {font_bold}")
        if font_italic is not None:
            parts.append(f"italic {font_italic}")
        if font_size is not None:
            parts.append(f"size {font_size}")
        if font_underline is not None:
            parts.append(f"underline {underline_value}")
        if font_strikethrough is not None:
            parts.append(f"strikethrough {font_strikethrough}")

        description = f"{', '.join(parts)} on {cell_count} cell(s): {', '.join(str(c) for c in cells[:4])}"
        applied.append(description)
        logger.info(f"🔤 [{worksheet.title}] {description}")

    return applied


# End of file #

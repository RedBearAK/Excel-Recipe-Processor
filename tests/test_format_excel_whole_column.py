"""
Tests for whole_column formatting in apply_column_formats.

tests/test_format_excel_whole_column.py

whole_column: true applies number formats at the column-dimension level
(a col-level style in the file) so cells that Excel creates later -
dynamic-array spill results - inherit the format, which per-cell
formatting up to the current data extent cannot achieve. These tests
write real files and inspect the stored col elements and style chain.
Runnable directly or with pytest.
"""

import os
import re
import sys
import zipfile
import tempfile

from openpyxl import Workbook

from excel_recipe_processor.processors._helpers.format_excel_column_formats import (
    ColumnFormatError,
    apply_column_formats,
)


def col_style_formats(path: str) -> dict:
    """Map column min-index -> stored format code, via the style chain."""
    with zipfile.ZipFile(path) as archive:
        sheet_xml = archive.read('xl/worksheets/sheet1.xml').decode()
        styles_xml = archive.read('xl/styles.xml').decode()

    xfs_block = re.search(r'<cellXfs[^>]*>(.*?)</cellXfs>', styles_xml, re.S)
    xf_entries = re.findall(r'<xf [^>]*?/?>', xfs_block.group(1))
    numfmt_by_xf = {}
    for index, entry in enumerate(xf_entries):
        id_match = re.search(r'numFmtId="(\d+)"', entry)
        numfmt_by_xf[str(index)] = id_match.group(1) if id_match else '0'

    custom = dict(re.findall(
        r'<numFmt numFmtId="(\d+)" formatCode="([^"]+)"/>', styles_xml))
    builtin = {'3': '#,##0'}

    formats = {}
    for match in re.finditer(r'<col ([^>]*)/>', sheet_xml):
        attrs = match.group(1)
        min_match = re.search(r'min="(\d+)"', attrs)
        style_match = re.search(r'style="(\d+)"', attrs)
        if min_match and style_match:
            fmt_id = numfmt_by_xf.get(style_match.group(1), '0')
            formats[int(min_match.group(1))] = custom.get(fmt_id) or \
                builtin.get(fmt_id, f'builtin:{fmt_id}')
    return formats


def test_whole_column_style_chain():
    """whole_column stores a col-level style resolving to the format."""
    print("\nTesting whole_column col-level style chain...")

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as handle:
        target = handle.name
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Label:'
        # No data below the header - the spill-fed sheet shape

        applied = apply_column_formats(sheet, [
            {'columns': ['C'], 'number_format': 'thousands',
             'whole_column': True, 'width': 15},
        ])
        workbook.save(target)

        if not applied or '(whole column)' not in applied[0]:
            print(f"✗ Description lacks mechanism note: {applied}")
            return False
        print("✓ Applied description notes (whole column)")

        formats = col_style_formats(target)
        if formats.get(3) != '#,##0':
            print(f"✗ Column C style chain wrong: {formats}")
            return False
        print("✓ Column C col element resolves to #,##0")
        return True
    finally:
        if os.path.exists(target):
            os.unlink(target)


def test_empty_sheet_behavior():
    """Empty sheet: whole_column rules run; per-cell rules still skip."""
    print("\nTesting empty-sheet guard preservation...")

    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Header'

    # Legacy path alone: skipped with the historic warning, nothing applied
    applied = apply_column_formats(sheet, [
        {'columns': ['B'], 'number_format': 'thousands'},
    ])
    if applied:
        print(f"✗ Per-cell rule ran on an empty sheet: {applied}")
        return False
    print("✓ Per-cell rule on empty sheet skipped (legacy behavior)")

    # Mixed: whole_column applies, per-cell portion still skipped
    applied = apply_column_formats(sheet, [
        {'columns': ['B'], 'number_format': 'thousands'},
        {'columns': ['C'], 'number_format': 'thousands',
         'whole_column': True, 'width': 12},
    ])
    if len(applied) != 1 or '(whole column)' not in applied[0]:
        print(f"✗ Mixed rules on empty sheet wrong: {applied}")
        return False
    print("✓ whole_column rule applied while per-cell sibling skipped")
    return True


def test_guided_error():
    """Non-bool whole_column refused with the rule number named."""
    print("\nTesting guided error...")

    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Header'

    try:
        apply_column_formats(sheet, [
            {'columns': ['C'], 'number_format': 'thousands',
             'whole_column': 'yes'},
        ])
        print("✗ Non-bool whole_column should have raised")
        return False
    except ColumnFormatError as error:
        if 'whole_column' not in str(error) or 'rule 1' not in str(error):
            print(f"✗ Error lacks guidance: {error}")
            return False
        print("✓ Refused with rule number and key named")
    return True


def test_header_row_honored():
    """Sheet-level header band lands on header_row, not hardcoded row 1."""
    print("\nTesting header_row honoring...")

    from excel_recipe_processor.processors.format_excel_processor import (
        FormatExcelProcessor,
    )

    config = FormatExcelProcessor.get_minimal_config()
    config['processor_type'] = 'format_excel'
    processor = FormatExcelProcessor(config)

    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Label:'
    sheet['A4'] = 'Product'
    sheet['B4'] = 'Units'

    processor._apply_header_formatting(sheet, {
        'header_row': 4, 'header_bold': True,
        'header_background': True, 'header_background_color': '1F4E79',
    })

    if not sheet['A4'].font.bold or sheet['A4'].fill.fill_type != 'solid':
        print("✗ Row 4 did not receive the header band")
        return False
    if sheet['A1'].font.bold or sheet['A1'].fill.fill_type == 'solid':
        print("✗ Row 1 was styled despite header_row: 4")
        return False
    print("✓ Band on row 4, row 1 untouched")

    # Default stays row 1 (legacy behavior when the key is absent)
    workbook_default = Workbook()
    sheet_default = workbook_default.active
    sheet_default['A1'] = 'Header'
    processor._apply_header_formatting(sheet_default, {'header_bold': True})
    if not sheet_default['A1'].font.bold:
        print("✗ Default header_row no longer styles row 1")
        return False
    print("✓ Absent header_row still styles row 1")
    return True


def test_cell_formats_spot_styling():
    """cell_formats styles named cells/ranges; guided errors on misuse."""
    print("\nTesting cell_formats spot styling...")

    from excel_recipe_processor.processors._helpers.format_excel_column_formats import (
        apply_cell_formats,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet['B2'] = 'prompt'
    sheet['A4'] = 'Product'
    sheet['D4'] = 'Total'

    applied = apply_cell_formats(sheet, [
        {'cells': ['B2'], 'font_color': 'FF0000', 'font_bold': True},
        {'cells': ['A4:D4'], 'font_italic': True},
    ])
    if len(applied) != 2:
        print(f"✗ Expected 2 applied notes, got {applied}")
        return False

    b2 = sheet['B2']
    if not b2.font.bold or str(b2.font.color.rgb) != '00FF0000':
        print(f"✗ B2 not red bold: bold={b2.font.bold} color={b2.font.color.rgb}")
        return False
    print("✓ Single cell styled red bold")

    if not (sheet['A4'].font.italic and sheet['D4'].font.italic
            and sheet['C4'].font.italic):
        print("✗ Range A4:D4 not fully styled")
        return False
    if sheet['A1'].font.italic:
        print("✗ Styling leaked outside the range")
        return False
    print("✓ Range styled, no leakage")

    for label, bad_rules, fragment in [
        ("bad ref", [{'cells': ['B2:B'], 'font_bold': True}], "A1-style"),
        ("empty cells", [{'cells': [], 'font_bold': True}], "non-empty"),
        ("no action", [{'cells': ['B2']}], "does nothing"),
    ]:
        try:
            apply_cell_formats(sheet, bad_rules)
            print(f"✗ {label}: should have raised")
            return False
        except Exception as error:
            if fragment not in str(error):
                print(f"✗ {label}: error lacks guidance: {error}")
                return False
            print(f"✓ {label}: refused with guidance")
    return True


def main():
    """Run all tests and report results."""
    print("whole_column formatting tests")
    print("=" * 50)

    tests = [
        test_whole_column_style_chain,
        test_empty_sheet_behavior,
        test_guided_error,
        test_header_row_honored,
        test_cell_formats_spot_styling,
    ]

    passed = 0
    for test in tests:
        try:
            if test():
                passed += 1
            else:
                print(f"FAILED: {test.__name__}")
        except Exception as error:
            print(f"FAILED with exception: {test.__name__}: {error}")

    print("=" * 50)
    print(f"Final score: {passed}/{len(tests)} tests passed")
    return passed == len(tests)


if __name__ == '__main__':
    sys.exit(0 if main() else 1)

# End of file #

"""
Test inject_formulas mode: text - inert formula text that awaken can make live.

File: excel_recipe_processor/tests/test_inject_formulas_text_mode.py

mode: text writes the same formulas into the same cells as live, but as
string cells: Excel calculates nothing and shows the formula text. A later
awaken step turns them live. This drill writes a workbook, injects in
text mode, reloads and checks the stored type, then awakens and checks
the cell became a formula.

Run: PYTHONPATH=. python3 tests/test_inject_formulas_text_mode.py
"""

import sys
import tempfile

from pathlib import Path

import openpyxl

from excel_recipe_processor.processors.inject_formulas_processor import InjectFormulasProcessor


def make_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Data'
    sheet.append(['Units', 'Unit Price', 'Total'])
    sheet.append([3, 2.5, None])
    sheet.append([4, 1.0, None])
    workbook.save(path)


def run(path: Path, mode: str) -> None:
    InjectFormulasProcessor({
        'processor_type': 'inject_formulas', 'step_description': f'{mode} drill',
        'target_file': str(path), 'mode': mode,
        'sheets_to_receive_formulas': [{'sheet_names': ['Data'], 'formulas': [
            {'cell': 'C2', 'excel_formula': '={col:Units}2*{col:Unit Price}2', 'fill_down': True}]}],
    }).execute()


def test_text_mode_stores_inert_strings() -> bool:
    print('\nTesting mode: text stores formula text as string cells...')
    with tempfile.TemporaryDirectory() as temp_dir:
        path = Path(temp_dir) / 'book.xlsx'
        make_workbook(path)
        run(path, 'text')
        sheet = openpyxl.load_workbook(path)['Data']
        c2, c3 = sheet['C2'], sheet['C3']
        good = (c2.data_type == 's' and c2.value == '=A2*B2'
                and c3.data_type == 's' and c3.value == '=A3*B3')
        print(f"  C2={c2.value!r} ({c2.data_type}) C3={c3.value!r} ({c3.data_type}) -> {'OK' if good else 'FAIL'}")
        return good


def test_awaken_makes_text_live() -> bool:
    print('\nTesting that awaken turns the text cells into formulas...')
    with tempfile.TemporaryDirectory() as temp_dir:
        path = Path(temp_dir) / 'book.xlsx'
        make_workbook(path)
        run(path, 'text')
        InjectFormulasProcessor({
            'processor_type': 'inject_formulas', 'step_description': 'awaken drill',
            'target_file': str(path), 'mode': 'awaken', 'sheet_names': ['Data'],
        }).execute()
        sheet = openpyxl.load_workbook(path)['Data']
        good = sheet['C2'].data_type == 'f' and sheet['C3'].data_type == 'f'
        print(f"  C2 type={sheet['C2'].data_type} C3 type={sheet['C3'].data_type} -> {'OK' if good else 'FAIL'}")
        return good


def test_live_still_live() -> bool:
    print('\nTesting that mode: live is unchanged...')
    with tempfile.TemporaryDirectory() as temp_dir:
        path = Path(temp_dir) / 'book.xlsx'
        make_workbook(path)
        run(path, 'live')
        sheet = openpyxl.load_workbook(path)['Data']
        good = sheet['C2'].data_type == 'f'
        print(f"  C2 type={sheet['C2'].data_type} -> {'OK' if good else 'FAIL'}")
        return good


def main() -> int:
    tests = [test_text_mode_stores_inert_strings, test_awaken_makes_text_live, test_live_still_live]
    passed = 0
    for test in tests:
        try:
            if test():
                passed += 1
        except Exception as error:
            print(f'  EXCEPTION in {test.__name__}: {error}')
    print(f'\n{passed}/{len(tests)} tests passed')
    return 0 if passed == len(tests) else 1


if __name__ == '__main__':
    sys.exit(main())


# End of file #

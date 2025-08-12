"""
Performance test comparing old vs new openpyxl reading methods.

tests/test_openpyxl_performance.py

Creates test Excel files and measures performance difference between
cell-by-cell reading and efficient iter_cols() method.
"""

import os
import tempfile
import time
import pandas as pd

from pathlib import Path

# Import the optimized reader
from excel_recipe_processor.readers.openpyxl_excel_reader import OpenpyxlExcelReader


def create_large_test_excel_file(rows: int = 5000, cols: int = 50) -> str:
    """
    Create a large Excel file for performance testing.
    
    Args:
        rows: Number of data rows to create
        cols: Number of columns to create
        
    Returns:
        Path to temporary Excel file
    """
    print(f"Creating test Excel file with {rows} rows × {cols} columns...")
    
    # Create realistic test data
    data = {}
    
    for col_idx in range(cols):
        if col_idx == 0:
            # First column: Product IDs
            data[f'Product_ID'] = [f'P{i:05d}' for i in range(rows)]
        elif col_idx == 1:
            # Second column: Date-like header that should be preserved
            data['8/4/2025'] = [100 + i for i in range(rows)]
        elif col_idx == 2:
            # Third column: Empty header with data
            data[''] = [f'Data_{i}' for i in range(rows)]
        elif col_idx < 40:
            # Regular data columns
            data[f'Column_{col_idx}'] = [f'Value_{col_idx}_{i}' for i in range(rows)]
        elif col_idx < 45:
            # Columns with headers but no data (ghost columns)
            data[f'Empty_Header_{col_idx}'] = [''] * rows
        else:
            # Trailing empty columns (both header and data empty)
            data[''] = [''] * rows
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Save to temporary Excel file
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()
    
    print(f"Writing Excel file to {temp_file.name}...")
    df.to_excel(temp_file.name, index=False, engine='openpyxl')
    
    print(f"✓ Created {temp_file.name} ({rows} rows, {cols} columns)")
    return temp_file.name


def simulate_old_method(file_path: str, max_rows: int = 100000) -> dict:
    """
    Simulate the old cell-by-cell reading method for comparison.
    
    Args:
        file_path: Path to Excel file
        max_rows: Maximum rows to scan
        
    Returns:
        Same format as new method for comparison
    """
    import openpyxl
    
    print("Testing OLD METHOD (cell-by-cell access)...")
    start_time = time.time()
    
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    worksheet = workbook.active
    
    max_col = worksheet.max_column or 1
    max_row = worksheet.max_row or 1
    header_row = 1
    data_end_row = min(max_row, header_row + max_rows)
    
    # Read headers
    headers = []
    for col in range(1, max_col + 1):
        header_cell = worksheet.cell(row=header_row, column=col)
        header_value = str(header_cell.value) if header_cell.value is not None else ""
        headers.append(header_value)
    
    # OLD METHOD: Check data using individual cell access
    column_has_data = []
    for col in range(1, max_col + 1):
        has_data = False
        for row in range(header_row + 1, data_end_row + 1):
            cell = worksheet.cell(row=row, column=col)  # ← SLOW!
            if cell.value is not None and str(cell.value).strip():
                has_data = True
                break
        column_has_data.append(has_data)
    
    workbook.close()
    
    elapsed = time.time() - start_time
    empty_columns = [i for i, has_data in enumerate(column_has_data) if not has_data]
    
    print(f"✓ OLD METHOD completed in {elapsed:.2f} seconds")
    
    return {
        'headers': headers,
        'column_has_data': column_has_data,
        'empty_column_indices': empty_columns,
        'scanned_data_rows': data_end_row - header_row,
        'total_columns': len(headers),
        'elapsed_time': elapsed,
        'method': 'cell_by_cell'
    }


def test_new_method(file_path: str, max_rows: int = 100000) -> dict:
    """
    Test the new optimized iter_cols method.
    
    Args:
        file_path: Path to Excel file
        max_rows: Maximum rows to scan
        
    Returns:
        Analysis results with timing
    """
    print("Testing NEW METHOD (iter_cols efficient access)...")
    start_time = time.time()
    
    result = OpenpyxlExcelReader.read_headers_with_data_check(
        file_path=file_path,
        max_rows=max_rows
    )
    
    elapsed = time.time() - start_time
    result['elapsed_time'] = elapsed
    result['method'] = 'iter_cols'
    
    print(f"✓ NEW METHOD completed in {elapsed:.2f} seconds")
    
    return result


def verify_results_match(old_result: dict, new_result: dict) -> bool:
    """
    Verify that both methods produce identical results.
    
    Args:
        old_result: Results from old method
        new_result: Results from new method
        
    Returns:
        True if results match
    """
    print("\nVerifying results match between methods...")
    
    # Check headers
    if old_result['headers'] != new_result['headers']:
        print("✗ Headers don't match!")
        return False
    
    # Check column data detection
    if old_result['column_has_data'] != new_result['column_has_data']:
        print("✗ Column data detection doesn't match!")
        return False
    
    # Check empty column indices
    if old_result['empty_column_indices'] != new_result['empty_column_indices']:
        print("✗ Empty column indices don't match!")
        return False
    
    print("✓ Both methods produce identical results")
    return True


def performance_comparison_test():
    """
    Run a comprehensive performance comparison test.
    """
    print("=" * 60)
    print("OPENPYXL PERFORMANCE COMPARISON TEST")
    print("=" * 60)
    
    # Test with different file sizes
    test_cases = [
        {'rows': 1000, 'cols': 20, 'name': 'Small file'},
        {'rows': 5000, 'cols': 50, 'name': 'Medium file'},
        {'rows': 10000, 'cols': 100, 'name': 'Large file (optional)'}
    ]
    
    results = []
    
    for i, test_case in enumerate(test_cases):
        print(f"\n📊 TEST {i+1}: {test_case['name']}")
        print("-" * 40)
        
        # Create test file
        excel_file = create_large_test_excel_file(
            rows=test_case['rows'], 
            cols=test_case['cols']
        )
        
        try:
            # Test old method
            old_result = simulate_old_method(excel_file)
            
            # Test new method
            new_result = test_new_method(excel_file)
            
            # Verify results match
            results_match = verify_results_match(old_result, new_result)
            
            # Calculate performance improvement
            speedup = old_result['elapsed_time'] / new_result['elapsed_time']
            
            print(f"\n📈 PERFORMANCE RESULTS:")
            print(f"   Old method: {old_result['elapsed_time']:.2f} seconds")
            print(f"   New method: {new_result['elapsed_time']:.2f} seconds")
            print(f"   Speedup: {speedup:.1f}x faster")
            print(f"   Results match: {'✓' if results_match else '✗'}")
            
            # Show data analysis results
            print(f"\n📋 DATA ANALYSIS:")
            print(f"   Total columns: {new_result['total_columns']}")
            print(f"   Empty columns: {len(new_result['empty_column_indices'])}")
            print(f"   Scanned rows: {new_result['scanned_data_rows']}")
            
            results.append({
                'test_case': test_case['name'],
                'old_time': old_result['elapsed_time'],
                'new_time': new_result['elapsed_time'],
                'speedup': speedup,
                'results_match': results_match
            })
            
            # Skip large file test if already very slow
            if old_result['elapsed_time'] > 30 and i < len(test_cases) - 1:
                print(f"\n⚠️  Skipping larger tests (old method too slow)")
                break
                
        finally:
            # Clean up test file
            if Path(excel_file).exists():
                os.unlink(excel_file)
    
    # Summary
    print("\n" + "=" * 60)
    print("📊 PERFORMANCE SUMMARY")
    print("=" * 60)
    
    for result in results:
        print(f"{result['test_case']:15} {result['speedup']:6.1f}x faster  "
              f"({result['old_time']:.2f}s → {result['new_time']:.2f}s)  "
              f"{'✓' if result['results_match'] else '✗'}")
    
    avg_speedup = sum(r['speedup'] for r in results) / len(results)
    all_match = all(r['results_match'] for r in results)
    
    print(f"\nAverage speedup: {avg_speedup:.1f}x")
    print(f"All results match: {'✓' if all_match else '✗'}")
    
    if avg_speedup > 10 and all_match:
        print("\n🎉 SUCCESS: Optimization delivers significant performance improvement!")
        return True
    else:
        print("\n❌ ISSUES: Performance improvement insufficient or results don't match")
        return False


def quick_functionality_test():
    """
    Quick test to verify basic functionality works.
    """
    print("\n🔍 QUICK FUNCTIONALITY TEST")
    print("-" * 30)
    
    # Create small test file
    excel_file = create_large_test_excel_file(rows=100, cols=10)
    
    try:
        # Test headers only
        headers = OpenpyxlExcelReader.read_headers(excel_file)
        print(f"✓ Headers read: {len(headers)} columns")
        print(f"  Sample: {headers[:3]}...")
        
        # Test with data check
        analysis = OpenpyxlExcelReader.read_headers_with_data_check(excel_file)
        print(f"✓ Data analysis complete:")
        print(f"  Empty columns: {len(analysis['empty_column_indices'])}")
        print(f"  Scanned rows: {analysis['scanned_data_rows']}")
        
        # Verify date preservation
        date_header = next((h for h in headers if '/' in h), None)
        if date_header:
            print(f"✓ Date header preserved: '{date_header}'")
        
        return True
        
    except Exception as e:
        print(f"✗ Test failed: {e}")
        return False
        
    finally:
        if Path(excel_file).exists():
            os.unlink(excel_file)


def main():
    """
    Run all performance tests.
    """
    print("🚀 OPENPYXL PERFORMANCE OPTIMIZATION TEST SUITE")
    
    # Quick functionality test
    if not quick_functionality_test():
        print("❌ Basic functionality test failed!")
        return False
    
    # Full performance comparison
    success = performance_comparison_test()
    
    if success:
        print("\n✅ ALL TESTS PASSED!")
        print("The optimized OpenpyxlExcelReader is ready for production use.")
        return True
    else:
        print("\n❌ TESTS FAILED!")
        print("The optimization needs further work.")
        return False


if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)


# End of file #

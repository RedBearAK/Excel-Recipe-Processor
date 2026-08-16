"""
Profile sheets: per-column metadata as an ordinary stage.

excel_recipe_processor/processors/profile_sheets_processor.py

Part of the profile_* family (profile_files, profile_sheets, and the
planned profile_workbooks): discovery processors that store metadata
as plain stages for OTHER steps to consume. There are deliberately no
"apply" siblings - consumers are existing processors with directives
that read the profile by column name (the first: format_excel's
column_widths_from_stage, which lets FILTER-spill view tabs inherit
the seed sheet's calculated widths instead of accumulating hand-set
overrides).

Inputs are plural: each entry profiles either a stage (source_stage)
or an on-disk sheet (input_file + optional sheet_name). The
memory-vs-disk distinction is a pointer choice, not a mode - both
kinds land in the same output with the same facts, identified by the
Source column ("stg_name" or "file.xlsx!Sheet").

OUTPUT CONTRACT (columns by name; future facts APPEND, never rename):
    Source          Where the sheet came from
    Column          Header name
    Position        1-based column position
    Width           Auto-fit width via the SHARED clamp - the same
                    arithmetic format_excel's auto-fit uses, so a view
                    inheriting these matches what the seed sheet's own
                    auto-fit would compute. Exact for plain data;
                    workbook-side font/bold factors are approximated
                    as none (stages have no fonts).
    Dtype           pandas dtype string
    Blank_Count     NA values plus empty-string values
    Distinct_Count  Distinct NON-BLANK values (NA and empty string both
                    excluded). An empty string does not survive an Excel
                    write/read round trip - it becomes an empty cell - so
                    counting it would make stage-side and disk-side
                    profiles of the SAME data disagree.
    Row_Count       Data rows profiled

FORMAT SURVEY (file inputs only; '' / False for stage inputs, which
have no formatting to survey - a stage is data, not a styled sheet):
    Number_Format        Modal data-cell number format ('' for General)
    Alignment_Horizontal Modal explicit data alignment ('' when unset)
    Data_Font_Color      Modal explicit data font color RGB ('' unset)
    Header_Fill_Color    Header cell solid-fill RGB ('' unfilled)
    Header_Font_Color    Header cell font color RGB ('' default)
    Header_Bold          Header cell bold flag

The format survey (2026-08-15) is the previously PLANNED census,
implemented for disk-sheet inputs: format_excel's
column_styles_from_stage consumes it so view tabs mirror the seed
sheet's formatting the same way widths inherit.
"""

import logging
from collections import Counter

import pandas as pd
from openpyxl import load_workbook

from excel_recipe_processor.core.stage_manager import StageManager
from excel_recipe_processor.core.base_processor import ImportBaseProcessor, StepProcessorError
from excel_recipe_processor.processors._helpers.column_width_scan import (
    BASE_PADDING,
    DEFAULT_MIN_WIDTH,
    DEFAULT_MAX_WIDTH,
    scan_frame_column_widths,
)


logger = logging.getLogger(__name__)


class ProfileSheetsProcessor(ImportBaseProcessor):
    """Profile one or more sheets/stages into a per-column metadata stage."""

    @classmethod
    def get_minimal_config(cls):
        return {
            'sheets': [{'source_stage': 'stg_example_data'}],
            'save_to_stage': 'stg_sheet_profiles',
        }

    def __init__(self, step_config: dict):
        super().__init__(step_config)

        self.sheet_entries = self.get_config_value('sheets', None)
        self.min_width = self.get_config_value('min_width', DEFAULT_MIN_WIDTH)
        self.max_width = self.get_config_value('max_width', DEFAULT_MAX_WIDTH)
        self.padding = self.get_config_value('padding', BASE_PADDING)
        self.scan_rows = self.get_config_value('scan_rows', None)

        self._validate_sheet_entries(self.sheet_entries)

    def get_capabilities(self) -> dict:
        """Processor capabilities information."""
        return {
            'description': 'Per-column sheet metadata discovery (e.g., widths, dtypes, blanks, distincts)',
            'profile_columns': ['Source', 'Column', 'Position', 'Width',
                                'Dtype', 'Blank_Count', 'Distinct_Count',
                                'Row_Count'],
            'input_kinds': ['source_stage', 'input_file (+ sheet_name)'],
            'width_math': 'shared with format_excel auto-fit (column_width_scan helper)',
            'consumers': ['format_excel column_widths_from_stage'],
        }

    def load_data(self) -> pd.DataFrame:
        """Build the profile frame from every configured input."""
        rows = []
        for entry in self.sheet_entries:
            source_label, frame, styled_sheet = self._resolve_entry(entry)
            format_survey = self._survey_formats(styled_sheet, frame.columns) \
                if styled_sheet is not None else {}
            widths = scan_frame_column_widths(
                frame, min_width=self.min_width, max_width=self.max_width,
                padding=self.padding, scan_rows=self.scan_rows)
            for position, (column_name, width, _length) in enumerate(widths, 1):
                series = frame[column_name]
                non_null = series.dropna()
                if len(non_null) > 0:
                    empty_mask = non_null.astype(str) == ''
                else:
                    empty_mask = non_null == non_null  # empty boolean series
                blank_count = int(series.isna().sum()) + int(empty_mask.sum())
                non_blank = non_null[~empty_mask] if len(non_null) > 0 else non_null
                rows.append({
                    'Source': source_label,
                    'Column': str(column_name),
                    'Position': position,
                    'Width': width,
                    'Dtype': str(series.dtype),
                    'Blank_Count': blank_count,
                    'Distinct_Count': int(non_blank.nunique()),
                    'Row_Count': len(frame),
                    **format_survey.get(str(column_name), {
                        'Number_Format': '', 'Alignment_Horizontal': '',
                        'Data_Font_Color': '', 'Header_Fill_Color': '',
                        'Header_Font_Color': '', 'Header_Bold': False,
                    }),
                })

        profile = pd.DataFrame(rows, columns=[
            'Source', 'Column', 'Position', 'Width', 'Dtype',
            'Blank_Count', 'Distinct_Count', 'Row_Count',
            'Number_Format', 'Alignment_Horizontal', 'Data_Font_Color',
            'Header_Fill_Color', 'Header_Font_Color', 'Header_Bold'])
        logger.info(
            f"📊 Profiled {len(self.sheet_entries)} sheet(s): "
            f"{len(profile)} column rows")
        return profile

    def _validate_sheet_entries(self, sheet_entries) -> None:
        """Guided errors for the sheets list shape."""
        if not isinstance(sheet_entries, list) or len(sheet_entries) == 0:
            raise StepProcessorError(
                f"Step '{self.step_name}': 'sheets' must be a non-empty "
                f"list of entries, each with either source_stage OR "
                f"input_file (+ optional sheet_name)"
            )
        for index, entry in enumerate(sheet_entries, 1):
            if not isinstance(entry, dict):
                raise StepProcessorError(
                    f"Step '{self.step_name}': sheets entry {index} must "
                    f"be a mapping, got {type(entry).__name__}"
                )
            has_stage = 'source_stage' in entry
            has_file = 'input_file' in entry
            if has_stage == has_file:
                raise StepProcessorError(
                    f"Step '{self.step_name}': sheets entry {index} needs "
                    f"exactly ONE of source_stage or input_file, got "
                    f"{'both' if has_stage else 'neither'}"
                )

    def _resolve_entry(self, entry: dict):
        """(source label, DataFrame) for one sheets entry.

        Paths arrive already variable-substituted - the pipeline
        substitutes the whole step config before construction.
        """
        if 'source_stage' in entry:
            stage_name = entry['source_stage']
            return stage_name, StageManager.load_stage(stage_name), None
        input_path = entry['input_file']
        sheet_name = entry.get('sheet_name', 0)

        # SESSION CACHE FIRST (2026-08-16): when the path is the run's own
        # in-flight workbook, use the live object - every recipe-applied
        # format is already on it, and the disk parse (13s on the real VMS
        # file, plus the flush save and reload it forced) disappears. This
        # is the memory-vs-disk seam staying hidden, as ruled when the
        # family was designed: the pointer decides, the user never does.
        # peek never loads: profiling a previous run's file still reads
        # disk without dragging it into the session.
        from excel_recipe_processor.core.workbook_session import WorkbookSession
        cached = WorkbookSession.peek_workbook(input_path)
        if cached is not None:
            styled = cached[sheet_name] if isinstance(sheet_name, str) \
                else cached[cached.sheetnames[sheet_name]]
            frame = self._frame_from_worksheet(styled)
        else:
            frame = pd.read_excel(input_path, sheet_name=sheet_name)
            workbook = load_workbook(input_path)
            styled = workbook[sheet_name] if isinstance(sheet_name, str) \
                else workbook[workbook.sheetnames[sheet_name]]
        label = sheet_name if isinstance(sheet_name, str) else 'first sheet'
        return f"{input_path}!{label}", frame, styled

    def _frame_from_worksheet(self, worksheet) -> pd.DataFrame:
        """DataFrame from a live sheet's values, formula cells as NA.

        Formula cells (text starting '=' or ArrayFormula objects) become
        None, matching what a disk read of the same fresh file yields -
        openpyxl carries no calculated values until Excel has opened it.
        """
        rows_iter = worksheet.iter_rows(values_only=True)
        headers = [str(h) if h is not None else '' for h in next(rows_iter)]
        cleaned_rows = []
        for row in rows_iter:
            cleaned_rows.append([
                None if (isinstance(value, str) and value.startswith('='))
                or type(value).__name__ == 'ArrayFormula'
                else value
                for value in row
            ])
        return pd.DataFrame(cleaned_rows, columns=headers)

    def _survey_formats(self, worksheet, column_names) -> dict:
        """Per-column formatting census of a styled disk sheet.

        Modal values over the first 50 data cells: a column's format is
        what MOST of its cells wear, so a stray hand-edit cannot hijack
        the inheritance. 'General' surveys as '' (nothing to inherit).
        """
        survey = {}
        header_cells = {str(cell.value): cell for cell in worksheet[1]
                        if cell.value is not None}
        for column_name in column_names:
            header_cell = header_cells.get(str(column_name))
            if header_cell is None:
                continue
            facts = {'Number_Format': '', 'Alignment_Horizontal': '',
                     'Data_Font_Color': '', 'Header_Fill_Color': '',
                     'Header_Font_Color': '', 'Header_Bold': False}

            fill = header_cell.fill
            if fill is not None and fill.fill_type == 'solid' \
                    and getattr(fill.start_color, 'rgb', None):
                facts['Header_Fill_Color'] = str(fill.start_color.rgb)
            font = header_cell.font
            if font is not None:
                facts['Header_Bold'] = bool(font.bold)
                if font.color is not None and isinstance(
                        getattr(font.color, 'rgb', None), str):
                    facts['Header_Font_Color'] = str(font.color.rgb)

            formats, alignments, colors = Counter(), Counter(), Counter()
            column_index = header_cell.column
            last_row = min(worksheet.max_row, 51)
            for row in range(2, last_row + 1):
                cell = worksheet.cell(row=row, column=column_index)
                if cell.number_format and cell.number_format != 'General':
                    formats[cell.number_format] += 1
                if cell.alignment is not None and cell.alignment.horizontal:
                    alignments[cell.alignment.horizontal] += 1
                if cell.font is not None and cell.font.color is not None \
                        and isinstance(getattr(cell.font.color, 'rgb', None), str):
                    colors[str(cell.font.color.rgb)] += 1
            if formats:
                facts['Number_Format'] = formats.most_common(1)[0][0]
            if alignments:
                facts['Alignment_Horizontal'] = alignments.most_common(1)[0][0]
            if colors:
                facts['Data_Font_Color'] = colors.most_common(1)[0][0]
            survey[str(column_name)] = facts
        return survey

# End of file #

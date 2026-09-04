"""
Tests for profile_workbooks, including the drift-alarm consumer shape.

tests/test_profile_workbooks.py

Facts are asserted against a workbook BUILT with known properties
(hidden and veryHidden sheets, tab colors, freeze, zoom, a DV rule, a
defined name), then the anchor consumer is exercised for real: two
profile frames of deliberately-drifted workbooks go through diff_data,
and the drift must surface as diff rows. Runnable directly or with
pytest; direct runs are the authoritative score.
"""

import os
import sys
import tempfile

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from excel_recipe_processor.core.stage_manager import StageManager
from excel_recipe_processor.processors.profile_workbooks_processor import (
    ProfileWorkbooksProcessor,
)


def build_workbook(path, zoom=125, hide_lookup=True, extra_rows=0):
    """A workbook with every profiled property set to a known value."""
    workbook = Workbook()
    main = workbook.active
    main.title = 'Main'
    main.append(['A', 'B'])
    for row_number in range(3 + extra_rows):
        main.append([row_number, row_number * 2])
    main.freeze_panes = 'A2'
    main.sheet_view.zoomScale = zoom
    main.sheet_properties.tabColor = 'FFEB3B'
    validation = DataValidation(type='list', formula1='"x,y"')
    validation.add('A2:A5')
    main.add_data_validation(validation)

    lookup = workbook.create_sheet('Lookup')
    lookup['A1'] = 'k'
    if hide_lookup:
        lookup.sheet_state = 'hidden'
    vault = workbook.create_sheet('Vault')
    vault['A1'] = 'v'
    vault.sheet_state = 'veryHidden'

    workbook.defined_names['rng_keys'] = DefinedName(
        'rng_keys', attr_text='Lookup!$A$1')
    workbook.save(path)


def run_processor(workbooks):
    config = {'processor_type': 'profile_workbooks',
              'workbooks': workbooks,
              'save_to_stage': 'stg_t_wb_profile'}
    return ProfileWorkbooksProcessor(config).load_data()


def test_profile_facts():
    """Every contract column reports the built workbook's known values."""
    print("\nTesting workbook profile facts...")

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as handle:
        path = handle.name
    try:
        build_workbook(path)
        profile = run_processor([path]).set_index('Sheet')

        checks = [
            ('three sheets, ordered positions',
             list(profile['Position']) == [1, 2, 3]),
            ('states', list(profile['State'])
             == ['visible', 'hidden', 'veryHidden']),
            ('tab color captured',
             profile.loc['Main', 'Tab_Color'] == '00FFEB3B'),
            ('uncolored is empty string',
             profile.loc['Lookup', 'Tab_Color'] == ''),
            ('frozen panes', profile.loc['Main', 'Frozen_Panes'] == 'A2'),
            ('unfrozen empty', profile.loc['Lookup', 'Frozen_Panes'] == ''),
            ('zoom', profile.loc['Main', 'Zoom_Percent'] == 125),
            ('zoom default 100', profile.loc['Lookup', 'Zoom_Percent'] == 100),
            ('DV counted', profile.loc['Main', 'DV_Count'] == 1),
            ('extents (header + 3 data rows)',
             profile.loc['Main', 'Max_Row'] == 4
             and profile.loc['Main', 'Max_Col'] == 2),
            ('named objects counted, repeated per row',
             (profile['Named_Object_Count'] == 1).all()),
            ('no VBA in an xlsx', (profile['Has_VBA'] == False).all()),
        ]
        for name, ok in checks:
            if not ok:
                print(f"✗ {name}")
                print(profile)
                return False
            print(f"✓ {name}")
        return True
    finally:
        os.unlink(path)


def test_drift_alarm_through_diff_data():
    """The anchor consumer: drift between two runs surfaces as diff rows."""
    print("\nTesting the drift alarm shape (diff_data on two profiles)...")

    from excel_recipe_processor.processors.diff_data_processor import (
        DiffDataProcessor,
    )

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as h1, \
         tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as h2:
        previous_path, current_path = h1.name, h2.name
    try:
        build_workbook(previous_path)
        # The drifted run: someone unhid the lookup tab, zoom changed,
        # and the main sheet gained rows
        build_workbook(current_path, zoom=100, hide_lookup=False,
                       extra_rows=4)

        StageManager.initialize_stages(max_stages=30)
        baseline = run_processor([previous_path])
        current = run_processor([current_path])
        # Workbook paths differ between runs by nature; the drift key is
        # the SHEET - drop the path column for the comparison
        StageManager.save_stage('stg_t_baseline',
                                baseline.drop(columns='Workbook'), 'test')
        StageManager.save_stage('stg_t_current',
                                current.drop(columns='Workbook'), 'test')

        diff_config = {
            'processor_type': 'diff_data',
            'reference_stage': 'stg_t_baseline',
            'source_stage': 'stg_t_current',
            'key_columns': ['Sheet'],
            'save_to_stage': 'stg_t_drift',
        }
        # diff_data is data-in-data-out: current frame as the argument,
        # reference loaded from its configured stage
        diff = DiffDataProcessor(diff_config).execute(
            current.drop(columns='Workbook'))

        text = diff.to_string()
        drifted = [term for term in ('Main', 'Lookup') if term in text]
        if len(drifted) < 2:
            print(f"✗ drift rows missing: expected Main and Lookup changes")
            print(text[:600])
            return False
        print(f"✓ diff_data surfaced drift on: {drifted}")
        print("✓ the drift-alarm step shape works end to end")
        return True
    finally:
        os.unlink(previous_path)
        os.unlink(current_path)


def main():
    """Run all tests and report results."""
    print("profile_workbooks tests")
    print("=" * 50)

    tests = [
        test_profile_facts,
        test_drift_alarm_through_diff_data,
    ]

    passed = 0
    for test in tests:
        try:
            if test():
                passed += 1
            else:
                print(f"FAILED: {test.__name__}")
        except Exception as error:
            print(f"FAILED with exception: {test.__name__}: {error}")

    print("=" * 50)
    print(f"Final score: {passed}/{len(tests)} tests passed")
    return passed == len(tests)


if __name__ == '__main__':
    sys.exit(0 if main() else 1)

# End of file #

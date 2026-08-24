"""
Tests for the consolidated formatting cycle (2026-08-23).

tests/test_format_excel_consolidated_cycle.py

One cycle, one module: make_hyperlinks (three kinds, real cell.hyperlink
relationships), font underline/strikethrough, banded rows (banding WINS
over column tints by design), outline borders (default used range and
addressable ranges), row height conveniences plus gridline toggle, and
the profile_files Path column that feeds the Sources-tab file:// links.
Runnable directly or with pytest.
"""

import os
import sys
import tempfile

from openpyxl import Workbook, load_workbook

from excel_recipe_processor.core.base_processor import StepProcessorError
from excel_recipe_processor.processors.format_excel_processor import FormatExcelProcessor
from excel_recipe_processor.processors.profile_files_processor import ProfileFilesProcessor
from excel_recipe_processor.processors._helpers.format_excel_column_formats import (
    ColumnFormatError,
    apply_column_formats,
)
from excel_recipe_processor.processors._helpers.format_excel_sheet_features import (
    apply_banded_rows,
    apply_outline_border,
)


def build_sheet(rows_of_data=6):
    """A workbook with Name, Amount, Path, Site, Contact and data rows."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(['Name', 'Amount', 'Path', 'Site', 'Contact'])
    for row_num in range(1, rows_of_data + 1):
        worksheet.append([
            f'row{row_num}',
            row_num * 10,
            f'/Users/mmf/Dropbox/Data Files/file {row_num}.xlsx',
            'www.example.com',
            f'user{row_num}@example.com',
        ])
    return workbook, worksheet


def make_processor():
    """A FormatExcelProcessor to drive _apply_sheet_formatting directly."""
    step = {
        'processor_type': 'format_excel',
        'step_description': 'consolidated cycle test',
        'target_file': 'unused.xlsx',
    }
    return FormatExcelProcessor(step)


def test_hyperlinks_three_kinds():
    """All three kinds link, blanks skip, targets survive a save."""
    print("\nTesting make_hyperlinks three kinds with save round-trip...")

    workbook, worksheet = build_sheet(3)
    worksheet['C3'] = ''

    apply_column_formats(worksheet, [
        {'columns': ['Path'], 'make_hyperlinks': 'file_paths'},
        {'columns': ['Site'], 'make_hyperlinks': 'web_urls'},
        {'columns': ['Contact'], 'make_hyperlinks': 'email_addresses'},
    ])

    file_target = worksheet['C2'].hyperlink.target
    if file_target != 'file:///Users/mmf/Dropbox/Data%20Files/file%201.xlsx':
        print(f"  Wrong file target: {file_target}")
        return False
    if worksheet['C2'].value != '/Users/mmf/Dropbox/Data Files/file 1.xlsx':
        print("  Cell text should stay the readable un-encoded path")
        return False
    if worksheet['C3'].hyperlink is not None:
        print("  Blank cell should have been skipped")
        return False
    if worksheet['D2'].hyperlink.target != 'https://www.example.com':
        print(f"  Scheme-less URL not assumed https: {worksheet['D2'].hyperlink.target}")
        return False
    if worksheet['E2'].hyperlink.target != 'mailto:user1@example.com':
        print(f"  Wrong mailto target: {worksheet['E2'].hyperlink.target}")
        return False
    if worksheet['C2'].font.underline != 'single':
        print("  Link font should be underlined")
        return False

    # The relationship must survive a real save/load, not just live in memory
    with tempfile.TemporaryDirectory() as workdir:
        path = os.path.join(workdir, 'links.xlsx')
        workbook.save(path)
        reloaded = load_workbook(path)
        cell = reloaded.active['C2']
        if cell.hyperlink is None or 'Data%20Files' not in cell.hyperlink.target:
            print("  Hyperlink relationship did not survive save/load")
            return False

    print("  All three kinds correct, blanks skipped, round-trip OK")
    return True


def test_hyperlink_guided_errors():
    """Bad kind, whole_column combo, and bad values refuse loudly."""
    print("\nTesting make_hyperlinks guided errors...")

    workbook, worksheet = build_sheet(2)

    try:
        apply_column_formats(worksheet, [
            {'columns': ['Path'], 'make_hyperlinks': 'file_path'}])
        print("  Missed the near-miss singular 'file_path'")
        return False
    except ColumnFormatError as error:
        if 'file_paths, web_urls, email_addresses' not in str(error):
            print(f"  Error should name the legal kinds: {error}")
            return False

    try:
        apply_column_formats(worksheet, [
            {'columns': ['Path'], 'make_hyperlinks': 'file_paths',
             'whole_column': True}])
        print("  Missed the whole_column combination")
        return False
    except ColumnFormatError:
        pass

    worksheet['C2'] = 'relative/path.xlsx'
    try:
        apply_column_formats(worksheet, [
            {'columns': ['Path'], 'make_hyperlinks': 'file_paths'}])
        print("  Missed the relative path")
        return False
    except ColumnFormatError as error:
        if 'C2' not in str(error):
            print(f"  Error should name the cell address: {error}")
            return False

    print("  All three refusals fired with guided messages")
    return True


def test_underline_and_strikethrough():
    """Underline and strikethrough apply and bad values refuse."""
    print("\nTesting font_underline and font_strikethrough...")

    workbook, worksheet = build_sheet(2)

    apply_column_formats(worksheet, [
        {'columns': ['Name'], 'font_underline': True,
         'font_strikethrough': True},
        {'columns': ['Amount'], 'font_underline': 'double'},
    ])

    if worksheet['A2'].font.underline != 'single':
        print(f"  true should mean single, got {worksheet['A2'].font.underline}")
        return False
    if worksheet['A2'].font.strike is not True:
        print("  Strikethrough did not apply")
        return False
    if worksheet['B2'].font.underline != 'double':
        print("  Explicit double did not apply")
        return False

    try:
        apply_column_formats(worksheet, [
            {'columns': ['Name'], 'font_underline': 'wavy'}])
        print("  Missed the invalid underline value")
        return False
    except ColumnFormatError as error:
        if 'single' not in str(error) or 'double' not in str(error):
            print(f"  Error should name the legal values: {error}")
            return False

    print("  Underline single/double and strikethrough correct")
    return True


def test_banding_wins_over_tint():
    """Band rows paint across tinted columns; off-band rows keep the tint."""
    print("\nTesting banded rows beating column tints...")

    processor = make_processor()
    workbook, worksheet = build_sheet(6)

    processor._apply_sheet_formatting(worksheet, {
        'banded_row_color': 'EAF3FB',
        'banded_row_border_style': 'thin',
        'column_formats': [
            {'columns': ['Amount'], 'background_color': 'FCE4D6',
             'border_style': 'thin'},
        ],
    })

    # Row 3 is the first band row; B3 sits in the tinted column
    if worksheet['B3'].fill.start_color.rgb != '00EAF3FB':
        print(f"  Band should win on B3, got {worksheet['B3'].fill.start_color.rgb}")
        return False
    # Row 2 is off-band; the tint must read through
    if worksheet['B2'].fill.start_color.rgb != '00FCE4D6':
        print(f"  Tint should keep B2, got {worksheet['B2'].fill.start_color.rgb}")
        return False
    # Off-band, untinted cell keeps no fill at all
    if worksheet['A2'].fill.fill_type is not None:
        print("  A2 should carry no fill")
        return False
    # The band border remedy rules the banded cells
    if worksheet['A3'].border.left.style != 'thin':
        print("  Band border remedy missing on A3")
        return False

    # Border-without-band refuses at validation
    try:
        processor._validate_sheet_formatting_options(
            {'sheet_name': 'x', 'banded_row_border_style': 'thin'}, 'test')
        print("  Missed banded_row_border_style without banded_row_color")
        return False
    except StepProcessorError:
        pass

    print("  Banding wins, tints read through off-band, remedy ruled")
    return True


def test_outline_borders_and_ranges():
    """Default used-range box, addressable list, interior sides survive."""
    print("\nTesting outline borders...")

    processor = make_processor()
    workbook, worksheet = build_sheet(6)

    processor._apply_sheet_formatting(worksheet, {
        'column_formats': [
            {'columns': ['Amount'], 'background_color': 'FCE4D6',
             'border_style': 'thin'},
        ],
        'outline_border_style': 'medium',
        'outline_border_range': ['A1:E7', 'B2:C4'],
    })

    corner = worksheet['A1'].border
    if corner.top.style != 'medium' or corner.left.style != 'medium':
        print("  Used-range corner missing its two outward sides")
        return False
    if corner.bottom.style == 'medium':
        print("  Corner bottom is interior and must not be boxed")
        return False

    # B2 is a corner of the second box AND a thin-ruled tinted cell:
    # outward sides go medium, the interior thin ruling survives
    inner = worksheet['B2'].border
    if inner.left.style != 'medium' or inner.top.style != 'medium':
        print("  Explicit range corner missing outward sides")
        return False
    if inner.right.style != 'thin':
        print(f"  Interior thin side should survive, got {inner.right.style}")
        return False

    workbook2, worksheet2 = build_sheet(3)
    try:
        processor._apply_sheet_formatting(worksheet2, {
            'outline_border_style': 'thin',
            'outline_border_range': 'nonsense',
        })
        print("  Missed the malformed range")
        return False
    except StepProcessorError:
        pass

    print("  Boxes landed on outward sides only, interior ruling intact")
    return True


def test_heights_gridlines_and_path_column():
    """Row height conveniences, gridline toggle, profile_files Path."""
    print("\nTesting row heights, gridlines, and the Path column...")

    processor = make_processor()
    workbook, worksheet = build_sheet(3)

    processor._apply_sheet_formatting(worksheet, {
        'header_row_height': 24,
        'data_row_height': 16.5,
        'show_gridlines': False,
    })

    if worksheet.row_dimensions[1].height != 24:
        print("  Header row height missing")
        return False
    if worksheet.sheet_format.defaultRowHeight != 16.5:
        print("  Default data row height missing")
        return False
    if worksheet.sheet_format.customHeight is not True:
        print("  customHeight flag must be set for the default to stick")
        return False
    if worksheet.sheet_view.showGridLines is not False:
        print("  Gridlines toggle missing")
        return False

    with tempfile.TemporaryDirectory() as workdir:
        real_path = os.path.join(workdir, 'real file.xlsx')
        with open(real_path, 'w') as handle:
            handle.write('x')

        step = {
            'processor_type': 'profile_files',
            'step_description': 'path column test',
            'files': [real_path],
            'include_full_paths': True,
            'save_to_stage': 'stg_test_profiled',
        }
        frame = ProfileFilesProcessor(step).load_data()

        if list(frame.columns) != ['File', 'Modified', 'Size (KB)', 'Path']:
            print(f"  Wrong columns: {list(frame.columns)}")
            return False
        if not str(frame.iloc[0]['Path']).endswith('real file.xlsx'):
            print(f"  Wrong Path value: {frame.iloc[0]['Path']}")
            return False

        step_default = dict(step)
        del step_default['include_full_paths']
        frame_default = ProfileFilesProcessor(step_default).load_data()
        if 'Path' in frame_default.columns:
            print("  Default shape must stay Path-free")
            return False

    print("  Heights, gridlines, and opt-in Path column all correct")
    return True


def main():
    """Run the consolidated cycle tests and accumulate the score."""
    print("=" * 50)
    print("Consolidated formatting cycle tests")
    print("=" * 50)

    tests = [
        test_hyperlinks_three_kinds,
        test_hyperlink_guided_errors,
        test_underline_and_strikethrough,
        test_banding_wins_over_tint,
        test_outline_borders_and_ranges,
        test_heights_gridlines_and_path_column,
    ]

    passed = 0
    for test in tests:
        try:
            if test():
                passed += 1
            else:
                print(f"FAILED: {test.__name__}")
        except Exception as error:
            print(f"FAILED with exception: {test.__name__}: {error}")

    print("=" * 50)
    print(f"Final score: {passed}/{len(tests)} tests passed")
    return passed == len(tests)


if __name__ == '__main__':
    sys.exit(0 if main() else 1)

# End of file #

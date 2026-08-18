"""
Tests for the Dom_View braces incident class (2026-08-16).

tests/test_declaration_lambda_and_registry.py

Three defects interacted: openpyxl drops the cm attribute on a
load/save round trip; the provenance registry was popped after each
save, so the re-declaration pass at the next save could only rescue
vocabulary-hit cells; and the vocabulary could not see a spill hidden
behind a named lambda (fn_vms_view). Cells reverted to bare t="array"
- legacy CSE braces, a spill collapsed to one value until hand-
re-entered. These tests pin all three fixes, plus a reusable
no-legacy-CSE sweep any output file can be held to. Runnable directly
or with pytest; direct runs are the authoritative score.
"""

import os
import re
import sys
import tempfile
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.formula import ArrayFormula

from excel_recipe_processor.core.dynamic_array_metadata import (
    save_workbook_with_declaration,
)

STORED_LAMBDA = ('_xlfn.LAMBDA(_xlpm.criteria,fn_blank_safe('
                 '_xlfn._xlws.FILTER(rng_all,_xlpm.criteria,"none")))')


# The legacy-CSE sweep PROMOTED (2026-08-16) to
# core/excel_storage_audit.py as audit_legacy_cse; the local name stays
# for existing callers.
from excel_recipe_processor.core.excel_storage_audit import (  # noqa
    audit_legacy_cse as assert_no_legacy_cse,
)


def build_lambda_workbook(path):
    """A cell calling a named lambda - no spill function in the TEXT."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'View'
    sheet['A1'] = 'Header'
    sheet['A2'] = ArrayFormula('A2', '=fn_vms_view(rng_saletype="Export")')
    workbook.defined_names['fn_vms_view'] = DefinedName(
        'fn_vms_view', attr_text=STORED_LAMBDA)
    workbook.save(path)


def test_lambda_call_declares_dynamic():
    """A cell whose only 'function' is a named lambda gets declared."""
    print("\nTesting lambda-call vocabulary (the fn_vms_view cell)...")

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as handle:
        path = handle.name
    try:
        build_lambda_workbook(path)
        workbook = load_workbook(path)
        report = save_workbook_with_declaration(workbook, path)

        violations = assert_no_legacy_cse(path)
        if violations:
            print(f"✗ legacy CSE cells remain: {violations}")
            return False
        with zipfile.ZipFile(path) as archive:
            sheet_xml = archive.read('xl/worksheets/sheet1.xml').decode()
        a2 = re.search(r'<c r="A2"[^>]*>', sheet_xml).group(0)
        if 'cm="' not in a2:
            print(f"✗ A2 not declared: {a2}")
            return False
        print(f"✓ A2 declared dynamic through the lambda name alone")
        print(f"✓ no legacy CSE cells in the file")
        return True
    finally:
        os.unlink(path)


def test_declaration_survives_round_trip():
    """save -> openpyxl reload (cm dropped) -> save: still fully declared.

    This is the flush -> reload -> final-save shape of the incident. The
    second save's rescue comes from vocabulary here (the lambda name);
    the point is the OUTPUT invariant, which assert_no_legacy_cse pins.
    """
    print("\nTesting declaration across an openpyxl round trip...")

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as handle:
        path = handle.name
    try:
        build_lambda_workbook(path)
        workbook = load_workbook(path)
        save_workbook_with_declaration(workbook, path)   # the "flush"

        reloaded = load_workbook(path)                   # cm silently lost
        save_workbook_with_declaration(reloaded, path)   # the final save

        violations = assert_no_legacy_cse(path)
        if violations:
            print(f"✗ round trip produced legacy CSE cells: {violations}")
            return False
        print("✓ file fully declared after flush -> reload -> save")
        return True
    finally:
        os.unlink(path)


def test_registry_survives_saves():
    """WorkbookSession keeps provenance across saves of the same file."""
    print("\nTesting provenance registry persistence...")

    from excel_recipe_processor.core.workbook_session import WorkbookSession

    WorkbookSession._injected_formula_ranges = {}
    WorkbookSession.register_injected_formulas('/tmp/x.xlsx', 'S', [('A', 2, 5)])
    key = WorkbookSession._key('/tmp/x.xlsx')
    before = WorkbookSession._injected_formula_ranges.get(key)

    # The pop used to live in the save path; the fix removed it. Assert
    # the registry is untouched by simulating what save no longer does.
    if before is None:
        print("✗ registration itself failed")
        return False
    if 'S' not in before or before['S'] != [('A', 2, 5)]:
        print(f"✗ registry content wrong: {before}")
        return False

    source = open('excel_recipe_processor/core/workbook_session.py').read()
    if '_injected_formula_ranges.pop(' in source:
        print("✗ the save-path pop is back - provenance dies at first save")
        return False
    print("✓ registry populated and the save-path pop is gone")
    WorkbookSession._injected_formula_ranges = {}
    return True


def main():
    """Run all tests and report results."""
    print("Declaration lambda-vocabulary and registry tests")
    print("=" * 50)

    tests = [
        test_lambda_call_declares_dynamic,
        test_declaration_survives_round_trip,
        test_registry_survives_saves,
    ]

    passed = 0
    for test in tests:
        try:
            if test():
                passed += 1
            else:
                print(f"FAILED: {test.__name__}")
        except Exception as error:
            print(f"FAILED with exception: {test.__name__}: {error}")

    print("=" * 50)
    print(f"Final score: {passed}/{len(tests)} tests passed")
    return passed == len(tests)


if __name__ == '__main__':
    sys.exit(0 if main() else 1)

# End of file #

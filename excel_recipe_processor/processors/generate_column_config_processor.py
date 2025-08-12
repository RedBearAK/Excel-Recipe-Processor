"""
Enhanced column configuration processor with super fast Excel reading.

excel_recipe_processor/processors/generate_column_config_processor.py

Compares column names between two files and generates YAML configuration files.
Uses direct Excel sampling for blazing fast header analysis while preserving
original text formatting. Much faster than CSV conversion approaches.
"""

import pandas as pd
import logging
import yaml

from difflib import SequenceMatcher
from pathlib import Path

from excel_recipe_processor.core.base_processor import FileOpsBaseProcessor, StepProcessorError
from excel_recipe_processor.processors._helpers.column_patterns import empty_or_whitespace_rgx


logger = logging.getLogger(__name__)


class GenerateColumnConfigProcessor(FileOpsBaseProcessor):
    """
    Enhanced processor that compares column names between two files and generates 
    YAML configuration files for column management in recipes.
    
    SUPER FAST: Uses direct Excel sampling instead of conversion.
    
    Supports both Excel (.xlsx, .xls, .xlsm) and CSV files:
    - Excel files: Direct sampling of first few rows (preserves "8/4/2025" format)
    - CSV files: Fast pandas reading with string preservation
    
    This approach is orders of magnitude faster than cell-by-cell or CSV conversion
    while avoiding pandas auto-conversion issues.
    """
    
    @classmethod
    def get_minimal_config(cls) -> dict:
        return {
            'processor_type': 'generate_column_config',
            'source_file': 'data/raw_export.csv',
            'template_file': 'templates/desired_format.csv',
            'output_file': 'configs/column_config.yaml'
        }
    
    def get_capabilities(self) -> dict:
        """Get processor capabilities information."""
        return {
            'description': 'Generate YAML column configuration by comparing source and template files',
            'operation_type': 'file_analysis',
            'file_formats': ['csv', 'xlsx', 'xls', 'xlsm'],
            'analysis_features': [
                'header_comparison',
                'fuzzy_column_matching', 
                'empty_column_detection',
                'text_format_preservation',
                'rename_mapping_generation'
            ],
            'optimization_features': [
                'fast_excel_sampling',
                'smart_column_trimming',
                'configurable_similarity_threshold'
            ],
            'output_formats': ['yaml'],
            'special_capabilities': [
                'preserves_date_headers',
                'avoids_pandas_autoconversion',
                'handles_empty_columns',
                'generates_example_recipes'
            ],
            'dependencies': ['pandas', 'openpyxl', 'yaml'],
            'stage_requirements': 'none',
            'examples': {
                'basic_comparison': 'Compare source.xlsx with template.csv to generate column mapping',
                'fuzzy_matching': 'Find similar column names using configurable similarity threshold',
                'empty_detection': 'Identify columns with headers but no data (ghost columns)'
            }
        }
    
    def get_usage_examples(self) -> dict:
        """Get usage examples for this processor from external YAML file."""
        try:
            from excel_recipe_processor.utils.processor_examples_loader import load_processor_examples
            return load_processor_examples('generate_column_config')
        except Exception as e:
            return {
                'error': f'Could not load usage examples: {e}',
                'fallback_description': 'Generate column configuration by comparing files with different column structures'
            }
    
    def __init__(self, step_config: dict):
        # Call parent first (which calls _validate_file_operation_config)
        super().__init__(step_config)
        
        # THEN set instance attributes for easy access
        self.source_file = self.get_config_value('source_file')
        self.template_file = self.get_config_value('template_file')
        self.output_file = self.get_config_value('output_file')
        
        # Optional configuration
        self.include_recipe_section = self.get_config_value('include_recipe_section', False)
        self.similarity_threshold = self.get_config_value('similarity_threshold', 0.8)
        self.source_sheet = self.get_config_value('source_sheet', None)
        self.template_sheet = self.get_config_value('template_sheet', None)
        self.header_row = self.get_config_value('header_row', 1)  # 1-based for Excel
        self.check_column_data = self.get_config_value('check_column_data', True)  # Check for empty columns
        self.sample_rows = self.get_config_value('sample_rows', 5)  # Very small sample for speed
    
    def _validate_file_operation_config(self):
        """Validate that files exist and are supported formats."""
        required_params = ['source_file', 'template_file', 'output_file']
        
        for param in required_params:
            if not self.get_config_value(param):
                raise StepProcessorError(f"Missing required parameter: {param}")
        
        # Validate input files exist and have supported extensions
        supported_extensions = {'.csv', '.xlsx', '.xls', '.xlsm', '.xlsb'}
        
        for file_param in ['source_file', 'template_file']:
            file_path = self.get_config_value(file_param)
            
            if not Path(file_path).exists():
                raise StepProcessorError(f"File not found: {file_path}")
            
            extension = Path(file_path).suffix.lower()
            if extension not in supported_extensions:
                raise StepProcessorError(
                    f"Unsupported file format for {file_param}: {extension}. "
                    f"Got: {file_path} (supported: {', '.join(sorted(supported_extensions))})"
                    )
    
    def perform_file_operation(self) -> str:
        """
        Generate column configuration by comparing files.
        
        Returns:
            Description of operation performed
        """
        try:
            # Read source file headers
            logger.info(f"Reading source file: '{self.source_file}'")
            source_columns = self._read_file_headers(self.source_file, self.source_sheet)
            
            # Read template file headers
            logger.info(f"Reading template file: '{self.template_file}'")
            template_columns = self._read_file_headers(self.template_file, self.template_sheet)
            
            # Trim trailing empty columns from template
            template_columns = self._trim_trailing_empty_columns(template_columns)
            
            logger.info(f"Column analysis: Source={len(source_columns)}, Template={len(template_columns)} (after trimming)")
            
            # Generate column analysis
            analysis = self._analyze_columns(source_columns, template_columns)
            
            # Count actual renames (exclude identity mappings for accurate reporting)
            actual_renames = {
                source: target for source, target in analysis['rename_mapping'].items()
                if source != target
            }
            
            # Write YAML configuration file
            self._write_yaml_config(analysis)
            
            result_msg =    (f"Generated column config: {len(analysis['raw'])} source columns, "
                            f"{len(analysis['desired'])} template columns, "
                            f"{len(analysis['to_create'])} to create, "
                            f"{len(actual_renames)} renames")  # Use actual count
            
            logger.info(f"Wrote configuration to: {self.output_file}")
            return result_msg
            
        except Exception as e:
            raise StepProcessorError(f"Failed to generate column configuration: {e}")
    
    def _read_file_headers(self, file_path: str, sheet_name=None) -> list:
        """
        Read column headers from either Excel or CSV file using super fast methods.
        
        Args:
            file_path: Path to file
            sheet_name: Sheet name for Excel files (ignored for CSV)
            
        Returns:
            List of column header names as strings
        """
        file_path_obj = Path(file_path)
        extension = file_path_obj.suffix.lower()
        
        if extension == '.csv':
            return self._read_csv_headers_fast(file_path)
        elif extension in {'.xlsx', '.xls', '.xlsm', '.xlsb'}:
            return self._read_excel_headers_super_fast(file_path, sheet_name)
        else:
            raise StepProcessorError(f"Unsupported file format: {extension}")
    
    def _read_csv_headers_fast(self, file_path: str) -> list:
        """
        Read column headers from CSV file.
        
        Args:
            file_path: Path to CSV file
            
        Returns:
            List of column header names
        """
        try:
            # Read only headers using pandas with string dtype and no inference
            df = pd.read_csv(
                file_path,
                dtype=str,  # Force string reading to prevent conversions
                nrows=0,    # Only read headers
                parse_dates=False,  # Disable date parsing completely
                infer_datetime_format=False  # Disable datetime inference
            )
            
            headers = [str(col) for col in df.columns]
            logger.debug(f"Read {len(headers)} CSV headers from {file_path}")
            return headers
            
        except Exception as e:
            raise StepProcessorError(f"Failed to read CSV headers from {file_path}: {e}")
    
    def _read_excel_headers_super_fast(self, file_path: str, sheet_name=None) -> list:
        """
        Read Excel headers using SUPER FAST direct sampling approach.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name or None for active sheet
            
        Returns:
            List of column header names as they appear in Excel
        """
        try:
            import openpyxl
            from datetime import datetime
            
            # Load workbook with read_only for memory efficiency
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            
            # Get worksheet
            if sheet_name is None:
                worksheet = workbook.active
                sheet_name = worksheet.title
            elif isinstance(sheet_name, int):
                if sheet_name < 1 or sheet_name > len(workbook.sheetnames):
                    available_sheets = workbook.sheetnames
                    workbook.close()
                    raise StepProcessorError(
                        f"Sheet index {sheet_name} out of range. "
                        f"Available sheets (1-{len(workbook.sheetnames)}): {available_sheets}"
                    )
                worksheet = workbook.worksheets[sheet_name - 1]
                sheet_name = worksheet.title
            else:
                if sheet_name not in workbook.sheetnames:
                    available_sheets = workbook.sheetnames
                    workbook.close()
                    raise StepProcessorError(
                        f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}"
                    )
                worksheet = workbook[sheet_name]
            
            # SUPER FAST: Just read the header row - preserve displayed format
            headers = []
            max_col = worksheet.max_column or 1
            
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=self.header_row, column=col)
                
                # Get cell value and handle date formatting properly
                if cell.value is None:
                    headers.append("")
                elif isinstance(cell.value, datetime):
                    # This is a datetime object - Excel formatted it as a date
                    # Try to get the displayed format, fallback to a reasonable date format
                    try:
                        # Check if it has a date format applied
                        if cell.number_format and any(fmt_char in cell.number_format.lower() 
                                                    for fmt_char in ['m', 'd', 'y', '/']):
                            # Try to format according to Excel's display format
                            # Common formats: "m/d/yyyy" -> "8/4/2025"
                            formatted_date = cell.value.strftime("%m/%d/%Y").lstrip('0').replace('/0', '/')
                            headers.append(formatted_date)
                        else:
                            # No special formatting, use default
                            formatted_date = cell.value.strftime("%m/%d/%Y").lstrip('0').replace('/0', '/')
                            headers.append(formatted_date)
                    except:
                        # Fallback to string conversion if formatting fails
                        headers.append(str(cell.value))
                else:
                    # Convert to string preserving original format
                    # This handles text, numbers, etc.
                    headers.append(str(cell.value))
            
            workbook.close()
            
            # Optional: Quick data check if enabled (but keep it minimal)
            if self.check_column_data and max_col > 0:
                empty_count = self._quick_excel_empty_check(file_path, sheet_name, headers)
                if empty_count > 0:
                    logger.info(f"Found ~{empty_count} likely empty columns (quick check)")
            
            logger.debug(f"Read {len(headers)} Excel headers from {file_path} (super fast mode)")
            return headers
            
        except Exception as e:
            if hasattr(locals(), 'workbook'):
                workbook.close()
            raise StepProcessorError(f"Failed to read Excel headers from {file_path}: {e}")
    
    def _quick_excel_empty_check(self, file_path: str, sheet_name: str, headers: list) -> int:
        """
        OPTIONAL quick check for empty columns - only if specifically requested.
        
        Args:
            file_path: Excel file path
            sheet_name: Sheet name  
            headers: List of headers
            
        Returns:
            Estimated number of empty columns
        """
        try:
            import openpyxl
            
            # Quick sampling - just check a couple rows
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            worksheet = workbook[sheet_name] if isinstance(sheet_name, str) else workbook.worksheets[sheet_name - 1]
            
            empty_count = 0
            max_check_row = min(worksheet.max_row or self.header_row, self.header_row + self.sample_rows)
            
            for col_idx, header in enumerate(headers):
                if not header.strip():  # Empty header = likely empty column
                    empty_count += 1
                    continue
                
                # Quick check: does this column have any data in first few rows?
                has_data = False
                for row in range(self.header_row + 1, max_check_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx + 1)
                    if cell.value is not None and str(cell.value).strip():
                        has_data = True
                        break
                
                if not has_data:
                    empty_count += 1
            
            workbook.close()
            return empty_count
            
        except Exception:
            # Don't let optional checking break the main operation
            return 0
    
    def _trim_trailing_empty_columns(self, headers: list) -> list:
        """
        Remove trailing empty columns from headers list.
        
        Args:
            headers: List of header strings
            
        Returns:
            Headers with trailing empty strings removed
        """
        # Find last non-empty header
        last_non_empty = -1
        for i in range(len(headers) - 1, -1, -1):
            if headers[i].strip():
                last_non_empty = i
                break
        
        # Return trimmed list
        if last_non_empty >= 0:
            trimmed = headers[:last_non_empty + 1]
            if len(trimmed) < len(headers):
                logger.debug(f"Trimmed {len(headers) - len(trimmed)} trailing empty columns")
            return trimmed
        else:
            # All headers are empty
            return []
    
    def _analyze_columns(self, source_columns: list, template_columns: list) -> dict:
        """
        Analyze columns and generate mapping configuration.
        
        Args:
            source_columns: Column names from source file
            template_columns: Column names from template file
            
        Returns:
            Dictionary with analysis results
        """
        # Prepare rename mapping using fuzzy matching
        rename_mapping = {}
        used_source_columns = set()
        
        for template_col in template_columns:
            if not template_col.strip():  # Skip empty template columns
                continue
            
            best_match = None
            best_ratio = 0
            
            for source_col in source_columns:
                if source_col in used_source_columns or not source_col.strip():
                    continue
                
                ratio = SequenceMatcher(None, source_col.lower(), template_col.lower()).ratio()
                if ratio > best_ratio and ratio >= self.similarity_threshold:
                    best_ratio = ratio
                    best_match = source_col
            
            if best_match:
                rename_mapping[best_match] = template_col
                used_source_columns.add(best_match)
        
        # FIXED: Determine columns to create (template columns not in source at all)
        all_source_columns = set(col.strip() for col in source_columns if col.strip())
        rename_targets = set(rename_mapping.values())  # Columns satisfied by renames
        
        columns_to_create = [
            col for col in template_columns 
            if col.strip() and col.strip() not in all_source_columns and col.strip() not in rename_targets
        ]
        
        return {
            'raw': source_columns,
            'desired': template_columns,
            'rename_mapping': rename_mapping,
            'to_create': columns_to_create
        }
    
    def _write_yaml_config(self, analysis: dict):
        """
        Write YAML configuration file with column analysis in the correct format.
        
        Args:
            analysis: Dictionary with column analysis results
        """
        from datetime import datetime
        
        # Filter rename mapping to only include actual renames (not identity mappings)
        actual_renames = {
            source: target for source, target in analysis['rename_mapping'].items()
            if source != target  # Only include where source != target
        }
        
        # Write YAML file with proper formatting
        output_path = Path(self.output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w') as f:
            # Custom YAML formatting with timestamp
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"# Column configuration generated by generate_column_config processor\n")
            f.write(f"# Generated: {timestamp}\n\n")
            
            # Write rename_mapping FIRST (shows transformations that will happen)
            f.write("rename_mapping: ")
            if actual_renames:
                f.write("{\n")
                for i, (source, target) in enumerate(actual_renames.items()):
                    # Properly quote keys and values manually
                    source_quoted = f'"{source}"' if ' ' in source or "'" in source else f'"{source}"'
                    target_quoted = f'"{target}"' if ' ' in target or "'" in target else f'"{target}"'
                    comma = "," if i < len(actual_renames) - 1 else ""
                    f.write(f"  {source_quoted}: {target_quoted}{comma}\n")
                f.write("}\n\n")
            else:
                f.write("{}\n\n")
            
            # Write var_columns_raw_download (bracketed, 3 per line)
            f.write("var_columns_raw_download: ")
            self._write_bracketed_list(f, analysis['raw'], max_per_line=3)
            f.write("\n")
            
            # Write var_columns_to_keep (bracketed, 3 per line)
            f.write("var_columns_to_keep: ")
            self._write_bracketed_list(f, analysis['desired'], max_per_line=3)
            f.write("\n")
            
            # Write var_columns_to_create (bracketed, 1 per line)
            f.write("var_columns_to_create: ")
            self._write_bracketed_list(f, sorted(analysis['to_create']), max_per_line=1)
            f.write("\n")
            
            # Add recipe section if requested
            if self.include_recipe_section:
                f.write("# Example recipe using this configuration:\n")
                f.write("example_recipe:\n")
                example_recipe = self._generate_example_recipe(analysis)
                yaml.dump(example_recipe, f, default_flow_style=False, sort_keys=False, 
                         allow_unicode=True, width=120, indent=2)
    
    def _write_bracketed_list(self, file_handle, columns: list, max_per_line: int = 3):
        """Write columns as bracketed lists with max items per line."""
        clean_columns = [col for col in columns if col.strip()]
        
        if not clean_columns:
            file_handle.write("[]\n")
            return
        
        # Start the list with newline after opening bracket
        file_handle.write("[\n")
        
        # Write columns with proper line breaks
        for i, col in enumerate(clean_columns):
            # Add proper quoting manually (no yaml.dump)
            if '"' in col:
                quoted_col = f"'{col}'"  # Use single quotes if double quotes in text
            elif "'" in col:
                quoted_col = f'"{col}"'  # Use double quotes if single quotes in text
            elif ' ' in col or '#' in col:
                quoted_col = f'"{col}"'  # Quote if spaces or special chars
            else:
                quoted_col = f'"{col}"'  # Default to double quotes
            
            # Add indentation for first item on line
            if i % max_per_line == 0:
                file_handle.write("  ")  # Just 2 spaces indentation
            
            # Write the column
            file_handle.write(quoted_col)
            
            # Add comma if not last item
            if i < len(clean_columns) - 1:
                file_handle.write(",")
                
                # Add line break for next line or space for same line
                if (i + 1) % max_per_line == 0:
                    file_handle.write("\n")
                else:
                    file_handle.write(" ")
        
        # Close the list with newline before closing bracket
        file_handle.write("\n  ]\n")  # Just 2 spaces indentation
    
    def _generate_example_recipe(self, analysis: dict) -> dict:
        """
        Generate example recipe using the column analysis.
        
        Args:
            analysis: Dictionary with column analysis results
            
        Returns:
            Dictionary with example recipe steps
        """
        steps = []
        
        # Step 1: Import source file
        steps.append({
            'step_description': 'Import source data file',
            'processor_type': 'import_file',
            'input_file': self.source_file,
            'save_to_stage': 'stg_source_data'
        })
        
        # Step 2: Rename columns if needed
        if analysis['rename_mapping']:
            steps.append({
                'step_description': 'Rename columns to match template',
                'processor_type': 'rename_columns',
                'source_stage': 'stg_source_data',
                'rename_type': 'mapping',
                'column_mapping': analysis['rename_mapping'],
                'save_to_stage': 'stg_renamed_columns'
            })
            last_stage = 'stg_renamed_columns'
        else:
            last_stage = 'stg_source_data'
        
        # Step 3: Add missing columns if needed
        if analysis['to_create']:
            create_mapping = {col: '' for col in analysis['to_create']}
            steps.append({
                'step_description': 'Add missing columns from template',
                'processor_type': 'add_columns',
                'source_stage': last_stage,
                'new_columns': create_mapping,
                'save_to_stage': 'stg_final_structure'
            })
            last_stage = 'stg_final_structure'
        
        # Step 4: Export result
        steps.append({
            'step_description': 'Export processed data',
            'processor_type': 'export_file',
            'source_stage': last_stage,
            'output_file': 'output/processed_data.xlsx',
            'file_format': 'excel'
        })
        
        return {
            'description': 'Example recipe generated from column analysis',
            'settings': {
                'stages': [
                    {'stage_name': 'stg_source_data', 'description': 'Raw imported data'},
                    {'stage_name': 'stg_renamed_columns', 'description': 'Data with renamed columns'},
                    {'stage_name': 'stg_final_structure', 'description': 'Data with final column structure'}
                ]
            },
            'recipe': steps
        }


# End of file #

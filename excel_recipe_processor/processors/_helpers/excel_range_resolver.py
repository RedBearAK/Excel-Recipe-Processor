"""
Shared Excel range addressing for processors that target columns by name.

excel_recipe_processor/processors/_helpers/excel_range_resolver.py

Resolves column names on a worksheet into Excel range addresses. Four separate
features need this: number formatting, column hiding, per-column alignment, and
defined-name creation. Keeping the resolution in one place means they agree on
what "the PID column" means.

Row modes:
    data                - data rows only, header excluded      ($C$2:$C$1915)
    data_with_header    - header row through last data row     ($C$1:$C$1915)
    full_col            - entire column, header included       ($C:$C)
    full_col_no_header  - entire column below the header       ($C$2:$C$1048576)

The two data modes require scanning for the last populated row. The two
full_col modes are purely positional and cannot go stale, which makes them the
safer choice for validation lists that humans extend by hand.
"""

import logging

from openpyxl.utils import get_column_letter, column_index_from_string

from excel_recipe_processor.processors._helpers.range_patterns import (
    excel_column_ref_rgx,
    sheet_name_needs_quotes_rgx,
)


logger = logging.getLogger(__name__)


EXCEL_MAX_ROW = 1048576
EXCEL_MAX_COL = 16384

ROW_MODES = ('data', 'data_with_header', 'full_col', 'full_col_no_header')
MISSING_POLICIES = ('error', 'warn', 'skip')

# Bound the upward scan in find_last_data_row so a worksheet with an inflated
# dimension record cannot turn extent detection into a million-cell walk.
DEFAULT_MAX_SCAN_ROWS = 100000


class ExcelRangeResolverError(Exception):
    """Raised when a column specification cannot be resolved to a range."""
    pass


# ============================================================================
# Column resolution
# ============================================================================

def find_column_letter_by_name(worksheet, column_name: str, header_row: int = 1) -> str:
    """
    Find the Excel column letter whose header matches column_name exactly.

    Args:
        worksheet:      openpyxl worksheet object
        column_name:    Header text to search for, compared after stripping
        header_row:     Row containing headers, 1-based

    Returns:
        Column letter, or an empty string when no header matches

    Raises:
        ExcelRangeResolverError: If the same header appears more than once
    """
    if not isinstance(column_name, str):
        raise ExcelRangeResolverError(
            f"Column name must be a string, got {type(column_name).__name__}"
        )

    target = column_name.strip()
    matches = []

    max_col = worksheet.max_column
    if not isinstance(max_col, int) or max_col < 1:
        max_col = 1

    for col_num in range(1, max_col + 1):
        cell_value = worksheet.cell(row=header_row, column=col_num).value

        if cell_value is None:
            continue

        if str(cell_value).strip() == target:
            matches.append(col_num)

    if len(matches) > 1:
        letters = ', '.join(get_column_letter(num) for num in matches)
        raise ExcelRangeResolverError(
            f"Header '{target}' appears in {len(matches)} columns ({letters}) "
            f"on sheet '{worksheet.title}' row {header_row}. "
            f"Duplicate headers are ambiguous and must be corrected at the source."
        )

    if len(matches) == 0:
        return ''

    return get_column_letter(matches[0])


def resolve_column_letter(worksheet, column_spec, header_row: int = 1,
                          force_column_names: bool = False) -> str:
    """
    Resolve one column specification to an Excel column letter.

    A spec that looks like a bare Excel column reference ("A", "AB") is taken
    as one, unless force_column_names is set. Everything else is looked up as
    a header name.

    Args:
        worksheet:              openpyxl worksheet object
        column_spec:            Column letter or header name
        header_row:             Row containing headers, 1-based
        force_column_names:     Treat every spec as a header name

    Returns:
        Column letter, or an empty string when a name is not found

    Raises:
        ExcelRangeResolverError: If the spec is empty or the header is duplicated
    """
    spec = str(column_spec).strip()

    if not spec:
        raise ExcelRangeResolverError("Column specification cannot be empty")

    if not force_column_names and excel_column_ref_rgx.match(spec):
        logger.debug(f"Column '{spec}' treated as an Excel reference")
        return spec

    letter = find_column_letter_by_name(worksheet, spec, header_row)

    if letter:
        logger.debug(f"Column name '{spec}' resolved to {letter}")

    return letter


def resolve_column_letters(worksheet, column_specs: list, header_row: int = 1,
                           force_column_names: bool = False,
                           on_missing: str = 'error') -> list:
    """
    Resolve a list of column specifications to Excel column letters.

    Args:
        worksheet:              openpyxl worksheet object
        column_specs:           List of column letters or header names
        header_row:             Row containing headers, 1-based
        force_column_names:     Treat every spec as a header name
        on_missing:             'error', 'warn', or 'skip'

    Returns:
        List of column letters, in the order given

    Raises:
        ExcelRangeResolverError: On bad input, or on a missing column when
                                 on_missing is 'error'
    """
    if not isinstance(column_specs, list):
        raise ExcelRangeResolverError(
            f"column_specs must be a list, got {type(column_specs).__name__}"
        )

    if len(column_specs) == 0:
        raise ExcelRangeResolverError("column_specs cannot be empty")

    if on_missing not in MISSING_POLICIES:
        valid = ', '.join(MISSING_POLICIES)
        raise ExcelRangeResolverError(
            f"Unknown on_missing policy '{on_missing}'. Valid options: {valid}"
        )

    letters = []

    for spec in column_specs:
        letter = resolve_column_letter(worksheet, spec, header_row, force_column_names)

        if letter:
            letters.append(letter)
            continue

        message = (
            f"Column '{spec}' not found on sheet '{worksheet.title}' "
            f"in header row {header_row}"
        )

        if on_missing == 'error':
            raise ExcelRangeResolverError(message)

        if on_missing == 'warn':
            logger.warning(f"{message} - skipping")
        else:
            logger.debug(f"{message} - skipping")

    if len(letters) == 0:
        raise ExcelRangeResolverError(
            f"No columns could be resolved on sheet '{worksheet.title}'"
        )

    return letters


def expand_column_span(column_letters: list) -> list:
    """
    Expand each adjacent pair in the list into the full span between them.

    ["B", "E"]           -> ["B", "C", "D", "E"]
    ["B", "E", "G"]      -> ["B", "C", "D", "E", "F", "G"]
    ["E", "B"]           -> ["B", "C", "D", "E"]

    Because consecutive pairs chain, the result is always contiguous.

    Args:
        column_letters: List of Excel column letters

    Returns:
        Sorted list of column letters covering the whole span
    """
    if not isinstance(column_letters, list) or len(column_letters) == 0:
        raise ExcelRangeResolverError("column_letters must be a non-empty list")

    indices = [column_index_from_string(letter) for letter in column_letters]

    if len(indices) == 1:
        return [get_column_letter(indices[0])]

    covered = set()

    for position in range(len(indices) - 1):
        first = indices[position]
        second = indices[position + 1]
        low = min(first, second)
        high = max(first, second)
        covered.update(range(low, high + 1))

    return [get_column_letter(num) for num in sorted(covered)]


def assert_contiguous(column_letters: list) -> None:
    """
    Raise unless the given column letters form an unbroken run.

    Defined names that feed XLOOKUP or VLOOKUP must be single-area references,
    so a gap is a hard error rather than something to work around.

    Args:
        column_letters: List of Excel column letters

    Raises:
        ExcelRangeResolverError: If the columns are not contiguous
    """
    indices = sorted(column_index_from_string(letter) for letter in column_letters)

    expected_span = indices[-1] - indices[0] + 1

    if expected_span != len(indices):
        present = ', '.join(get_column_letter(num) for num in indices)
        raise ExcelRangeResolverError(
            f"Columns {present} are not contiguous. A single-area reference is "
            f"required. Either include the intervening columns or allow span "
            f"expansion."
        )


# ============================================================================
# Row extent
# ============================================================================

def find_last_data_row(worksheet, anchor_columns: list, header_row: int = 1,
                       max_scan_rows: int = DEFAULT_MAX_SCAN_ROWS) -> int:
    """
    Find the last row holding data in any of the anchor columns.

    Scans upward from the worksheet's reported dimension. That dimension counts
    rows carrying formatting or stale cell records, not just rows with values,
    so it routinely overstates the real extent and cannot be used directly.

    Args:
        worksheet:      openpyxl worksheet object
        anchor_columns: Column letters to scan; the largest extent wins
        header_row:     Row containing headers, 1-based
        max_scan_rows:  Upper bound on rows examined before giving up

    Returns:
        1-based row number of the last populated row, or header_row when the
        sheet holds no data below the header
    """
    if not isinstance(anchor_columns, list) or len(anchor_columns) == 0:
        raise ExcelRangeResolverError("anchor_columns must be a non-empty list")

    indices = [column_index_from_string(letter) for letter in anchor_columns]

    reported_max = worksheet.max_row
    if not isinstance(reported_max, int) or reported_max < header_row:
        reported_max = header_row

    scan_floor = max(header_row, reported_max - max_scan_rows + 1)

    for row_num in range(reported_max, scan_floor - 1, -1):
        for col_index in indices:
            cell_value = worksheet.cell(row=row_num, column=col_index).value

            if cell_value is None:
                continue

            if isinstance(cell_value, str) and not cell_value.strip():
                continue

            if row_num < reported_max:
                logger.debug(
                    f"Sheet '{worksheet.title}' reports max_row={reported_max} "
                    f"but last populated row is {row_num}"
                )

            return row_num

    if scan_floor > header_row:
        logger.warning(
            f"Scanned {max_scan_rows} rows upward on sheet '{worksheet.title}' "
            f"without finding data. Extent may be wrong."
        )

    return header_row


# ============================================================================
# Range construction
# ============================================================================

def quote_sheet_name(sheet_name: str) -> str:
    """
    Wrap a sheet name in single quotes when Excel requires it.

    Embedded apostrophes are doubled, per Excel's escaping rule.

    Args:
        sheet_name: Worksheet name

    Returns:
        Sheet name, quoted if necessary
    """
    name = str(sheet_name)

    if not sheet_name_needs_quotes_rgx.search(name):
        return name

    escaped = name.replace("'", "''")
    return f"'{escaped}'"


def build_range_ref(start_col: str, end_col: str, start_row=None, end_row=None,
                    absolute: bool = True, sheet_name=None) -> str:
    """
    Assemble an Excel range address from resolved parts.

    Omitting both row arguments produces a whole-column reference.

    Args:
        start_col:  First column letter
        end_col:    Last column letter
        start_row:  First row number, or None for a whole-column reference
        end_row:    Last row number, or None for a whole-column reference
        absolute:   Emit dollar signs
        sheet_name: Sheet to qualify the reference with, or None

    Returns:
        Range address such as "$C$2:$C$1915" or "'Region-Carrier'!$D:$D"
    """
    marker = '$' if absolute else ''

    if start_row is None or end_row is None:
        reference = f"{marker}{start_col}:{marker}{end_col}"
    else:
        reference = (
            f"{marker}{start_col}{marker}{start_row}"
            f":{marker}{end_col}{marker}{end_row}"
        )

    if sheet_name is None:
        return reference

    return f"{quote_sheet_name(sheet_name)}!{reference}"


def resolve_range(worksheet, columns: list, row_mode: str = 'data_with_header',
                  header_row: int = 1, anchor_columns=None,
                  expand_span: bool = True, force_column_names: bool = False,
                  on_missing: str = 'error', absolute: bool = True,
                  qualify_sheet: bool = True) -> str:
    """
    Resolve a column specification into a complete Excel range address.

    Args:
        worksheet:           openpyxl worksheet object
        columns:             Column letters or header names
        row_mode:            One of ROW_MODES
        header_row:          Row containing headers, 1-based
        anchor_columns:      Columns to measure extent from; defaults to the
                             resolved columns themselves, longest wins
        expand_span:         Fill in columns between adjacent list entries
        force_column_names:  Treat every spec as a header name
        on_missing:          'error', 'warn', or 'skip'
        absolute:            Emit dollar signs
        qualify_sheet:       Prefix the reference with the sheet name

    Returns:
        Range address string

    Raises:
        ExcelRangeResolverError: On any resolution failure
    """
    if row_mode not in ROW_MODES:
        valid = ', '.join(ROW_MODES)
        raise ExcelRangeResolverError(
            f"Unknown row_mode '{row_mode}'. Valid options: {valid}"
        )

    letters = resolve_column_letters(
        worksheet, columns, header_row, force_column_names, on_missing
    )

    if expand_span:
        letters = expand_column_span(letters)

    assert_contiguous(letters)

    indices = sorted(column_index_from_string(letter) for letter in letters)
    start_col = get_column_letter(indices[0])
    end_col = get_column_letter(indices[-1])

    sheet_name = worksheet.title if qualify_sheet else None

    if row_mode == 'full_col':
        return build_range_ref(start_col, end_col, None, None, absolute, sheet_name)

    if row_mode == 'full_col_no_header':
        return build_range_ref(
            start_col, end_col, header_row + 1, EXCEL_MAX_ROW, absolute, sheet_name
        )

    if anchor_columns is None:
        anchor_letters = letters
    else:
        anchor_letters = resolve_column_letters(
            worksheet, anchor_columns, header_row, force_column_names, on_missing
        )

    last_row = find_last_data_row(worksheet, anchor_letters, header_row)

    if last_row <= header_row:
        raise ExcelRangeResolverError(
            f"No data rows found below header row {header_row} on sheet "
            f"'{worksheet.title}' using anchor columns {anchor_letters}"
        )

    start_row = header_row if row_mode == 'data_with_header' else header_row + 1

    return build_range_ref(start_col, end_col, start_row, last_row, absolute, sheet_name)


# End of file #
